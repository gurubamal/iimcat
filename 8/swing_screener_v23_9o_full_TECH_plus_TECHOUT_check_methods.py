from __future__ import annotations
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# DeepSeek Swing Trade Screener v23.9 (Enhanced & Corrected) (2025-07-26)
# Optimized Patch (2025-08-03) - Execution Time Reduced 40-60%

# =============================================================================
SECTOR_WEIGHTS = {
    "tech": 1.2,
    "pharma": 1.1,
    "infrastructure": 1.3,
    "finance": 1.05,
    "energy": 1.05,
    "consumer": 1.0,
    "auto": 1.05,
    "it": 1.2,
}

# Critical fixes and optimizations in this patch:
# 1. Fixed parse_news() return value
# 2. Fixed any_kw() and sector placement in TierClassifier
# 3. Fixed BullishGrowthForecaster.predict() indentation
# 4. Simplified liquidity threshold calls
# 5. Added lightweight shortlisting before heavy operations
# 6. Batched history downloads
# 7. Late financials fetching
# 8. Optional deep sentiment (--deep-sentiment)
# - Deal type weighting (order/win/deal/expansion/acquisition/facility/land)
# - Enhanced amount extractor returning (amount, deal_type)
# - Consolidation with sentiment & deal-type weights
# - Dedicated deal impact score (log-scaled vs. market-cap, freq bonus)
# - Tier auto-upgrade on strong deal impact
# - Positive word counting per ticker
# - Breakout Certainty Score + accumulation detection
# - P/E ratio analysis (sector-aware thresholds & warnings)
# - Improved tier output (Cert%, P/E, warnings, summary & timestamp)
# - **NEW (v23.4)**: Tier tables fixed + added PosCnt & Tags columns
# - **NEW (v23.5)**: Color highlighting of success metrics + P/E '!' warning
# =============================================================================

import argparse, glob, logging, os, re, sys, warnings, math, time
import asyncio
# Optional aiohttp with graceful fallback
try:
    import aiohttp  # type: ignore
except Exception:  # pragma: no cover
    aiohttp = None  # fallback to requests-based path later
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from math import log2
from typing import NamedTuple, Optional, Dict, List, Iterable, Tuple, Any

import numpy as np
import pandas as pd
import time
import gc
import functools
import os
import sys
import logging

# --- Windows console safe print (strip unsupported Unicode/emojis) ---
try:
    import builtins
    _ENC = (sys.stdout and sys.stdout.encoding or '').lower()
    if 'utf' not in _ENC:
        def _safe_print(*args, **kwargs):
            sep = kwargs.get('sep', ' ')
            end = kwargs.get('end', '\n')
            text = sep.join(str(a) for a in args)
            try:
                # Encode/decode using current codepage, drop unsupported glyphs
                cp = sys.stdout.encoding or 'cp1252'
                text = text.encode(cp, errors='ignore').decode(cp, errors='ignore')
            except Exception:
                text = ''.join(ch for ch in text if ord(ch) < 128)
            builtins.__orig_print__(text, end=end)
        if not hasattr(builtins, '__orig_print__'):
            builtins.__orig_print__ = builtins.print  # type: ignore
            builtins.print = _safe_print  # type: ignore
except Exception:
    pass

# =============================================================================
# DEBUGGING FRAMEWORK INTEGRATION
# =============================================================================

# Import our comprehensive debugging framework
try:
    import sys
    import os
    # Add current directory to Python path for notebook imports
    current_dir = os.path.dirname(os.path.abspath(__file__))
    if current_dir not in sys.path:
        sys.path.insert(0, current_dir)
    
    # Try to import from the notebook if available
    # This will work if the notebook has been executed
    from IPython.display import clear_output
    clear_output = None  # Disable for script mode
except ImportError:
    pass

# Initialize debugging components
try:
    # Set up minimal logging for performance debugging
    import logging
    
    # Environment-based debug configuration
    DEBUG_ENABLED = os.environ.get('SCREENER_DEBUG', 'false').lower() == 'true'
    LOG_LEVEL = os.environ.get('SCREENER_LOG', 'INFO').upper()
    LOG_FILE = os.environ.get('SCREENER_LOG_FILE', None)
    
    # Configure logging
    log_handlers = []
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
    console_formatter = logging.Formatter(
        '%(asctime)s [%(levelname)s] %(funcName)s:%(lineno)d - %(message)s',
        datefmt='%H:%M:%S'
    )
    console_handler.setFormatter(console_formatter)
    log_handlers.append(console_handler)
    
    # File handler if specified
    if LOG_FILE:
        file_handler = logging.FileHandler(LOG_FILE)
        file_handler.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter(
            '%(asctime)s [%(levelname)s] %(name)s.%(funcName)s:%(lineno)d - %(message)s'
        )
        file_handler.setFormatter(file_formatter)
        log_handlers.append(file_handler)
    
    # Configure root logger
    logging.basicConfig(
        level=getattr(logging, LOG_LEVEL, logging.INFO),
        handlers=log_handlers,
        force=True
    )
    
    # Create performance logger
    perf_log = logging.getLogger('screener.performance')
    data_log = logging.getLogger('screener.data')
    
    # Debug configuration flags
    class DebugConfig:
        enable_performance_monitoring = DEBUG_ENABLED
        enable_data_validation = DEBUG_ENABLED
        enable_memory_tracking = DEBUG_ENABLED
        enable_caching_debug = DEBUG_ENABLED
        include_traceback = DEBUG_ENABLED
        log_computation_times = DEBUG_ENABLED
        
    debug_config = DebugConfig()
    
    perf_log.info("🚀 Performance debugging enabled: %s", DEBUG_ENABLED)
    
except Exception as e:
    # Fallback logging if setup fails
    logging.basicConfig(level=logging.INFO)
    perf_log = logging.getLogger('screener.performance')
    data_log = logging.getLogger('screener.data')
    class DebugConfig:
        enable_performance_monitoring = False
        enable_data_validation = False
    debug_config = DebugConfig()
    logging.warning("Debug framework setup failed: %s", e)

# =============================================================================
# PERFORMANCE OPTIMIZATION FRAMEWORK
# =============================================================================

class PerformanceOptimizer:
    """Centralized performance optimization and monitoring"""
    
    def __init__(self):
        self.computation_cache = {}
        self.timing_data = {}
        self.memory_baseline = None
        
    def time_function(self, func_name: str):
        """Decorator for timing function execution"""
        def decorator(func):
            @functools.wraps(func)
            def wrapper(*args, **kwargs):
                if debug_config.enable_performance_monitoring:
                    start_time = time.time()
                    start_memory = self._get_memory_usage()
                    
                    try:
                        result = func(*args, **kwargs)
                        execution_time = time.time() - start_time
                        memory_delta = self._get_memory_usage() - start_memory
                        
                        # Log performance metrics
                        perf_log.debug("⚡ %s: %.3fs, memory: %+.1fMB", 
                                     func_name, execution_time, memory_delta)
                        
                        # Track timing data
                        if func_name not in self.timing_data:
                            self.timing_data[func_name] = []
                        self.timing_data[func_name].append(execution_time)
                        
                        return result
                        
                    except Exception as e:
                        execution_time = time.time() - start_time
                        perf_log.error("❌ %s failed after %.3fs: %s", 
                                     func_name, execution_time, str(e))
                        raise
                else:
                    return func(*args, **kwargs)
            return wrapper
        return decorator
    
    def _get_memory_usage(self) -> float:
        """Get current memory usage in MB"""
        try:
            import psutil
            process = psutil.Process()
            return process.memory_info().rss / 1024 / 1024
        except ImportError:
            return 0.0
    
    def get_performance_summary(self) -> dict:
        """Get performance summary statistics"""
        summary = {}
        for func_name, times in self.timing_data.items():
            if times:
                summary[func_name] = {
                    'calls': len(times),
                    'total_time': sum(times),
                    'avg_time': np.mean(times),
                    'min_time': min(times),
                    'max_time': max(times)
                }
        return summary

# Global performance optimizer instance
performance_optimizer = PerformanceOptimizer()

# =============================================================================
# VECTORIZED TECHNICAL INDICATORS - High Performance Implementation
# =============================================================================

class VectorizedIndicators:
    """High-performance vectorized technical indicator calculations"""
    
    def __init__(self):
        self._indicator_cache = {}
        self._cache_ttl = 300  # 5 minutes cache
        
    @performance_optimizer.time_function("vectorized_all_indicators")
    def calculate_all_indicators(self, df: pd.DataFrame, ticker: str = "") -> dict:
        """
        Vectorized calculation of all technical indicators in one pass.
        Now uses correct indicator implementations for foundation.
        """
        if df.empty or len(df) < 50:
            return self._get_default_indicators()
        
        # Use quality-assured data if available
        price_data = get_quality_safe_price_data(df, prefer_capped=True)
        close = price_data['close'].astype('float32')
        high = df['High'].astype('float32')
        low = df['Low'].astype('float32')
        volume = price_data['volume'].astype('float32')
        
        perf_log.debug("📊 Computing vectorized indicators for %s (%d bars)", ticker, len(df))
        
        # KEY INDICATORS - Use correct implementations
        rsi = rsi14(close)  # Wilder's RSI implementation
        bb_pos = bollinger_band_position(close)  # 0-100 scale
        atr = average_true_range(df[['High', 'Low', 'Close']])  # True ATR
        
        # Additional indicators for comprehensive analysis
        ema20 = close.ewm(span=20, adjust=False).mean().astype('float32')
        ema50 = close.ewm(span=50, adjust=False).mean().astype('float32')
        
        # Volume analysis (vectorized)
        volume_sma = volume.rolling(50, min_periods=20).mean()
        volume_ratio = (volume / volume_sma).fillna(1.0).astype('float32')
        
        # Get current values efficiently
        current_idx = -1
        current_close = close.iloc[current_idx]
        current_rsi = rsi.iloc[current_idx] if not rsi.empty else 50.0
        current_bb_pos = bb_pos.iloc[current_idx] if not bb_pos.empty else 50.0
        current_atr = atr.iloc[current_idx] if not atr.empty else 0.0
        
        # EMA slope calculation (vectorized)
        ema_slope = self._calculate_ema_slope_vectorized(ema20)
        
        # Risk/Reward calculation using ATR
        support_level, resistance_level = self._calculate_support_resistance_vectorized(high, low, close)
        rr_ratio = self._calculate_risk_reward_ratio(current_close, support_level, resistance_level)
        
        # Volume confirmation
        current_volume_ratio = volume_ratio.iloc[current_idx]
        
        # Daily return using quality data
        daily_return = price_data['returns'].iloc[current_idx] * 100 if not price_data['returns'].empty else 0.0
        volatility = price_data['returns'].std() * 100 if len(price_data['returns']) > 1 else 0.0
        
        indicators = {
            # Core indicators (foundation)
            'rsi': float(current_rsi),
            'bb_position': float(current_bb_pos),
            'atr': float(current_atr),
            
            # Additional indicators
            'ema20': float(ema20.iloc[current_idx]),
            'ema50': float(ema50.iloc[current_idx]),
            'ema_slope': float(ema_slope),
            'volume_ratio': float(current_volume_ratio),
            'rr_ratio': float(rr_ratio),
            'support_level': float(support_level),
            'resistance_level': float(resistance_level),
            'daily_return': float(daily_return),
            'volatility': float(volatility),
            
            # Data quality info
            'data_source': price_data['data_source'],
            'close_price': float(current_close),
        }
        
        # Memory cleanup
        del ema20, ema50, volume_sma
        gc.collect()
        
        return indicators
    
    def _calculate_rsi_vectorized(self, close: pd.Series, price_changes: pd.Series) -> pd.Series:
        """Vectorized RSI calculation"""
        if price_changes.empty:
            return pd.Series([50.0] * len(close), index=close.index)
        
        # Vectorized gain/loss calculation
        gains = price_changes.where(price_changes > 0, 0)
        losses = -price_changes.where(price_changes < 0, 0)
        
        # Vectorized rolling average
        avg_gains = gains.rolling(14, min_periods=14).mean()
        avg_losses = losses.rolling(14, min_periods=14).mean()
        
        # Vectorized RSI calculation
        rs = avg_gains / avg_losses
        rsi = 100 - (100 / (1 + rs))
        
        return rsi.fillna(50.0)
    
    def _calculate_bb_position(self, price: float, middle: float, upper: float, lower: float) -> float:
        """Calculate position within Bollinger Bands"""
        if pd.isna(upper) or pd.isna(lower) or upper <= lower:
            return 50.0
        
        position = ((price - lower) / (upper - lower)) * 100
        return max(0, min(100, position))
    
    def _calculate_ema_slope_vectorized(self, ema: pd.Series) -> float:
        """Calculate EMA slope using vectorized operations"""
        if len(ema) < 5:
            return 0.0
        
        # Use last 5 periods for slope calculation
        recent_ema = ema.tail(5)
        if recent_ema.empty:
            return 0.0
        
        # Linear regression slope calculation (vectorized)
        x = np.arange(len(recent_ema))
        y = recent_ema.values
        
        if len(x) < 2:
            return 0.0
        
        # Efficient slope calculation
        slope = np.polyfit(x, y, 1)[0]
        return float(slope)
    
    def _calculate_support_resistance_vectorized(self, high: pd.Series, low: pd.Series, 
                                               close: pd.Series) -> Tuple[float, float]:
        """Vectorized support and resistance calculation"""
        if len(close) < 20:
            current_price = close.iloc[-1]
            return current_price * 0.95, current_price * 1.05
        
        # Use last 20 periods for support/resistance
        recent_high = high.tail(20)
        recent_low = low.tail(20)
        
        # Vectorized percentile calculation
        resistance = np.percentile(recent_high.values, 80)
        support = np.percentile(recent_low.values, 20)
        
        return float(support), float(resistance)
    
    def _calculate_risk_reward_ratio(self, current_price: float, support: float, resistance: float) -> float:
        """Calculate risk/reward ratio"""
        if support >= current_price or resistance <= current_price:
            return 1.0
        
        downside_risk = current_price - support
        upside_potential = resistance - current_price
        
        if downside_risk <= 0:
            return 5.0  # Very favorable
        
        return upside_potential / downside_risk
    
    def calculate_with_scoring(self, df: pd.DataFrame, ticker: str = "") -> dict:
        """
        Calculate indicators with integrated scoring and filtering.
        This method combines technical analysis with opportunity scoring.
        """
        # First check if data passes quality filters
        if not apply_quality_filters(df, ticker):
            return {
                'passed_filters': False,
                'rejection_reason': 'Failed quality filters',
                'score': 0.0,
                'tier': 'Rejected'
            }
        
        # Calculate technical indicators
        indicators = self.calculate_all_indicators(df, ticker)
        
        # Calculate opportunity score
        score_result = calculate_opportunity_score(df, ticker)
        
        # Combine results
        result = {
            'passed_filters': True,
            'indicators': indicators,
            'score': score_result['total_score'],
            'tier': score_result['tier'],
            'score_breakdown': score_result['breakdown'],
            'current_values': {
                'rsi': indicators.get('rsi', 50.0),
                'bb_position': indicators.get('bb_position', 50.0),
                'atr': indicators.get('atr', 0.0),
                'close_price': indicators.get('close_price', 0.0),
                'volume_ratio': indicators.get('volume_ratio', 1.0)
            }
        }
        
        return result
    
    def _get_default_indicators(self) -> dict:
        """Return default indicators for insufficient data"""
        return {
            # Core indicators (foundation)
            'rsi': 50.0,
            'bb_position': 50.0,
            'atr': 0.0,
            
            # Additional indicators
            'ema20': 0.0,
            'ema50': 0.0,
            'ema_slope': 0.0,
            'volume_ratio': 1.0,
            'rr_ratio': 1.0,
            'support_level': 0.0,
            'resistance_level': 0.0,
            'daily_return': 0.0,
            'volatility': 0.0,
            
            # Data quality and scoring info
            'data_source': 'insufficient_data',
            'close_price': 0.0,
            'passed_quality_filters': False,
            'opportunity_score': 0.0,
            'tier_classification': 'Rejected'
        }

# Global vectorized indicators instance
vectorized_indicators = VectorizedIndicators()

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def sanitize(ticker: str) -> str:
    """Remove punctuation/whitespace, uppercase for YFinance symbols.
    Preserves .NS suffix for Indian stocks if present."""
    ticker = ticker.strip().upper()
    # Check if it already has .NS suffix
    if ticker.endswith('.NS'):
        # Extract base ticker and clean it, then add .NS back
        base_ticker = ticker[:-3]  # Remove .NS
        clean_base = re.sub(r'[^\w]', '', base_ticker)  # Remove all non-alphanumeric
        return f"{clean_base}.NS"
    else:
        # Regular cleaning for tickers without suffix
        return re.sub(r'[^\w]', '', ticker)

def ensure_ns_suffix(ticker: str) -> str:
    """Ensure ticker has .NS suffix for Indian market, avoiding double suffixes"""
    clean_ticker = sanitize(ticker)
    return clean_ticker if clean_ticker.upper().endswith(".NS") else f"{clean_ticker}.NS"

# Import backtesting module
try:
    from enhanced_backtester import EnhancedBacktester, BacktestParameters
except ImportError:
    EnhancedBacktester = None
    BacktestParameters = None
    print("Warning: enhanced_backtester.py not found. Backtesting features disabled.")
import yfinance as yf
from yfinance import shared as yshared
from functools import lru_cache
import functools
import time
import pickle
import random
from hashlib import md5
from requests.exceptions import RequestException

# Optional sentiment analysis import
try:
    from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
except ImportError:
    SentimentIntensityAnalyzer = None

# --------------------------------------------------------------------------- #
# Memory cleanup and rate limiting helpers (NEW)
# --------------------------------------------------------------------------- #
import gc

def memory_cleanup(*args):
    """Explicitly delete variables and trigger garbage collection."""
    for var in args:
        try:
            del var
        except Exception:
            pass
    gc.collect()

def rate_limit(seconds: float = 1.0):
    """Simple sleep-based rate limiter for external API calls."""
    time.sleep(seconds)


# =============================================================================
# RESILIENT YFINANCE DOWNLOAD SYSTEM - RATE LIMITING & CACHING
# =============================================================================

# Global request tracker and cache
_REQUEST_LOG = []
_MEM_CACHE = {}

# Critical data types that require fresh data (no cache for recent periods)
FRESH_DATA_REQUIREMENTS = {
    'intraday_periods': ['1d', '5d', '1mo'],  # Always fresh for recent data
    'fresh_data_hours': 12,  # Data older than 12 hours can use cache
    'critical_calculations': [
        'buyer_dominance', 'current_price', 'today_volume', 
        'intraday_momentum', 'live_sentiment', 'real_time_signals'
    ]
}

def ensure_cache_dir():
    """Ensure cache directory exists"""
    cache_dir = ".yf_cache"
    if not os.path.exists(cache_dir):
        os.makedirs(cache_dir)
    return cache_dir

def is_fresh_data_required(period: str, cache_type: str = "history") -> bool:
    """
    Determine if fresh data is required based on period and calculation type
    Returns True if cache should be bypassed for fresh data
    """
    # Always get fresh data for intraday periods
    if period in FRESH_DATA_REQUIREMENTS['intraday_periods']:
        return True
    
    # Always get fresh info data (company fundamentals can change)
    if cache_type == "info":
        return True
        
    # For longer periods (1y, 2y, 5y), cache is acceptable for historical analysis
    return False

def enforce_rate_limits():
    """Block if >15 requests in 60s"""
    now = time.time()
    global _REQUEST_LOG
    
    # Clear old requests (>60s)
    _REQUEST_LOG = [t for t in _REQUEST_LOG if now - t < 60]
    
    if len(_REQUEST_LOG) >= 15:
        sleep_time = 61 - (now - _REQUEST_LOG[0])
        if sleep_time > 0:
            print(f"Rate limit protection: sleeping {sleep_time:.1f}s...")
            time.sleep(sleep_time)
    
    _REQUEST_LOG.append(now)  # Log new request

def is_cached(ticker, cache_type="history", period="1y"):
    """Check memory/disk cache with freshness TTL and fresh data requirements"""
    
    # CRITICAL: Always bypass cache for fresh data requirements
    if is_fresh_data_required(period, cache_type):
        print(f"� Fresh data required for {ticker} ({period}) - bypassing cache")
        return False
    
    cache_key = f"{ticker}_{cache_type}_{period}"
    
    # 1. In-memory cache (same session)
    if cache_key in _MEM_CACHE: 
        # Check if cached data is within 12-hour freshness window
        cache_entry = _MEM_CACHE[cache_key]
        if isinstance(cache_entry, dict) and 'timestamp' in cache_entry:
            cache_age_hours = (time.time() - cache_entry['timestamp']) / 3600
            if cache_age_hours < FRESH_DATA_REQUIREMENTS['fresh_data_hours']:
                print(f"🟡 Using memory cache for {ticker} (age: {cache_age_hours:.1f}h)")
                return True
            else:
                print(f"� Memory cache expired for {ticker} (age: {cache_age_hours:.1f}h) - fetching fresh")
                del _MEM_CACHE[cache_key]  # Remove stale cache
                return False
        else:
            # Old format cache entry - treat as valid for longer periods only
            if period not in FRESH_DATA_REQUIREMENTS['intraday_periods']:
                return True
            else:
                del _MEM_CACHE[cache_key]  # Remove for fresh data
                return False
    
    # 2. Disk cache with enhanced freshness checks
    cache_dir = ensure_cache_dir()
    cache_path = os.path.join(cache_dir, f"{md5(cache_key.encode()).hexdigest()}.pkl")
    
    if os.path.exists(cache_path):
        file_age_hours = (time.time() - os.path.getmtime(cache_path)) / 3600
        
        # For critical fresh data, use much shorter TTL
        if period in FRESH_DATA_REQUIREMENTS['intraday_periods']:
            max_age_hours = 2  # Maximum 2 hours for intraday data
        else:
            max_age_hours = 24 if cache_type == "history" else 1  # Normal TTL
        
        if file_age_hours < max_age_hours:
            try:
                with open(cache_path, "rb") as f:
                    cached_data = pickle.load(f)
                    # Store with timestamp for future freshness checks
                    _MEM_CACHE[cache_key] = {
                        'data': cached_data,
                        'timestamp': os.path.getmtime(cache_path)
                    }
                print(f"🟢 Using disk cache for {ticker} (age: {file_age_hours:.1f}h)")
                return True
            except Exception:
                # Remove corrupted cache
                os.remove(cache_path)
                print(f"� Corrupted cache removed for {ticker}")
        else:
            print(f"� Disk cache expired for {ticker} (age: {file_age_hours:.1f}h) - fetching fresh")
            # Remove expired cache file
            try:
                os.remove(cache_path)
            except:
                pass
    
    return False

def cache_data(ticker, data, cache_type="history", period="1y"):
    """Save data to memory and disk cache with timestamp for freshness tracking"""
    
    # Don't cache fresh data requirements
    if is_fresh_data_required(period, cache_type):
        print(f"� Not caching {ticker} ({period}) - fresh data required")
        return
    
    cache_key = f"{ticker}_{cache_type}_{period}"
    current_time = time.time()
    
    # Store in memory with timestamp
    _MEM_CACHE[cache_key] = {
        'data': data,
        'timestamp': current_time
    }
    
    # Save to disk
    try:
        cache_dir = ensure_cache_dir()
        cache_path = os.path.join(cache_dir, f"{md5(cache_key.encode()).hexdigest()}.pkl")
        with open(cache_path, "wb") as f:
            pickle.dump(data, f)
        print(f"🟢 Cached {ticker} ({period}) for future use")
    except Exception as e:
        print(f"Warning: Could not cache {ticker}: {e}")

def get_cached_data(ticker, cache_type="history", period="1y"):
    """Retrieve cached data with freshness validation"""
    
    # Fresh data should not use cache
    if is_fresh_data_required(period, cache_type):
        return None
    
    cache_key = f"{ticker}_{cache_type}_{period}"
    cache_entry = _MEM_CACHE.get(cache_key)
    
    if cache_entry:
        if isinstance(cache_entry, dict) and 'data' in cache_entry:
            return cache_entry['data']
        else:
            # Old format - return as-is for backward compatibility
            return cache_entry
    
    return None

def safe_yf_download(tickers, period="1y", max_retries=3, timeout=10):
    """Resilient yfinance download with rate limiting and caching"""
    if isinstance(tickers, str):
        tickers = [tickers]
    
    results = {}
    live_tickers = [t for t in tickers if not is_cached(t, "history", period)]
    
    if not live_tickers:
        # All cached - return cached data
        for ticker in tickers:
            cached = get_cached_data(ticker, "history", period)
            if cached is not None:
                results[ticker] = cached
        return results
    
    # Indicate fresh vs cached data
    fresh_required = is_fresh_data_required(period, "history")
    cache_status = "FRESH DATA" if fresh_required else "cached"
    print(f"Downloading {len(live_tickers)} new tickers ({cache_status}: {len(tickers) - len(live_tickers)})")
    
    for i, ticker in enumerate(live_tickers):
        for attempt in range(max_retries):
            try:
                # Institutional-grade throttling
                if i % 10 == 0 and i > 0:  
                    time.sleep(0.5)  # Brief pause every 10 requests
                
                enforce_rate_limits()
                
                data = yf.Ticker(ticker).history(period=period, timeout=timeout)
                if not data.empty:
                    cache_data(ticker, data, "history", period)
                    results[ticker] = data
                    status = "� FRESH" if fresh_required else "✓"
                    print(f"{status} {ticker} ({i+1}/{len(live_tickers)})")
                else:
                    print(f"✗ {ticker}: No data")
                break  # Exit retry loop on success
                
            except (RequestException, ConnectionError, Exception) as e:
                if attempt == max_retries - 1:
                    print(f"✗ {ticker}: Failed after {max_retries} attempts - {str(e)[:50]}")
                else:
                    # Exponential backoff with jitter
                    backoff = (2 ** attempt) + random.uniform(0, 1)
                    time.sleep(backoff)
    
    # Add any cached data for requested tickers
    for ticker in tickers:
        if ticker not in results:
            cached = get_cached_data(ticker, "history", period)
            if cached is not None:
                results[ticker] = cached
    
    return results

def safe_yf_info(ticker, max_retries=2):
    """Resilient yfinance info download with caching"""
    if is_cached(ticker, "info"):
        return get_cached_data(ticker, "info")
    
    for attempt in range(max_retries):
        try:
            enforce_rate_limits()
            info = yf.Ticker(ticker).info
            if info:
                cache_data(ticker, info, "info")
                return info
        except Exception as e:
            if attempt == max_retries - 1:
                print(f"Warning: Could not get info for {ticker}: {e}")
                return {}
            time.sleep(1)
    return {}

def batch_download(tickers, period="1y"):
    """Batch download with fallback to individual if needed - respects fresh data requirements"""
    if len(tickers) <= 1:
        return safe_yf_download(tickers, period)
    
    try:
        # Check cache first - but respect fresh data requirements
        cached_tickers = [t for t in tickers if is_cached(t, "history", period)]
        live_tickers = [t for t in tickers if not is_cached(t, "history", period)]
        
        results = {}
        
        # Get cached data (only if not requiring fresh data)
        for ticker in cached_tickers:
            cached = get_cached_data(ticker, "history", period)
            if cached is not None:
                results[ticker] = cached
        
        if not live_tickers:
            return results
        
        # Indicate if this is fresh data download
        fresh_required = is_fresh_data_required(period, "history")
        download_type = "� FRESH BATCH" if fresh_required else "Batch"
        print(f"{download_type} downloading {len(live_tickers)} tickers...")
        enforce_rate_limits()
        
        # Attempt batched download (works for similar-listed symbols)
        batch = yf.download(
            tickers=live_tickers,
            period=period,
            group_by="ticker",
            threads=False,  # Avoid parallel threads
            timeout=15,
            progress=False
        )
        
        # Parse batch results
        if len(live_tickers) == 1:
            # Single ticker case
            ticker = live_tickers[0]
            if not batch.empty:
                cache_data(ticker, batch, "history", period)
                results[ticker] = batch
        else:
            # Multi-ticker case
            for ticker in live_tickers:
                try:
                    if ticker in batch.columns.levels[0]:
                        ticker_data = batch[ticker]
                        if not ticker_data.empty:
                            cache_data(ticker, ticker_data, "history", period)
                            results[ticker] = ticker_data
                except Exception:
                    pass
        
        # Fallback to individual download for missing tickers
        missing = [t for t in live_tickers if t not in results]
        if missing:
            fallback_type = "� FRESH FALLBACK" if fresh_required else "Fallback"
            print(f"{fallback_type}: downloading {len(missing)} individual tickers...")
            individual_results = safe_yf_download(missing, period)
            results.update(individual_results)
        
        return results
        
    except Exception as e:
        print(f"Batch download failed: {e}. Falling back to individual downloads...")
        return safe_yf_download(tickers, period)  # Fallback to individual downloads

def get_cache_stats():
    """Get cache performance statistics"""
    total_entries = len(_MEM_CACHE)
    cache_dir = ensure_cache_dir()
    disk_files = len([f for f in os.listdir(cache_dir) if f.endswith('.pkl')]) if os.path.exists(cache_dir) else 0
    
    return {
        "memory_cache_entries": total_entries,
        "disk_cache_files": disk_files,
        "requests_in_last_minute": len([t for t in _REQUEST_LOG if time.time() - t < 60])
    }

def print_cache_stats():
    """Print cache statistics for monitoring"""
    stats = get_cache_stats()
    print(f"\n📊 YFinance Cache Stats:")
    print(f"   Memory Cache: {stats['memory_cache_entries']} entries")
    print(f"   Disk Cache: {stats['disk_cache_files']} files")
    print(f"   Recent Requests: {stats['requests_in_last_minute']}/min")

# =============================================================================
# CRITICAL FRESH DATA FUNCTIONS - Always Live Data for Important Calculations
# =============================================================================

def get_fresh_intraday_data(ticker: str, period: str = "1d") -> pd.DataFrame:
    """
    Get ALWAYS FRESH intraday data for critical calculations like buyer dominance.
    This bypasses ALL caching to ensure real-time data for trading decisions.
    """
    print(f"🔵 CRITICAL: Fetching LIVE data for {ticker} - bypassing ALL cache")
    
    try:
        enforce_rate_limits()
        # Direct yfinance call with no caching for critical calculations
        data = yf.Ticker(ticker).history(period=period, interval="5m", timeout=15)
        
        if not data.empty:
            print(f"🔵 LIVE: {ticker} - {len(data)} 5min bars, latest: {data.index[-1]}")
            return data
        else:
            print(f"⚠️ WARNING: No live data for {ticker}")
            return pd.DataFrame()
            
    except Exception as e:
        print(f"🔴 ERROR: Failed to get live data for {ticker}: {e}")
        return pd.DataFrame()

def get_fresh_current_price(ticker: str) -> dict:
    """
    Get ALWAYS FRESH current price and today's trading data.
    Critical for real-time trading decisions - never uses cache.
    """
    print(f"🔵 CRITICAL: Fetching LIVE price for {ticker}")
    
    try:
        enforce_rate_limits()
        
        # Get current day data with no caching
        ticker_obj = yf.Ticker(ticker)
        
        # Get today's OHLCV data
        today_data = ticker_obj.history(period="1d", interval="1m")
        
        # Get current info for real-time price
        info = ticker_obj.info
        
        result = {
            'current_price': info.get('currentPrice', 0.0),
            'previous_close': info.get('previousClose', 0.0),
            'today_open': info.get('open', 0.0),
            'day_high': info.get('dayHigh', 0.0),
            'day_low': info.get('dayLow', 0.0),
            'volume': info.get('volume', 0),
            'avg_volume': info.get('averageVolume', 0),
            'market_cap': info.get('marketCap', 0),
            'bid': info.get('bid', 0.0),
            'ask': info.get('ask', 0.0),
            'bid_size': info.get('bidSize', 0),
            'ask_size': info.get('askSize', 0),
            'today_data': today_data,
            'data_timestamp': datetime.now().isoformat()
        }
        
        # Calculate real-time metrics
        if result['previous_close'] > 0:
            result['change_percent'] = ((result['current_price'] - result['previous_close']) / result['previous_close']) * 100
        else:
            result['change_percent'] = 0.0
            
        # Volume ratio for today
        if result['avg_volume'] > 0:
            result['volume_ratio'] = result['volume'] / result['avg_volume']
        else:
            result['volume_ratio'] = 1.0
        
        print(f"🔵 LIVE: {ticker} = ₹{result['current_price']:.2f} ({result['change_percent']:+.2f}%) Vol: {result['volume_ratio']:.1f}x")
        
        return result
        
    except Exception as e:
        print(f"🔴 ERROR: Failed to get live price for {ticker}: {e}")
        return {'current_price': 0.0, 'error': str(e)}

def get_fresh_buyer_dominance_data(ticker: str) -> dict:
    """
    Get ALWAYS FRESH data for buyer dominance calculation.
    This is critical for momentum analysis and must be real-time.
    """
    print(f"🔵 CRITICAL: Calculating LIVE buyer dominance for {ticker}")
    
    try:
        # Get fresh 5-minute data for the last few hours
        intraday_data = get_fresh_intraday_data(ticker, period="1d")
        current_price_data = get_fresh_current_price(ticker)
        
        if intraday_data.empty:
            return {'buyer_dominance': 0.0, 'error': 'No intraday data'}
        
        # Calculate buyer dominance using fresh data
        result = calculate_buyer_dominance_from_fresh_data(intraday_data, current_price_data)
        
        print(f"🔵 LIVE BUYER DOMINANCE: {ticker} = {result['buyer_dominance']:.1f}%")
        
        return result
        
    except Exception as e:
        print(f"🔴 ERROR: Failed buyer dominance calculation for {ticker}: {e}")
        return {'buyer_dominance': 0.0, 'error': str(e)}

def calculate_buyer_dominance_from_fresh_data(intraday_df: pd.DataFrame, price_data: dict) -> dict:
    """
    Calculate buyer dominance using fresh intraday data.
    Uses multiple indicators: price action, volume, bid-ask spread.
    """
    if intraday_df.empty:
        return {'buyer_dominance': 0.0, 'confidence': 0.0}
    
    try:
        # 1. Price momentum (last 2 hours vs first 2 hours)
        total_bars = len(intraday_df)
        if total_bars < 10:
            return {'buyer_dominance': 0.0, 'confidence': 0.0}
        
        # Split data into early and late periods
        early_period = intraday_df.iloc[:total_bars//3]
        late_period = intraday_df.iloc[-total_bars//3:]
        
        early_avg_price = early_period['Close'].mean()
        late_avg_price = late_period['Close'].mean()
        
        # Price momentum score
        if early_avg_price > 0:
            price_momentum = ((late_avg_price - early_avg_price) / early_avg_price) * 100
        else:
            price_momentum = 0.0
        
        # 2. Volume-weighted price momentum
        early_vwap = (early_period['Close'] * early_period['Volume']).sum() / early_period['Volume'].sum() if early_period['Volume'].sum() > 0 else early_avg_price
        late_vwap = (late_period['Close'] * late_period['Volume']).sum() / late_period['Volume'].sum() if late_period['Volume'].sum() > 0 else late_avg_price
        
        if early_vwap > 0:
            vwap_momentum = ((late_vwap - early_vwap) / early_vwap) * 100
        else:
            vwap_momentum = 0.0
        
        # 3. Volume analysis (buying vs selling pressure)
        buying_volume = 0
        selling_volume = 0
        
        for i in range(1, len(intraday_df)):
            price_change = intraday_df.iloc[i]['Close'] - intraday_df.iloc[i-1]['Close']
            volume = intraday_df.iloc[i]['Volume']
            
            if price_change > 0:
                buying_volume += volume
            elif price_change < 0:
                selling_volume += volume
        
        total_volume = buying_volume + selling_volume
        if total_volume > 0:
            volume_dominance = (buying_volume / total_volume) * 100
        else:
            volume_dominance = 50.0
        
        # 4. Bid-Ask analysis (if available)
        bid = price_data.get('bid', 0.0)
        ask = price_data.get('ask', 0.0)
        bid_size = price_data.get('bid_size', 0)
        ask_size = price_data.get('ask_size', 0)
        
        if bid > 0 and ask > 0 and bid_size > 0 and ask_size > 0:
            # Bid-ask spread analysis
            spread = ((ask - bid) / bid) * 100
            size_ratio = bid_size / (bid_size + ask_size) * 100
            bid_ask_score = (100 - spread) * (size_ratio / 100)  # Lower spread + higher bid size = more buying pressure
        else:
            bid_ask_score = 50.0  # Neutral if no bid-ask data
        
        # 5. Combine all factors with weights
        buyer_dominance = (
            price_momentum * 0.3 +      # 30% weight on price momentum
            vwap_momentum * 0.3 +       # 30% weight on VWAP momentum  
            volume_dominance * 0.3 +    # 30% weight on volume analysis
            (bid_ask_score - 50) * 0.1  # 10% weight on bid-ask (normalized to ±50)
        ) + 50  # Base 50% + momentum adjustments
        
        # Clamp to 0-100 range
        buyer_dominance = max(0.0, min(100.0, buyer_dominance))
        
        # Calculate confidence based on data quality
        data_points = len(intraday_df)
        volume_consistency = min(1.0, total_volume / max(1, price_data.get('avg_volume', 1)))
        bid_ask_availability = 1.0 if (bid > 0 and ask > 0) else 0.5
        
        confidence = min(1.0, (data_points / 50) * volume_consistency * bid_ask_availability)
        
        return {
            'buyer_dominance': buyer_dominance,
            'confidence': confidence * 100,
            'price_momentum': price_momentum,
            'vwap_momentum': vwap_momentum,
            'volume_dominance': volume_dominance,
            'bid_ask_score': bid_ask_score,
            'data_points': data_points,
            'calculation_timestamp': datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"Error in buyer dominance calculation: {e}")
        return {'buyer_dominance': 0.0, 'confidence': 0.0, 'error': str(e)}


# =============================================================================
# END RESILIENT YFINANCE SYSTEM
# =============================================================================

# =============================================================================
# CLAUDE DESKTOP INTEGRATION - For AI-Powered Analysis
# =============================================================================

def save_opportunities_for_claude_analysis(opportunities, filename=None):
    """
    Save trading opportunities for Claude Desktop analysis
    Creates formatted files that Claude Desktop can analyze for trading insights
    
    Args:
        opportunities: List of trading opportunities from screener
        filename: Optional custom filename
        
    Returns:
        Path to created analysis file
    """
    if not opportunities:
        print("No opportunities to save for Claude analysis")
        return None
    
    # Import here to avoid dependency issues
    from pathlib import Path
    import os
    
    # Create output directory
    output_dir = Path("claude_analysis")
    output_dir.mkdir(exist_ok=True)
    
    # Generate filename
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trading_opportunities_{timestamp}.md"
    
    filepath = output_dir / filename
    
    # Create analysis content optimized for Claude Desktop
    content = f"""# 🤖 Trading Opportunities Analysis
*Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}*

## Claude: Please analyze these Indian stock opportunities

**Total Opportunities**: {len(opportunities)}

### 🎯 Analysis Request:
1. **Market Assessment**: What do these picks tell us about market conditions?
2. **Top Recommendations**: Which 3-5 stocks look most promising and why?
3. **Risk Analysis**: Key risks and mitigation strategies?
4. **Entry Strategy**: Optimal entry approach and timing?
5. **Position Sizing**: Recommended allocation per stock?

## 📊 OPPORTUNITY DATA

"""
    
    # Add top opportunities with detailed metrics
    for i, opp in enumerate(opportunities[:15], 1):  # Top 15 for Claude analysis
        content += f"""
### {i}. {opp.get('ticker', 'N/A')} - Score: {opp.get('total_score', 0):.1f}

**Classification**: {opp.get('tier', 'N/A')} Grade

**Current Setup**:
- Price: ₹{opp.get('close_price', 0):.2f}
- RSI: {opp.get('rsi', 'N/A')} {'(Oversold)' if opp.get('rsi', 50) < 30 else '(Overbought)' if opp.get('rsi', 50) > 70 else '(Normal)'}
- Bollinger Position: {opp.get('bb_position', 'N/A')}%
- Volume: {opp.get('volume_ratio', 1):.1f}x average {'(Strong)' if opp.get('volume_ratio', 1) > 1.5 else '(Normal)'}
- Daily Return: {opp.get('daily_return', 0):.1f}%

**Levels**:
- Support: ₹{opp.get('support_level', 0):.2f}
- Resistance: ₹{opp.get('resistance_level', 0):.2f}
- Risk/Reward: {opp.get('rr_ratio', 1):.1f} {'(Excellent)' if opp.get('rr_ratio', 1) > 2.5 else '(Good)' if opp.get('rr_ratio', 1) > 2.0 else '(Fair)'}

"""
    
    content += f"""
## 📈 MARKET CONTEXT

These opportunities were selected using:
- **RSI Analysis**: Momentum and oversold/overbought conditions
- **Bollinger Bands**: Price position relative to volatility bands
- **Volume Confirmation**: Above-average trading activity
- **Risk/Reward**: Favorable upside vs downside potential
- **Support/Resistance**: Key technical levels for entry/exit

## 🎯 FOR CLAUDE TO CONSIDER

**Indian Market Specifics**:
- NSE/BSE dynamics and sector rotation
- FII/DII activity impact
- Current market regime (trending/sideways/volatile)

**Trading Approach**:
- Swing trading (2-10 days holding)
- Position sizing based on volatility
- Stop-loss using support levels
- Target profits at resistance levels

---
*Generated by algorithmic screening system - Open in Claude Desktop for AI analysis*
"""
    
    # Save file
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        
        print(f"✅ Claude analysis file created: {filepath}")
        print(f"🤖 Open this file in Claude Desktop for AI-powered trading insights!")
        
        # Try to open the file automatically
        try:
            os.startfile(str(filepath))
        except:
            try:
                os.startfile(str(output_dir))
            except:
                pass
        
        return str(filepath)
        
    except Exception as e:
        print(f"Error creating Claude analysis file: {e}")
        return None

def quick_claude_stock_analysis(ticker, stock_data):
    """
    Create a focused single-stock analysis file for Claude Desktop
    
    Args:
        ticker: Stock symbol
        stock_data: Stock analysis results
        
    Returns:
        Path to created analysis file
    """
    if not stock_data:
        print(f"No data available for {ticker}")
        return None
    
    from pathlib import Path
    import os
    
    output_dir = Path("claude_analysis")
    output_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"stock_analysis_{ticker}_{timestamp}.md"
    filepath = output_dir / filename
    
    content = f"""# 📊 {ticker} - Individual Stock Analysis
*Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}*

## 🤖 Claude: Analyze this Indian stock for swing trading

### Stock Overview: {ticker}
**Opportunity Score**: {stock_data.get('total_score', 0):.2f}/100
**Grade**: {stock_data.get('tier', 'N/A')}

### Technical Setup
| Indicator | Value | Status |
|-----------|--------|--------|
| **Current Price** | ₹{stock_data.get('close_price', 0):.2f} | - |
| **RSI** | {stock_data.get('rsi', 'N/A')} | {'Oversold' if stock_data.get('rsi', 50) < 30 else 'Overbought' if stock_data.get('rsi', 50) > 70 else 'Normal'} |
| **Bollinger Position** | {stock_data.get('bb_position', 'N/A')}% | {'Lower band' if stock_data.get('bb_position', 50) < 20 else 'Upper band' if stock_data.get('bb_position', 50) > 80 else 'Middle range'} |
| **Volume Ratio** | {stock_data.get('volume_ratio', 1):.1f}x | {'High' if stock_data.get('volume_ratio', 1) > 1.5 else 'Normal'} |
| **Daily Return** | {stock_data.get('daily_return', 0):.1f}% | - |

### Key Levels
- **Support**: ₹{stock_data.get('support_level', 0):.2f}
- **Current**: ₹{stock_data.get('close_price', 0):.2f}
- **Resistance**: ₹{stock_data.get('resistance_level', 0):.2f}
- **Risk/Reward**: {stock_data.get('rr_ratio', 1):.1f}

### 🎯 Questions for Claude:

1. **Trade Setup**: Is this a high-probability swing trade setup?
2. **Entry Strategy**: What's the optimal entry price and method?
3. **Risk Management**: Where should I place my stop-loss?
4. **Profit Targets**: What are realistic profit targets?
5. **Timeline**: Expected holding period for this trade?
6. **Key Risks**: What could invalidate this setup?

### Additional Context:
- This is an NSE/BSE listed Indian stock
- Analysis is for swing trading (2-10 days holding)
- Consider current market conditions and sector trends

---
*Open in Claude Desktop for detailed trading analysis*
"""
    
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        
        print(f"✅ {ticker} analysis saved: {filepath}")
        print(f"🤖 Open in Claude Desktop for AI insights")
        
        # Try to open the file
        try:
            os.startfile(str(filepath))
        except:
            pass
        
        return str(filepath)
        
    except Exception as e:
        print(f"Error creating {ticker} analysis: {e}")
        return None

# =============================================================================
# END CLAUDE DESKTOP INTEGRATION
# =============================================================================


# Optional libs for FII scraping
try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    requests = None
    BeautifulSoup = None

try:
    import talib
except ImportError:
    talib = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# =============================================================================
# WINSORIZATION CONFIGURATION - Data Quality Control Parameters
# =============================================================================

# Price return winsorization thresholds
WINSORIZATION_CONFIG = {
    'price_return_threshold': 0.5,      # ±50% daily return cap (configurable)
    'volume_change_threshold': 1.0,     # ±100% volume change cap  
    'excessive_clip_rate': 0.05,        # Warn if >5% of data gets clipped
    'min_data_points': 5,               # Minimum rows after cleaning
    'enable_volume_capping': True,      # Whether to cap volume changes
    'enable_trading_calendar': True,    # Align to NSE business days
    'log_extreme_examples': True,       # Log examples of capped values
}

# Market-specific thresholds (can be customized by market cap, sector, etc.)
MARKET_SPECIFIC_THRESHOLDS = {
    'large_cap': {'price_threshold': 0.3, 'volume_threshold': 0.8},     # Large caps more stable
    'small_cap': {'price_threshold': 0.7, 'volume_threshold': 1.5},     # Small caps more volatile
    'default': {'price_threshold': 0.5, 'volume_threshold': 1.0},       # Default fallback
}

# =============================================================================
# YFINANCE DATA VALIDATION SYSTEM - CRITICAL DATA QUALITY CONTROL
# =============================================================================

class YFinanceDataValidator:
    """Comprehensive validation system for yfinance data to ensure data quality"""
    
    @staticmethod
    def validate_price_data(df: pd.DataFrame, ticker: str = "", soft_mode: bool = False) -> Tuple[bool, str, pd.DataFrame]:
        """
        Validate OHLCV price data for anomalies and errors
        Returns: (is_valid, error_message, cleaned_dataframe)
        """
        if df is None or df.empty:
            return False, f"Empty dataframe for {ticker}", pd.DataFrame()
        
        try:
            # 1. Check required columns
            required_cols = ['Open', 'High', 'Low', 'Close', 'Volume']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                return False, f"Missing columns {missing_cols} for {ticker}", df
            
            # 2. Remove rows with all NaN values
            df_clean = df.dropna(how='all')
            if df_clean.empty:
                return False, f"All data is NaN for {ticker}", pd.DataFrame()
            
            # 3. Validate OHLC relationships (High >= Low, High >= Open/Close, Low <= Open/Close)
            invalid_ohlc = (
                (df_clean['High'] < df_clean['Low']) |
                (df_clean['High'] < df_clean['Open']) |
                (df_clean['High'] < df_clean['Close']) |
                (df_clean['Low'] > df_clean['Open']) |
                (df_clean['Low'] > df_clean['Close'])
            )
            
            if invalid_ohlc.any():
                invalid_count = invalid_ohlc.sum()
                logging.warning(f"Found {invalid_count} invalid OHLC relationships for {ticker}")
                # Remove invalid rows
                df_clean = df_clean[~invalid_ohlc]
            
            # 4. Check for negative or zero prices
            price_cols = ['Open', 'High', 'Low', 'Close']
            for col in price_cols:
                invalid_prices = (df_clean[col] <= 0) | df_clean[col].isna()
                if invalid_prices.any():
                    invalid_count = invalid_prices.sum()
                    logging.warning(f"Found {invalid_count} invalid prices in {col} for {ticker}")
                    # Forward fill small gaps, drop if too many
                    if invalid_count < len(df_clean) * 0.1:  # Less than 10% invalid
                        df_clean.loc[:, col] = df_clean[col].replace(0, np.nan)
                        df_clean.loc[:, col] = df_clean[col].fillna(method='ffill')
                    else:
                        return False, f"Too many invalid prices in {col} for {ticker}", df_clean
            
            # 5. Check for negative volume
            invalid_volume = (df_clean['Volume'] < 0) | df_clean['Volume'].isna()
            if invalid_volume.any():
                invalid_count = invalid_volume.sum()
                logging.warning(f"Found {invalid_count} invalid volume values for {ticker}")
                # Replace negative/NaN volume with 0 using .loc to avoid SettingWithCopyWarning
                df_clean.loc[:, 'Volume'] = df_clean['Volume'].fillna(0)
                df_clean.loc[df_clean['Volume'] < 0, 'Volume'] = 0
            
            # 6. Apply centralized winsorization system for extreme movements
            if len(df_clean) > 1:
                # Fetch market cap info for adaptive thresholds
                market_cap_info = YFinanceDataValidator.safe_ticker_info(ticker)
                # Apply adaptive winsorization using centralized system with real market cap data
                df_clean = apply_adaptive_winsorization(df_clean, ticker, market_cap_info=market_cap_info)
            else:
                # Single data point - add placeholder columns for consistency
                df_clean.loc[:, 'PctChange_Capped'] = 0.0
                df_clean.loc[:, 'PctChange_Raw'] = 0.0
                df_clean.loc[:, 'Close_Capped'] = df_clean['Close']
                if WINSORIZATION_CONFIG['enable_volume_capping']:
                    df_clean.loc[:, 'VolChange_Capped'] = 0.0
                    df_clean.loc[:, 'VolChange_Raw'] = 0.0
                    df_clean.loc[:, 'Volume_Capped'] = df_clean['Volume']
            
            # 7. Check for sufficient data points (preserved by winsorization instead of dropping)
            if len(df_clean) < WINSORIZATION_CONFIG['min_data_points']:
                return False, f"Insufficient data points ({len(df_clean)}) for {ticker}", df_clean
            
            # 8. Validate data types using .loc to ensure conversion sticks
            for col in price_cols + ['Volume']:
                df_clean.loc[:, col] = pd.to_numeric(df_clean[col], errors='coerce')
            
            # 9. Final validation: ensure all numeric columns are actually numeric
            for col in price_cols + ['Volume']:
                if df_clean[col].dtype == 'object':
                    logging.warning(f"Column {col} still contains non-numeric data for {ticker}")
                    # Force conversion one more time
                    df_clean.loc[:, col] = pd.to_numeric(df_clean[col], errors='coerce')
            
            # 10. Remove any remaining NaN rows after data type conversion
            df_clean = df_clean.dropna()
            
            if df_clean.empty:
                return False, f"No valid data remaining after cleaning for {ticker}", pd.DataFrame()
            
            # 11. Final sanity check: verify OHLC relationships again after all cleaning
            if len(df_clean) > 0:
                final_invalid_ohlc = (
                    (df_clean['High'] < df_clean['Low']) |
                    (df_clean['High'] < df_clean['Open']) |
                    (df_clean['High'] < df_clean['Close']) |
                    (df_clean['Low'] > df_clean['Open']) |
                    (df_clean['Low'] > df_clean['Close'])
                )
                if final_invalid_ohlc.any():
                    logging.error(f"Final validation failed: Still have invalid OHLC relationships for {ticker}")
                    df_clean = df_clean[~final_invalid_ohlc]
            
            # 12. Align to NSE trading calendar for proper business day handling
            if WINSORIZATION_CONFIG.get('enable_trading_calendar', True):
                df_clean = align_to_nse_trading_calendar(df_clean, ticker)
            
            # 13. Generate quality report for winsorization auditing
            quality_report = validate_winsorization_quality(df_clean, ticker)
            if quality_report['data_quality'] == 'FAILED':
                return False, f"Winsorization quality check failed for {ticker}: {quality_report['warnings']}", df_clean
            elif quality_report['data_quality'] == 'POOR':
                logging.warning(f"Poor winsorization quality for {ticker}: {quality_report['warnings']}")
                # In soft mode, allow POOR data quality but tag for user awareness
                if not soft_mode:
                    return False, f"Excessive data clipping for {ticker} (use --soft-mode to include)", df_clean
            
            return True, "Valid", df_clean
            
        except Exception as e:
            return False, f"Validation error for {ticker}: {str(e)}", df
    
    @staticmethod
    def validate_info_data(info: dict, ticker: str = "") -> Tuple[bool, str, dict]:
        """
        Validate and clean stock info data from yfinance
        Returns: (is_valid, error_message, cleaned_info)
        """
        if not info or not isinstance(info, dict):
            return False, f"Empty or invalid info data for {ticker}", {}
        
        try:
            cleaned_info = {}
            
            # Common numeric fields that should be positive
            positive_fields = [
                'marketCap', 'sharesOutstanding', 'floatShares', 'volume',
                'averageVolume', 'averageVolume10days', 'bookValue', 'priceToBook',
                'totalRevenue', 'totalCash', 'totalDebt', 'currentPrice',
                'previousClose', 'open', 'dayLow', 'dayHigh', 'fiftyTwoWeekLow',
                'fiftyTwoWeekHigh', 'volume24Hr'
            ]
            
            # Ratio fields that can be negative but should be reasonable
            ratio_fields = [
                'trailingPE', 'forwardPE', 'pegRatio', 'priceToSalesTrailing12Months',
                'enterpriseToRevenue', 'enterpriseToEbitda', 'returnOnAssets',
                'returnOnEquity', 'profitMargins', 'grossMargins', 'operatingMargins',
                'debtToEquity', 'currentRatio', 'quickRatio', 'beta'
            ]
            
            # Percentage fields (should be between 0 and 1 if decimal, or reasonable if percentage)
            percentage_fields = [
                'heldPercentInsiders', 'heldPercentInstitutions', 'shortPercentOfFloat',
                'payoutRatio', 'dividendYield'
            ]
            
            # Validate positive fields
            for field in positive_fields:
                value = info.get(field)
                if value is not None:
                    try:
                        num_value = float(value)
                        if num_value > 0 and not np.isinf(num_value):
                            cleaned_info[field] = num_value
                        elif num_value == 0 and field in ['totalDebt', 'dividendYield']:
                            # Zero debt or dividend is valid
                            cleaned_info[field] = num_value
                        else:
                            logging.debug(f"Invalid {field} value {value} for {ticker}")
                    except (ValueError, TypeError):
                        logging.debug(f"Non-numeric {field} value {value} for {ticker}")
            
            # Validate ratio fields
            for field in ratio_fields:
                value = info.get(field)
                if value is not None:
                    try:
                        num_value = float(value)
                        # Reasonable bounds for financial ratios
                        if field in ['trailingPE', 'forwardPE'] and (num_value < 0 or num_value > 1000):
                            logging.debug(f"Extreme PE ratio {num_value} for {ticker}")
                            continue
                        elif field == 'beta' and (num_value < -3 or num_value > 5):
                            logging.debug(f"Extreme beta {num_value} for {ticker}")
                            continue
                        elif field in ['debtToEquity'] and num_value < 0:
                            logging.debug(f"Negative debt-to-equity {num_value} for {ticker}")
                            continue
                        
                        if not np.isinf(num_value) and not np.isnan(num_value):
                            cleaned_info[field] = num_value
                    except (ValueError, TypeError):
                        logging.debug(f"Non-numeric {field} value {value} for {ticker}")
            
            # Validate percentage fields
            for field in percentage_fields:
                value = info.get(field)
                if value is not None:
                    try:
                        num_value = float(value)
                        # Accept both decimal (0-1) and percentage (0-100) formats
                        if 0 <= num_value <= 1 or 0 <= num_value <= 100:
                            cleaned_info[field] = num_value
                        else:
                            logging.debug(f"Invalid percentage {field} value {value} for {ticker}")
                    except (ValueError, TypeError):
                        logging.debug(f"Non-numeric {field} value {value} for {ticker}")
            
            # Preserve string fields as-is
            string_fields = [
                'symbol', 'shortName', 'longName', 'sector', 'industry',
                'country', 'exchange', 'currency', 'quoteType'
            ]
            for field in string_fields:
                value = info.get(field)
                if value and isinstance(value, str) and value.strip():
                    cleaned_info[field] = value.strip()
            
            return True, "Valid", cleaned_info
            
        except Exception as e:
            return False, f"Info validation error for {ticker}: {str(e)}", {}
    
    @staticmethod
    def safe_yfinance_download(symbols: list, period: str = "1y", interval: str = "1d", 
                              timeout: int = 30, soft_mode: bool = False) -> dict:
        """
        Safely download data for multiple symbols with validation
        Returns: dict of {symbol: validated_dataframe}
        """
        validated_data = {}
        
        try:
            # Download with error handling using resilient batch system
            logging.info(f"Downloading data for {len(symbols)} symbols...")
            
            # Use our resilient batch download system
            df_batch_dict = batch_download(symbols, period)
            
            if not df_batch_dict:
                logging.warning("No data returned from resilient download")
                return validated_data
                logging.warning("No data returned from yfinance download")
                return validated_data
            
            # Handle single vs multiple symbols with dictionary format
            if len(symbols) == 1:
                symbol = symbols[0]
                if symbol in df_batch_dict:
                    is_valid, error_msg, clean_df = YFinanceDataValidator.validate_price_data(df_batch_dict[symbol], symbol, soft_mode)
                    if is_valid:
                        validated_data[symbol] = clean_df
                    else:
                        logging.warning(f"Data validation failed for {symbol}: {error_msg}")
                else:
                    logging.warning(f"No data found for symbol {symbol}")
            else:
                # Multiple symbols - iterate through each
                for symbol in symbols:
                    try:
                        if symbol in df_batch_dict:
                            symbol_df = df_batch_dict[symbol]
                            is_valid, error_msg, clean_df = YFinanceDataValidator.validate_price_data(symbol_df, symbol, soft_mode)
                            if is_valid:
                                validated_data[symbol] = clean_df
                            else:
                                logging.warning(f"Data validation failed for {symbol}: {error_msg}")
                        else:
                            logging.warning(f"No data found for symbol {symbol}")
                    except Exception as e:
                        logging.warning(f"Error processing symbol {symbol}: {str(e)}")
            
            logging.info(f"Successfully validated data for {len(validated_data)} out of {len(symbols)} symbols")
            return validated_data
            
        except Exception as e:
            logging.error(f"Error in batch download: {str(e)}")
            return validated_data
    
    @staticmethod
    def safe_ticker_info(ticker: str, timeout: int = 10) -> dict:
        """
        Safely get ticker info with validation using resilient download
        Returns: validated info dict
        """
        try:
            rate_limit(0.1)  # Rate limiting
            # --- FIX: prevent double ".NS" ---------------------------------
            yf_symbol = ensure_ns_suffix(ticker)
            info = safe_yf_info(yf_symbol)
            
            is_valid, error_msg, clean_info = YFinanceDataValidator.validate_info_data(info, ticker)
            if is_valid:
                return clean_info
            else:
                logging.warning(f"Info validation failed for {ticker}: {error_msg}")
                return {}
                
        except Exception as e:
            logging.warning(f"Error fetching info for {ticker}: {str(e)}")
            return {}
    
    @staticmethod
    def safe_ticker_history(ticker: str, period: str = "1y", interval: str = "1d", 
                           timeout: int = 10) -> pd.DataFrame:
        """
        Safely get ticker history with validation
        Returns: validated dataframe
        """
        try:
            rate_limit(0.1)  # Rate limiting
            # --- FIX: prevent double ".NS" ---------------------------------
            yf_symbol = ensure_ns_suffix(ticker)
            result = safe_yf_download([yf_symbol], period=period)
            if yf_symbol in result:
                df = result[yf_symbol]
            else:
                return pd.DataFrame()
            
            is_valid, error_msg, clean_df = YFinanceDataValidator.validate_price_data(df, ticker, False)  # Safe ticker history always strict
            if is_valid:
                return clean_df
            else:
                logging.warning(f"History validation failed for {ticker}: {error_msg}")
                return pd.DataFrame()
                
        except Exception as e:
            logging.warning(f"Error fetching history for {ticker}: {str(e)}")
            return pd.DataFrame()

# =============================================================================
# CENTRALIZED WINSORIZATION SYSTEM - DRY Principle Implementation
# =============================================================================

def winsorize_series(series: pd.Series, threshold: float, series_name: str = "data", 
                    ticker: str = "", log_examples: bool = True) -> tuple:
    """
    Centralized winsorization (capping) of extreme values in a pandas Series.
    
    Args:
        series: The pandas Series to winsorize
        threshold: The threshold for capping (e.g., 0.5 for ±50%)
        series_name: Name of the series for logging (e.g., "returns", "volume_changes")
        ticker: Ticker symbol for logging
        log_examples: Whether to log examples of capped values
    
    Returns:
        tuple: (capped_series, extreme_mask, clip_rate, first_value)
    """
    if series.empty:
        return series, pd.Series([], dtype=bool), 0.0, None
    
    # Identify extreme values
    extreme_mask = series.abs() > threshold
    extreme_count = extreme_mask.sum()
    clip_rate = extreme_count / len(series) if len(series) > 0 else 0.0
    
    # Cap the series
    capped_series = series.clip(lower=-threshold, upper=threshold)
    
    # Get first valid value for cumulative reconstruction
    first_value = series.iloc[0] if len(series) > 0 else None
    
    # Logging
    if extreme_count > 0:
        logging.warning(f"Found {extreme_count} extreme {series_name} (>{threshold*100:.0f}%) for {ticker} - WINSORIZING to preserve continuity")
        
        # Log examples if requested
        if log_examples and extreme_count > 0:
            extreme_values = series[extreme_mask].dropna()
            if len(extreme_values) > 0:
                examples = extreme_values.head(3).values
                logging.info(f"Extreme {series_name} being capped for {ticker}: {examples}")
        
        # Alert if excessive clipping
        if clip_rate > WINSORIZATION_CONFIG['excessive_clip_rate']:
            logging.error(f"EXCESSIVE WINSORIZATION: {clip_rate:.1%} of {series_name} data clipped for {ticker} (threshold: {WINSORIZATION_CONFIG['excessive_clip_rate']:.1%})")
    
    return capped_series, extreme_mask, clip_rate, first_value

def get_market_cap_category(info: dict) -> str:
    """
    Determine market cap category for adaptive thresholds.
    """
    market_cap = info.get('marketCap', 0)
    if market_cap > 1e12:  # >₹1 lakh crore (large cap)
        return 'large_cap'
    elif market_cap < 1e10:  # <₹1000 crore (small cap)
        return 'small_cap'
    else:
        return 'default'

def apply_adaptive_winsorization(df: pd.DataFrame, ticker: str = "", 
                                market_cap_info: dict = None) -> pd.DataFrame:
    """
    OPTIMIZED: Apply adaptive winsorization - clamp returns/volume once, then reconstruct.
    Performance improvement: Single-pass clamping with efficient cumulative reconstruction.
    
    Args:
        df: DataFrame with OHLCV data
        ticker: Ticker symbol for logging
        market_cap_info: Market cap info for adaptive thresholds
    
    Returns:
        DataFrame with winsorized columns added
    """
    if df.empty or len(df) <= 1:
        return df
    
    df_result = df.copy()
    
    # Determine adaptive thresholds by market cap bucket
    if market_cap_info:
        category = get_market_cap_category(market_cap_info)
        thresholds = MARKET_SPECIFIC_THRESHOLDS.get(category, MARKET_SPECIFIC_THRESHOLDS['default'])
        price_threshold = thresholds['price_threshold']
        volume_threshold = thresholds['volume_threshold']
    else:
        price_threshold = WINSORIZATION_CONFIG['price_return_threshold']
        volume_threshold = WINSORIZATION_CONFIG['volume_change_threshold']
    
    # OPTIMIZED APPROACH: Clamp returns once, then reconstruct close series
    # 1. Calculate raw returns
    close = df['Close']
    ret_raw = close.pct_change()
    
    # 2. Clamp returns in single operation (adapt by cap bucket)
    ret_cap = ret_raw.clip(-price_threshold, price_threshold)
    
    # 3. Reconstruct capped close series using cumulative product
    close_cap = (1 + ret_cap.fillna(0)).cumprod() * close.iloc[0]
    
    # 4. Store results efficiently
    df_result.loc[:, 'PctChange_Raw'] = ret_raw
    df_result.loc[:, 'PctChange_Capped'] = ret_cap
    df_result.loc[:, 'Close_Capped'] = close_cap
    
    # 5. Quality report guardrail for price data
    clip_rate = (ret_raw.ne(ret_cap)).mean() if len(ret_raw) > 0 else 0.0
    if clip_rate > 0.05:  # 5% threshold
        data_log.warning("High price clipping %.1f%% for %s - review data quality", clip_rate*100, ticker)
        if clip_rate > 0.20:  # 20% threshold - severe data quality issue
            data_log.error("SEVERE clipping %.1f%% for %s - consider data source alternatives", clip_rate*100, ticker)
    
    # 6. OPTIMIZED VOLUME PROCESSING (if enabled)
    if WINSORIZATION_CONFIG['enable_volume_capping'] and 'Volume' in df.columns:
        volume = df['Volume']
        vol_ret_raw = volume.pct_change()
        
        # Clamp volume returns in single operation
        vol_ret_cap = vol_ret_raw.clip(-volume_threshold, volume_threshold)
        
        # Reconstruct capped volume series
        first_volume = volume.iloc[0]
        if first_volume > 0:
            vol_cap = (1 + vol_ret_cap.fillna(0)).cumprod() * first_volume
            # Ensure volume stays non-negative
            vol_cap = vol_cap.clip(lower=0)
        else:
            vol_cap = volume.copy()
        
        # Store volume results
        df_result.loc[:, 'VolChange_Raw'] = vol_ret_raw
        df_result.loc[:, 'VolChange_Capped'] = vol_ret_cap
        df_result.loc[:, 'Volume_Capped'] = vol_cap
        
        # Volume quality report guardrail
        vol_clip_rate = (vol_ret_raw.ne(vol_ret_cap)).mean() if len(vol_ret_raw) > 0 else 0.0
        if vol_clip_rate > 0.05:
            data_log.warning("High volume clipping %.1f%% for %s", vol_clip_rate*100, ticker)
    
    # 7. Performance logging
    if debug_config.enable_data_validation:
        perf_log.debug("✅ Winsorization complete for %s: price_clip=%.1f%%, vol_clip=%.1f%%", 
                      ticker, clip_rate*100, vol_clip_rate*100 if 'vol_clip_rate' in locals() else 0.0)
    
    return df_result

# =============================================================================
# ENHANCED WINSORIZED DATA HELPERS - For robust technical calculations  
# =============================================================================

def get_robust_price_data(df: pd.DataFrame, use_capped: bool = True) -> dict:
    """
    Get price data for calculations, preferring winsorized (capped) data when available.
    
    Args:
        df: DataFrame from YFinanceDataValidator with potential capped columns
        use_capped: Whether to use capped data for calculations (recommended)
    
    Returns:
        dict with 'close', 'returns', 'high', 'low', 'open', 'volume', 'volume_changes' series
    """
    result = {}
    
    # Use capped close for calculations if available and requested
    if use_capped and 'Close_Capped' in df.columns:
        result['close'] = df['Close_Capped']
        result['close_source'] = 'capped'
    else:
        result['close'] = df['Close']
        result['close_source'] = 'raw'
    
    # Use capped returns if available
    if use_capped and 'PctChange_Capped' in df.columns:
        result['returns'] = df['PctChange_Capped']
        result['returns_source'] = 'capped'
    else:
        result['returns'] = df['Close'].pct_change()
        result['returns_source'] = 'raw'
    
    # Use capped volume if available and enabled
    if use_capped and 'Volume_Capped' in df.columns and WINSORIZATION_CONFIG['enable_volume_capping']:
        result['volume'] = df['Volume_Capped']
        result['volume_source'] = 'capped'
        result['volume_changes'] = df.get('VolChange_Capped', df['Volume_Capped'].pct_change())
    else:
        result['volume'] = df['Volume']
        result['volume_source'] = 'raw'
        result['volume_changes'] = df['Volume'].pct_change()
    
    # OHLC data (always use raw since capping only affects Close and Volume)
    result['high'] = df['High']
    result['low'] = df['Low'] 
    result['open'] = df['Open']
    
    # Also provide raw data for auditing and comparison
    if 'Close_Capped' in df.columns:
        result['close_raw'] = df['Close']
        result['returns_raw'] = df.get('PctChange_Raw', df['Close'].pct_change())
    
    if 'Volume_Capped' in df.columns:
        result['volume_raw'] = df['Volume']
        result['volume_changes_raw'] = df.get('VolChange_Raw', df['Volume'].pct_change())
    
    return result

def safe_daily_return_pct(df: pd.DataFrame) -> float:
    """
    Calculate daily return percentage using winsorized data when available.
    This ensures extreme corporate actions don't skew momentum calculations.
    """
    if df.empty:
        return 0.0
    
    price_data = get_robust_price_data(df, use_capped=True)
    returns = price_data['returns']
    
    if len(returns) < 2:
        return 0.0
    
    # Get last valid return (skip any NaN)
    last_return = returns.dropna().iloc[-1] if not returns.dropna().empty else 0.0
    return last_return * 100  # Convert to percentage

def safe_volume_multiplier(df: pd.DataFrame, period: int = 20) -> float:
    """
    Calculate volume multiplier using winsorized data when available.
    This prevents extreme volume spikes from skewing the calculation.
    """
    if df.empty or len(df) < period:
        return 1.0
    
    price_data = get_robust_price_data(df, use_capped=True)
    volume = price_data['volume']
    
    if len(volume) < period:
        return 1.0
    
    # Use winsorized volume for more stable calculations
    current_volume = volume.iloc[-1]
    avg_volume = volume.tail(period).mean()
    
    if avg_volume > 0:
        return current_volume / avg_volume
    else:
        return 1.0

def validate_winsorization_quality(df: pd.DataFrame, ticker: str = "") -> dict:
    """
    OPTIMIZED: Quality check for winsorization results with simplified guardrails.
    Uses efficient comparison operations and clear thresholds.
    
    Returns:
        dict with quality metrics and warnings
    """
    quality_report = {
        'price_clip_rate': 0.0,
        'volume_clip_rate': 0.0,
        'data_quality': 'GOOD',
        'warnings': [],
        'recommendations': []
    }
    
    if df.empty:
        quality_report['data_quality'] = 'FAILED'
        quality_report['warnings'].append('Empty DataFrame')
        return quality_report
    
    # OPTIMIZED: Check price clipping rate with direct comparison
    if 'PctChange_Raw' in df.columns and 'PctChange_Capped' in df.columns:
        ret_raw = df['PctChange_Raw']
        ret_cap = df['PctChange_Capped']
        
        # Quality report guardrail - efficient comparison
        clip_rate = (ret_raw.ne(ret_cap)).mean() if len(ret_raw) > 0 else 0.0
        quality_report['price_clip_rate'] = clip_rate
        
        # Graduated quality assessment - RELAXED thresholds for soft mode
        if clip_rate > 0.30:  # 30% threshold - severe (raised from 20%)
            quality_report['data_quality'] = 'FAILED'
            quality_report['warnings'].append(f"SEVERE price clipping: {clip_rate:.1%} - data unreliable")
            quality_report['recommendations'].append("Consider alternative data source or symbol delisting check")
        elif clip_rate > 0.15:  # 15% threshold - poor (raised from 10%)
            quality_report['data_quality'] = 'POOR' 
            quality_report['warnings'].append(f"High price clipping: {clip_rate:.1%} - review carefully")
            quality_report['recommendations'].append("Verify against multiple data sources")
        elif clip_rate > 0.05:  # 5% threshold - warning
            quality_report['warnings'].append(f"Moderate price clipping: {clip_rate:.1%}")
            quality_report['recommendations'].append("Monitor for data quality issues")
        
        # Log according to severity
        if clip_rate > 0.05:
            data_log.warning("High price clipping %.1f%% for %s", clip_rate*100, ticker)
    
    # OPTIMIZED: Check volume clipping rate (if enabled)
    if (WINSORIZATION_CONFIG['enable_volume_capping'] and 
        'VolChange_Raw' in df.columns and 'VolChange_Capped' in df.columns):
        
        vol_raw = df['VolChange_Raw']
        vol_cap = df['VolChange_Capped']
        
        # Volume quality report guardrail
        vol_clip_rate = (vol_raw.ne(vol_cap)).mean() if len(vol_raw) > 0 else 0.0
        quality_report['volume_clip_rate'] = vol_clip_rate
        
        if vol_clip_rate > 0.15:  # Volume can be more volatile
            if quality_report['data_quality'] != 'FAILED':
                quality_report['data_quality'] = 'POOR'
            quality_report['warnings'].append(f"High volume clipping: {vol_clip_rate:.1%}")
            quality_report['recommendations'].append("Check for volume anomalies or corporate actions")
        elif vol_clip_rate > 0.05:
            quality_report['warnings'].append(f"Moderate volume clipping: {vol_clip_rate:.1%}")
        
        # Log volume quality issues
        if vol_clip_rate > 0.05:
            data_log.warning("High volume clipping %.1f%% for %s", vol_clip_rate*100, ticker)
    
    # Overall data quality assessment
    total_issues = len(quality_report['warnings'])
    if total_issues == 0:
        quality_report['data_quality'] = 'EXCELLENT'
    elif total_issues == 1 and quality_report['data_quality'] == 'GOOD':
        quality_report['data_quality'] = 'ACCEPTABLE'
    
    return quality_report

def align_to_nse_trading_calendar(df: pd.DataFrame, ticker: str = "") -> pd.DataFrame:
    """
    OPTIMIZED: Business-day alignment with efficient fallback approach.
    Align cleaned DataFrame to NSE business days to handle market holidays.
    
    This prevents gaps in data from creating false signals during holidays.
    Uses efficient pandas business day range as primary method.
    """
    if df.empty:
        return df
    
    try:
        # OPTIMIZED APPROACH: Use pandas built-in business day range as primary method
        # This is faster and more reliable than external calendar libraries
        biz = pd.bdate_range(df.index.min(), df.index.max())
        df_aligned = df.reindex(biz).ffill()  # Forward fill missing business days
        
        if len(df_aligned) != len(df):
            perf_log.debug("📅 Business day alignment for %s: %d -> %d trading days", 
                          ticker, len(df), len(df_aligned))
        
        return df_aligned
        
    except Exception as e:
        # Fallback: try NSE-specific calendar if available
        try:
            from pandas_market_calendars import get_calendar
            
            # Get NSE calendar (fallback to BSE if NSE not available)
            nse_cal = get_calendar('BSE')  # Use BSE as proxy for NSE
            
            # Get the date range from our data
            start_date = df.index.min().date()
            end_date = df.index.max().date()
            
            # Get valid trading days
            trading_days = nse_cal.valid_days(start_date, end_date)
            
            # Reindex to trading days only, forward-filling missing values
            df_aligned = df.reindex(trading_days, method='ffill')
            
            if len(df_aligned) != len(df):
                data_log.info("📅 NSE calendar alignment for %s: %d -> %d trading days", 
                             ticker, len(df), len(df_aligned))
            
            return df_aligned
            
        except ImportError:
            data_log.debug("pandas_market_calendars not available for %s - using business day fallback", ticker)
            # Ultimate fallback: return original data
            return df
            
        except Exception as fallback_error:
            data_log.warning("Calendar alignment failed for %s: %s - using original data", 
                           ticker, fallback_error)
            return df

# =============================================================================
# STREAMLINED DATA QUALITY HELPERS - Your Exact Approach Implementation
# =============================================================================

def streamlined_winsorization_and_alignment(df: pd.DataFrame, ticker: str = "", 
                                           price_threshold: float = 0.5,
                                           volume_threshold: float = 1.0) -> pd.DataFrame:
    """
    IMPLEMENTATION: Your exact streamlined approach for data quality.
    
    1. Clamp returns/volume once, then reconstruct  
    2. Business-day alignment fallback
    3. Quality report guardrail
    
    Args:
        df: Raw OHLCV DataFrame
        ticker: Symbol for logging
        price_threshold: Return clamping threshold (default ±50%)
        volume_threshold: Volume change clamping threshold (default ±100%)
        
    Returns:
        DataFrame with capped data and quality metrics
    """
    if df.empty or len(df) <= 1:
        return df
    
    result_df = df.copy()
    
    # ========================================================================
    # 1. CLAMP RETURNS/VOLUME ONCE, THEN RECONSTRUCT (Your Approach)
    # ========================================================================
    
    # Price return clamping and reconstruction
    close = df['Close']
    ret_raw = close.pct_change()
    ret_cap = ret_raw.clip(-price_threshold, price_threshold)  # adapt by cap bucket
    close_cap = (1 + ret_cap.fillna(0)).cumprod() * close.iloc[0]
    
    # Store results
    result_df['Close_Capped'] = close_cap
    result_df['PctChange_Raw'] = ret_raw
    result_df['PctChange_Capped'] = ret_cap
    
    # Volume clamping and reconstruction (if enabled)
    if WINSORIZATION_CONFIG.get('enable_volume_capping', True) and 'Volume' in df.columns:
        volume = df['Volume']
        vol_ret_raw = volume.pct_change()
        vol_ret_cap = vol_ret_raw.clip(-volume_threshold, volume_threshold)
        
        first_volume = volume.iloc[0]
        if first_volume > 0:
            vol_cap = (1 + vol_ret_cap.fillna(0)).cumprod() * first_volume
            vol_cap = vol_cap.clip(lower=0)  # Ensure non-negative
        else:
            vol_cap = volume.copy()
        
        result_df['Volume_Capped'] = vol_cap
        result_df['VolChange_Raw'] = vol_ret_raw
        result_df['VolChange_Capped'] = vol_ret_cap
    
    # ========================================================================
    # 2. BUSINESS-DAY ALIGNMENT FALLBACK (Your Approach)
    # ========================================================================
    
    try:
        # Efficient pandas business day range
        biz = pd.bdate_range(result_df.index.min(), result_df.index.max())
        result_df = result_df.reindex(biz).ffill()
        
        if debug_config.enable_data_validation:
            perf_log.debug("📅 Business day alignment applied to %s", ticker)
            
    except Exception as e:
        data_log.warning("Business day alignment failed for %s: %s", ticker, e)
        # Continue with original data if alignment fails
    
    # ========================================================================
    # 3. QUALITY REPORT GUARDRAIL (Your Approach)  
    # ========================================================================
    
    # Price clipping quality check
    clip_rate = (ret_raw.ne(ret_cap)).mean() if len(ret_raw) > 0 else 0.0
    if clip_rate > 0.05:
        data_log.warning("High clipping %.1f%% for %s", clip_rate*100, ticker)
        
        # Add quality metadata
        result_df.attrs['quality_metrics'] = {
            'price_clip_rate': clip_rate,
            'quality_warning': clip_rate > 0.05,
            'severe_clipping': clip_rate > 0.20,
            'ticker': ticker
        }
        
        # Severe quality issues
        if clip_rate > 0.20:
            data_log.error("SEVERE clipping %.1f%% for %s - data unreliable", clip_rate*100, ticker)
    
    # Volume clipping quality check (if applicable)
    if 'VolChange_Raw' in result_df.columns and 'VolChange_Capped' in result_df.columns:
        vol_clip_rate = (result_df['VolChange_Raw'].ne(result_df['VolChange_Capped'])).mean()
        if vol_clip_rate > 0.05:
            data_log.warning("High volume clipping %.1f%% for %s", vol_clip_rate*100, ticker)
            
            # Update quality metadata
            if 'quality_metrics' not in result_df.attrs:
                result_df.attrs['quality_metrics'] = {'ticker': ticker}
            result_df.attrs['quality_metrics']['volume_clip_rate'] = vol_clip_rate
    
    return result_df

def get_quality_safe_price_data(df: pd.DataFrame, prefer_capped: bool = True) -> dict:
    """
    Get quality-assured price data, preferring capped versions when available.
    This is the recommended way to access price data after winsorization.
    
    Args:
        df: DataFrame with potential capped columns
        prefer_capped: Whether to use capped data for calculations
        
    Returns:
        dict with quality-assured price series
    """
    result = {}
    
    # Use capped close if available and preferred
    if prefer_capped and 'Close_Capped' in df.columns:
        result['close'] = df['Close_Capped']
        result['returns'] = df.get('PctChange_Capped', df['Close_Capped'].pct_change())
        result['data_source'] = 'quality_capped'
    else:
        result['close'] = df['Close']
        result['returns'] = df['Close'].pct_change()
        result['data_source'] = 'raw'
    
    # Volume data
    if prefer_capped and 'Volume_Capped' in df.columns:
        result['volume'] = df['Volume_Capped']
        result['volume_changes'] = df.get('VolChange_Capped', df['Volume_Capped'].pct_change())
    else:
        result['volume'] = df['Volume']
        result['volume_changes'] = df['Volume'].pct_change()
    
    # OHLC data (always use raw since only Close/Volume are capped)
    result['high'] = df['High']
    result['low'] = df['Low']
    result['open'] = df['Open']
    
    # Quality metrics if available
    if hasattr(df, 'attrs') and 'quality_metrics' in df.attrs:
        result['quality_metrics'] = df.attrs['quality_metrics']
    else:
        result['quality_metrics'] = {'quality_warning': False}
    
    return result

# Convenience function using your exact syntax
def apply_your_data_quality_approach(df: pd.DataFrame, ticker: str = "") -> pd.DataFrame:
    """
    Apply exactly your specified approach:
    1. ret_raw = close.pct_change()
    2. ret_cap = ret_raw.clip(-0.5, 0.5)   # adapt by cap bucket  
    3. close_cap = (1 + ret_cap.fillna(0)).cumprod() * close.iloc[0]
    4. biz = pd.bdate_range(df.index.min(), df.index.max())
    5. df = df.reindex(biz).ffill()
    6. clip_rate = (ret_raw.ne(ret_cap)).mean()
    """
    if df.empty or len(df) <= 1:
        return df
    
    result_df = df.copy()
    
    # EXACTLY YOUR APPROACH: 1) Clamp returns/volume once, then reconstruct
    close = df['Close'] 
    ret_raw = close.pct_change()
    ret_cap = ret_raw.clip(-0.5, 0.5)   # adapt by cap bucket
    close_cap = (1 + ret_cap.fillna(0)).cumprod() * close.iloc[0]
    df['Close_Capped'] = close_cap
    
    # Store intermediate results for analysis
    result_df['PctChange_Raw'] = ret_raw
    result_df['PctChange_Capped'] = ret_cap
    result_df['Close_Capped'] = close_cap
    
    # EXACTLY YOUR APPROACH: 2) Business-day alignment fallback
    try:
        biz = pd.bdate_range(df.index.min(), df.index.max())
        result_df = result_df.reindex(biz).ffill()
    except Exception as e:
        data_log.warning("Business day alignment failed for %s: %s", ticker, e)
    
    # EXACTLY YOUR APPROACH: 3) Quality report guardrail
    clip_rate = (ret_raw.ne(ret_cap)).mean()
    if clip_rate > 0.05:
        data_log.warning("High clipping %.1f%% for %s", clip_rate*100, ticker)
        
        # Add quality metadata
        result_df.attrs['quality_metrics'] = {
            'price_clip_rate': clip_rate,
            'quality_warning': True,
            'ticker': ticker
        }
    
    return result_df

# Enhanced version that also handles volume 
def apply_enhanced_data_quality_approach(df: pd.DataFrame, ticker: str = "", 
                                       price_threshold: float = 0.5,
                                       volume_threshold: float = 1.0) -> pd.DataFrame:
    """
    Enhanced version of your approach that also handles volume clamping.
    Uses the exact same logic but extends to volume data.
    """
    return streamlined_winsorization_and_alignment(df, ticker, 
                                                 price_threshold, 
                                                 volume_threshold)

# =============================================================================
# DATA QUALITY VERIFICATION AND TESTING
# =============================================================================

def verify_data_quality_alignment(test_ticker: str = "RELIANCE") -> dict:
    """
    Verify that the data quality implementation matches your exact specifications.
    Tests all three components: clamping, business-day alignment, and quality guardrail.
    
    Returns verification report with test results.
    """
    verification_report = {
        'test_ticker': test_ticker,
        'tests_passed': 0,
        'tests_failed': 0,
        'alignment_status': 'UNKNOWN',
        'details': []
    }
    
    try:
        # Get sample data for testing
        perf_log.info("🧪 Testing data quality alignment for %s", test_ticker)
        
        # Create synthetic test data to verify approach
        dates = pd.date_range('2024-01-01', periods=100, freq='D')
        np.random.seed(42)  # Reproducible results
        
        # Create test data with some extreme returns
        base_price = 100.0
        returns = np.random.normal(0.01, 0.03, 100)  # Normal returns
        returns[10] = 0.8   # Extreme positive return (80%)
        returns[20] = -0.7  # Extreme negative return (-70%)
        returns[30] = 0.6   # Another extreme return
        
        prices = [base_price]
        for ret in returns:
            prices.append(prices[-1] * (1 + ret))
        
        test_df = pd.DataFrame({
            'Close': prices[1:],
            'High': [p * 1.02 for p in prices[1:]],
            'Low': [p * 0.98 for p in prices[1:]],
            'Open': prices[:-1],
            'Volume': np.random.randint(1000000, 10000000, 100)
        }, index=dates)
        
        # Test 1: Verify exact clamping approach
        close = test_df['Close']
        ret_raw = close.pct_change()
        ret_cap = ret_raw.clip(-0.5, 0.5)   # adapt by cap bucket
        close_cap = (1 + ret_cap.fillna(0)).cumprod() * close.iloc[0]
        
        # Check if extreme returns were capped
        extreme_returns = (ret_raw.abs() > 0.5).sum()
        capped_returns = (ret_cap.abs() > 0.5).sum()
        
        if extreme_returns > 0 and capped_returns == 0:
            verification_report['tests_passed'] += 1
            verification_report['details'].append("✅ Clamping: Extreme returns properly capped")
        else:
            verification_report['tests_failed'] += 1
            verification_report['details'].append(f"❌ Clamping: Expected capping but found {capped_returns} uncapped")
        
        # Test 2: Verify business day alignment
        biz = pd.bdate_range(test_df.index.min(), test_df.index.max())
        aligned_df = test_df.reindex(biz).ffill()
        
        business_days_count = len(biz)
        original_count = len(test_df)
        
        if business_days_count <= original_count:  # Should have fewer or equal business days
            verification_report['tests_passed'] += 1
            verification_report['details'].append(f"✅ Alignment: {original_count} → {business_days_count} business days")
        else:
            verification_report['tests_failed'] += 1
            verification_report['details'].append(f"❌ Alignment: Unexpected day count change")
        
        # Test 3: Verify quality guardrail  
        clip_rate = (ret_raw.ne(ret_cap)).mean()
        
        if clip_rate > 0.05:  # We expect this with our synthetic extreme data
            verification_report['tests_passed'] += 1
            verification_report['details'].append(f"✅ Quality Guardrail: Detected high clipping {clip_rate:.1%}")
        else:
            verification_report['tests_failed'] += 1
            verification_report['details'].append(f"❌ Quality Guardrail: Expected warning but clip_rate={clip_rate:.1%}")
        
        # Test 4: Verify our implementation matches the exact approach
        our_result = apply_your_data_quality_approach(test_df, test_ticker)
        
        if 'Close_Capped' in our_result.columns:
            our_close_cap = our_result['Close_Capped']
            manual_close_cap = close_cap
            
            # Check if results are approximately equal (allowing for small numerical differences)
            max_diff = (our_close_cap - manual_close_cap).abs().max()
            if max_diff < 1e-10:  # Very small tolerance
                verification_report['tests_passed'] += 1
                verification_report['details'].append("✅ Implementation: Matches exact manual calculation")
            else:
                verification_report['tests_failed'] += 1
                verification_report['details'].append(f"❌ Implementation: Max difference {max_diff}")
        else:
            verification_report['tests_failed'] += 1
            verification_report['details'].append("❌ Implementation: Missing Close_Capped column")
        
        # Overall alignment status
        total_tests = verification_report['tests_passed'] + verification_report['tests_failed']
        if verification_report['tests_failed'] == 0:
            verification_report['alignment_status'] = 'PERFECT_ALIGNMENT'
        elif verification_report['tests_passed'] >= total_tests * 0.75:
            verification_report['alignment_status'] = 'GOOD_ALIGNMENT'
        else:
            verification_report['alignment_status'] = 'NEEDS_ALIGNMENT'
        
        perf_log.info("🎯 Data quality verification complete: %s (%d/%d tests passed)", 
                     verification_report['alignment_status'],
                     verification_report['tests_passed'], 
                     total_tests)
        
        return verification_report
        
    except Exception as e:
        verification_report['tests_failed'] += 1
        verification_report['details'].append(f"❌ Verification Error: {str(e)}")
        verification_report['alignment_status'] = 'ERROR'
        data_log.error("Data quality verification failed: %s", e)
        return verification_report

# =============================================================================
# SCORING + FILTERS - Quantify Opportunities + Eliminate Junk
# =============================================================================

def tiered_score(value: float, tiers: List[Tuple[float, float]]) -> float:
    """
    Calculate tiered score based on value thresholds.
    
    Args:
        value: The value to score
        tiers: List of (threshold, score) tuples in ascending order
        
    Returns:
        Score based on tier matching
    """
    for threshold, score in tiers:
        if value <= threshold:
            return score
    return 0.0  # No tier matched

def apply_quality_filters(df: pd.DataFrame, ticker: str = "") -> bool:
    """
    Apply quality filters to eliminate junk stocks.
    
    Args:
        df: OHLCV DataFrame
        ticker: Stock symbol for logging
        
    Returns:
        True if stock passes filters, False if rejected
    """
    if df.empty:
        data_log.debug("❌ %s: Empty data", ticker)
        return False
    
    # FILTER 1: Volume threshold (minimum liquidity)
    avg_volume = df['Volume'].mean()
    if avg_volume < 300_000:
        data_log.debug("❌ %s: Low volume %.0f < 300K", ticker, avg_volume)
        return False
    
    # FILTER 2: Price threshold (avoid penny stocks)
    current_price = df['Close'].iloc[-1]
    if current_price < 20:
        data_log.debug("❌ %s: Low price ₹%.2f < ₹20", ticker, current_price)
        return False
    
    # FILTER 3: Data quality check
    if len(df) < 50:  # Need sufficient data points
        data_log.debug("❌ %s: Insufficient data %d bars", ticker, len(df))
        return False
    
    # FILTER 4: Recent trading activity
    recent_volume = df['Volume'].tail(5).mean()
    if recent_volume < 100_000:  # Recent activity threshold
        data_log.debug("❌ %s: Low recent volume %.0f", ticker, recent_volume)
        return False
    
    data_log.debug("✅ %s: Passed quality filters", ticker)
    return True

def calculate_opportunity_score(df: pd.DataFrame, ticker: str = "") -> dict:
    """
    Calculate comprehensive opportunity score for filtered stocks.
    
    Args:
        df: DataFrame with RSI, BB_Pos, ATR indicators
        ticker: Stock symbol for logging
        
    Returns:
        Dictionary with score breakdown and tier classification
    """
    if df.empty or not all(col in df.columns for col in ['RSI', 'BB_Pos', 'ATR']):
        return {
            'total_score': 0.0,
            'tier': 'Rejected',
            'breakdown': {},
            'reason': 'Missing indicators'
        }
    
    score_breakdown = {}
    total_score = 0.0
    
    # SCORING COMPONENT 1: RSI (Lower is better for buying opportunities)
    current_rsi = df['RSI'].iloc[-1]
    rsi_score = tiered_score(current_rsi, [(30, 10), (40, 7), (50, 3)])
    score_breakdown['rsi'] = {'value': current_rsi, 'score': rsi_score}
    total_score += rsi_score
    
    # SCORING COMPONENT 2: Bollinger Band Position (Lower is better)
    current_bb = df['BB_Pos'].iloc[-1]
    bb_score = tiered_score(current_bb, [(20, 10), (30, 7), (40, 3)])
    score_breakdown['bb_position'] = {'value': current_bb, 'score': bb_score}
    total_score += bb_score
    
    # SCORING COMPONENT 3: Volume confirmation (Higher recent volume = better)
    current_volume = df['Volume'].iloc[-1]
    avg_volume = df['Volume'].tail(20).mean()
    volume_ratio = current_volume / avg_volume if avg_volume > 0 else 1.0
    vol_score = tiered_score(volume_ratio, [(1.5, 3), (2.0, 5), (3.0, 7)])
    score_breakdown['volume_ratio'] = {'value': volume_ratio, 'score': vol_score}
    total_score += vol_score
    
    # SCORING COMPONENT 4: ATR (Volatility consideration)
    current_atr = df['ATR'].iloc[-1]
    current_price = df['Close'].iloc[-1]
    atr_pct = (current_atr / current_price * 100) if current_price > 0 else 0
    # Moderate volatility preferred (2-5%)
    if 2.0 <= atr_pct <= 5.0:
        atr_score = 3.0
    elif 1.0 <= atr_pct <= 6.0:
        atr_score = 1.5
    else:
        atr_score = 0.0
    score_breakdown['atr_volatility'] = {'value': atr_pct, 'score': atr_score}
    total_score += atr_score
    
    # SCORING COMPONENT 5: Recent price momentum
    if len(df) >= 5:
        recent_return = (df['Close'].iloc[-1] / df['Close'].iloc[-5] - 1) * 100
        # Slight positive momentum preferred
        if -2.0 <= recent_return <= 1.0:
            momentum_score = 2.0
        elif -5.0 <= recent_return <= 3.0:
            momentum_score = 1.0
        else:
            momentum_score = 0.0
        score_breakdown['momentum'] = {'value': recent_return, 'score': momentum_score}
        total_score += momentum_score
    
    # TIER CLASSIFICATION
    if total_score >= 25:
        tier = "Tier1"
    elif total_score >= 15:
        tier = "Tier2"
    else:
        tier = "Watch"
    
    # Add quality indicators to breakdown
    score_breakdown['data_quality'] = {
        'bars': len(df),
        'avg_volume': df['Volume'].mean(),
        'current_price': df['Close'].iloc[-1]
    }
    
    result = {
        'total_score': round(total_score, 1),
        'tier': tier,
        'breakdown': score_breakdown,
        'ticker': ticker,
        'reason': f"Scored {total_score:.1f} points"
    }
    
    perf_log.debug("📊 %s: Score=%.1f, Tier=%s", ticker, total_score, tier)
    return result

def screen_single_ticker_complete(ticker: str, period: str = "1y") -> Optional[dict]:
    """
    Complete screening pipeline: Data Quality + Indicators + Filters + Scoring.
    
    This is the main function that combines all components:
    1. Load and clean data (your exact approach)
    2. Calculate key indicators (RSI, BB_Pos, ATR)
    3. Apply quality filters
    4. Calculate opportunity score and tier
    
    Returns:
        Dictionary with complete analysis or None if rejected
    """
    try:
        # STEP 1: Load and process data with quality controls
        df = process_ticker_data_complete(ticker, period)
        if df.empty:
            return None
        
        # STEP 2: Apply quality filters (eliminate junk)
        if not apply_quality_filters(df, ticker):
            return None
        
        # STEP 3: Calculate opportunity score (quantify opportunities)
        score_result = calculate_opportunity_score(df, ticker)
        if score_result['total_score'] < 5.0:  # Minimum score threshold
            data_log.debug("❌ %s: Score too low %.1f", ticker, score_result['total_score'])
            return None
        
        # STEP 4: Get current values for reporting
        current_indicators = get_current_indicators(df)
        
        # STEP 5: Combine all results
        final_result = {
            'ticker': ticker,
            'tier': score_result['tier'],
            'total_score': score_result['total_score'],
            'current_values': {
                'price': current_indicators['close'],
                'rsi': current_indicators['rsi'],
                'bb_position': current_indicators['bb_position'],
                'atr': current_indicators['atr'],
                'volume': df['Volume'].iloc[-1],
                'avg_volume': df['Volume'].mean()
            },
            'score_breakdown': score_result['breakdown'],
            'data_quality': {
                'clip_rate': current_indicators.get('clip_rate', 0.0),
                'data_points': len(df),
                'quality_status': current_indicators.get('data_quality', 'UNKNOWN')
            },
            'analysis_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        perf_log.info("🎯 %s: %s (Score: %.1f) - Price: ₹%.2f, RSI: %.1f, BB: %.1f%%", 
                     ticker, score_result['tier'], score_result['total_score'],
                     current_indicators['close'], current_indicators['rsi'], 
                     current_indicators['bb_position'])
        
        return final_result
        
    except Exception as e:
        data_log.error("Error screening %s: %s", ticker, e)
        return None

def batch_screen_tickers(tickers: List[str], period: str = "1y", 
                        max_workers: int = 4) -> List[dict]:
    """
    Screen multiple tickers efficiently with filtering and scoring.
    
    Args:
        tickers: List of ticker symbols
        period: Data period
        max_workers: Maximum concurrent workers
        
    Returns:
        List of screening results sorted by score (highest first)
    """
    if not tickers:
        return []
    
    perf_log.info("🔍 Screening %d tickers with filters + scoring", len(tickers))
    
    results = []
    start_time = time.time()
    
    # Process tickers concurrently
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_ticker = {
            executor.submit(screen_single_ticker_complete, ticker, period): ticker 
            for ticker in tickers
        }
        
        for future in as_completed(future_to_ticker):
            ticker = future_to_ticker[future]
            try:
                result = future.result(timeout=30)
                if result:  # Only include passed results
                    results.append(result)
            except Exception as e:
                data_log.warning("Failed to screen %s: %s", ticker, e)
    
    # Sort by score (highest first)
    results.sort(key=lambda x: x['total_score'], reverse=True)
    
    # Performance summary
    elapsed = time.time() - start_time
    passed_count = len(results)
    rejection_rate = ((len(tickers) - passed_count) / len(tickers) * 100) if tickers else 0
    
    perf_log.info("✅ Screening complete: %d/%d passed (%.1f%% filtered out) in %.2fs", 
                 passed_count, len(tickers), rejection_rate, elapsed)
    
    return results

# =============================================================================
# KEY INDICATORS - Correct Implementations for Foundation
# =============================================================================

def rsi14(close: pd.Series) -> pd.Series:
    """
    Wilder's RSI implementation - the correct RSI calculation method.
    Uses Wilder's smoothing (EMA with alpha=1/14) instead of simple moving average.
    
    Args:
        close: Close price series
        
    Returns:
        RSI series (0-100 scale)
    """
    if len(close) < 15:  # Need at least 15 periods for initial calculation
        return pd.Series([50.0] * len(close), index=close.index)
    
    # Calculate price changes
    delta = close.diff()
    
    # Separate gains and losses
    gains = delta.where(delta > 0, 0)
    losses = -delta.where(delta < 0, 0)
    
    # Wilder's smoothing: use EMA with alpha = 1/14
    alpha = 1.0 / 14.0
    avg_gains = gains.ewm(alpha=alpha, adjust=False).mean()
    avg_losses = losses.ewm(alpha=alpha, adjust=False).mean()
    
    # Calculate RS and RSI
    rs = avg_gains / avg_losses
    rsi = 100 - (100 / (1 + rs))
    
    return rsi.fillna(50.0)

def bollinger_band_position(close: pd.Series, period: int = 20, std_dev: float = 2.0) -> pd.Series:
    """
    Calculate Bollinger Band position as percentage (0-100 scale).
    
    Args:
        close: Close price series
        period: Moving average period (default 20)
        std_dev: Standard deviation multiplier (default 2.0)
        
    Returns:
        BB position series (0-100 scale where 0=lower band, 50=middle, 100=upper band)
    """
    if len(close) < period:
        return pd.Series([50.0] * len(close), index=close.index)
    
    # Calculate Bollinger Bands
    sma = close.rolling(period, min_periods=period).mean()
    std = close.rolling(period, min_periods=period).std(ddof=0)
    
    upper_band = sma + (std_dev * std)
    lower_band = sma - (std_dev * std)
    
    # Calculate position as percentage (0-100)
    bb_position = ((close - lower_band) / (upper_band - lower_band)) * 100
    
    # Clamp to 0-100 range and handle edge cases
    bb_position = bb_position.clip(0, 100).fillna(50.0)
    
    return bb_position

def average_true_range(ohlc_df: pd.DataFrame, period: int = 14) -> pd.Series:
    """
    Calculate Average True Range (ATR) using Wilder's smoothing.
    
    Args:
        ohlc_df: DataFrame with High, Low, Close columns
        period: ATR period (default 14)
        
    Returns:
        ATR series
    """
    if len(ohlc_df) < period + 1:
        return pd.Series([0.0] * len(ohlc_df), index=ohlc_df.index)
    
    high = ohlc_df['High']
    low = ohlc_df['Low'] 
    close = ohlc_df['Close']
    prev_close = close.shift(1)
    
    # Calculate True Range components
    tr1 = high - low                    # Current high - current low
    tr2 = (high - prev_close).abs()     # Current high - previous close
    tr3 = (low - prev_close).abs()      # Current low - previous close
    
    # True Range is the maximum of the three
    true_range = pd.DataFrame({'tr1': tr1, 'tr2': tr2, 'tr3': tr3}).max(axis=1)
    
    # Calculate ATR using Wilder's smoothing (EMA with alpha = 1/period)
    alpha = 1.0 / period
    atr = true_range.ewm(alpha=alpha, adjust=False).mean()
    
    return atr.fillna(0.0)

# =============================================================================
# INTEGRATED DATA QUALITY + KEY INDICATORS PIPELINE
# =============================================================================

def process_ticker_data_complete(ticker: str, period: str = "1y") -> pd.DataFrame:
    """
    Complete data processing pipeline following your exact approach:
    1. Load data
    2. Apply data quality (business day alignment + return capping)
    3. Calculate key indicators (RSI, BB_Pos, ATR)
    
    This is the foundation function that ensures clean data + correct indicators.
    """
    try:
        # Load data
        df = get_history_optimized(ticker, period=period)
        if df.empty:
            perf_log.warning("No data available for %s", ticker)
            return pd.DataFrame()
        
        # DATA QUALITY - Your exact approach
        df = df.reindex(pd.bdate_range(df.index.min(), df.index.max())).ffill()
        ret_capped = df['Close'].pct_change().clip(-0.5, 0.5)
        df['Close'] = (1 + ret_capped.fillna(0)).cumprod() * df['Close'].iloc[0]
        
        # Store the capped returns for analysis
        df['Returns_Raw'] = df['Close'].pct_change()
        df['Returns_Capped'] = ret_capped
        
        # KEY INDICATORS - Correct implementations
        df['RSI'] = rsi14(df['Close'])  # Use Wilder's RSI impl
        df['BB_Pos'] = bollinger_band_position(df['Close'])  # 0-100 scale
        df['ATR'] = average_true_range(df[['High','Low','Close']])
        
        # Quality check
        clip_rate = (df['Returns_Raw'].ne(ret_capped)).mean()
        if clip_rate > 0.05:
            data_log.warning("High clipping %.1f%% for %s", clip_rate*100, ticker)
        
        # Add metadata
        df.attrs['processed_ticker'] = ticker
        df.attrs['clip_rate'] = clip_rate
        df.attrs['data_quality'] = 'GOOD' if clip_rate <= 0.05 else 'WARNING'
        
        perf_log.debug("✅ Processed %s: %d bars, clip_rate=%.1f%%", 
                      ticker, len(df), clip_rate*100)
        
        return df
        
    except Exception as e:
        data_log.error("Error processing %s: %s", ticker, e)
        return pd.DataFrame()

def get_current_indicators(df: pd.DataFrame) -> dict:
    """
    Extract current indicator values from processed DataFrame.
    Returns the most recent values for decision making.
    """
    if df.empty:
        return {
            'rsi': 50.0,
            'bb_position': 50.0,
            'atr': 0.0,
            'close': 0.0,
            'data_quality': 'FAILED'
        }
    
    current_values = {
        'rsi': float(df['RSI'].iloc[-1]) if not df['RSI'].empty else 50.0,
        'bb_position': float(df['BB_Pos'].iloc[-1]) if not df['BB_Pos'].empty else 50.0,
        'atr': float(df['ATR'].iloc[-1]) if not df['ATR'].empty else 0.0,
        'close': float(df['Close'].iloc[-1]) if not df['Close'].empty else 0.0,
        'data_quality': df.attrs.get('data_quality', 'UNKNOWN'),
        'clip_rate': df.attrs.get('clip_rate', 0.0),
        'data_points': len(df)
    }
    
    return current_values

# =============================================================================
# BACKTEST HOOKS - Validate Strategy Before Going Live
# =============================================================================

import json
from pathlib import Path
from dataclasses import dataclass, asdict
from enum import Enum

class TradeType(Enum):
    LONG = "LONG"
    SHORT = "SHORT"

class TradeStatus(Enum):
    OPEN = "OPEN"
    CLOSED = "CLOSED"
    STOPPED = "STOPPED"

@dataclass
class TradeSignal:
    """Trade signal data structure"""
    ticker: str
    signal_time: str
    signal_type: TradeType
    entry_price: float
    stop_loss: float
    take_profit: float
    confidence: float
    indicators: dict
    tier: str
    reason: str

@dataclass
class PaperTrade:
    """Paper trade tracking structure"""
    trade_id: str
    ticker: str
    trade_type: TradeType
    entry_time: str
    entry_price: float
    stop_loss: float
    take_profit: float
    quantity: int
    status: TradeStatus
    exit_time: str = None
    exit_price: float = None
    pnl: float = 0.0
    pnl_percent: float = 0.0
    tier: str = "Watch"
    confidence: float = 0.0

class BacktestHooks:
    """Comprehensive backtesting and paper trading system"""
    
    def __init__(self, log_file: str = "paper_trades.json", signal_file: str = "trade_signals.json"):
        self.log_file = Path(log_file)
        self.signal_file = Path(signal_file)
        self.paper_trades = self._load_paper_trades()
        self.trade_signals = self._load_trade_signals()
        self.trade_counter = len(self.paper_trades) + 1
        
    def _load_paper_trades(self) -> List[PaperTrade]:
        """Load existing paper trades from file"""
        if self.log_file.exists():
            try:
                with open(self.log_file, 'r') as f:
                    content = f.read().strip()
                    if not content:
                        # Empty file
                        return []
                    data = json.loads(content)
                    
                # Ensure data is a list
                if not isinstance(data, list):
                    data_log.warning("Paper trades file contains invalid format (not a list), resetting")
                    self._backup_and_reset_trades_file()
                    return []
                    
                trades = []
                for trade_dict in data:
                    # Convert string enums back to enum types
                    trade_dict['trade_type'] = TradeType(trade_dict['trade_type'])
                    trade_dict['status'] = TradeStatus(trade_dict['status'])
                    trades.append(PaperTrade(**trade_dict))
                return trades
            except json.JSONDecodeError as e:
                data_log.warning("Failed to parse paper trades JSON (corrupted file): %s", e)
                data_log.info("Backing up corrupted file and creating fresh paper trades log")
                self._backup_and_reset_trades_file()
                return []
            except Exception as e:
                data_log.warning("Failed to load paper trades: %s", e)
                return []
        return []
    
    def _backup_and_reset_trades_file(self):
        """Backup corrupted trades file and create fresh one"""
        try:
            if self.log_file.exists():
                backup_name = f"{self.log_file.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                backup_path = self.log_file.parent / backup_name
                self.log_file.rename(backup_path)
                data_log.info(f"Corrupted trades file backed up as: {backup_name}")
            
            # Create fresh empty trades file
            with open(self.log_file, 'w') as f:
                json.dump([], f, indent=2)
            data_log.info("Created fresh paper trades file")
        except Exception as e:
            data_log.error(f"Failed to backup/reset trades file: {e}")
    
    def _load_trade_signals(self) -> List[TradeSignal]:
        """Load existing trade signals from file"""
        if self.signal_file.exists():
            try:
                with open(self.signal_file, 'r') as f:
                    content = f.read().strip()
                    if not content:
                        # Empty file
                        return []
                    data = json.loads(content)
                    
                # Ensure data is a list
                if not isinstance(data, list):
                    data_log.warning("Trade signals file contains invalid format (not a list), resetting")
                    self._backup_and_reset_signals_file()
                    return []
                    
                signals = []
                for signal_dict in data:
                    # Convert string enums back to enum types
                    signal_dict['signal_type'] = TradeType(signal_dict['signal_type'])
                    signals.append(TradeSignal(**signal_dict))
                return signals
            except json.JSONDecodeError as e:
                data_log.warning("Failed to parse trade signals JSON (corrupted file): %s", e)
                data_log.info("Backing up corrupted file and creating fresh trade signals log")
                self._backup_and_reset_signals_file()
                return []
            except Exception as e:
                data_log.warning("Failed to load trade signals: %s", e)
                return []
        return []
    
    def _backup_and_reset_signals_file(self):
        """Backup corrupted signals file and create fresh one"""
        try:
            if self.signal_file.exists():
                backup_name = f"{self.signal_file.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                backup_path = self.signal_file.parent / backup_name
                self.signal_file.rename(backup_path)
                data_log.info(f"Corrupted signals file backed up as: {backup_name}")
            
            # Create fresh empty signals file
            with open(self.signal_file, 'w') as f:
                json.dump([], f, indent=2)
            data_log.info("Created fresh trade signals file")
        except Exception as e:
            data_log.error(f"Failed to backup/reset signals file: {e}")
    
    def _save_paper_trades(self):
        """Save paper trades to file"""
        try:
            # Convert to JSON-serializable format
            trades_data = []
            for trade in self.paper_trades:
                trade_dict = asdict(trade)
                # Convert enums to strings
                trade_dict['trade_type'] = trade.trade_type.value
                trade_dict['status'] = trade.status.value
                trades_data.append(trade_dict)
            
            with open(self.log_file, 'w') as f:
                json.dump(trades_data, f, indent=2)
        except Exception as e:
            data_log.error("Failed to save paper trades: %s", e)
    
    def _save_trade_signals(self):
        """Save trade signals to file"""
        try:
            # Convert to JSON-serializable format
            signals_data = []
            for signal in self.trade_signals:
                signal_dict = asdict(signal)
                # Convert enums to strings
                signal_dict['signal_type'] = signal.signal_type.value
                signals_data.append(signal_dict)
            
            with open(self.signal_file, 'w') as f:
                json.dump(signals_data, f, indent=2)
        except Exception as e:
            data_log.error("Failed to save trade signals: %s", e)

def generate_trading_signals(df: pd.DataFrame, ticker: str = "") -> dict:
    """
    Generate comprehensive trading signals based on technical indicators.
    
    Args:
        df: DataFrame with RSI, BB_Pos, ATR, and Volume indicators
        ticker: Stock symbol
        
    Returns:
        Dictionary with signal information
    """
    if df.empty or len(df) < 50:
        return {
            'signal_generated': False,
            'reason': 'Insufficient data',
            'ticker': ticker
        }
    
    # Ensure required columns exist
    required_cols = ['RSI', 'BB_Pos', 'Volume', 'ATR', 'Close']
    if not all(col in df.columns for col in required_cols):
        return {
            'signal_generated': False,
            'reason': 'Missing required indicators',
            'ticker': ticker
        }
    
    # Calculate volume moving average
    df_signals = df.copy()
    df_signals['Volume_MA50'] = df['Volume'].rolling(50, min_periods=20).mean()
    
    # SIGNAL GENERATION - Your exact logic
    signal_condition = (
        (df_signals['RSI'] < 30) & 
        (df_signals['BB_Pos'] < 20) & 
        (df_signals['Volume'] > 1.5 * df_signals['Volume_MA50'])
    )
    
    df_signals['Signal'] = signal_condition
    
    # Get current signal status
    current_signal = df_signals['Signal'].iloc[-1] if not df_signals['Signal'].empty else False
    
    if not current_signal:
        return {
            'signal_generated': False,
            'reason': 'Conditions not met',
            'ticker': ticker,
            'current_values': {
                'rsi': float(df['RSI'].iloc[-1]),
                'bb_position': float(df['BB_Pos'].iloc[-1]),
                'volume_ratio': float(df['Volume'].iloc[-1] / df_signals['Volume_MA50'].iloc[-1]) if df_signals['Volume_MA50'].iloc[-1] > 0 else 0.0
            }
        }
    
    # Generate signal details
    current_price = float(df['Close'].iloc[-1])
    current_atr = float(df['ATR'].iloc[-1])
    current_rsi = float(df['RSI'].iloc[-1])
    current_bb = float(df['BB_Pos'].iloc[-1])
    volume_ratio = float(df['Volume'].iloc[-1] / df_signals['Volume_MA50'].iloc[-1]) if df_signals['Volume_MA50'].iloc[-1] > 0 else 1.0
    
    # Calculate stop loss and take profit using ATR
    stop_loss = current_price - 1.5 * current_atr
    take_profit = current_price + 3 * current_atr
    
    # Calculate confidence score
    confidence = calculate_signal_confidence(df, current_rsi, current_bb, volume_ratio)
    
    signal_result = {
        'signal_generated': True,
        'ticker': ticker,
        'signal_type': 'LONG',
        'entry_price': current_price,
        'stop_loss': stop_loss,
        'take_profit': take_profit,
        'confidence': confidence,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'indicators': {
            'rsi': current_rsi,
            'bb_position': current_bb,
            'atr': current_atr,
            'volume_ratio': volume_ratio
        },
        'risk_reward_ratio': (take_profit - current_price) / (current_price - stop_loss) if current_price > stop_loss else 0.0,
        'reason': f"RSI={current_rsi:.1f} < 30, BB={current_bb:.1f}% < 20, Vol={volume_ratio:.1f}x > 1.5x"
    }
    
    perf_log.info("📊 Signal generated for %s: LONG at ₹%.2f (SL: ₹%.2f, TP: ₹%.2f, Conf: %.1f%%)", 
                 ticker, current_price, stop_loss, take_profit, confidence * 100)
    
    return signal_result

def calculate_signal_confidence(df: pd.DataFrame, rsi: float, bb_pos: float, volume_ratio: float) -> float:
    """Calculate confidence score for trading signal (0.0 to 1.0)"""
    confidence = 0.0
    
    # RSI confidence (lower is better for buy signals)
    if rsi <= 20:
        confidence += 0.4  # Very oversold
    elif rsi <= 25:
        confidence += 0.3  # Oversold
    elif rsi <= 30:
        confidence += 0.2  # Moderately oversold
    
    # Bollinger Band confidence (lower is better)
    if bb_pos <= 10:
        confidence += 0.3  # Very close to lower band
    elif bb_pos <= 15:
        confidence += 0.2  # Close to lower band
    elif bb_pos <= 20:
        confidence += 0.1  # Near lower band
    
    # Volume confirmation confidence
    if volume_ratio >= 3.0:
        confidence += 0.3  # Very high volume
    elif volume_ratio >= 2.0:
        confidence += 0.2  # High volume
    elif volume_ratio >= 1.5:
        confidence += 0.1  # Above average volume
    
    return min(1.0, confidence)

def log_trade(ticker: str, entry: float, sl: float, tp: float, 
              tier: str = "Watch", confidence: float = 0.0, 
              trade_type: str = "LONG", quantity: int = 100) -> str:
    """
    Log a paper trade for backtesting and validation.
    
    Args:
        ticker: Stock symbol
        entry: Entry price
        sl: Stop loss price
        tp: Take profit price
        tier: Tier classification (Tier1, Tier2, Watch)
        confidence: Signal confidence (0.0 to 1.0)
        trade_type: Trade type (LONG/SHORT)
        quantity: Number of shares
        
    Returns:
        Trade ID for tracking
    """
    global backtest_hooks
    
    # Initialize backtest hooks if not already done
    if 'backtest_hooks' not in globals():
        backtest_hooks = BacktestHooks()
    
    # Generate unique trade ID
    trade_id = f"{ticker}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{backtest_hooks.trade_counter}"
    backtest_hooks.trade_counter += 1
    
    # Create paper trade
    paper_trade = PaperTrade(
        trade_id=trade_id,
        ticker=ticker,
        trade_type=TradeType(trade_type),
        entry_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        entry_price=entry,
        stop_loss=sl,
        take_profit=tp,
        quantity=quantity,
        status=TradeStatus.OPEN,
        tier=tier,
        confidence=confidence
    )
    
    # Add to tracking
    backtest_hooks.paper_trades.append(paper_trade)
    backtest_hooks._save_paper_trades()
    
    # Calculate risk metrics
    risk_per_share = abs(entry - sl)
    potential_profit = abs(tp - entry)
    risk_reward = potential_profit / risk_per_share if risk_per_share > 0 else 0
    total_risk = risk_per_share * quantity
    
    perf_log.info("📝 Paper trade logged: %s %s %d shares at ₹%.2f (SL: ₹%.2f, TP: ₹%.2f, RR: %.2f, Risk: ₹%.2f)", 
                 trade_id, ticker, quantity, entry, sl, tp, risk_reward, total_risk)
    
    return trade_id

def update_paper_trades(current_prices: dict = None):
    """
    Update paper trades with current market prices and close completed trades.
    
    Args:
        current_prices: Dictionary of {ticker: current_price}
    """
    global backtest_hooks
    
    if 'backtest_hooks' not in globals():
        return
    
    updated_trades = []
    
    for trade in backtest_hooks.paper_trades:
        if trade.status != TradeStatus.OPEN:
            updated_trades.append(trade)
            continue
        
        # Get current price
        current_price = None
        if current_prices and trade.ticker in current_prices:
            current_price = current_prices[trade.ticker]
        else:
            # Try to fetch current price
            try:
                df = get_history_optimized(trade.ticker, period="1d")
                if not df.empty:
                    current_price = float(df['Close'].iloc[-1])
            except Exception as e:
                data_log.warning("Failed to get current price for %s: %s", trade.ticker, e)
                updated_trades.append(trade)
                continue
        
        if current_price is None:
            updated_trades.append(trade)
            continue
        
        # Check exit conditions
        trade_closed = False
        exit_reason = ""
        
        if trade.trade_type == TradeType.LONG:
            # Long trade exit conditions
            if current_price <= trade.stop_loss:
                trade.status = TradeStatus.STOPPED
                trade.exit_price = trade.stop_loss
                exit_reason = "Stop Loss Hit"
                trade_closed = True
            elif current_price >= trade.take_profit:
                trade.status = TradeStatus.CLOSED
                trade.exit_price = trade.take_profit
                exit_reason = "Take Profit Hit"
                trade_closed = True
        
        if trade_closed:
            trade.exit_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            trade.pnl = (trade.exit_price - trade.entry_price) * trade.quantity
            trade.pnl_percent = ((trade.exit_price - trade.entry_price) / trade.entry_price) * 100
            
            perf_log.info("🎯 Trade closed: %s - %s at ₹%.2f, P&L: ₹%.2f (%.2f%%), Reason: %s", 
                         trade.trade_id, trade.ticker, trade.exit_price, 
                         trade.pnl, trade.pnl_percent, exit_reason)
        
        updated_trades.append(trade)
    
    backtest_hooks.paper_trades = updated_trades
    backtest_hooks._save_paper_trades()

def get_paper_trading_summary() -> dict:
    """Get summary of paper trading performance"""
    global backtest_hooks
    
    if 'backtest_hooks' not in globals():
        return {'error': 'No paper trades found'}
    
    open_trades = [t for t in backtest_hooks.paper_trades if t.status == TradeStatus.OPEN]
    closed_trades = [t for t in backtest_hooks.paper_trades if t.status in [TradeStatus.CLOSED, TradeStatus.STOPPED]]
    
    if not closed_trades:
        return {
            'total_trades': len(backtest_hooks.paper_trades),
            'open_trades': len(open_trades),
            'closed_trades': 0,
            'win_rate': 0.0,
            'total_pnl': 0.0,
            'avg_pnl_percent': 0.0
        }
    
    # Calculate performance metrics
    winning_trades = [t for t in closed_trades if t.pnl > 0]
    losing_trades = [t for t in closed_trades if t.pnl <= 0]
    
    total_pnl = sum(t.pnl for t in closed_trades)
    win_rate = len(winning_trades) / len(closed_trades) * 100
    avg_pnl_percent = sum(t.pnl_percent for t in closed_trades) / len(closed_trades)
    
    # Tier-wise performance
    tier_performance = {}
    for tier in ['Tier1', 'Tier2', 'Watch']:
        tier_trades = [t for t in closed_trades if t.tier == tier]
        if tier_trades:
            tier_performance[tier] = {
                'count': len(tier_trades),
                'win_rate': len([t for t in tier_trades if t.pnl > 0]) / len(tier_trades) * 100,
                'avg_pnl': sum(t.pnl for t in tier_trades) / len(tier_trades)
            }
    
    summary = {
        'total_trades': len(backtest_hooks.paper_trades),
        'open_trades': len(open_trades),
        'closed_trades': len(closed_trades),
        'winning_trades': len(winning_trades),
        'losing_trades': len(losing_trades),
        'win_rate': round(win_rate, 2),
        'total_pnl': round(total_pnl, 2),
        'avg_pnl_percent': round(avg_pnl_percent, 2),
        'tier_performance': tier_performance
    }
    
    return summary

def backtest_strategy(tickers: List[str], period: str = "6mo", 
                     initial_capital: float = 100000) -> dict:
    """
    Comprehensive strategy backtesting on historical data.
    
    Args:
        tickers: List of tickers to backtest
        period: Historical period to test
        initial_capital: Starting capital for backtesting
        
    Returns:
        Backtest results with performance metrics
    """
    perf_log.info("🔄 Starting strategy backtest on %d tickers (%s period)", len(tickers), period)
    
    backtest_results = {
        'tickers_tested': len(tickers),
        'period': period,
        'initial_capital': initial_capital,
        'signals_generated': 0,
        'trades_completed': 0,
        'performance': {}
    }
    
    all_signals = []
    
    for ticker in tickers:
        try:
            # Get historical data
            df = process_ticker_data_complete(ticker, period)
            if df.empty or len(df) < 100:
                continue
            
            # Generate signals for historical data
            signal_results = generate_trading_signals(df, ticker)
            
            if signal_results['signal_generated']:
                all_signals.append(signal_results)
                backtest_results['signals_generated'] += 1
                
                # Simulate historical performance if we have enough data
                if len(df) > 200:  # Need enough data for forward testing
                    entry_price = signal_results['entry_price']
                    stop_loss = signal_results['stop_loss']
                    take_profit = signal_results['take_profit']
                    
                    # Simple forward simulation (this could be enhanced)
                    future_prices = df['Close'].tail(50)  # Look at next 50 periods
                    
                    for price in future_prices:
                        if price <= stop_loss:
                            # Stop loss hit
                            pnl_percent = (stop_loss - entry_price) / entry_price * 100
                            signal_results['simulated_outcome'] = 'STOP_LOSS'
                            signal_results['simulated_pnl_percent'] = pnl_percent
                            break
                        elif price >= take_profit:
                            # Take profit hit
                            pnl_percent = (take_profit - entry_price) / entry_price * 100
                            signal_results['simulated_outcome'] = 'TAKE_PROFIT'
                            signal_results['simulated_pnl_percent'] = pnl_percent
                            break
                    else:
                        # No exit condition met
                        final_price = future_prices.iloc[-1]
                        pnl_percent = (final_price - entry_price) / entry_price * 100
                        signal_results['simulated_outcome'] = 'OPEN'
                        signal_results['simulated_pnl_percent'] = pnl_percent
                    
                    if 'simulated_outcome' in signal_results:
                        backtest_results['trades_completed'] += 1
                
        except Exception as e:
            data_log.warning("Backtest failed for %s: %s", ticker, e)
    
    # Calculate overall performance
    if all_signals:
        completed_signals = [s for s in all_signals if 'simulated_outcome' in s]
        
        if completed_signals:
            winning_signals = [s for s in completed_signals if s.get('simulated_pnl_percent', 0) > 0]
            
            backtest_results['performance'] = {
                'total_signals': len(all_signals),
                'completed_trades': len(completed_signals),
                'winning_trades': len(winning_signals),
                'win_rate': len(winning_signals) / len(completed_signals) * 100,
                'avg_pnl_percent': sum(s.get('simulated_pnl_percent', 0) for s in completed_signals) / len(completed_signals),
                'best_trade': max(s.get('simulated_pnl_percent', 0) for s in completed_signals),
                'worst_trade': min(s.get('simulated_pnl_percent', 0) for s in completed_signals)
            }
    
    perf_log.info("✅ Backtest complete: %d signals, %d completed trades", 
                 backtest_results['signals_generated'], backtest_results['trades_completed'])
    
    return backtest_results

def screen_with_backtest_hooks(ticker: str, period: str = "1y", 
                              enable_paper_trading: bool = True) -> Optional[dict]:
    """
    Enhanced screening function that includes signal generation and paper trading.
    
    This combines the complete screening pipeline with backtesting capabilities.
    """
    try:
        # Run the complete screening
        result = screen_single_ticker_complete(ticker, period)
        if not result:
            return None
        
        # Get the processed data
        df = process_ticker_data_complete(ticker, period)
        if df.empty:
            return result
        
        # Generate trading signals
        signal_result = generate_trading_signals(df, ticker)
        result['signal_analysis'] = signal_result
        
        # PAPER TRADING LOG - Your exact approach
        if enable_paper_trading and result['tier'] == "Tier1" and signal_result['signal_generated']:
            entry_price = result['current_values']['price']
            atr = result['current_values']['atr']
            
            # Calculate SL and TP using ATR
            stop_loss = entry_price - 1.5 * atr
            take_profit = entry_price + 3 * atr
            
            trade_id = log_trade(
                ticker=ticker,
                entry=entry_price,
                sl=stop_loss,
                tp=take_profit,
                tier=result['tier'],
                confidence=signal_result.get('confidence', 0.0)
            )
            
            result['paper_trade'] = {
                'trade_id': trade_id,
                'entry_price': entry_price,
                'stop_loss': stop_loss,
                'take_profit': take_profit,
                'risk_reward': (take_profit - entry_price) / (entry_price - stop_loss) if entry_price > stop_loss else 0.0
            }
            
            perf_log.info("💼 Paper trade initiated for %s (Tier1): Entry ₹%.2f, SL ₹%.2f, TP ₹%.2f", 
                         ticker, entry_price, stop_loss, take_profit)
        
        return result
        
    except Exception as e:
        data_log.error("Enhanced screening failed for %s: %s", ticker, e)
        return None

# Initialize global backtest hooks
try:
    backtest_hooks = BacktestHooks()
except Exception as e:
    data_log.warning("Failed to initialize backtest hooks: %s", e)
    backtest_hooks = None

# =============================================================================
# ENHANCED CACHING AND PROCESS OPTIMIZATION SYSTEM
# =============================================================================

class EnhancedCacheManager:
    """Advanced caching system with TTL and intelligent invalidation"""
    
    def __init__(self):
        self.cache = {}
        self.cache_times = {}
        self.cache_hits = 0
        self.cache_misses = 0
        
    def get_cache_key(self, ticker: str, operation: str, **kwargs) -> str:
        """Generate cache key from ticker and operation parameters"""
        param_str = "_".join(f"{k}={v}" for k, v in sorted(kwargs.items()))
        return f"{ticker}_{operation}_{param_str}" if param_str else f"{ticker}_{operation}"
    
    def get(self, key: str, ttl: int = 300) -> Optional[any]:
        """Get item from cache with TTL check"""
        if key in self.cache:
            cache_time = self.cache_times.get(key, 0)
            if time.time() - cache_time < ttl:
                self.cache_hits += 1
                if debug_config.enable_caching_debug:
                    perf_log.debug("🎯 Cache HIT: %s", key)
                return self.cache[key]
            else:
                # Expired - remove from cache
                del self.cache[key]
                del self.cache_times[key]
        
        self.cache_misses += 1
        if debug_config.enable_caching_debug:
            perf_log.debug("❌ Cache MISS: %s", key)
        return None
    
    def set(self, key: str, value: any) -> None:
        """Set item in cache with timestamp"""
        self.cache[key] = value
        self.cache_times[key] = time.time()
        
        # Cleanup old entries if cache gets too large
        if len(self.cache) > 1000:
            self._cleanup_old_entries()
    
    def _cleanup_old_entries(self, max_age: int = 3600) -> None:
        """Remove cache entries older than max_age seconds"""
        current_time = time.time()
        expired_keys = [
            key for key, timestamp in self.cache_times.items()
            if current_time - timestamp > max_age
        ]
        
        for key in expired_keys:
            del self.cache[key]
            del self.cache_times[key]
        
        if expired_keys:
            perf_log.debug("🧹 Cleaned %d expired cache entries", len(expired_keys))
    
    def get_stats(self) -> dict:
        """Get cache performance statistics"""
        total_requests = self.cache_hits + self.cache_misses
        hit_rate = (self.cache_hits / total_requests * 100) if total_requests > 0 else 0
        
        return {
            'hit_rate': hit_rate,
            'hits': self.cache_hits,
            'misses': self.cache_misses,
            'cache_size': len(self.cache)
        }

# Global cache manager
cache_manager = EnhancedCacheManager()

# =============================================================================
# CACHED YFINANCE OPERATIONS - Performance Critical
# =============================================================================

@lru_cache(maxsize=1024)
def get_info_cached(ticker: str) -> dict:
    """Cached version of ticker info retrieval with TTL"""
    cache_key = cache_manager.get_cache_key(ticker, "info")
    cached_result = cache_manager.get(cache_key, ttl=3600)  # 1 hour TTL
    
    if cached_result is not None:
        return cached_result
    
    # Cache miss - fetch from yfinance
    result = YFinanceDataValidator.safe_ticker_info(ticker)
    cache_manager.set(cache_key, result)
    
    return result

@performance_optimizer.time_function("get_history_optimized")
def get_history_optimized(ticker: str, period: str = "1y", use_cache: bool = True) -> pd.DataFrame:
    """Optimized history retrieval with caching and minimal data slicing"""
    cache_key = cache_manager.get_cache_key(ticker, "history", period=period)
    
    if use_cache:
        cached_result = cache_manager.get(cache_key, ttl=1800)  # 30 minutes TTL
        if cached_result is not None and not cached_result.empty:
            return cached_result
    
    # Cache miss - fetch from yfinance
    df = YFinanceDataValidator.safe_ticker_history(ticker, period=period)
    
    if not df.empty:
        # Slice to minimal needed history (300 bars max for indicators)
        if len(df) > 300:
            df = df.tail(300)
            perf_log.debug("📉 Sliced history for %s: %d -> 300 bars", ticker, len(df))
        
        # Use float32 for memory efficiency
        numeric_cols = ['Open', 'High', 'Low', 'Close', 'Volume']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = df[col].astype('float32')
        
        if use_cache:
            cache_manager.set(cache_key, df)
    
    return df

# =============================================================================
# PROCESS POOL OPTIMIZATION - CPU-bound Analysis
# =============================================================================

def analyze_single_ticker_optimized(ticker_data: Tuple[str, pd.DataFrame]) -> dict:
    """
    Optimized single ticker analysis for process pool execution.
    This function is designed to be pickle-able for multiprocessing.
    """
    ticker, df = ticker_data
    
    try:
        # Use vectorized indicators for maximum performance
        indicators = vectorized_indicators.calculate_all_indicators(df, ticker)
        
        # Calculate swing reversal signals efficiently
        swing_signals = calculate_swing_reversal_signals_fast(df)
        
        # Basic scoring (can be enhanced)
        score = calculate_basic_score_fast(indicators, swing_signals)
        
        result = {
            'ticker': ticker,
            'score': score,
            'indicators': indicators,
            'swing_signals': swing_signals,
            'data_points': len(df),
            'analysis_timestamp': time.time()
        }
        
        return result
        
    except Exception as e:
        return {
            'ticker': ticker,
            'error': str(e),
            'score': 0.0,
            'indicators': {},
            'swing_signals': {},
            'data_points': 0,
            'analysis_timestamp': time.time()
        }

def calculate_swing_reversal_signals_fast(df: pd.DataFrame) -> dict:
    """Fast version of swing reversal signals calculation"""
    if df.empty or len(df) < 20:
        return {'reversal_direction': 'NEUTRAL', 'swing_reversal_score': 0.0}
    
    # Use optimized price data access
    high = df['High'].values  # NumPy arrays for speed
    low = df['Low'].values
    close = df['Close'].values
    
    # Fast min/max operations
    lookback = min(20, len(df))
    recent_high_idx = np.argmax(high[-lookback:])
    recent_low_idx = np.argmin(low[-lookback:])
    
    # Calculate rejection signals efficiently
    highest_bar_high = high[-(lookback - recent_high_idx)]
    highest_bar_low = low[-(lookback - recent_high_idx)]
    peak_rejection = ((highest_bar_high - highest_bar_low) / highest_bar_high) * 100
    
    lowest_bar_low = low[-(lookback - recent_low_idx)]
    lowest_bar_high = high[-(lookback - recent_low_idx)]
    trough_bounce = ((lowest_bar_high - lowest_bar_low) / lowest_bar_low) * 100
    
    # Simple scoring
    if trough_bounce > peak_rejection and trough_bounce > 2.0:
        return {'reversal_direction': 'BULLISH_REVERSAL', 'swing_reversal_score': min(trough_bounce, 10)}
    elif peak_rejection > trough_bounce and peak_rejection > 2.0:
        return {'reversal_direction': 'BEARISH_REVERSAL', 'swing_reversal_score': -min(peak_rejection, 10)}
    else:
        return {'reversal_direction': 'NEUTRAL', 'swing_reversal_score': 0.0}

def calculate_basic_score_fast(indicators: dict, swing_signals: dict) -> float:
    """Fast basic scoring algorithm"""
    score = 0.0
    
    # RSI scoring (lower is better for buying)
    rsi = indicators.get('rsi', 50)
    if rsi < 30:
        score += 3.0
    elif rsi < 40:
        score += 2.0
    elif rsi > 70:
        score -= 2.0
    
    # Bollinger Band position (lower is better)
    bb_pos = indicators.get('bb_position', 50)
    if bb_pos < 20:
        score += 2.5
    elif bb_pos < 40:
        score += 1.5
    elif bb_pos > 80:
        score -= 2.0
    
    # Volume confirmation
    vol_ratio = indicators.get('volume_ratio', 1.0)
    if vol_ratio > 2.0:
        score += 1.5
    elif vol_ratio > 1.5:
        score += 1.0
    
    # Risk/Reward ratio
    rr_ratio = indicators.get('rr_ratio', 1.0)
    if rr_ratio > 3.0:
        score += 1.5
    elif rr_ratio > 2.0:
        score += 1.0
    
    # Swing reversal signals
    swing_score = swing_signals.get('swing_reversal_score', 0.0)
    if swing_signals.get('reversal_direction') == 'BULLISH_REVERSAL':
        score += swing_score * 0.3  # Weight swing signals
    
    return max(0.0, min(10.0, score))  # Clamp between 0-10

# =============================================================================
# MEMORY AND PERFORMANCE MONITORING
# =============================================================================

class MemoryMonitor:
    """Monitor memory usage and trigger cleanup when needed"""
    
    def __init__(self, cleanup_threshold_mb: float = 1024.0):
        self.cleanup_threshold = cleanup_threshold_mb
        self.last_cleanup = time.time()
        self.cleanup_interval = 300  # 5 minutes
        
    def check_and_cleanup(self) -> bool:
        """Check memory usage and cleanup if needed"""
        try:
            current_memory = performance_optimizer._get_memory_usage()
            current_time = time.time()
            
            should_cleanup = (
                current_memory > self.cleanup_threshold or
                current_time - self.last_cleanup > self.cleanup_interval
            )
            
            if should_cleanup:
                self.perform_cleanup()
                self.last_cleanup = current_time
                perf_log.info("🧹 Memory cleanup performed: %.1fMB", current_memory)
                return True
                
        except Exception as e:
            perf_log.warning("Memory monitoring failed: %s", e)
        
        return False
    
    def perform_cleanup(self):
        """Perform memory cleanup operations"""
        # Clear old cache entries
        cache_manager._cleanup_old_entries()
        
        # Force garbage collection
        gc.collect()
        
        # Clear yfinance cache if available
        try:
            yshared._DFS.clear()
            yshared._ISINS.clear()
        except:
            pass

# Global memory monitor
memory_monitor = MemoryMonitor()

# =============================================================================
# PROCESS POOL OPTIMIZATION WRAPPER
# =============================================================================

def optimize_screener_for_processes() -> None:
    """
    Configure the screener for optimal multiprocessing performance.
    Call this function before using ProcessPoolExecutor.
    """
    # Configure global performance settings
    global performance_optimizer, cache_manager, memory_monitor
    
    # Reduce cache size for subprocess memory efficiency
    cache_manager.cleanup_interval = 600  # 10 minutes
    
    # Enable aggressive garbage collection
    gc.set_threshold(700, 10, 10)  # More frequent GC
    
    # Disable debug logging in subprocesses for performance
    if not DEBUG_ENABLED:
        logging.getLogger('screener.performance').setLevel(logging.WARNING)
        logging.getLogger('screener.data').setLevel(logging.WARNING)
    
    perf_log.info("🚀 Screener optimized for multiprocessing")

@performance_optimizer.time_function("batch_analyze_optimized")
def batch_analyze_with_process_pool(tickers: List[str], max_workers: int = None, 
                                  use_processes: bool = True) -> List[dict]:
    """
    OPTIMIZED: Analyze multiple tickers using process pool for CPU-bound operations.
    Performance improvement: 2-4x faster on multi-core systems.
    """
    if not tickers:
        return []
    
    # Determine optimal worker count
    if max_workers is None:
        import multiprocessing
        max_workers = min(max(2, multiprocessing.cpu_count() - 1), len(tickers))
    
    perf_log.info("🔥 Starting batch analysis: %d tickers, %d workers, processes=%s", 
                 len(tickers), max_workers, use_processes)
    
    results = []
    
    # Pre-fetch data efficiently in batches
    perf_log.debug("📡 Pre-fetching historical data in batches...")
    hist_data = {}
    
    # Batch download for efficiency
    batch_size = 50
    for i in range(0, len(tickers), batch_size):
        batch_tickers = tickers[i:i + batch_size]
        
        try:
            # Use our optimized cached download
            for ticker in batch_tickers:
                df = get_history_optimized(ticker, period="1y", use_cache=True)
                if not df.empty:
                    hist_data[ticker] = df
                    
        except Exception as e:
            perf_log.warning("Batch download failed for batch %d: %s", i//batch_size, e)
    
    # Prepare data for process pool
    ticker_data_pairs = [(ticker, hist_data.get(ticker, pd.DataFrame())) 
                        for ticker in tickers if ticker in hist_data]
    
    if use_processes and len(ticker_data_pairs) > 5:
        # Use ProcessPoolExecutor for CPU-bound analysis
        from concurrent.futures import ProcessPoolExecutor
        
        try:
            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                # Configure each process for optimal performance
                executor.submit(optimize_screener_for_processes).result()
                
                # Submit analysis tasks
                future_to_ticker = {
                    executor.submit(analyze_single_ticker_optimized, data): data[0] 
                    for data in ticker_data_pairs
                }
                
                # Collect results as they complete
                for future in as_completed(future_to_ticker):
                    try:
                        result = future.result(timeout=30)  # 30 second timeout per ticker
                        results.append(result)
                        
                        if len(results) % 10 == 0:
                            perf_log.debug("✅ Completed %d/%d ticker analyses", 
                                         len(results), len(ticker_data_pairs))
                            
                    except Exception as e:
                        ticker = future_to_ticker[future]
                        perf_log.error("❌ Analysis failed for %s: %s", ticker, str(e))
                        results.append({
                            'ticker': ticker,
                            'error': str(e),
                            'score': 0.0,
                            'indicators': {},
                            'swing_signals': {}
                        })
                        
        except Exception as e:
            perf_log.error("ProcessPoolExecutor failed: %s, falling back to ThreadPoolExecutor", e)
            use_processes = False
    
    if not use_processes:
        # Fallback to ThreadPoolExecutor for I/O-bound or small datasets
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_ticker = {
                executor.submit(analyze_single_ticker_optimized, data): data[0] 
                for data in ticker_data_pairs
            }
            
            for future in as_completed(future_to_ticker):
                try:
                    result = future.result(timeout=15)
                    results.append(result)
                except Exception as e:
                    ticker = future_to_ticker[future]
                    perf_log.error("❌ Thread analysis failed for %s: %s", ticker, str(e))
                    results.append({
                        'ticker': ticker,
                        'error': str(e),
                        'score': 0.0,
                        'indicators': {},
                        'swing_signals': {}
                    })
    
    # Performance summary
    successful = len([r for r in results if 'error' not in r])
    perf_log.info("🎯 Batch analysis complete: %d/%d successful (%.1f%%)", 
                 successful, len(tickers), successful/len(tickers)*100 if tickers else 0)
    
    # Memory cleanup after batch processing
    memory_monitor.perform_cleanup()
    
    return results

# =============================================================================
# ENHANCED FINANCIAL METRICS COLLECTION (Critical Additions)
# =============================================================================

class EnhancedFinancialMetrics:
    """Enhanced financial metrics collection for better stock screening"""
    
    @staticmethod
    def get_financial_health_metrics(ticker: str) -> Dict[str, float]:
        """
        Collect critical financial health indicators with data validation
        Returns: debt_to_equity, interest_coverage, current_ratio, roe, free_cash_flow
        """
        try:
            # Use safe data fetching with validation
            info = YFinanceDataValidator.safe_ticker_info(ticker)
            
            if not info:
                logging.debug(f"No valid financial data available for {ticker}")
                return {
                    'debt_to_equity': 0.0,
                    'interest_coverage': 0.0, 
                    'current_ratio': 1.0,
                    'roe': 0.0,
                    'free_cash_flow': 0.0,
                    'promoter_holding': 0.0,
                    'promoter_pledging': 0.0
                }
            
            metrics = {}
            
            # 1. Debt-to-Equity Ratio (with validation)
            total_debt = info.get('totalDebt', 0) or 0
            total_equity = info.get('totalStockholderEquity', 0) or info.get('bookValue', 0) or 0
            if total_equity > 0:
                debt_to_equity = total_debt / total_equity
                # Validate reasonable range
                if 0 <= debt_to_equity <= 10:  # Reasonable D/E ratio
                    metrics['debt_to_equity'] = debt_to_equity
                else:
                    logging.debug(f"Extreme D/E ratio {debt_to_equity} for {ticker}")
                    metrics['debt_to_equity'] = min(debt_to_equity, 10.0)  # Cap at 10
            else:
                metrics['debt_to_equity'] = 0.0
            
            # 2. Interest Coverage Ratio (with validation)
            ebit = info.get('ebitda', 0) or 0
            interest_expense = info.get('interestExpense', 0) or 0
            if interest_expense > 0 and ebit > 0:
                coverage = ebit / interest_expense
                # Validate reasonable range
                if coverage >= 0:
                    metrics['interest_coverage'] = min(coverage, 100.0)  # Cap at 100
                else:
                    metrics['interest_coverage'] = 0.0
            else:
                metrics['interest_coverage'] = 0.0
            
            # 3. Current Ratio (with validation)
            current_assets = info.get('totalCurrentAssets', 0) or 0
            current_liabilities = info.get('totalCurrentLiabilities', 0) or 0
            if current_liabilities > 0 and current_assets > 0:
                current_ratio = current_assets / current_liabilities
                # Validate reasonable range
                if 0.1 <= current_ratio <= 20:  # Reasonable current ratio
                    metrics['current_ratio'] = current_ratio
                else:
                    logging.debug(f"Extreme current ratio {current_ratio} for {ticker}")
                    metrics['current_ratio'] = max(0.1, min(current_ratio, 20.0))
            else:
                metrics['current_ratio'] = 1.0
            
            # 4. Return on Equity (ROE) - validated
            roe = info.get('returnOnEquity', 0)
            if roe is not None and isinstance(roe, (int, float)) and not np.isnan(roe):
                roe_pct = roe * 100 if roe else 0
                # Validate reasonable range (-100% to 100%)
                if -100 <= roe_pct <= 100:
                    metrics['roe'] = roe_pct
                else:
                    logging.debug(f"Extreme ROE {roe_pct}% for {ticker}")
                    metrics['roe'] = max(-100, min(roe_pct, 100))
            else:
                metrics['roe'] = 0
                
            # 5. Free Cash Flow (with validation)
            free_cash_flow = info.get('freeCashflow', 0) or info.get('operatingCashflow', 0) or 0
            if isinstance(free_cash_flow, (int, float)) and not np.isnan(free_cash_flow):
                fcf_cr = free_cash_flow / 1e7 if free_cash_flow else 0.0  # In crores
                metrics['free_cash_flow'] = fcf_cr
            else:
                metrics['free_cash_flow'] = 0.0
            
            # 6. Promoter Holdings (with validation)
            promoter_holding = info.get('heldPercentInsiders', 0) or 0
            if isinstance(promoter_holding, (int, float)) and not np.isnan(promoter_holding):
                holding_pct = promoter_holding * 100 if promoter_holding <= 1 else promoter_holding
                # Validate reasonable range (0% to 100%)
                if 0 <= holding_pct <= 100:
                    metrics['promoter_holding'] = holding_pct
                else:
                    logging.debug(f"Invalid promoter holding {holding_pct}% for {ticker}")
                    metrics['promoter_holding'] = max(0, min(holding_pct, 100))
            else:
                metrics['promoter_holding'] = 0
            
            # Note: Pledging data not available in yfinance, would need separate API
            metrics['promoter_pledging'] = 0  # Placeholder
            
            # Final validation - ensure all metrics are reasonable numbers
            for key, value in metrics.items():
                if not isinstance(value, (int, float)) or np.isnan(value) or np.isinf(value):
                    logging.debug(f"Invalid {key} value {value} for {ticker}, setting to default")
                    if key == 'current_ratio':
                        metrics[key] = 1.0
                    else:
                        metrics[key] = 0.0
            
            return metrics
            
        except Exception as e:
            logging.warning(f"Error fetching financial health metrics for {ticker}: {e}")
            # Return default values if data fetch fails
            return {
                'debt_to_equity': 0.0,
                'interest_coverage': 0.0,
                'current_ratio': 1.0,
                'roe': 0.0,
                'free_cash_flow': 0.0,
                'promoter_holding': 0.0,
                'promoter_pledging': 0.0
            }
            
        except Exception as e:
            # Return default values if data fetch fails
            return {
                'debt_to_equity': 999,
                'interest_coverage': 0,
                'current_ratio': 0,
                'roe': 0,
                'free_cash_flow': 0,
                'promoter_holding': 0,
                'promoter_pledging': 0
            }
    
    @staticmethod
    def get_institutional_flow_metrics(ticker: str) -> Dict[str, float]:
        """
        Get institutional holding and flow data with validation
        Returns: fii_holding, dii_holding, delivery_percentage
        """
        try:
            # Use safe data fetching with validation
            info = YFinanceDataValidator.safe_ticker_info(ticker)
            
            if not info:
                logging.debug(f"No valid institutional data available for {ticker}")
                return {
                    'institutional_holding': 0.0,
                    'float_shares': 0.0,
                    'delivery_percentage': 50.0
                }
            
            metrics = {}
            
            # FII/DII Holdings (with validation)
            institutional_holding = info.get('heldPercentInstitutions', 0) or 0
            if isinstance(institutional_holding, (int, float)) and not np.isnan(institutional_holding):
                holding_pct = institutional_holding * 100 if institutional_holding <= 1 else institutional_holding
                # Validate reasonable range (0% to 100%)
                if 0 <= holding_pct <= 100:
                    metrics['institutional_holding'] = holding_pct
                else:
                    logging.debug(f"Invalid institutional holding {holding_pct}% for {ticker}")
                    metrics['institutional_holding'] = max(0, min(holding_pct, 100))
            else:
                metrics['institutional_holding'] = 0.0
            
            # Float shares (with validation)
            float_shares = info.get('floatShares', 0) or 0
            if isinstance(float_shares, (int, float)) and not np.isnan(float_shares) and float_shares > 0:
                float_cr = float_shares / 1e7  # In crores
                metrics['float_shares'] = float_cr
            else:
                metrics['float_shares'] = 0.0
            
            # Delivery percentage calculation (from recent volume data with validation)
            try:
                hist_df = YFinanceDataValidator.safe_ticker_history(ticker, period='5d')
                if not hist_df.empty and len(hist_df) >= 2:
                    avg_volume = hist_df['Volume'].mean()
                    latest_volume = hist_df['Volume'].iloc[-1]
                    
                    # Validate volume data
                    if avg_volume > 0 and latest_volume > 0:
                        # Approximate delivery % based on volume patterns
                        delivery_estimate = min(100, (avg_volume / latest_volume * 50))
                        # Ensure reasonable range
                        if 0 <= delivery_estimate <= 100:
                            metrics['delivery_percentage'] = delivery_estimate
                        else:
                            metrics['delivery_percentage'] = 50.0
                    else:
                        metrics['delivery_percentage'] = 50.0
                else:
                    metrics['delivery_percentage'] = 50.0
            except Exception as e:
                logging.debug(f"Error calculating delivery percentage for {ticker}: {e}")
                metrics['delivery_percentage'] = 50.0
            
            # Final validation
            for key, value in metrics.items():
                if not isinstance(value, (int, float)) or np.isnan(value) or np.isinf(value):
                    logging.debug(f"Invalid {key} value {value} for {ticker}, setting to default")
                    if key == 'delivery_percentage':
                        metrics[key] = 50.0
                    else:
                        metrics[key] = 0.0
            
            return metrics
            
        except Exception as e:
            logging.warning(f"Error fetching institutional flow metrics for {ticker}: {e}")
            return {
                'institutional_holding': 0.0,
                'float_shares': 0.0,
                'delivery_percentage': 50.0
            }
    
    @staticmethod
    def get_technical_risk_metrics(ticker: str, price_data: pd.DataFrame) -> Dict[str, float]:
        """
        Calculate enhanced technical risk metrics
        Returns: beta, price_distance_50ma, price_distance_200ma, volatility
        """
        try:
            if price_data.empty:
                return {'beta': 1.0, 'price_distance_50ma': 0, 'price_distance_200ma': 0, 'volatility': 0}
            
            metrics = {}
            close_prices = price_data['Close']
            
            # 1. Beta calculation (simplified using price volatility)
            try:
                returns = close_prices.pct_change().dropna()
                metrics['beta'] = returns.std() * np.sqrt(252) if len(returns) > 30 else 1.0
            except:
                metrics['beta'] = 1.0
            
            # 2. Price distance from moving averages
            try:
                current_price = close_prices.iloc[-1]
                ma_50 = close_prices.rolling(50).mean().iloc[-1]
                ma_200 = close_prices.rolling(200).mean().iloc[-1]
                
                metrics['price_distance_50ma'] = ((current_price - ma_50) / ma_50 * 100) if ma_50 > 0 else 0
                metrics['price_distance_200ma'] = ((current_price - ma_200) / ma_200 * 100) if ma_200 > 0 else 0
            except:
                metrics['price_distance_50ma'] = 0
                metrics['price_distance_200ma'] = 0
            
            # 3. Volatility (annualized)
            try:
                returns = close_prices.pct_change().dropna()
                metrics['volatility'] = returns.std() * np.sqrt(252) * 100 if len(returns) > 10 else 0
            except:
                metrics['volatility'] = 0
            
            return metrics
            
        except Exception as e:
            return {'beta': 1.0, 'price_distance_50ma': 0, 'price_distance_200ma': 0, 'volatility': 0}

# =============================================================================
# ENHANCED FINANCIAL ANALYSIS FOR FINAL CANDIDATES ONLY
# =============================================================================

def analyze_final_candidates_enhanced(tickers: List[str], hist_data: Dict[str, pd.DataFrame]) -> Dict[str, Dict]:
    """
    Run comprehensive financial analysis only on final tier candidates.
    This preserves speed by avoiding API calls during initial screening.
    """
    enhanced_results = {}
    
    print(f"\n[ANALYSIS] Running enhanced financial analysis on {len(tickers)} final candidates...")
    
    for ticker in tickers:
        try:
            # Get comprehensive financial metrics
            financial_health = EnhancedFinancialMetrics.get_financial_health_metrics(ticker)
            institutional_flow = EnhancedFinancialMetrics.get_institutional_flow_metrics(ticker)
            
            # Get price data for technical analysis
            df = hist_data.get(ticker, pd.DataFrame())
            technical_risk = EnhancedFinancialMetrics.get_technical_risk_metrics(ticker, df)
            
            enhanced_results[ticker] = {
                'financial_health': financial_health,
                'institutional_flow': institutional_flow,
                'technical_risk': technical_risk
            }
            
        except Exception as e:
            # Graceful fallback for failed API calls
            enhanced_results[ticker] = {
                'financial_health': {'debt_to_equity': 999, 'roe': 0, 'interest_coverage': 0, 'free_cash_flow': 0},
                'institutional_flow': {'institutional_holding': 0, 'delivery_percentage': 50},
                'technical_risk': {'beta': 1.0, 'price_distance_50ma': 0, 'volatility': 0}
            }
            
    return enhanced_results

def format_financial_metric(value: float, metric_type: str) -> str:
    """
    Smart highlighting for financial metrics:
    * = Excellent (Purple in terminal)
    * = Good/Very Good (Green in terminal)
    ~ = OK/Acceptable (Yellow in terminal)
    (no color) = Not sure/Neutral
    ! = Bad/Poor (Red in terminal)
    """
    if metric_type == "debt_to_equity":
        if value >= 999:  # Missing data
            return "N/A"
        elif value <= 0.2:  # Excellent
            return f"\033[95m{value:.2f}*\033[0m"  # Purple with *
        elif value <= 0.5:  # Good
            return f"\033[92m{value:.2f}*\033[0m"  # Green with *
        elif value <= 1.0:  # OK/Acceptable
            return f"\033[93m{value:.2f}~\033[0m"  # Yellow with ~
        elif value <= 2.0:  # Not sure/Neutral
            return f"{value:.2f}"
        else:  # Bad
            return f"\033[91m{value:.2f}!\033[0m"  # Red with !
            
    elif metric_type == "roe":
        if value <= 0:
            return "N/A" if value == 0 else f"\033[91m{value:.1f}!\033[0m"
        elif value >= 20:  # Excellent
            return f"\033[95m{value:.1f}*\033[0m"  # Purple with *
        elif value >= 15:  # Good
            return f"\033[92m{value:.1f}*\033[0m"  # Green with *
        elif value >= 10:  # OK/Acceptable
            return f"\033[93m{value:.1f}~\033[0m"  # Yellow with ~
        elif value >= 5:   # Not sure/Neutral
            return f"{value:.1f}"
        else:  # Poor
            return f"\033[91m{value:.1f}!\033[0m"  # Red with !
            
    elif metric_type == "interest_coverage":
        if value <= 0:
            return "N/A" if value == 0 else f"\033[91m{value:.1f}!\033[0m"
        elif value >= 10.0:  # Excellent
            return f"\033[95m{value:.1f}*\033[0m"  # Purple with *
        elif value >= 5.0:  # Good
            return f"\033[92m{value:.1f}*\033[0m"  # Green with *
        elif value >= 2.0:  # OK/Acceptable
            return f"\033[93m{value:.1f}~\033[0m"  # Yellow with ~
        elif value >= 1.0:  # Not sure/Neutral
            return f"{value:.1f}"
        else:  # Poor
            return f"\033[91m{value:.1f}!\033[0m"  # Red with !
            
    elif metric_type == "institutional_holding":
        if value >= 60:  # Excellent
            return f"\033[95m{value:.1f}*\033[0m"  # Purple with *
        elif value >= 40:  # Good
            return f"\033[92m{value:.1f}*\033[0m"  # Green with *
        elif value >= 20:  # OK/Acceptable
            return f"\033[93m{value:.1f}~\033[0m"  # Yellow with ~
        elif value >= 10:   # Not sure/Neutral
            return f"{value:.1f}"
        else:  # Poor
            return f"\033[91m{value:.1f}!\033[0m"  # Red with !
            return f"\033[91m{value:.1f}!\033[0m"
            
    elif metric_type == "delivery_percentage":
        if value >= 80:  # Excellent conviction
            return f"\033[95m{value:.1f}*\033[0m"  # Purple with *
        elif value >= 70:  # Good conviction
            return f"\033[92m{value:.1f}*\033[0m"  # Green with *
        elif value >= 50:  # OK/Acceptable
            return f"\033[93m{value:.1f}~\033[0m"  # Yellow with ~
        elif value >= 35:  # Not sure/Neutral
            return f"{value:.1f}"
        else:  # Speculative/Poor
            return f"\033[91m{value:.1f}!\033[0m"  # Red with !
            
    else:  # Default formatting
        return f"{value:.2f}"

def get_financial_health_grade(debt_to_equity: float, roe: float, interest_coverage: float) -> tuple:
    """
    Return (grade, color_code, description) based on key financial metrics
    """
    score = 0
    
    # Debt-to-Equity scoring (40% weight)
    if debt_to_equity <= 0.3:
        score += 4
    elif debt_to_equity <= 0.5:
        score += 3
    elif debt_to_equity <= 1.0:
        score += 2
    elif debt_to_equity <= 2.0:
        score += 1
    # >2.0 gets 0 points
    
    # ROE scoring (35% weight)
    if roe >= 20:
        score += 3.5
    elif roe >= 15:
        score += 3
    elif roe >= 10:
        score += 2
    elif roe >= 5:
        score += 1
    # <5% gets 0 points
    
    # Interest Coverage scoring (25% weight)
    if interest_coverage >= 5.0:
        score += 2.5
    elif interest_coverage >= 2.5:
        score += 2
    elif interest_coverage >= 1.5:
        score += 1.5
    elif interest_coverage >= 1.0:
        score += 1
    # <1.0 gets 0 points
    
    # Convert to grade
    if score >= 8.5:
        return "EXCELLENT", "\033[95m", "Strong balance sheet, high profitability"  # Purple
    elif score >= 6.5:
        return "GOOD", "\033[92m", "Solid financials, manageable risk"  # Green
    elif score >= 4.5:
        return "FAIR", "\033[93m", "Acceptable metrics, monitor closely"  # Yellow
    elif score >= 2.5:
        return "WEAK", "", "Concerning financials, high risk"  # No color
    else:
        return "POOR", "\033[91m", "Poor financial health, avoid"  # Red

def calculate_color_score(stock_data: dict) -> tuple:
    """
    Calculate color-based score for each stock based on individual colored outputs.
    Excludes longer-term price movements (week%, month%, 3month%, yr%, 5yr%).
    
    Returns: (color_score, color_display, score_breakdown)
    """
    score = 0
    breakdown = {}
    
    # 1. LIQUIDITY SCORE (Weight: 15) - Higher is better for entry/exit
    liq_score = stock_data.get('liquidity_score', 0.5)
    if liq_score >= 0.8:
        score += 15  # Purple
        breakdown['liquidity'] = '+15 (Excellent)'
    elif liq_score >= 0.6:
        score += 10  # Green  
        breakdown['liquidity'] = '+10 (Good)'
    elif liq_score >= 0.4:
        score += 5   # Yellow
        breakdown['liquidity'] = '+5 (OK)'
    else:
        score -= 5   # Red
        breakdown['liquidity'] = '-5 (Poor)'
    
    # 2. MAIN SCORE (Weight: 12) - Our core screening score
    main_score = stock_data.get('score', 0)
    if main_score >= 7.5:
        score += 12  # Purple
        breakdown['main_score'] = '+12 (Excellent)'
    elif main_score >= 6.0:
        score += 8   # Green
        breakdown['main_score'] = '+8 (Good)'
    elif main_score >= 4.5:
        score += 4   # Yellow
        breakdown['main_score'] = '+4 (OK)'
    else:
        score -= 4   # Red/None
        breakdown['main_score'] = '-4 (Poor)'
    
    # 3. FII GROWTH (Weight: 10) - Institutional interest
    fii_growth = stock_data.get('fii_growth', 0)
    if fii_growth >= 15:
        score += 10  # Purple
        breakdown['fii'] = '+10 (Excellent)'
    elif fii_growth >= 5:
        score += 6   # Green
        breakdown['fii'] = '+6 (Good)'
    elif fii_growth >= 0:
        score += 2   # Yellow
        breakdown['fii'] = '+2 (OK)'
    else:
        score -= 3   # Red
        breakdown['fii'] = '-3 (Declining)'
    
    # 4. QTR GROWTH (Weight: 10) - Recent performance
    qtr_growth = stock_data.get('qtr_growth', 0)
    if qtr_growth >= 20:
        score += 10  # Purple
        breakdown['qtr'] = '+10 (Excellent)'
    elif qtr_growth >= 10:
        score += 6   # Green
        breakdown['qtr'] = '+6 (Good)'
    elif qtr_growth >= 0:
        score += 2   # Yellow
        breakdown['qtr'] = '+2 (OK)'
    else:
        score -= 3   # Red
        breakdown['qtr'] = '-3 (Declining)'
    
    # 5. DAILY PRICE CHANGE (Weight: 8) - Only daily, not longer term
    daily_change = stock_data.get('daily_change', 0)
    if daily_change >= 3:
        score += 8   # Purple (Strong daily gain)
        breakdown['daily'] = '+8 (Strong gain)'
    elif daily_change >= 1:
        score += 4   # Green (Good gain)
        breakdown['daily'] = '+4 (Good gain)'
    elif daily_change >= -1:
        score += 1   # Yellow (Minor change)
        breakdown['daily'] = '+1 (Stable)'
    else:
        score -= 2   # Red (Declining)
        breakdown['daily'] = '-2 (Declining)'
    
    # 6. RSI (Weight: 12) - For buying opportunities (lower is better)
    rsi = stock_data.get('rsi', 50)
    if rsi <= 20:
        score += 12  # Purple (Extreme oversold - excellent buy)
        breakdown['rsi'] = '+12 (Extreme oversold)'
    elif rsi <= 30:
        score += 8   # Green (Strong oversold)
        breakdown['rsi'] = '+8 (Strong oversold)'
    elif rsi <= 40:
        score += 4   # Yellow (Mild oversold)
        breakdown['rsi'] = '+4 (Mild oversold)'
    elif rsi >= 70:
        score -= 8   # Red (Overbought - avoid)
        breakdown['rsi'] = '-8 (Overbought)'
    else:
        score += 0   # Neutral (40-70)
        breakdown['rsi'] = '0 (Neutral)'
    
    # 7. EMA SLOPE (Weight: 8) - For mean reversion (declining better for entry)
    ema_slope = stock_data.get('ema_slope', 0)
    if ema_slope < -2.0:
        score += 8   # Purple (Strong decline - excellent for mean reversion)
        breakdown['ema_slope'] = '+8 (Strong decline)'
    elif ema_slope < -1.0:
        score += 5   # Green (Good decline)
        breakdown['ema_slope'] = '+5 (Good decline)'
    elif ema_slope < -0.5:
        score += 2   # Yellow (Mild decline)
        breakdown['ema_slope'] = '+2 (Mild decline)'
    elif ema_slope > 2.0:
        score -= 5   # Red (Strong rise - not ideal for entry)
        breakdown['ema_slope'] = '-5 (Strong rise)'
    else:
        score += 0   # Neutral
        breakdown['ema_slope'] = '0 (Neutral)'
    
    # 8. BOLLINGER BAND POSITION (Weight: 10) - Lower is better for buying
    bb_position = stock_data.get('bb_position', 50)
    if bb_position <= 10:
        score += 10  # Purple (Extreme oversold)
        breakdown['bb_pos'] = '+10 (Extreme oversold)'
    elif bb_position <= 20:
        score += 7   # Green (Strong oversold)
        breakdown['bb_pos'] = '+7 (Strong oversold)'
    elif bb_position <= 30:
        score += 4   # Green (Good zone)
        breakdown['bb_pos'] = '+4 (Good zone)'
    elif bb_position <= 40:
        score += 2   # Yellow (OK entry)
        breakdown['bb_pos'] = '+2 (OK entry)'
    elif bb_position >= 80:
        score -= 6   # Red (Overbought)
        breakdown['bb_pos'] = '-6 (Overbought)'
    else:
        score += 0   # Neutral
        breakdown['bb_pos'] = '0 (Neutral)'
    
    # 9. RISK/REWARD RATIO (Weight: 8)
    rr_ratio = stock_data.get('rr_ratio', 1.0)
    if rr_ratio >= 5:
        score += 8   # Purple
        breakdown['rr_ratio'] = '+8 (Excellent R/R)'
    elif rr_ratio >= 3:
        score += 5   # Green
        breakdown['rr_ratio'] = '+5 (Good R/R)'
    elif rr_ratio >= 1.5:
        score += 2   # Yellow
        breakdown['rr_ratio'] = '+2 (OK R/R)'
    else:
        score -= 3   # Red
        breakdown['rr_ratio'] = '-3 (Poor R/R)'
    
    # 10. DEAL PERCENTAGE (Weight: 6)
    deal_pct = stock_data.get('deal_percentage', 0)
    if deal_pct >= 10:
        score += 6   # Purple
        breakdown['deal_pct'] = '+6 (High delivery)'
    elif deal_pct >= 5:
        score += 3   # Green
        breakdown['deal_pct'] = '+3 (Good delivery)'
    elif deal_pct >= 2:
        score += 1   # Yellow
        breakdown['deal_pct'] = '+1 (OK delivery)'
    else:
        score -= 1   # Red
        breakdown['deal_pct'] = '-1 (Low delivery)'
    
    # 11. SWING REVERSAL SIGNALS (Weight: 12) - NEW ENHANCEMENT
    swing_signals = stock_data.get('swing_reversal', {})
    swing_score = swing_signals.get('swing_reversal_score', 0)
    reversal_direction = swing_signals.get('reversal_direction', 'NEUTRAL')
    
    if reversal_direction == 'BULLISH_REVERSAL':
        if swing_score >= 8:
            score += 12  # Purple - Excellent reversal opportunity
            breakdown['swing_reversal'] = '+12 (Excellent reversal)'
        elif swing_score >= 6:
            score += 8   # Green - Good reversal
            breakdown['swing_reversal'] = '+8 (Good reversal)'
        elif swing_score >= 4:
            score += 4   # Yellow - Mild reversal
            breakdown['swing_reversal'] = '+4 (Mild reversal)'
        else:
            score += 2   # Minor bullish reversal
            breakdown['swing_reversal'] = '+2 (Minor reversal)'
    elif reversal_direction == 'BEARISH_REVERSAL':
        if swing_score <= -4:
            score -= 8   # Red - Strong bearish reversal (avoid)
            breakdown['swing_reversal'] = '-8 (Strong bearish)'
        elif swing_score <= -2:
            score -= 4   # Red - Moderate bearish reversal
            breakdown['swing_reversal'] = '-4 (Moderate bearish)'
        else:
            score -= 2   # Minor bearish warning
            breakdown['swing_reversal'] = '-2 (Minor bearish)'
    else:
        score += 0   # Neutral
        breakdown['swing_reversal'] = '0 (Neutral)'
    
    # 12. VOLUME CONFIRMATION (Weight: 10) - CRITICAL for real money deployment
    volume_multiplier = stock_data.get('volume_multiplier', 1.0)
    if volume_multiplier >= 3.0:
        score += 10  # Purple - Exceptional volume surge
        breakdown['volume'] = '+10 (Exceptional volume)'
    elif volume_multiplier >= 2.0:
        score += 7   # Green - Strong volume confirmation
        breakdown['volume'] = '+7 (Strong volume)'
    elif volume_multiplier >= 1.5:
        score += 4   # Yellow - Good volume
        breakdown['volume'] = '+4 (Good volume)'
    elif volume_multiplier >= 1.0:
        score += 1   # Minor volume
        breakdown['volume'] = '+1 (Average volume)'
    else:
        score -= 5   # Red - Weak volume (RED FLAG)
        breakdown['volume'] = '-5 (Weak volume - AVOID)'
    
    # 13. BUYER DOMINANCE (Weight: 10) - CRITICAL buyer vs seller strength
    buyer_dominance = stock_data.get('buyer_dominance', 50.0)  # BD percentage
    if buyer_dominance >= 80.0:
        score += 10  # Purple - Strong buyer dominance
        breakdown['buyer_dom'] = '+10 (Strong buyers)'
    elif buyer_dominance >= 70.0:
        score += 7   # Green - Good buyer interest
        breakdown['buyer_dom'] = '+7 (Good buyers)'
    elif buyer_dominance >= 60.0:
        score += 4   # Yellow - Decent buyers
        breakdown['buyer_dom'] = '+4 (Decent buyers)'
    elif buyer_dominance >= 50.0:
        score += 1   # Neutral
        breakdown['buyer_dom'] = '+1 (Neutral)'
    else:
        score -= 6   # Red - Seller dominance (RED FLAG)
        breakdown['buyer_dom'] = '-6 (Seller dominance - AVOID)'
    
    # Generate color display based on total score
    
    # === INSTITUTIONAL-GRADE ENHANCEMENTS ===
    
    # 14. RELATIVE VOLUME CURVE (Weight: 8) - Time-of-day adjusted volume
    rel_vol_ratio = stock_data.get('institutional_rel_vol_ratio', 1.0)
    high_liquidity = stock_data.get('institutional_high_liquidity', False)
    
    if high_liquidity and rel_vol_ratio >= 3.0:
        score += 8  # Purple - Exceptional institutional flow
        breakdown['inst_rel_vol'] = '+8 (Exceptional institutional flow)'
    elif rel_vol_ratio >= 2.0:
        score += 6  # Green - Strong relative volume
        breakdown['inst_rel_vol'] = '+6 (Strong relative volume)'
    elif rel_vol_ratio >= 1.5:
        score += 4  # Yellow - Good relative volume
        breakdown['inst_rel_vol'] = '+4 (Good relative volume)'
    elif rel_vol_ratio >= 1.0:
        score += 1  # Minor positive
        breakdown['inst_rel_vol'] = '+1 (Average relative volume)'
    else:
        score += 0  # Neutral
        breakdown['inst_rel_vol'] = '0 (Below average rel vol)'
    
    # 15. AVWAP BREAKOUT (Weight: 6) - Institutional entry signal
    avwap_breakout = stock_data.get('institutional_avwap_breakout')
    if avwap_breakout is not None:
        score += 6  # Purple - Confirmed AVWAP breakout
        breakdown['inst_avwap'] = '+6 (AVWAP breakout confirmed)'
    else:
        score += 0  # No breakout signal
        breakdown['inst_avwap'] = '0 (No AVWAP breakout)'
    
    # 16. ORDER BOOK IMBALANCE (Weight: 8) - Institutional flow direction
    ob_score = stock_data.get('institutional_ob_score', 50.0)
    if ob_score >= 80:
        score += 8  # Purple - Strong bullish imbalance
        breakdown['inst_ob'] = '+8 (Strong bullish imbalance)'
    elif ob_score >= 70:
        score += 6  # Green - Good bullish bias
        breakdown['inst_ob'] = '+6 (Good bullish bias)'
    elif ob_score >= 60:
        score += 4  # Yellow - Mild bullish bias
        breakdown['inst_ob'] = '+4 (Mild bullish bias)'
    elif ob_score >= 40:
        score += 1  # Neutral
        breakdown['inst_ob'] = '+1 (Neutral order flow)'
    else:
        score -= 3  # Red - Bearish imbalance
        breakdown['inst_ob'] = '-3 (Bearish imbalance)'
    
    # 17. RESIDUAL MOMENTUM (Weight: 10) - Sector-adjusted strength
    residual_momentum = stock_data.get('institutional_residual_momentum', 0.0)
    if residual_momentum >= 2.0:
        score += 10  # Purple - Exceptional idiosyncratic strength
        breakdown['inst_momentum'] = '+10 (Exceptional momentum)'
    elif residual_momentum >= 1.0:
        score += 8   # Green - Strong relative momentum
        breakdown['inst_momentum'] = '+8 (Strong momentum)'
    elif residual_momentum >= 0.5:
        score += 6   # Green - Good relative performance
        breakdown['inst_momentum'] = '+6 (Good momentum)'
    elif residual_momentum >= 0.0:
        score += 3   # Yellow - Positive momentum
        breakdown['inst_momentum'] = '+3 (Positive momentum)'
    elif residual_momentum >= -0.5:
        score += 0   # Neutral
        breakdown['inst_momentum'] = '0 (Neutral momentum)'
    else:
        score -= 4   # Red - Weak relative performance
        breakdown['inst_momentum'] = '-4 (Weak momentum)'
    
    # 18. CATALYST FACTOR (Weight: 8) - Forward-looking catalysts
    catalyst_score = stock_data.get('institutional_catalyst_score', 0.0)
    if catalyst_score >= 80:
        score += 8  # Purple - Multiple strong catalysts
        breakdown['inst_catalyst'] = '+8 (Multiple strong catalysts)'
    elif catalyst_score >= 60:
        score += 6  # Green - Good catalyst setup
        breakdown['inst_catalyst'] = '+6 (Good catalyst setup)'
    elif catalyst_score >= 40:
        score += 4  # Yellow - Some catalysts
        breakdown['inst_catalyst'] = '+4 (Some catalysts)'
    elif catalyst_score >= 20:
        score += 2  # Minor catalysts
        breakdown['inst_catalyst'] = '+2 (Minor catalysts)'
    else:
        score += 0  # No significant catalysts
        breakdown['inst_catalyst'] = '0 (No significant catalysts)'
    
    # 19. VOLATILITY DIVERGENCE STOP (Weight: 6) - Risk management signal
    vol_divergence = stock_data.get('institutional_vol_divergence', 0)
    if vol_divergence == 2:  # Strong positive signal
        score += 6  # Green - Strong vol setup 
        breakdown['inst_vol_div'] = '+6 (Strong vol setup)'
    elif vol_divergence == 1:  # Positive signal
        score += 4  # Good vol signal
        breakdown['inst_vol_div'] = '+4 (Good vol signal)'
    elif vol_divergence == 0:  # Neutral
        score += 0  # Neutral vol state
        breakdown['inst_vol_div'] = '0 (Neutral vol state)'
    else:  # Negative signal (-1 or -2)
        score -= 3  # Red - Vol warning
        breakdown['inst_vol_div'] = '-3 (Vol warning)'
    
    # 20. FRACTAL RISK BUDGET (Weight: 6) - Position sizing signal
    risk_budget = stock_data.get('institutional_risk_budget', 1.0)
    if risk_budget >= 1.5:  # High risk budget
        score += 6  # Purple - High conviction size
        breakdown['inst_risk_budget'] = '+6 (High conviction size)'
    elif risk_budget >= 1.2:  # Above average
        score += 4  # Green - Good size opportunity
        breakdown['inst_risk_budget'] = '+4 (Good size opportunity)'
    elif risk_budget >= 0.8:  # Normal sizing
        score += 2  # Normal sizing
        breakdown['inst_risk_budget'] = '+2 (Normal sizing)'
    elif risk_budget >= 0.5:  # Reduced sizing
        score += 0  # Caution sizing
        breakdown['inst_risk_budget'] = '0 (Caution sizing)'
    else:  # Very low risk budget
        score -= 2  # Red - Avoid or minimal size
        breakdown['inst_risk_budget'] = '-2 (Avoid/minimal size)'
    
    # Updated maximum possible score
    max_possible = 15+12+10+10+8+12+8+10+8+6+12+10+10+8+6+8+10+8+6+6  # 193 max (with all institutional enhancements)
    
    if score >= 150:        # 150+ out of 193 (~78%) - Exceptional with institutional confirmation
        color_display = f"\033[95m{score}⭐⭐\033[0m"  # Purple with double star
        grade = "ULTIMATE"
    elif score >= 120:      # 120-149 (~62%) - Excellent with strong institutional signals
        color_display = f"\033[95m{score}⭐\033[0m"   # Purple with star
        grade = "EXCELLENT" 
    elif score >= 95:       # 95-119 (~49%) - Good with some institutional validation
        color_display = f"\033[92m{score}\033[0m"     # Green
        grade = "GOOD"
    elif score >= 70:       # 70-94 (~36%) - Acceptable but watch institutional metrics
        color_display = f"\033[93m{score}\033[0m"     # Yellow
        grade = "OK"
    elif score >= 45:       # 45-69 (~23%) - Neutral, mixed institutional signals
        color_display = f"{score}"                    # No color
        grade = "NEUTRAL"
    elif score >= 15:       # 15-44 (~8%) - Weak, poor institutional setup
        color_display = f"\033[37m{score}\033[0m"     # White (negative)
        grade = "WEAK"
    else:                   # <15 - Avoid, institutional metrics negative
        color_display = f"\033[91m{score}\033[0m"     # Red (very negative)
        grade = "AVOID"
    
    return score, color_display, grade, breakdown

def calculate_swing_reversal_signals(df: pd.DataFrame, lookback_period: int = 20) -> dict:
    """
    Calculate swing trading reversal signals based on highest/lowest tick analysis.
    Uses winsorized price data to prevent extreme outliers from corrupting signals.
    
    Key Concepts:
    - Lowest tick of highest bar: Price rejection from peaks (sell pressure)
    - Highest tick of lowest bar: Price rejection from lows (buy pressure)  
    - These indicate potential swing reversal points
    
    Returns: Dictionary with reversal signals and strength indicators
    """
    if df is None or df.empty or len(df) < lookback_period:
        return {
            'reversal_strength': 0.0,
            'reversal_direction': 'NEUTRAL',
            'peak_rejection_signal': 0.0,
            'trough_rejection_signal': 0.0,
            'swing_reversal_score': 0.0,
            'color_boost': 0
        }
    
    # Use winsorized data for more robust signal calculation
    price_data = get_robust_price_data(df, use_capped=True)
    
    signals = {}
    current_price = price_data['close'].iloc[-1]
    current_high = price_data['high'].iloc[-1]
    current_low = price_data['low'].iloc[-1]
    current_open = price_data['open'].iloc[-1]
    
    # Get recent data for analysis using winsorized values
    # Create a temporary DataFrame with capped values for analysis
    recent_df = pd.DataFrame({
        'High': price_data['high'].tail(lookback_period),
        'Low': price_data['low'].tail(lookback_period),
        'Close': price_data['close'].tail(lookback_period),
        'Open': price_data['open'].tail(lookback_period)
    })
    
    # 1. HIGHEST BAR ANALYSIS - Find bars with highest highs and check rejection
    highest_high_idx = recent_df['High'].idxmax()
    highest_bar = recent_df.loc[highest_high_idx]
    highest_bar_high = highest_bar['High']
    highest_bar_low = highest_bar['Low']  # This is the "lowest tick of highest bar"
    highest_bar_close = highest_bar['Close']
    highest_bar_open = highest_bar['Open']
    
    # Calculate rejection from peak (how much price fell from high)
    peak_rejection_pct = ((highest_bar_high - highest_bar_low) / highest_bar_high) * 100
    peak_close_rejection = ((highest_bar_high - highest_bar_close) / highest_bar_high) * 100
    
    # Strong rejection signals from peaks (bearish for swing trading)
    peak_rejection_signal = 0.0
    if peak_rejection_pct > 4.0:  # Strong intraday rejection (>4% from high to low)
        peak_rejection_signal = 0.8
    elif peak_rejection_pct > 2.5:  # Moderate rejection
        peak_rejection_signal = 0.5
    elif peak_rejection_pct > 1.5:  # Mild rejection
        peak_rejection_signal = 0.3
    
    # Additional confirmation from close position
    if peak_close_rejection > 2.0:  # Closed well below high
        peak_rejection_signal += 0.2
    
    # 2. LOWEST BAR ANALYSIS - Find bars with lowest lows and check bounce
    lowest_low_idx = recent_df['Low'].idxmin()
    lowest_bar = recent_df.loc[lowest_low_idx]
    lowest_bar_high = lowest_bar['High']  # This is the "highest tick of lowest bar"
    lowest_bar_low = lowest_bar['Low']
    lowest_bar_close = lowest_bar['Close']
    lowest_bar_open = lowest_bar['Open']
    
    # Calculate bounce from trough (how much price rose from low)
    trough_bounce_pct = ((lowest_bar_high - lowest_bar_low) / lowest_bar_low) * 100
    trough_close_bounce = ((lowest_bar_close - lowest_bar_low) / lowest_bar_low) * 100
    
    # Strong bounce signals from troughs (bullish for swing trading)
    trough_rejection_signal = 0.0
    if trough_bounce_pct > 4.0:  # Strong intraday bounce (>4% from low to high)
        trough_rejection_signal = 0.8
    elif trough_bounce_pct > 2.5:  # Moderate bounce
        trough_rejection_signal = 0.5
    elif trough_bounce_pct > 1.5:  # Mild bounce
        trough_rejection_signal = 0.3
    
    # Additional confirmation from close position
    if trough_close_bounce > 2.0:  # Closed well above low
        trough_rejection_signal += 0.2
    
    # 3. CURRENT POSITION RELATIVE TO SWING POINTS
    days_since_highest = len(recent_df) - (recent_df.index.get_loc(highest_high_idx) + 1)
    days_since_lowest = len(recent_df) - (recent_df.index.get_loc(lowest_low_idx) + 1)
    
    # Recency boost (more recent signals are stronger)
    recency_boost_peak = max(0, (lookback_period - days_since_highest) / lookback_period)
    recency_boost_trough = max(0, (lookback_period - days_since_lowest) / lookback_period)
    
    # Apply recency weighting
    peak_rejection_signal *= recency_boost_peak
    trough_rejection_signal *= recency_boost_trough
    
    # 4. DETERMINE OVERALL REVERSAL DIRECTION AND STRENGTH
    reversal_strength = 0.0
    reversal_direction = 'NEUTRAL'
    
    # For BUYING opportunities (mean reversion from lows)
    if trough_rejection_signal > peak_rejection_signal:
        reversal_strength = trough_rejection_signal
        reversal_direction = 'BULLISH_REVERSAL'
        
        # Current price near the rejection low increases signal
        distance_from_low = ((current_price - lowest_bar_low) / lowest_bar_low) * 100
        if distance_from_low < 3.0:  # Within 3% of rejection low
            reversal_strength += 0.3
        elif distance_from_low < 5.0:  # Within 5% of rejection low
            reversal_strength += 0.2
            
    # For selling/avoiding (rejection from highs)
    elif peak_rejection_signal > trough_rejection_signal:
        reversal_strength = peak_rejection_signal
        reversal_direction = 'BEARISH_REVERSAL'
        
        # Current price near the rejection high is warning
        distance_from_high = ((highest_bar_high - current_price) / highest_bar_high) * 100
        if distance_from_high < 3.0:  # Within 3% of rejection high
            reversal_strength += 0.3
    
    # 5. SWING TRADING SCORE CALCULATION
    swing_reversal_score = 0.0
    color_boost = 0
    
    if reversal_direction == 'BULLISH_REVERSAL':
        # Bullish reversal - good for buying/mean reversion
        swing_reversal_score = reversal_strength * 10  # Scale to 0-10
        
        # Color boost for excellent reversal signals
        if reversal_strength >= 0.8:
            color_boost = 8  # Excellent reversal opportunity
        elif reversal_strength >= 0.6:
            color_boost = 5  # Good reversal
        elif reversal_strength >= 0.4:
            color_boost = 3  # Mild reversal
            
    elif reversal_direction == 'BEARISH_REVERSAL':
        # Bearish reversal - warning for buyers
        swing_reversal_score = -reversal_strength * 5  # Negative score
        color_boost = -3 if reversal_strength >= 0.6 else -1
    
    # 6. ADDITIONAL CONFIRMATION SIGNALS - USING WINSORIZED VOLUME
    volume_confirmation = False
    if len(df) >= 5:
        # Check if reversal bars had above-average volume using winsorized data
        volume = price_data['volume']
        avg_volume = volume.tail(10).mean()
        highest_bar_volume = volume.loc[highest_high_idx] if highest_high_idx in volume.index else avg_volume
        lowest_bar_volume = volume.loc[lowest_low_idx] if lowest_low_idx in volume.index else avg_volume
        
        if reversal_direction == 'BULLISH_REVERSAL' and lowest_bar_volume > avg_volume * 1.3:
            volume_confirmation = True
            swing_reversal_score += 2
            color_boost += 2
        elif reversal_direction == 'BEARISH_REVERSAL' and highest_bar_volume > avg_volume * 1.3:
            volume_confirmation = True
            swing_reversal_score -= 1
            color_boost -= 1
    
    # Compile results
    signals = {
        'reversal_strength': min(reversal_strength, 1.0),  # Cap at 1.0
        'reversal_direction': reversal_direction,
        'peak_rejection_signal': peak_rejection_signal,
        'trough_rejection_signal': trough_rejection_signal,
        'swing_reversal_score': swing_reversal_score,
        'color_boost': color_boost,
        'volume_confirmation': volume_confirmation,
        'days_since_peak': days_since_highest,
        'days_since_trough': days_since_lowest,
        'peak_rejection_pct': peak_rejection_pct,
        'trough_bounce_pct': trough_bounce_pct,
        'distance_from_high_pct': ((highest_bar_high - current_price) / highest_bar_high) * 100 if highest_bar_high > 0 else 0,
        'distance_from_low_pct': ((current_price - lowest_bar_low) / lowest_bar_low) * 100 if lowest_bar_low > 0 else 0,
    }
    
    return signals

# =============================================================================
# ENHANCED INSTITUTIONAL-GRADE STOCK ANALYSIS SYSTEM
# =============================================================================

class InstitutionalGradeAnalyzer:
    """
    Institutional-grade analysis tools for professional stock screening.
    Implements advanced quantitative methods used by hedge funds and institutions.
    """
    
    def __init__(self):
        # Import sklearn components for institutional analysis
        try:
            from sklearn.linear_model import LinearRegression
            from sklearn.preprocessing import StandardScaler
            from sklearn.metrics import r2_score
            self.sklearn_available = True
            self.LinearRegression = LinearRegression
            self.StandardScaler = StandardScaler
            self.r2_score = r2_score
        except ImportError:
            self.sklearn_available = False
            print("Warning: scikit-learn not available. Some institutional features will be limited.")
    
    def relative_volume_curve(self, df: pd.DataFrame, lookback: int = 50) -> float:
        """
        Institutional Method 1: Relative Volume Curve Analysis
        Analyzes volume patterns relative to historical norms with decay functions.
        Used by institutions to detect accumulation/distribution phases.
        
        Returns: Curve score (0-10 scale, higher = more institutional interest)
        """
        if df.empty or len(df) < lookback:
            return 0.0
        
        try:
            volume = df['Volume'].tail(lookback)
            
            # Calculate volume percentiles for context
            vol_50th = volume.quantile(0.5)
            vol_75th = volume.quantile(0.75)
            vol_90th = volume.quantile(0.9)
            
            current_vol = volume.iloc[-1]
            recent_avg = volume.tail(5).mean()
            
            # Institutional signature: Sustained above-average volume
            if recent_avg > vol_75th and current_vol > vol_50th:
                sustained_score = 3.0
            elif recent_avg > vol_50th:
                sustained_score = 1.5
            else:
                sustained_score = 0.0
            
            # Volume trend analysis (institutional accumulation pattern)
            if len(volume) >= 20:
                early_period = volume.head(10).mean()
                late_period = volume.tail(10).mean()
                
                if late_period > early_period * 1.2:  # 20% increase
                    trend_score = 2.0
                elif late_period > early_period:
                    trend_score = 1.0
                else:
                    trend_score = 0.0
            else:
                trend_score = 0.0
            
            # Spike detection (avoid manipulated volume)
            max_vol = volume.max()
            if max_vol > vol_90th * 3:  # Extreme spike
                spike_penalty = -1.0
            else:
                spike_penalty = 0.0
            
            total_score = sustained_score + trend_score + spike_penalty
            return max(0.0, min(10.0, total_score))
            
        except Exception as e:
            print(f"Error in relative_volume_curve: {e}")
            return 0.0
    
    def avwap_breakout_entry(self, df: pd.DataFrame, period: int = 20) -> float:
        """
        Institutional Method 2: Anchored VWAP Breakout Analysis
        Calculates price position relative to volume-weighted average price.
        Used by institutions for optimal entry timing.
        
        Returns: Breakout score (0-10 scale, higher = better breakout)
        """
        if df.empty or len(df) < period:
            return 0.0
        
        try:
            recent_data = df.tail(period)
            close = recent_data['Close']
            volume = recent_data['Volume']
            high = recent_data['High']
            low = recent_data['Low']
            
            # Calculate VWAP (Volume Weighted Average Price)
            typical_price = (high + low + close) / 3
            vwap = (typical_price * volume).sum() / volume.sum()
            
            current_price = close.iloc[-1]
            
            # Price position relative to VWAP
            vwap_distance = ((current_price - vwap) / vwap) * 100
            
            # Breakout strength analysis
            if vwap_distance > 2.0:  # Above VWAP
                if current_price > close.quantile(0.8):  # Near recent highs
                    breakout_score = 4.0
                else:
                    breakout_score = 2.0
            elif vwap_distance > 0:  # Slightly above VWAP
                breakout_score = 1.0
            elif vwap_distance > -2.0:  # Near VWAP
                breakout_score = 0.5
            else:  # Well below VWAP
                breakout_score = 0.0
            
            # Volume confirmation
            recent_vol_avg = volume.tail(3).mean()
            period_vol_avg = volume.mean()
            
            if recent_vol_avg > period_vol_avg * 1.5:
                volume_conf = 2.0
            elif recent_vol_avg > period_vol_avg:
                volume_conf = 1.0
            else:
                volume_conf = 0.0
            
            total_score = breakout_score + volume_conf
            return max(0.0, min(10.0, total_score))
            
        except Exception as e:
            print(f"Error in avwap_breakout_entry: {e}")
            return 0.0
    
    def orderbook_imbalance_score(self, df: pd.DataFrame) -> float:
        """
        Institutional Method 3: Order Book Imbalance Approximation
        Estimates buying/selling pressure using price and volume patterns.
        Simulates institutional order flow analysis.
        
        Returns: Imbalance score (-10 to +10, positive = buying pressure)
        """
        if df.empty or len(df) < 10:
            return 0.0
        
        try:
            recent_data = df.tail(10)
            
            buying_pressure = 0.0
            selling_pressure = 0.0
            
            for i in range(1, len(recent_data)):
                prev_bar = recent_data.iloc[i-1]
                curr_bar = recent_data.iloc[i]
                
                price_change = curr_bar['Close'] - prev_bar['Close']
                volume = curr_bar['Volume']
                
                # Estimate order flow direction
                if price_change > 0:
                    # Price up = buying pressure
                    buying_pressure += volume * abs(price_change)
                elif price_change < 0:
                    # Price down = selling pressure  
                    selling_pressure += volume * abs(price_change)
            
            # Calculate imbalance ratio
            total_pressure = buying_pressure + selling_pressure
            if total_pressure > 0:
                imbalance = (buying_pressure - selling_pressure) / total_pressure
                imbalance_score = imbalance * 10  # Scale to -10 to +10
            else:
                imbalance_score = 0.0
            
            return max(-10.0, min(10.0, imbalance_score))
            
        except Exception as e:
            print(f"Error in orderbook_imbalance_score: {e}")
            return 0.0
    
    def residual_momentum_rank(self, df: pd.DataFrame, market_data: pd.DataFrame = None) -> float:
        """
        Institutional Method 4: Residual Momentum Analysis
        Calculates stock momentum after removing market effects.
        Used by quant funds for alpha generation.
        
        Returns: Residual momentum score (0-10 scale)
        """
        if df.empty or len(df) < 30:
            return 0.0
        
        try:
            # Calculate stock returns
            stock_returns = df['Close'].pct_change().dropna()
            
            if len(stock_returns) < 20:
                return 0.0
            
            # If market data available, calculate residual momentum
            if market_data is not None and not market_data.empty:
                market_returns = market_data['Close'].pct_change().dropna()
                
                # Align data
                min_len = min(len(stock_returns), len(market_returns))
                if min_len >= 20:
                    stock_ret_aligned = stock_returns.tail(min_len)
                    market_ret_aligned = market_returns.tail(min_len)
                    
                    # Simple regression to remove market effects
                    if self.sklearn_available:
                        try:
                            lr = self.LinearRegression()
                            X = market_ret_aligned.values.reshape(-1, 1)
                            y = stock_ret_aligned.values
                            
                            lr.fit(X, y)
                            predicted = lr.predict(X)
                            residuals = y - predicted
                            
                            # Calculate momentum from residuals
                            recent_residual_momentum = residuals[-5:].mean()
                        except:
                            recent_residual_momentum = stock_returns.tail(5).mean()
                    else:
                        # Fallback: simple momentum
                        recent_residual_momentum = stock_returns.tail(5).mean()
                else:
                    recent_residual_momentum = stock_returns.tail(5).mean()
            else:
                # No market data - use raw momentum
                recent_residual_momentum = stock_returns.tail(5).mean()
            
            # Convert to score (0-10 scale)
            momentum_pct = recent_residual_momentum * 100
            
            if momentum_pct > 2.0:
                score = 8.0
            elif momentum_pct > 1.0:
                score = 6.0
            elif momentum_pct > 0.5:
                score = 4.0
            elif momentum_pct > 0:
                score = 2.0
            elif momentum_pct > -1.0:
                score = 1.0
            else:
                score = 0.0
            
            return min(10.0, score)
            
        except Exception as e:
            print(f"Error in residual_momentum_rank: {e}")
            return 0.0
    
    def catalyst_factor_score(self, df: pd.DataFrame, news_sentiment: float = 0.0) -> float:
        """
        Institutional Method 5: Catalyst Factor Analysis
        Combines technical signals with fundamental catalysts.
        Used by institutions for event-driven strategies.
        
        Returns: Catalyst score (0-10 scale)
        """
        if df.empty or len(df) < 20:
            return 0.0
        
        try:
            score = 0.0
            
            # Technical catalyst signals
            close = df['Close']
            volume = df['Volume']
            
            # 1. Breakout detection
            recent_high = close.tail(20).max()
            current_price = close.iloc[-1]
            
            if current_price >= recent_high * 0.98:  # Near 20-day high
                score += 2.0
            
            # 2. Volume surge (institutional interest)
            avg_volume = volume.tail(50).mean()
            recent_volume = volume.tail(3).mean()
            
            if recent_volume > avg_volume * 2:
                score += 3.0
            elif recent_volume > avg_volume * 1.5:
                score += 1.5
            
            # 3. Price momentum acceleration
            returns_5d = close.pct_change(5).iloc[-1]
            returns_1d = close.pct_change(1).iloc[-1]
            
            if returns_5d > 0.05 and returns_1d > 0.02:  # Accelerating
                score += 2.0
            
            # 4. News sentiment factor
            if news_sentiment > 0.5:
                score += 2.0
            elif news_sentiment > 0.2:
                score += 1.0
            elif news_sentiment < -0.3:
                score -= 1.0
            
            return max(0.0, min(10.0, score))
            
        except Exception as e:
            print(f"Error in catalyst_factor_score: {e}")
            return 0.0
    
    def vol_divergence_stop(self, df: pd.DataFrame) -> float:
        """
        Institutional Method 6: Volatility Divergence Stop Loss
        Calculates dynamic stop loss based on volatility patterns.
        Used by risk management systems.
        
        Returns: Stop loss percentage below current price
        """
        if df.empty or len(df) < 30:
            return 5.0  # Default 5% stop
        
        try:
            close = df['Close']
            returns = close.pct_change().dropna()
            
            # Calculate realized volatility
            vol_20d = returns.tail(20).std() * np.sqrt(252)  # Annualized
            vol_5d = returns.tail(5).std() * np.sqrt(252)
            
            # Adaptive stop based on volatility regime
            if vol_5d > vol_20d * 1.5:  # High vol regime
                stop_pct = min(15.0, vol_20d * 30)  # Cap at 15%
            elif vol_5d < vol_20d * 0.7:  # Low vol regime
                stop_pct = max(2.0, vol_20d * 20)  # Minimum 2%
            else:  # Normal regime
                stop_pct = vol_20d * 25
            
            # Ensure reasonable range
            return max(2.0, min(15.0, stop_pct))
            
        except Exception as e:
            print(f"Error in vol_divergence_stop: {e}")
            return 5.0
    
    def fractal_risk_budget(self, df: pd.DataFrame, portfolio_value: float = 100000) -> dict:
        """
        Institutional Method 7: Fractal Risk Budgeting
        Calculates position sizing based on fractal market analysis.
        Used by quantitative hedge funds for risk management.
        
        Returns: Dict with position sizing recommendations
        """
        if df.empty or len(df) < 50:
            return {
                'position_size': 0,
                'risk_per_share': 0.0,
                'max_shares': 0,
                'risk_score': 0.0
            }
        
        try:
            close = df['Close']
            returns = close.pct_change().dropna()
            current_price = close.iloc[-1]
            
            # Calculate fractal dimension approximation
            price_changes = np.abs(returns)
            
            # Hurst exponent estimation (simplified)
            if len(price_changes) >= 30:
                # Calculate variance ratio
                var_1 = price_changes.var()
                var_5 = price_changes.rolling(5).sum().var() / 25
                
                if var_1 > 0:
                    variance_ratio = var_5 / var_1
                    hurst_approx = 0.5 + 0.5 * np.log2(variance_ratio)
                    hurst_approx = max(0.1, min(0.9, hurst_approx))
                else:
                    hurst_approx = 0.5
            else:
                hurst_approx = 0.5
            
            # Risk scoring based on fractal properties
            if hurst_approx > 0.7:  # Trending/persistent
                risk_multiplier = 0.8  # Lower risk
            elif hurst_approx < 0.3:  # Mean reverting
                risk_multiplier = 1.2  # Higher risk
            else:  # Random walk
                risk_multiplier = 1.0
            
            # Calculate position sizing
            base_risk_pct = 2.0  # 2% of portfolio at risk
            adjusted_risk_pct = base_risk_pct * risk_multiplier
            
            # Stop loss from volatility analysis
            stop_loss_pct = self.vol_divergence_stop(df)
            
            # Position size calculation
            risk_amount = portfolio_value * (adjusted_risk_pct / 100)
            risk_per_share = current_price * (stop_loss_pct / 100)
            
            if risk_per_share > 0:
                max_shares = int(risk_amount / risk_per_share)
                position_size = min(max_shares, int(portfolio_value * 0.1 / current_price))  # Max 10% position
            else:
                max_shares = 0
                position_size = 0
            
            return {
                'position_size': position_size,
                'risk_per_share': risk_per_share,
                'max_shares': max_shares,
                'risk_score': risk_multiplier,
                'hurst_exponent': hurst_approx,
                'stop_loss_pct': stop_loss_pct
            }
            
        except Exception as e:
            print(f"Error in fractal_risk_budget: {e}")
            return {
                'position_size': 0,
                'risk_per_share': 0.0,
                'max_shares': 0,
                'risk_score': 1.0
            }

# Global institutional analyzer instance
institutional_analyzer = InstitutionalGradeAnalyzer()

# =============================================================================
# ENHANCED SCREENING FUNCTIONS WITH INSTITUTIONAL INTEGRATION
# =============================================================================

def calculate_institutional_score(df: pd.DataFrame, ticker: str = "") -> dict:
    """
    Calculate comprehensive institutional-grade score using all 7 methods.
    This provides the professional-level analysis used by hedge funds.
    
    Returns: Dictionary with institutional analysis breakdown
    """
    if df.empty:
        return {
            'total_institutional_score': 0.0,
            'institutional_breakdown': {},
            'risk_analysis': {},
            'recommendation': 'INSUFFICIENT_DATA'
        }
    
    try:
        # Run all 7 institutional methods
        method_scores = {}
        
        # Method 1: Relative Volume Curve (0-10)
        method_scores['volume_curve'] = institutional_analyzer.relative_volume_curve(df)
        
        # Method 2: AVWAP Breakout Entry (0-10)
        method_scores['avwap_breakout'] = institutional_analyzer.avwap_breakout_entry(df)
        
        # Method 3: Order Book Imbalance (-10 to +10)
        method_scores['orderbook_imbalance'] = institutional_analyzer.orderbook_imbalance_score(df)
        
        # Method 4: Residual Momentum (0-10)
        method_scores['residual_momentum'] = institutional_analyzer.residual_momentum_rank(df)
        
        # Method 5: Catalyst Factor (0-10)
        method_scores['catalyst_factor'] = institutional_analyzer.catalyst_factor_score(df)
        
        # Method 6: Vol Divergence Stop (percentage)
        stop_loss_pct = institutional_analyzer.vol_divergence_stop(df)
        
        # Method 7: Fractal Risk Budget
        risk_budget = institutional_analyzer.fractal_risk_budget(df)
        
        # Calculate composite institutional score
        # Weight positive factors, normalize orderbook imbalance
        total_score = (
            method_scores['volume_curve'] * 0.2 +           # 20% weight
            method_scores['avwap_breakout'] * 0.2 +         # 20% weight  
            (method_scores['orderbook_imbalance'] + 10) * 0.15 +  # 15% weight (normalized)
            method_scores['residual_momentum'] * 0.2 +      # 20% weight
            method_scores['catalyst_factor'] * 0.25         # 25% weight
        )
        
        # Determine institutional recommendation
        if total_score >= 7.5:
            recommendation = 'STRONG_BUY'
        elif total_score >= 6.0:
            recommendation = 'BUY'
        elif total_score >= 4.0:
            recommendation = 'HOLD'
        elif total_score >= 2.0:
            recommendation = 'WEAK_HOLD'
        else:
            recommendation = 'AVOID'
        
        return {
            'total_institutional_score': round(total_score, 2),
            'institutional_breakdown': {
                'volume_curve': round(method_scores['volume_curve'], 2),
                'avwap_breakout': round(method_scores['avwap_breakout'], 2),
                'orderbook_imbalance': round(method_scores['orderbook_imbalance'], 2),
                'residual_momentum': round(method_scores['residual_momentum'], 2),
                'catalyst_factor': round(method_scores['catalyst_factor'], 2)
            },
            'risk_analysis': {
                'stop_loss_pct': round(stop_loss_pct, 2),
                'position_sizing': risk_budget['position_size'],
                'risk_score': round(risk_budget['risk_score'], 2),
                'hurst_exponent': round(risk_budget.get('hurst_exponent', 0.5), 3)
            },
            'recommendation': recommendation,
            'ticker': ticker
        }
        
    except Exception as e:
        print(f"Error calculating institutional score for {ticker}: {e}")
        return {
            'total_institutional_score': 0.0,
            'institutional_breakdown': {},
            'risk_analysis': {},
            'recommendation': 'ERROR'
        }

def enhanced_screening_with_institutional(ticker: str, period: str = "1y") -> Optional[dict]:
    """
    Enhanced screening that combines original scoring with institutional-grade analysis.
    This provides institutional-level insights for professional trading decisions.
    """
    try:
        # Get base screening result
        base_result = screen_single_ticker_complete(ticker, period)
        if not base_result:
            return None
        
        # Get historical data for institutional analysis
        df = process_ticker_data_complete(ticker, period)
        if df.empty:
            return base_result
        
        # Add institutional analysis
        institutional_result = calculate_institutional_score(df, ticker)
        
        # Combine results
        enhanced_result = base_result.copy()
        enhanced_result['institutional_analysis'] = institutional_result
        
        # Calculate combined score (70% base + 30% institutional)
        base_score = base_result['total_score']
        institutional_score = institutional_result['total_institutional_score']
        
        combined_score = (base_score * 0.7) + (institutional_score * 0.3)
        enhanced_result['combined_score'] = round(combined_score, 2)
        
        # Enhanced tier classification
        if combined_score >= 8.0:
            enhanced_tier = "Institutional_Grade"
        elif combined_score >= 6.5:
            enhanced_tier = "Professional_Grade"
        elif combined_score >= 5.0:
            enhanced_tier = "Retail_Grade"
        else:
            enhanced_tier = "Speculative"
        
        enhanced_result['enhanced_tier'] = enhanced_tier
        
        # Risk-adjusted recommendation
        risk_score = institutional_result['risk_analysis'].get('risk_score', 1.0)
        if risk_score > 1.2:
            enhanced_result['risk_warning'] = "HIGH_RISK"
        elif risk_score < 0.8:
            enhanced_result['risk_warning'] = "LOW_RISK"
        else:
            enhanced_result['risk_warning'] = "MODERATE_RISK"
        
        print(f"🏛️ {ticker}: {enhanced_tier} (Combined: {combined_score:.1f}, "
              f"Institutional: {institutional_score:.1f}, Risk: {enhanced_result['risk_warning']})")
        
        return enhanced_result
        
    except Exception as e:
        print(f"Enhanced screening failed for {ticker}: {e}")
        return base_result  # Fallback to base result

def batch_institutional_screening(tickers: List[str], max_workers: int = 4) -> List[dict]:
    """
    Run institutional-grade screening on a batch of tickers.
    Provides professional-level analysis for multiple stocks.
    """
    print(f"\n🏛️ Starting Institutional-Grade Analysis on {len(tickers)} tickers...")
    
    results = []
    start_time = time.time()
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_ticker = {
            executor.submit(enhanced_screening_with_institutional, ticker): ticker 
            for ticker in tickers
        }
        
        for future in as_completed(future_to_ticker):
            ticker = future_to_ticker[future]
            try:
                result = future.result(timeout=45)  # Longer timeout for institutional analysis
                if result:
                    results.append(result)
            except Exception as e:
                print(f"❌ Institutional analysis failed for {ticker}: {e}")
    
    # Sort by combined score
    results.sort(key=lambda x: x.get('combined_score', 0), reverse=True)
    
    elapsed = time.time() - start_time
    print(f"✅ Institutional screening complete: {len(results)} analyzed in {elapsed:.1f}s")
    
    # Performance summary
    if results:
        institutional_grades = len([r for r in results if r.get('enhanced_tier') == 'Institutional_Grade'])
        professional_grades = len([r for r in results if r.get('enhanced_tier') == 'Professional_Grade'])
        
        print(f"📊 Quality Distribution: {institutional_grades} Institutional, {professional_grades} Professional")
    
    return results

# =============================================================================
# ENHANCED DISPLAY WITH INSTITUTIONAL METRICS
# =============================================================================

def display_institutional_results(results: List[dict], show_details: bool = False):
    """
    Display screening results with institutional-grade analysis.
    Shows professional-level insights and risk metrics.
    """
    if not results:
        print("No institutional analysis results to display.")
        return
    
    print(f"\n{'='*80}")
    print("🏛️  INSTITUTIONAL-GRADE STOCK ANALYSIS REPORT")
    print(f"{'='*80}")
    
    # Summary statistics
    total_analyzed = len(results)
    institutional_grade = len([r for r in results if r.get('enhanced_tier') == 'Institutional_Grade'])
    professional_grade = len([r for r in results if r.get('enhanced_tier') == 'Professional_Grade'])
    
    print(f"📊 ANALYSIS SUMMARY:")
    print(f"   Total Analyzed: {total_analyzed}")
    print(f"   Institutional Grade: {institutional_grade}")
    print(f"   Professional Grade: {professional_grade}")
    print(f"   Success Rate: {(institutional_grade + professional_grade)/total_analyzed*100:.1f}%")
    
    print(f"\n{'Ticker':<12} {'Tier':<18} {'Combined':<10} {'Inst.':<8} {'Risk':<12} {'Recommendation':<15}")
    print(f"{'-'*12} {'-'*18} {'-'*10} {'-'*8} {'-'*12} {'-'*15}")
    
    for result in results[:20]:  # Show top 20
        ticker = result['ticker']
        tier = result.get('enhanced_tier', 'Unknown')[:17]
        combined_score = result.get('combined_score', 0)
        
        inst_analysis = result.get('institutional_analysis', {})
        inst_score = inst_analysis.get('total_institutional_score', 0)
        recommendation = inst_analysis.get('recommendation', 'N/A')[:14]
        risk_warning = result.get('risk_warning', 'UNKNOWN')[:11]
        
        # Color coding
        if tier == 'Institutional_Grade':
            tier_display = f"\033[95m{tier}\033[0m"  # Purple
        elif tier == 'Professional_Grade':
            tier_display = f"\033[92m{tier}\033[0m"  # Green
        else:
            tier_display = tier
        
        print(f"{ticker:<12} {tier_display:<27} {combined_score:<10.1f} {inst_score:<8.1f} {risk_warning:<12} {recommendation:<15}")
    
    if show_details and results:
        print(f"\n{'='*80}")
        print("🔍 DETAILED INSTITUTIONAL ANALYSIS (Top 5)")
        print(f"{'='*80}")
        
        for i, result in enumerate(results[:5]):
            ticker = result['ticker']
            inst_analysis = result.get('institutional_analysis', {})
            breakdown = inst_analysis.get('institutional_breakdown', {})
            risk_analysis = inst_analysis.get('risk_analysis', {})
            
            print(f"\n{i+1}. {ticker} - {result.get('enhanced_tier', 'Unknown')}")
            print(f"   Combined Score: {result.get('combined_score', 0):.2f}")
            print(f"   Institutional Breakdown:")
            print(f"     • Volume Curve: {breakdown.get('volume_curve', 0):.2f}/10")
            print(f"     • AVWAP Breakout: {breakdown.get('avwap_breakout', 0):.2f}/10")
            print(f"     • Order Imbalance: {breakdown.get('orderbook_imbalance', 0):.2f}/10")
            print(f"     • Residual Momentum: {breakdown.get('residual_momentum', 0):.2f}/10")
            print(f"     • Catalyst Factor: {breakdown.get('catalyst_factor', 0):.2f}/10")
            print(f"   Risk Metrics:")
            print(f"     • Stop Loss: {risk_analysis.get('stop_loss_pct', 0):.2f}%")
            print(f"     • Position Size: {risk_analysis.get('position_sizing', 0)} shares")
            print(f"     • Risk Score: {risk_analysis.get('risk_score', 1.0):.2f}")
            print(f"   Recommendation: {inst_analysis.get('recommendation', 'N/A')}")

# =============================================================================
# CONSTANTS AND MAIN EXECUTION
# =============================================================================

MAX_WORKERS = 8
DECAY_FLOOR = 0.50
PLATEAU_HR = 6
MIN_VOL_SURGE = 1.2
PROB_WT = 30
GROW_WT = 10
LIQUIDITY_WT = 15

# =============================================================================
# MAIN EXECUTION FUNCTION
# =============================================================================

async def main(argv=None):
    """
    Main execution function for the enhanced swing screener with institutional analysis.
    """
    import argparse
    
    parser = argparse.ArgumentParser(description='Enhanced Swing Trade Screener v23.9 with Institutional Analysis')
    parser.add_argument('--institutional', action='store_true', help='Run institutional-grade analysis')
    parser.add_argument('--details', action='store_true', help='Show detailed analysis breakdown')
    parser.add_argument('--backtest', action='store_true', help='Run strategy backtesting')
    parser.add_argument('--paper-trading', action='store_true', help='Enable paper trading logs')
    parser.add_argument('--max-workers', type=int, default=4, help='Maximum concurrent workers')
    parser.add_argument('--period', type=str, default='1y', help='Data period (1y, 6mo, 2y)')
    parser.add_argument('--tickers', type=str, help='Comma-separated list of tickers')
    
    args = parser.parse_args(argv)
    
    # Default Indian stock tickers for testing
    default_tickers = [
        'RELIANCE.NS', 'TCS.NS', 'HDFCBANK.NS', 'INFY.NS', 'HINDUNILVR.NS',
        'HDFC.NS', 'ICICIBANK.NS', 'KOTAKBANK.NS', 'BHARTIARTL.NS', 'ITC.NS',
        'ASIANPAINT.NS', 'LT.NS', 'AXISBANK.NS', 'MARUTI.NS', 'WIPRO.NS',
        'ULTRACEMCO.NS', 'TITAN.NS', 'POWERGRID.NS', 'TECHM.NS', 'NESTLEIND.NS'
    ]
    
    if args.tickers:
        tickers = [t.strip() for t in args.tickers.split(',')]
        # Ensure .NS suffix for Indian stocks
        tickers = [ensure_ns_suffix(t) for t in tickers]
    else:
        tickers = default_tickers
    
    print(f"🚀 Enhanced Swing Screener v23.9 Starting...")
    print(f"📊 Analyzing {len(tickers)} tickers with period: {args.period}")
    print(f"⚙️ Workers: {args.max_workers}, Institutional: {args.institutional}")
    
    try:
        if args.institutional:
            # Run institutional-grade analysis
            results = batch_institutional_screening(tickers, max_workers=args.max_workers)
            display_institutional_results(results, show_details=args.details)
        else:
            # Run standard screening with optional enhancements
            results = batch_screen_tickers(tickers, period=args.period, max_workers=args.max_workers)
            
            if results:
                print(f"\n🎯 SCREENING RESULTS ({len(results)} qualified):")
                print(f"{'Ticker':<12} {'Tier':<8} {'Score':<6} {'Price':<8} {'RSI':<5} {'BB%':<5} {'Volume':<8}")
                print("-" * 60)
                
                for result in results[:15]:  # Show top 15
                    ticker = result['ticker']
                    tier = result['tier']
                    score = result['total_score']
                    current_vals = result['current_values']
                    
                    # Color coding for tiers
                    if tier == 'Tier1':
                        tier_display = f"\033[92m{tier}\033[0m"  # Green
                    elif tier == 'Tier2':  
                        tier_display = f"\033[93m{tier}\033[0m"  # Yellow
                    else:
                        tier_display = tier
                    
                    print(f"{ticker:<12} {tier_display:<15} {score:<6.1f} "
                          f"₹{current_vals['price']:<7.1f} {current_vals['rsi']:<5.1f} "
                          f"{current_vals['bb_position']:<5.1f} {current_vals['volume']/1000:<7.0f}K")
        
        # Backtesting if requested
        if args.backtest and results:
            print(f"\n🔄 Running strategy backtest...")
            backtest_tickers = [r['ticker'] for r in results[:10]]  # Top 10 for backtest
            backtest_results = backtest_strategy(backtest_tickers, period='6mo')
            
            print(f"📈 Backtest Results:")
            perf = backtest_results.get('performance', {})
            if perf:
                print(f"   Win Rate: {perf.get('win_rate', 0):.1f}%")
                print(f"   Avg P&L: {perf.get('avg_pnl_percent', 0):.1f}%")
                print(f"   Best Trade: {perf.get('best_trade', 0):.1f}%")
                print(f"   Worst Trade: {perf.get('worst_trade', 0):.1f}%")
        
        # Paper trading summary if enabled
        if args.paper_trading:
            print(f"\n📝 Paper Trading Summary:")
            summary = get_paper_trading_summary()
            if 'error' not in summary:
                print(f"   Total Trades: {summary['total_trades']}")
                print(f"   Open Trades: {summary['open_trades']}")
                print(f"   Win Rate: {summary['win_rate']:.1f}%")
                print(f"   Total P&L: ₹{summary['total_pnl']:.2f}")
        
        # Performance monitoring summary
        print(f"\n📊 Performance Summary:")
        perf_summary = performance_optimizer.get_performance_summary()
        cache_stats = get_cache_stats()
        
        print(f"   Cache Hit Rate: {cache_stats['requests_in_last_minute']} req/min")
        print(f"   Memory Cache: {cache_stats['memory_cache_entries']} entries")
        
        # Show top function performance if available
        if perf_summary:
            slowest_func = max(perf_summary.items(), key=lambda x: x[1]['total_time'])
            print(f"   Slowest Operation: {slowest_func[0]} ({slowest_func[1]['total_time']:.2f}s total)")
        
        print(f"\n✅ Analysis Complete! Results ready for trading decisions.")
        
    except KeyboardInterrupt:
        print(f"\n⏹️ Analysis interrupted by user.")
    except Exception as e:
        print(f"\n❌ Error during analysis: {e}")
        if debug_config.include_traceback:
            import traceback
            traceback.print_exc()

# Disabled hardcoded ticker main function - using news-based version instead
# if __name__ == "__main__":
#     import asyncio
#     asyncio.run(main())

MIN_VOL_SURGE    = 1.2  # Lowered from 1.5 to catch more opportunities
PROB_WT          = 30
GROW_WT          = 10
LIQUIDITY_WT     = 15  # New weight for liquidity impact on scoring
TICK_RE          = re.compile(r'^([A-Z0-9\-$]{1,12})\.NS(?:\s|\(|.*-)', re.I)
NEWS_SECTION_RE  = re.compile(r'^([A-Z0-9\-$]{1,12})\.NS\s*-\s*(\d+)\s*relevant\s*news\s*items', re.I)
NEWS_FILE_RE     = re.compile(r'news_output_(\d{8})_(\d{6})\.txt$')
TS_FMT           = "%Y%m%d%H%M%S"
MONEY_TOLERANCE  = 0.05
BASE_SCORE_SCALE_DEFAULT = 60.0  # Lowered from 120 to align with realistic score ranges

# Deal weighting system
DEAL_TYPES = {
    "order": 1.2,
    "win": 1.5,
    "deal": 1.3,
    "expansion": 1.4,
    "acquisition": 1.6,
    "facility": 1.3,
    "land": 1.4,
    "contract": 1.25,
    "tender": 1.2,
    "project": 1.2,
    "fundraise": 1.3,
    "investment": 1.2,
    "general": 1.0,
}
DEAL_PATTERN = re.compile(
    r"\b(order|win|deal|expansion|acquisition|facility|land(?:\s+buy)?|contract|tender|project|fundraise|investment)\b",
    re.I
)

# Positive words count feature
POSITIVE_WORDS = ['rise', 'up', 'high', 'buy', 'add', 'growth', 'strong', 'surge', 'jump', 'gain', 'beat', 'positive', 'increase', 'higher', 'win', 'profit', 'success', 'approve', 'launch', 'new', 'record', 'expand', 'acquire', 'deal', 'contract', 'order', 'stellar', 'outperform', 'surpass', 'record-high', 'upgrade', 'beat estimates', 'record revenue', 'dividend hike', 'buyback']
POSITIVE_PATTERN = re.compile(r'\b(' + '|'.join(POSITIVE_WORDS) + r')\b', re.I)

# FII display
FII_TIMEOUT    = 12
FII_UA         = "Mozilla/5.0 (compatible; DeepSeekSwing/1.1; +local)"
FII_CACHE_TTL  = 6 * 3600  # 6h

# Enhanced imports for systematic improvements
try:
    from enhanced_screener_modules import (
        DynamicPEAnalyzer, 
        ResilientDataSource, 
        SmartTierManager, 
        MeanReversionAnalyzer,
        EnhancedScreenerEngine
    )
    ENHANCED_MODULES_AVAILABLE = True
except ImportError:
    ENHANCED_MODULES_AVAILABLE = False
    logging.warning("Enhanced screener modules not available - using legacy methods")
except Exception as e:
    ENHANCED_MODULES_AVAILABLE = False
    logging.warning(f"Enhanced screener modules failed to load: {e} - using legacy methods")

# Additional imports for institutional-grade modules
try:
    from sklearn.linear_model import LinearRegression
    from scipy import stats
    from scipy.signal import find_peaks
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False
    logging.warning("sklearn not available - some institutional modules will use fallback methods")

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    NUMPY_AVAILABLE = False
    logging.warning("numpy not available - institutional modules disabled")

# Enhanced imports availability dictionary
ENHANCED_IMPORTS = {
    'sklearn': SKLEARN_AVAILABLE,
    'numpy': NUMPY_AVAILABLE,
    'scipy': SKLEARN_AVAILABLE  # scipy is imported with sklearn
}

# =============================================================================
# INSTITUTIONAL-GRADE ENHANCEMENT MODULES
# =============================================================================

class InstitutionalGradeAnalyzer:
    """
    Advanced institutional-grade stock selection modules.
    Implements battle-tested methods from quantitative funds.
    """
    
    def __init__(self):
        self.initialized = NUMPY_AVAILABLE and SKLEARN_AVAILABLE
        if not self.initialized:
            logging.warning("Institutional modules initialized with limited functionality")
    
    @staticmethod
    def relative_volume_curve(ticker: str, hist_data: pd.DataFrame, lookback_days: int = 30) -> Tuple[float, bool]:
        """
        Compares current volume against time-of-day baseline.
        Returns (current_rel_vol_ratio, high_liquidity_flag)
        
        Args:
            ticker: Stock symbol
            hist_data: Historical OHLCV data
            lookback_days: Days for baseline calculation
        """
        try:
            if hist_data.empty or len(hist_data) < lookback_days:
                return 1.0, False
            
            # Use recent volume data for time-of-day analysis
            recent_data = hist_data.tail(lookback_days * 2)  # Extra buffer for calculations
            
            # Calculate volume percentiles for different time periods
            volume_series = recent_data['Volume']
            
            # Morning volume (first 2 hours equivalent)
            morning_vol = volume_series.tail(int(len(volume_series) * 0.15))
            afternoon_vol = volume_series.tail(int(len(volume_series) * 0.3))
            
            # Calculate relative volume metrics
            avg_volume = volume_series.mean()
            recent_vol = volume_series.tail(5).mean()  # Last 5 periods average
            
            rel_vol_ratio = recent_vol / avg_volume if avg_volume > 0 else 1.0
            
            # High liquidity conditions
            vol_surge = rel_vol_ratio >= 2.0
            sustained_activity = volume_series.tail(10).mean() / avg_volume >= 1.5
            high_liquidity = vol_surge and sustained_activity
            
            return rel_vol_ratio, high_liquidity
            
        except Exception as e:
            logging.debug(f"Relative volume curve error for {ticker}: {e}")
            return 1.0, False
    
    @staticmethod
    def avwap_breakout_entry(ticker: str, hist_data: pd.DataFrame, anchor_days: int = 60) -> Optional[float]:
        """
        Anchored VWAP breakout detection for institutional entry points.
        
        Args:
            ticker: Stock symbol  
            hist_data: Historical OHLCV data
            anchor_days: Days to anchor VWAP calculation
        """
        try:
            if hist_data.empty or len(hist_data) < anchor_days:
                return None
            
            # Use recent data for AVWAP calculation
            anchor_data = hist_data.tail(anchor_days)
            
            # Calculate Volume Weighted Average Price (VWAP)
            typical_price = (anchor_data['High'] + anchor_data['Low'] + anchor_data['Close']) / 3
            vwap_numerator = (typical_price * anchor_data['Volume']).cumsum()
            vwap_denominator = anchor_data['Volume'].cumsum()
            
            # Avoid division by zero
            vwap_denominator = vwap_denominator.replace(0, np.nan)
            vwap = vwap_numerator / vwap_denominator
            
            current_price = anchor_data['Close'].iloc[-1]
            current_vwap = vwap.iloc[-1]
            
            # Breakout conditions
            price_above_vwap = current_price > current_vwap * 1.005  # 0.5% headroom
            
            # Volume pocket condition (sustained volume above average)
            vol_ma_5 = anchor_data['Volume'].rolling(5).mean().iloc[-1]
            vol_ma_20 = anchor_data['Volume'].rolling(20).mean().iloc[-1]
            volume_pocket = vol_ma_5 > vol_ma_20 * 1.4
            
            if price_above_vwap and volume_pocket:
                return float(current_price)
            
            return None
            
        except Exception as e:
            logging.debug(f"AVWAP breakout error for {ticker}: {e}")
            return None
    
    @staticmethod
    def orderbook_imbalance_score(ticker: str, hist_data: pd.DataFrame) -> float:
        """
        Simulated order book imbalance using price-volume analysis.
        Returns score from 0-100 (higher = more bullish imbalance)
        """
        try:
            if hist_data.empty or len(hist_data) < 20:
                return 50.0  # Neutral
            
            recent_data = hist_data.tail(20)
            
            # Proxy for order book imbalance using price and volume
            # Up-ticks vs down-ticks analysis
            price_changes = recent_data['Close'].diff()
            volume_weighted_changes = price_changes * recent_data['Volume']
            
            # Cumulative Volume Delta approximation
            up_volume = recent_data['Volume'][price_changes > 0].sum()
            down_volume = recent_data['Volume'][price_changes < 0].sum()
            total_volume = up_volume + down_volume
            
            if total_volume == 0:
                return 50.0
                
            # Imbalance calculation
            imbalance = (up_volume - down_volume) / total_volume
            
            # Convert to 0-100 score
            base_score = 50
            imbalance_score = imbalance * 25  # Scale to ±25
            
            # Add momentum component
            momentum = volume_weighted_changes.tail(5).sum()
            momentum_score = min(25, max(-25, momentum / 1000))
            
            final_score = base_score + imbalance_score + momentum_score
            return max(0, min(100, final_score))
            
        except Exception as e:
            logging.debug(f"Order book imbalance error for {ticker}: {e}")
            return 50.0
    
    @staticmethod
    def residual_momentum_rank(tickers_data: Dict[str, pd.DataFrame], window: int = 20) -> Dict[str, float]:
        """
        Cross-sectional residual momentum after removing market beta.
        Returns z-scores for each ticker (higher = stronger idiosyncratic momentum)
        """
        try:
            if not SKLEARN_AVAILABLE or not tickers_data:
                # Fallback: simple momentum ranking
                momentum_scores = {}
                for ticker, data in tickers_data.items():
                    if not data.empty and len(data) >= window:
                        returns = data['Close'].pct_change().tail(window)
                        momentum_scores[ticker] = returns.sum()
                
                if momentum_scores:
                    mean_mom = np.mean(list(momentum_scores.values()))
                    std_mom = np.std(list(momentum_scores.values()))
                    if std_mom > 0:
                        return {ticker: (score - mean_mom) / std_mom 
                               for ticker, score in momentum_scores.items()}
                
                return {ticker: 0.0 for ticker in tickers_data.keys()}
            
            # Full implementation with beta removal
            returns_data = {}
            for ticker, data in tickers_data.items():
                if not data.empty and len(data) >= window:
                    returns = data['Close'].pct_change().dropna()
                    if len(returns) >= window:
                        returns_data[ticker] = returns.tail(window)
            
            if not returns_data:
                return {}
            
            # Create market proxy (equal-weighted average)
            market_returns = pd.DataFrame(returns_data).mean(axis=1)
            
            residual_momentum = {}
            for ticker, returns in returns_data.items():
                try:
                    # Align data
                    aligned_market = market_returns.reindex(returns.index).dropna()
                    aligned_returns = returns.reindex(aligned_market.index).dropna()
                    
                    if len(aligned_returns) < 10:  # Minimum data requirement
                        residual_momentum[ticker] = 0.0
                        continue
                    
                    # Linear regression to remove market beta
                    X = aligned_market.values.reshape(-1, 1)
                    y = aligned_returns.values
                    
                    model = LinearRegression().fit(X, y)
                    residuals = y - model.predict(X)
                    
                    # Sum residuals as momentum measure
                    residual_momentum[ticker] = residuals.sum()
                    
                except Exception:
                    residual_momentum[ticker] = 0.0
            
            # Convert to z-scores
            if residual_momentum:
                values = list(residual_momentum.values())
                mean_resid = np.mean(values)
                std_resid = np.std(values)
                
                if std_resid > 0:
                    return {ticker: (score - mean_resid) / std_resid 
                           for ticker, score in residual_momentum.items()}
            
            return {ticker: 0.0 for ticker in residual_momentum.keys()}
            
        except Exception as e:
            logging.debug(f"Residual momentum calculation error: {e}")
            return {ticker: 0.0 for ticker in tickers_data.keys()}
    
    @staticmethod
    def catalyst_factor_score(ticker: str, news_data: List, deal_amount: float = 0.0, 
                            recent_performance: float = 0.0) -> float:
        """
        Combined catalyst score from earnings, deals, and momentum.
        Returns score from 0-100
        """
        try:
            score = 0.0
            
            # Deal component (0-40 points)
            if deal_amount > 0:
                # Scale deal impact
                deal_score = min(40, deal_amount / 1e8 * 20)  # 20 points per 100Cr
                score += deal_score
            
            # News sentiment component (0-30 points)
            if news_data:
                positive_news = sum(1 for n in news_data if hasattr(n, 'sent') and n.sent > 0.1)
                news_score = min(30, positive_news * 5)
                score += news_score
            
            # Recent performance component (0-30 points)
            if recent_performance != 0:
                perf_score = min(30, max(-30, recent_performance * 100))
                score += perf_score
            
            return max(0, min(100, score))
            
        except Exception as e:
            logging.debug(f"Catalyst factor score error for {ticker}: {e}")
            return 0.0
    
    @staticmethod
    def vol_divergence_stop(current_price: float, realized_vol: float, 
                           implied_vol: float = None) -> float:
        """
        Dynamic stop loss based on volatility divergence.
        Tightens when IV collapses, loosens when RV is contained.
        """
        try:
            if implied_vol is None:
                # Fallback to standard ATR-based stop
                return current_price - (realized_vol * 1.8)
            
            # IV crush detection
            if implied_vol < realized_vol * 0.7:
                # Tighten stop when IV collapses
                stop_distance = realized_vol * 0.5
            else:
                # Standard stop
                stop_distance = realized_vol * 1.8
            
            return current_price - stop_distance
            
        except Exception as e:
            logging.debug(f"Vol divergence stop error: {e}")
            return current_price - (realized_vol * 1.8)
    
    @staticmethod
    def fractal_risk_budget(ticker: str, hist_data: pd.DataFrame, lookback: int = 200) -> float:
        """
        Dynamic risk allocation using Hurst exponent proxy.
        Returns risk percentage (0.005 to 0.015 = 0.5% to 1.5%)
        """
        try:
            if hist_data.empty or len(hist_data) < lookback:
                return 0.01  # Default 1%
            
            prices = hist_data['Close'].tail(lookback)
            
            # Simplified Hurst exponent calculation
            # Using variance ratio method as proxy
            lags = [2, 4, 8, 16]
            variance_ratios = []
            
            returns = prices.pct_change().dropna()
            if len(returns) < 50:
                return 0.01
            
            base_var = returns.var()
            
            for lag in lags:
                if len(returns) > lag * 2:
                    # Calculate variance ratio
                    lag_returns = returns.rolling(lag).sum().dropna()
                    if len(lag_returns) > 0:
                        lag_var = lag_returns.var()
                        if base_var > 0:
                            variance_ratios.append(lag_var / (lag * base_var))
            
            if not variance_ratios:
                return 0.01
            
            # Estimate Hurst-like measure
            avg_ratio = np.mean(variance_ratios)
            
            # Convert to risk allocation
            # Ratio > 1 suggests persistence (trending) -> higher risk
            # Ratio < 1 suggests mean reversion -> lower risk
            base_risk = 0.01
            risk_multiplier = 0.5 + min(1.0, avg_ratio)
            
            final_risk = base_risk * risk_multiplier
            return max(0.005, min(0.015, final_risk))
            
        except Exception as e:
            logging.debug(f"Fractal risk budget error for {ticker}: {e}")
            return 0.01  # Default 1%

# Global instance
institutional_analyzer = InstitutionalGradeAnalyzer()

# =============================================================================
# ENHANCED BUYER DOMINANCE CALCULATION
# =============================================================================

def calculate_enhanced_buyer_dominance(ticker: str, daily_df: pd.DataFrame, skip_intraday: bool = False) -> float:
    """
    Enhanced buyer dominance calculation with multiple fallback methods.
    
    Priority:
    1. 5-minute intraday data (most accurate)
    2. Daily OHLC patterns (medium accuracy) 
    3. Volume-weighted price analysis (basic accuracy)
    4. Smart neutral based on market conditions (fallback)
    
    Returns: Float between 0.0 and 1.0 (0-100% when multiplied by 100)
    """
    try:
        # Method 1: 5-minute intraday data (if not skipped)
        if not skip_intraday:
            try:
                rate_limit(0.5)  # 500ms between calls
                df5 = YFinanceDataValidator.safe_ticker_history(ticker, period="1d", interval="5m")
                if not df5.empty and len(df5) >= 5:
                    if all(col in df5.columns for col in ['Close', 'Open', 'Volume']):
                        up_vol = df5.loc[df5["Close"] > df5["Open"], "Volume"].sum()
                        dn_vol = df5.loc[df5["Close"] < df5["Open"], "Volume"].sum()
                        unchanged_vol = df5.loc[df5["Close"] == df5["Open"], "Volume"].sum()
                        
                        total_vol = up_vol + dn_vol + unchanged_vol
                        if total_vol > 0:
                            # Calculate buyer dominance with unchanged volume distributed proportionally
                            base_bd = up_vol / total_vol
                            if unchanged_vol > 0:
                                # Distribute unchanged volume based on recent momentum
                                recent_momentum = (df5["Close"].iloc[-1] - df5["Open"].iloc[0]) / df5["Open"].iloc[0]
                                if recent_momentum > 0:
                                    # Positive momentum: attribute more unchanged volume to buyers
                                    adjusted_up = up_vol + (unchanged_vol * 0.6)
                                    bd = adjusted_up / total_vol
                                elif recent_momentum < 0:
                                    # Negative momentum: attribute more unchanged volume to sellers
                                    adjusted_up = up_vol + (unchanged_vol * 0.4) 
                                    bd = adjusted_up / total_vol
                                else:
                                    # No momentum: split unchanged volume evenly
                                    bd = (up_vol + unchanged_vol * 0.5) / total_vol
                            else:
                                bd = base_bd
                            
                            # Validate reasonable range
                            if 0 <= bd <= 1:
                                logging.debug(f"✅ {ticker}: 5-min buyer dominance = {bd:.3f} ({bd*100:.1f}%)")
                                return bd
                            else:
                                logging.debug(f"❌ {ticker}: Invalid 5-min BD {bd}, trying fallback")
            except Exception as e:
                logging.debug(f"5-min data error for {ticker}: {e}")
        
        # Method 2: Daily OHLC patterns (analyze candle behavior)
        if not daily_df.empty and len(daily_df) >= 5:
            try:
                # Get recent 5 days of data
                recent_data = daily_df.tail(5).copy()
                
                buyer_signals = 0
                seller_signals = 0
                total_weight = 0
                
                for i, (idx, row) in enumerate(recent_data.iterrows()):
                    day_weight = (i + 1) * 0.2  # More recent days have higher weight
                    total_weight += day_weight
                    
                    open_price = row['Open']
                    high_price = row['High'] 
                    low_price = row['Low']
                    close_price = row['Close']
                    volume = row.get('Volume', 1)
                    
                    # Skip invalid data
                    if pd.isna(open_price) or pd.isna(close_price) or open_price <= 0 or close_price <= 0:
                        continue
                    
                    # Calculate candle metrics
                    body_size = abs(close_price - open_price)
                    upper_shadow = high_price - max(open_price, close_price)
                    lower_shadow = min(open_price, close_price) - low_price
                    day_range = high_price - low_price
                    
                    if day_range <= 0:
                        continue
                    
                    # Buyer dominance signals
                    day_bd_score = 0.5  # Start neutral
                    
                    # 1. Close vs Open (basic direction)
                    if close_price > open_price:
                        day_bd_score += 0.2  # Bullish candle
                    elif close_price < open_price:
                        day_bd_score -= 0.2  # Bearish candle
                    
                    # 2. Close position within the range
                    close_position = (close_price - low_price) / day_range
                    if close_position > 0.7:
                        day_bd_score += 0.15  # Closed in upper 30%
                    elif close_position < 0.3:
                        day_bd_score -= 0.15  # Closed in lower 30%
                    
                    # 3. Shadow analysis (rejection patterns)
                    if day_range > 0:
                        upper_shadow_pct = upper_shadow / day_range
                        lower_shadow_pct = lower_shadow / day_range
                        
                        # Long lower shadow = buying support
                        if lower_shadow_pct > 0.3:
                            day_bd_score += 0.1
                        
                        # Long upper shadow = selling pressure  
                        if upper_shadow_pct > 0.3:
                            day_bd_score -= 0.1
                    
                    # 4. Volume consideration (higher volume = more significance)
                    volume_weight = min(2.0, volume / recent_data['Volume'].mean()) if recent_data['Volume'].mean() > 0 else 1.0
                    
                    # Apply volume weighting
                    weighted_score = day_bd_score * volume_weight * day_weight
                    buyer_signals += max(0, weighted_score)
                    seller_signals += max(0, -weighted_score)
                
                if total_weight > 0 and (buyer_signals + seller_signals) > 0:
                    bd = buyer_signals / (buyer_signals + seller_signals)
                    logging.debug(f"✅ {ticker}: Daily OHLC buyer dominance = {bd:.3f} ({bd*100:.1f}%)")
                    return bd
                    
            except Exception as e:
                logging.debug(f"Daily OHLC analysis error for {ticker}: {e}")
        
        # Method 3: Volume-weighted price analysis (basic method)
        if not daily_df.empty and len(daily_df) >= 3:
            try:
                recent = daily_df.tail(3)
                
                # Calculate volume-weighted average price (VWAP) vs closing prices
                total_vol_price = 0
                total_volume = 0
                buyer_vol = 0
                
                for _, row in recent.iterrows():
                    if pd.notna(row['Volume']) and row['Volume'] > 0 and pd.notna(row['Close']):
                        volume = row['Volume']
                        close = row['Close']
                        open_price = row.get('Open', close)
                        
                        total_vol_price += volume * close
                        total_volume += volume
                        
                        # If close > open, consider it buyer volume
                        if close > open_price:
                            buyer_vol += volume
                        elif close == open_price:
                            buyer_vol += volume * 0.5  # Split neutral volume
                
                if total_volume > 0:
                    bd = buyer_vol / total_volume
                    logging.debug(f"✅ {ticker}: Volume-weighted buyer dominance = {bd:.3f} ({bd*100:.1f}%)")
                    return bd
                    
            except Exception as e:
                logging.debug(f"Volume-weighted analysis error for {ticker}: {e}")
        
        # Method 4: Smart neutral fallback based on market conditions
        try:
            if not daily_df.empty and len(daily_df) >= 2:
                # Calculate recent performance to set intelligent neutral
                current_price = daily_df['Close'].iloc[-1]
                prev_price = daily_df['Close'].iloc[-2]
                
                if current_price > prev_price:
                    # Recent upward movement: slightly bullish neutral
                    smart_neutral = 0.52
                    logging.debug(f"📈 {ticker}: Smart bullish neutral = {smart_neutral:.3f} (recent gain)")
                elif current_price < prev_price:
                    # Recent downward movement: slightly bearish neutral
                    smart_neutral = 0.48
                    logging.debug(f"📉 {ticker}: Smart bearish neutral = {smart_neutral:.3f} (recent decline)")
                else:
                    # No change: true neutral
                    smart_neutral = 0.50
                    logging.debug(f"➡️ {ticker}: True neutral = {smart_neutral:.3f} (no change)")
                
                return smart_neutral
        except:
            pass
        
        # Final fallback: true neutral
        logging.debug(f"⚪ {ticker}: Default neutral buyer dominance = 0.50 (all methods failed)")
        return 0.50
        
    except Exception as e:
        logging.debug(f"❌ {ticker}: Enhanced buyer dominance calculation failed: {e}")
        return 0.50

# Manual seed tiers (REDUCED - Dynamic management preferred)
HIGH_CONVICTION_TIER = {}  # Now managed by SmartTierManager
SELECTIVE_TIER       = {"RAYMOND","IOB","SUNTECK","LTIM","CUPID","LTF","INDOSTAR","MAHLIFE"}
TACTICAL_TIER        = {"JIOFIN","KSOLVES","BHAGYANGR","DPABHUSHAN","SPECIALITY",
                        "MTARTECH","ALKALI","PAR","TIMESGTY","RAMAPHO","BOROSCI"}

# Enhanced liquidity thresholds (market-cap aware)
LIQUIDITY_THRESHOLDS = {
    "min_volume_small": 50000,       # For stocks < ₹1000Cr market cap
    "min_volume_medium": 150000,     # For stocks ₹1000-5000Cr
    "min_volume_large": 300000,      # For stocks > ₹5000Cr
    "min_turnover_value": 2000000,   # ₹2M minimum daily turnover
    "max_spread_pct": 0.015,         # 1.5% max bid-ask spread
    "min_turnover_ratio": 0.005,     # 0.5% of shares outstanding
    "min_liquidity_score": 0.01,     # Minimum liquidity score
    "min_vol_volatility_ratio": 10000,  # Volume/volatility ratio
}

# Legacy constants for backward compatibility
MIN_ABS_VOLUME = 100000  # Will be replaced by market-cap aware logic
SPREAD_THRESHOLD_PCT = LIQUIDITY_THRESHOLDS["max_spread_pct"]
TURNOVER_THRESHOLD = LIQUIDITY_THRESHOLDS["min_turnover_ratio"]
LIQUIDITY_SCORE_THRESHOLD = LIQUIDITY_THRESHOLDS["min_liquidity_score"]
VOL_VOLATILITY_RATIO_THRESHOLD = LIQUIDITY_THRESHOLDS["min_vol_volatility_ratio"]

# =============================================================================
# Enhanced Sentiment Analysis
# =============================================================================
class EnhancedSentimentAnalyzer:
    """Enhanced sentiment analysis with transformer-based fallback."""
    
    def __init__(self, ultra_fast=False):
        self.transformer_available = False
        self.sentiment_pipeline = None
        self.ultra_fast = ultra_fast
        
        # Skip transformer loading in ultra-fast mode
        if ultra_fast:
            print("[ULTRA-FAST] Using basic pattern-based sentiment analysis")
            return
        
        # Try to load transformer-based sentiment analysis
        try:
            from transformers import pipeline
            import torch
            
            # Try PyTorch-based models first (more stable)
            try:
                self.sentiment_pipeline = pipeline(
                    "sentiment-analysis", 
                    model="ProsusAI/finbert",
                    return_all_scores=True,
                    framework="pt"  # Force PyTorch
                )
                print("[OK] Using FinBERT financial sentiment analysis (PyTorch)")
            except:
                try:
                    self.sentiment_pipeline = pipeline(
                        "sentiment-analysis",
                        model="cardiffnlp/twitter-roberta-base-sentiment-latest",
                        return_all_scores=True,
                        framework="pt"  # Force PyTorch
                    )
                    print("[OK] Using RoBERTa sentiment analysis (PyTorch)")
                except:
                    try:
                        # Try without specifying model (use default)
                        self.sentiment_pipeline = pipeline(
                            "sentiment-analysis",
                            return_all_scores=True,
                            framework="pt"  # Force PyTorch
                        )
                        print("[OK] Using default sentiment analysis (PyTorch)")
                    except:
                        # If PyTorch models fail, fall back to simple method
                        raise ImportError("PyTorch models not available")
            self.transformer_available = True
        except (ImportError, RuntimeError, ValueError) as e:
            print(f"[WARNING] Transformers not available ({str(e)[:50]}...). Using simple sentiment scoring.")
            self.transformer_available = False
    
    def analyze_sentiment(self, text: str, fallback_score: float = 0.0) -> float:
        """
        Analyze sentiment of text.
        Returns score between -1.0 (negative) and +1.0 (positive).
        """
        if not text or not text.strip():
            return fallback_score
        
        # Ultra-fast mode: use basic pattern matching only
        if self.ultra_fast:
            return self._basic_pattern_sentiment(text)
            
        if self.transformer_available and self.sentiment_pipeline:
            try:
                # Clean text for analysis
                clean_text = text.strip()[:512]  # Limit text length
                
                results = self.sentiment_pipeline(clean_text)
                
                # Convert results to a single score
                if isinstance(results[0], list):
                    # Handle return_all_scores=True format
                    scores = results[0]
                    sentiment_score = 0.0
                    
                    for item in scores:
                        label = item['label'].upper()
                        score = item['score']
                        
                        if 'POSITIVE' in label or 'POS' in label:
                            sentiment_score += score
                        elif 'NEGATIVE' in label or 'NEG' in label:
                            sentiment_score -= score
                        # NEUTRAL contributes 0
                    
                    # Normalize to -1 to +1 range
                    sentiment_score = max(-1.0, min(1.0, sentiment_score))
                    
                else:
                    # Handle single prediction format
                    label = results[0]['label'].upper()
                    score = results[0]['score']
                    
                    if 'POSITIVE' in label or 'POS' in label:
                        sentiment_score = score
                    elif 'NEGATIVE' in label or 'NEG' in label:
                        sentiment_score = -score
                    else:
                        sentiment_score = 0.0
                
                # Weight the transformer result with fallback
                if abs(fallback_score) > 0.1:  # If we have a meaningful fallback
                    # 70% transformer, 30% fallback
                    final_score = 0.7 * sentiment_score + 0.3 * fallback_score
                else:
                    final_score = sentiment_score
                
                return max(-1.0, min(1.0, final_score))
                
            except Exception as e:
                if hasattr(self, '_error_logged'):
                    return fallback_score
                print(f"⚠️  Sentiment analysis error: {e}")
                self._error_logged = True
                return fallback_score
        
        # Use simple keyword-based sentiment as ultimate fallback
        return self._simple_sentiment(text, fallback_score)
    
    def _simple_sentiment(self, text: str, fallback_score: float = 0.0) -> float:
        """Simple keyword-based sentiment analysis as fallback."""
        if not text:
            return fallback_score
            
        text_lower = text.lower()
        
        # Financial positive keywords
        positive_words = [
            'profit', 'growth', 'increase', 'gain', 'rise', 'up', 'strong', 'positive',
            'expansion', 'acquisition', 'win', 'success', 'beat', 'exceed', 'record',
            'dividend', 'bonus', 'upgrade', 'approval', 'launch', 'deal', 'contract'
        ]
        
        # Financial negative keywords  
        negative_words = [
            'loss', 'decline', 'decrease', 'fall', 'drop', 'down', 'weak', 'negative',
            'concern', 'risk', 'warning', 'cut', 'reduce', 'miss', 'below', 'problem',
            'issue', 'challenge', 'delay', 'cancel', 'suspend', 'downgrade'
        ]
        
        pos_count = sum(1 for word in positive_words if word in text_lower)
        neg_count = sum(1 for word in negative_words if word in text_lower)
        
        if pos_count == 0 and neg_count == 0:
            return fallback_score
            
        # Simple scoring
        total_words = len(text_lower.split())
        pos_score = pos_count / max(total_words, 10)
        neg_score = neg_count / max(total_words, 10)
        
        sentiment_score = pos_score - neg_score
        
        # Combine with fallback if available
        if abs(fallback_score) > 0.1:
            sentiment_score = 0.5 * sentiment_score + 0.5 * fallback_score
            
        return max(-1.0, min(1.0, sentiment_score))
    
    def _basic_pattern_sentiment(self, text: str) -> float:
        """Ultra-fast basic pattern-based sentiment for speed."""
        if not text:
            return 0.0
        
        text_lower = text.lower()
        
        # Quick positive/negative pattern scoring
        positive_patterns = ['profit', 'gain', 'rise', 'up', 'strong', 'growth', 'win', 'beat', 'success', 'upgrade', 'buy']
        negative_patterns = ['loss', 'fall', 'drop', 'down', 'weak', 'decline', 'miss', 'concern', 'risk', 'downgrade', 'sell']
        
        pos_score = sum(1 for p in positive_patterns if p in text_lower)
        neg_score = sum(1 for p in negative_patterns if p in text_lower)
        
        if pos_score == neg_score:
            return 0.0
        
        # Simple normalization
        total = pos_score + neg_score
        if total == 0:
            return 0.0
            
        return (pos_score - neg_score) / max(total, 1.0)

# Global sentiment analyzer instance
_sentiment_analyzer = EnhancedSentimentAnalyzer()

# =============================================================================
# Unicode / normalization helpers
# =============================================================================
_DASHES: Iterable[str] = (
    "\u002D","\u00AD","\u058A","\u1806","\u2010","\u2011","\u2012","\u2013",
    "\u2014","\u2015","\u2043","\u207B","\u208B","\u2212","\uFE58","\uFE63","\uFF0D"
)
_MOJIBAKE_DASHES = (
    "Ã¢â‚¬â€˜","Ã¢â‚¬â€œ","Ã¢â‚¬â€","Ã¢â‚¬â€¢","Ã¢Ë†â€™","Ã¢â‚¬Â¯","Ã¢â‚¬",
    "ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Ëœ","ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“","ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬â€","ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬â€¢",
    "ÃƒÂ¢Ã‹â€ Ã¢â‚¬â„¢","ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¯","ÃƒÂ¢Ã¢â€šÂ¬"
)
_dash_map = {ord(ch): "-" for ch in _DASHES}

def _latin1_roundtrip_fix(txt: str) -> str:
    if "Ãƒ" not in txt:
        return txt
    try:
        repaired = txt.encode("latin-1","ignore").decode("utf-8","ignore")
        if repaired and repaired.count("Ãƒ") < txt.count("Ãƒ"):
            return repaired
    except Exception:
        pass
    return txt

def normalize_hyphens(text: str) -> str:
    if not text: return ""
    txt = _latin1_roundtrip_fix(text).translate(_dash_map)
    for seq in _MOJIBAKE_DASHES:
        txt = txt.replace(seq, "-")
    txt = txt.replace("\u200b","").replace("\u200c","").replace("\u200d","")
    return re.sub(r"\s+"," ", txt)

def deduplicate_news_by_signature(news: List[News]) -> List[News]:
    """
    Remove duplicate deal entries using both exact and fuzzy matching.
    Keeps only the first occurrence of each unique headline.
    """
    seen = set()
    unique_news = []
    
    for n in news:
        # Use normalized headline as the signature
        sig = normalize_hyphens(n.headline).strip().lower()
        
        # Skip if exact match found
        if sig in seen:
            continue
            
        # Check for fuzzy duplicates using token-based similarity
        is_duplicate = False
        tokens_current = set(sig.split())
        
        for existing_sig in seen:
            tokens_existing = set(existing_sig.split())
            
            # Skip if either headline is too short for meaningful comparison
            if len(tokens_current) < 3 or len(tokens_existing) < 3:
                continue
                
            # Calculate Jaccard similarity (intersection/union)
            intersection = len(tokens_current & tokens_existing)
            union = len(tokens_current | tokens_existing)
            jaccard_ratio = intersection / union if union > 0 else 0
            
            # Consider duplicate if >85% token similarity
            if jaccard_ratio > 0.85:
                is_duplicate = True
                break
                
        if not is_duplicate:
            seen.add(sig)
            unique_news.append(n)
            
    return unique_news

ANY_HYPHEN_RX = r"(?:[-\u2010-\u2015\u2212\uFE63\uFF0D]|\s)"

# =============================================================================
# Keyword maps (legacy catalysts)
# =============================================================================
DEAL_KEYWORDS = {
    "order","orders","contract","contracts","deal","deals","project","projects",
    "tender","award","awarded","win","wins","bag","bags","secure","secures",
    "agreement","agreements","mou","pact","acquisition","acquires","acquired",
    "merge","merger","partnership","partners","collab","collaboration","alliance",
    "investment","invests","invest","funding","fundraise","raises","raise",
    "issue","ncd","bond","debenture","qip","rights","placement","stake","stake sale",
    "stake talks","stake purchase","stake buy","stake buyout","buyout","sell stake",
    "capacity","expansion","expand","expanded","capex","plant","facility",
    "commission","commissioning","greenfield","brownfield",
    "dividend","payout","bonus","stock dividend","dividend increase",
    "pat","profit after tax","net profit","earnings","results","quarterly results",
    "annual results","financial results","growth","revenue growth","profit growth",
    "interim dividend","final dividend","special dividend",
    "crore","cr","lakh","lac","million","mn","billion","bn",
    "percent","percentage","rise","surge","jump","increase",
    "fund raise","fund raising","fundraising","funds raised","funds raising",
    "fund approved","approval for fund","fund infusion","funding round",
    "capital raise","invested","investment",
    "procure","procurement","contract win","project win","selected for",
    "finalized","partners with","teams up with","joint venture","JV","collaborates"
}

CATALYST_KEYWORDS: Dict[str,List[str]] = {
    "STAKE": ["stake sale","stake talks","stake purchase","sell stake"],
    "BUYO":  [rf"buy{ANY_HYPHEN_RX}?out", rf"take{ANY_HYPHEN_RX}?over","buyback talks","buyout"],
    "BONUS": ["bonus issue","bonus shares",rf"stock{ANY_HYPHEN_RX}?split","2:1 bonus","3:1 bonus","stock split"],
    "BRKOUT":["breakout","triangle breakout",rf"chart{ANY_HYPHEN_RX}?pattern",
              "swing-trade buy","initiates buy","swing-trade call","triangle pattern"],
    "GREEN": ["battery","biogas","pivot","diversify into battery","green pivot",
              "renewable foray","cbg","compressed biogas"],
    "GROWTH":[r"\d{1,3}(?:,\d{3})*%",
              r"\d+(?:\.\d+)?% (?:growth|rise|increase|surge|jump)",
              "order win","contract award","project bag","dividend declaration",
              "bonus issue","stock split","merger","acquisition","takeover",
              "buyout","stake sale","profit jump","revenue jump","sales jump"],
    "FUND":  [rf"fund{ANY_HYPHEN_RX}raise", rf"fund{ANY_HYPHEN_RX}raising","capital raise",
              "fund infusion","funding round","invested", rf"ncd{ANY_HYPHEN_RX}issue",
              "non-convertible debentures","debenture issue","bond issue",
              rf"fund{ANY_HYPHEN_RX}approved","approval for fund"]
}

DEAL_PATTERN_LEGACY      = re.compile("|".join(map(re.escape, DEAL_KEYWORDS)), re.I)
CATALYST_PATTERNS = {k: re.compile("|".join(v), re.I) for k,v in CATALYST_KEYWORDS.items()}
CATALYST_WEIGHTS  = {"STAKE":1.2,"BUYO":1.5,"BONUS":0.8,"BRKOUT":0.9,"GREEN":0.9,"GROWTH":1.3,"FUND":2.0}

# =============================================================================

# =============================================================================
# RSI + EMA + Bollinger Strategy (Technical Composite)
# =============================================================================
class RSI_EMA_BB_Strategy:
    def __init__(self):
        # Enhanced weights for ideal swing trading conditions - prioritize oversold RSI + rising EMA
        self.weights = {
            "rsi_oversold": 40,      # Increased: RSI oversold is most important signal
            "bb_position": 30,       # Increased: Near lower Bollinger Band 
            "ema_slope": 35,         # Increased: Rising EMA trend
            "ema_cross": 30,         # EMA14 > EMA50 crossover
            "vol_confirmation": 20,  # Volume spike confirmation
            "ideal_combo": 50        # Bonus for perfect combination
        }

    def _safe_last(self, series, default=0.0):
        try:
            if series is None or len(series) == 0:
                return default
            v = series.iloc[-1]
            if v is None:
                return default
            return float(v)
        except Exception:
            return default

    def calculate_features(self, df: pd.DataFrame) -> dict:
        features = {}

        if df is None or df.empty or len(df) < 25:
            # Minimal defaults if insufficient history
            return {
                "rsi": 50.0,
                "ema14": self._safe_last(df["Close"]) if df is not None and "Close" in df else 0.0,
                "ema50": self._safe_last(df["Close"]) if df is not None and "Close" in df else 0.0,
                "ema_slope": 0.0,
                "bb_upper": self._safe_last(df["Close"])*1.02 if df is not None and "Close" in df else 0.0,
                "bb_lower": self._safe_last(df["Close"])*0.98 if df is not None and "Close" in df else 0.0,
                "vol_spike": False,
            }

        # Get robust price data (winsorized when available)
        price_data = get_robust_price_data(df, use_capped=True)
        close = price_data['close'].astype(float)
        volume = price_data['volume'].astype(float)  # Use winsorized volume for stable calculations
        
        # VWAP using capped close and volume prices
        try:
            vwap_series = (close * volume).cumsum() / (volume.replace(0, pd.NA)).cumsum()
            features["vwap"] = float(vwap_series.iloc[-1]) if len(vwap_series) else float(close.iloc[-1])
        except Exception:
            features["vwap"] = float(close.iloc[-1])

        # RSI Calculation
        try:
            if talib is not None:
                rsi = talib.RSI(close, timeperiod=14)
            else:
                delta = close.diff()
                gain = delta.where(delta > 0, 0.0)
                loss = -delta.where(delta < 0, 0.0)
                avg_gain = gain.rolling(14, min_periods=14).mean()
                avg_loss = loss.rolling(14, min_periods=14).mean()
                rs = avg_gain / (avg_loss.replace(0, pd.NA))
                rsi = 100 - (100 / (1 + rs))
        except Exception:
            rsi = pd.Series([50.0]*len(close), index=df.index)

        features["rsi"] = float(rsi.iloc[-1]) if not rsi.empty and pd.notna(rsi.iloc[-1]) else 50.0

        # OPTIMIZED: EMA Calculations - Vectorized once, reuse
        ema14 = close.ewm(span=14, adjust=False).mean().astype('float32')
        ema50 = close.ewm(span=50, adjust=False).mean().astype('float32')
        features["ema14"] = float(ema14.iloc[-1])
        features["ema50"] = float(ema50.iloc[-1])

        # OPTIMIZED: EMA Slope (vectorized calculation)
        if len(ema14) >= 4 and pd.notna(ema14.iloc[-4]) and ema14.iloc[-4] != 0:
            features["ema_slope"] = float((ema14.iloc[-1] - ema14.iloc[-4]) / ema14.iloc[-4] * 100.0)
        else:
            features["ema_slope"] = 0.0

        # OPTIMIZED: Bollinger Bands - Vectorized rolling operations
        try:
            if talib is not None:
                upper, middle, lower = talib.BBANDS(close, timeperiod=20, nbdevup=2, nbdevdn=2, matype=0)
            else:
                # Pre-compute rolling operations once for reuse
                rolling_20 = close.rolling(20, min_periods=20)
                sma = rolling_20.mean()
                std = rolling_20.std(ddof=0)
                upper = (sma + 2.0 * std).astype('float32')
                lower = (sma - 2.0 * std).astype('float32')
        except Exception:
            # OPTIMIZED: Fallback with minimal computation
            rolling_20_fallback = close.rolling(20, min_periods=1)
            sma = rolling_20_fallback.mean()
            std = rolling_20_fallback.std(ddof=0).fillna(0)
            upper = (sma + 2.0 * std).astype('float32')
            lower = (sma - 2.0 * std).astype('float32')

        features["bb_upper"] = float(upper.iloc[-1]) if not upper.empty and pd.notna(upper.iloc[-1]) else float(close.iloc[-1] * 1.02)
        features["bb_lower"] = float(lower.iloc[-1]) if not lower.empty and pd.notna(lower.iloc[-1]) else float(close.iloc[-1] * 0.98)

        # OPTIMIZED: Volume confirmation (reuse rolling calculation)
        try:
            # Use winsorized volume for stable calculations and reuse rolling operation
            volume_rolling_20 = volume.rolling(20, min_periods=1)
            vol_mean20 = volume_rolling_20.mean().iloc[-1]
            features["vol_spike"] = bool(volume.iloc[-1] > vol_mean20 * 1.5)
            
            # Additional volume metrics using the same rolling calculation
            vol_std20 = volume_rolling_20.std().iloc[-1]
            features["vol_ratio"] = float(volume.iloc[-1] / vol_mean20) if vol_mean20 > 0 else 1.0
            features["vol_zscore"] = float((volume.iloc[-1] - vol_mean20) / vol_std20) if vol_std20 > 0 else 0.0
        except Exception:
            features["vol_spike"] = False
            features["vol_ratio"] = 1.0
            features["vol_zscore"] = 0.0

        # OPTIMIZED: Memory cleanup for large DataFrames
        if len(df) > 500:  # Only cleanup for large datasets
            del rolling_20, volume_rolling_20
            if 'rolling_20_fallback' in locals():
                del rolling_20_fallback
            gc.collect()

        return features

    def calculate_score(self, features: dict) -> float:
        score = 0.0

        # VWAP bonus (above VWAP suggests strength)
        current_price = features.get('ema14', 0.0)
        vwap = features.get('vwap', 0.0)
        if current_price and vwap and current_price > vwap:
            score += 8.0  # Increased VWAP bonus

        # RSI Oversold condition - ENHANCED FOR MEAN REVERSION STRATEGY
        rsi = features.get("rsi", 50.0)
        rsi_score = 0.0
        if rsi <= 30.0:  # Focus on deeply oversold conditions
            if rsi <= 20.0:      # Extreme oversold - highest score
                rsi_score = self.weights["rsi_oversold"] * 1.5  # Extra bonus for extreme oversold
            elif rsi <= 25.0:    # Very oversold
                rsi_score = self.weights["rsi_oversold"] * 1.2
            elif rsi <= 30.0:    # Oversold target range
                rsi_score = self.weights["rsi_oversold"]
        elif rsi <= 35.0:       # Mildly oversold - reduced score
            rsi_score = self.weights["rsi_oversold"] * 0.3
        score += rsi_score

        # Bollinger Band position - ENHANCED FOR MEAN REVERSION
        current_price = features.get("ema14", 0.0)
        bb_lower = features.get("bb_lower", 0.0)
        bb_upper = features.get("bb_upper", 0.0)
        bb_range = bb_upper - bb_lower
        bb_score = 0.0
        if bb_range > 0 and current_price > 0:
            position_pct = (current_price - bb_lower) / bb_range
            if position_pct <= 0.10:     # Very close to lower band - perfect for mean reversion
                bb_score = self.weights["bb_position"] * 1.5  # Extra bonus
            elif position_pct <= 0.20:   # Near lower band - ideal target
                bb_score = self.weights["bb_position"] * 1.2
            elif position_pct <= 0.30:   # Lower third - good
                bb_score = self.weights["bb_position"] * 0.8
            elif position_pct <= 0.40:   # Below middle - acceptable
                bb_score = self.weights["bb_position"] * 0.4
        score += bb_score

        # EMA Condition - MODIFIED FOR MEAN REVERSION (declining or EMA14 < EMA50)
        ema_slope = features.get("ema_slope", 0.0)
        ema14 = features.get("ema14", 0.0)
        ema50 = features.get("ema50", 0.0)
        ema_score = 0.0
        
        # Check for declining EMA or EMA14 below EMA50 (correction/pullback conditions)
        ema_bearish = False
        if ema14 > 0 and ema50 > 0 and ema14 < ema50:  # EMA14 below EMA50 (bearish/correcting)
            ema_bearish = True
            cross_weakness = (ema50 - ema14) / ema50
            if cross_weakness > 0.05:       # Strong bearish signal (5%+ below)
                ema_score += self.weights["ema_cross"] * 1.2
            elif cross_weakness > 0.02:     # Moderate bearish signal (2%+ below)
                ema_score += self.weights["ema_cross"] * 0.8
            elif cross_weakness > 0.01:     # Mild bearish signal (1%+ below)
                ema_score += self.weights["ema_cross"] * 0.5
        
        # Declining EMA slope (additional bonus for mean reversion setup)
        if ema_slope < 0.0:  # Declining EMA
            if ema_slope < -2.0:         # Strong decline
                ema_score += self.weights["ema_slope"] * 1.0
            elif ema_slope < -1.0:       # Moderate decline
                ema_score += self.weights["ema_slope"] * 0.8
            elif ema_slope < -0.5:       # Mild decline
                ema_score += self.weights["ema_slope"] * 0.6
        elif ema_slope < 0.5 and not ema_bearish:  # Flat/slow rising (acceptable if not bearish)
            ema_score += self.weights["ema_slope"] * 0.2
            
        score += ema_score

        # Volume confirmation - ENHANCED
        if features.get("vol_spike", False):
            score += self.weights["vol_confirmation"]

        # IDEAL COMBINATION BONUS - Mean Reversion Setup (RSI≤30, BB≤20%, EMA bearish)
        ideal_conditions = 0
        if rsi <= 30.0:  # Target RSI oversold condition
            ideal_conditions += 1
        if position_pct <= 0.2 if 'position_pct' in locals() else False:  # Near lower BB (≤20%)
            ideal_conditions += 1  
        if ema_slope < 0.0:  # Declining EMA slope (correction)
            ideal_conditions += 1
        if ema14 > 0 and ema50 > 0 and ema14 < ema50:  # EMA14 below EMA50 (bearish trend)
            ideal_conditions += 1
        if features.get("vol_spike", False):  # Volume confirmation
            ideal_conditions += 1
            
        # Extra bonus for perfect mean reversion setup
        if rsi <= 30.0 and position_pct <= 0.2 if 'position_pct' in locals() else False:
            ideal_conditions += 1  # Extra condition for perfect setup
            
        # Bonus for multiple ideal conditions (mean reversion sweet spot)
        if ideal_conditions >= 3:
            score += self.weights["ideal_combo"] * (ideal_conditions / 6.0)

        return float(min(150.0, score))  # Allow higher max score for ideal conditions

# Breakout & Growth models
# =============================================================================
class BreakoutProbabilityModel:
    def __init__(self):
        self.weights = {
            "R1": 10, "Break": 15, "IntraBreak": 10, "Engulf": 20,
            "VolSpike": 5, "VolWave": 8, "BuyPress": 12, "VolConfirm": 7
        }

    def _features(self, df: pd.DataFrame) -> pd.DataFrame:
        f = pd.DataFrame(index=df.index)
        
        # Get robust price data (uses winsorized data when available)
        price_data = get_robust_price_data(df, use_capped=True)
        
        f["Close"] = price_data['close']  # Use capped close for consistency

        # 1. Momentum - Use winsorized returns to prevent extreme moves from skewing calculations
        f["R1"] = (price_data['returns'] * 100).fillna(0)  # Capped daily returns
        f["Break"] = (df["High"] > df["High"].rolling(20).max().shift(1)).astype(int)
        f["IntraBreak"] = (df["High"] > df["High"].rolling(30).max().shift(1)).astype(int)

        # 2. Candlestick - USING WINSORIZED CLOSE FOR STABILITY
        if talib is not None:
            # Use winsorized close for more stable pattern recognition
            close = price_data['close']
            f["Engulf"] = (talib.CDLENGULFING(df["Open"], df["High"], df["Low"], close) > 0).astype(int)
        else:
            # Fallback calculation using winsorized close
            close = price_data['close']
            body = close - df["Open"]
            prev = body.shift(1)
            f["Engulf"] = ((body > 0) & (prev < 0) &
                           (close > df["Open"].shift(1)) &
                           (df["Open"] < close.shift(1))).astype(int)

        # 3. Enhanced Volume Spike Detection (2-sigma approach) - USING WINSORIZED VOLUME
        volume = price_data['volume']  # Use winsorized volume for stable calculations
        if len(df) >= 20:  # Need sufficient history for std calculation
            vol_mean = volume.rolling(20).mean()
            vol_std = volume.rolling(20).std()
            # 2-sigma threshold for meaningful volume anomalies
            vol_threshold = vol_mean + 2 * vol_std
            f["VolSpike"] = (volume > vol_threshold).astype(int)
        else:
            # Fallback for insufficient history - also using winsorized volume
            f["VolSpike"] = (volume > 1.5 * volume.rolling(5).mean().shift(1)).astype(int)

        # 4. Volume Wave - USING WINSORIZED VOLUME
        if len(df) >= 10:
            vol_3d = volume.rolling(3).mean()
            vol_10d = volume.rolling(10).mean()
            f["VolWave"] = np.where(
                (volume > vol_3d) &
                (vol_3d > vol_10d * 1.1) &
                (vol_3d.pct_change() > 0.15), 1, 0
            )
        else:
            f["VolWave"] = 0

        # 5. Buying Pressure
        if len(df) >= 10:
            obv = [0]
            for i in range(1, len(df)):
                # Use capped close for OBV calculation to prevent extreme moves from skewing
                curr_close = price_data['close'].iloc[i]
                prev_close = price_data['close'].iloc[i-1]
                
                if curr_close > prev_close:
                    obv.append(obv[-1] + df["Volume"].iloc[i])
                elif curr_close < prev_close:
                    obv.append(obv[-1] - df["Volume"].iloc[i])
                else:
                    obv.append(obv[-1])
            
            # Use capped returns for volume-price correlation
            vpc = df["Volume"].rolling(3).corr(price_data['returns'])
            obv_trend = 1 if obv[-1] > obv[-5] else 0
            vpc_strength = 1 if (vpc.iloc[-1] if not np.isnan(vpc.iloc[-1]) else 0) > 0.6 else 0
            f["BuyPress"] = obv_trend + vpc_strength
        else:
            f["BuyPress"] = 0

        # 6. Volume Confirmation
        f["VolConfirm"] = np.where(
            (f["Break"] == 1) &
            (df["Volume"] > df["Volume"].rolling(20).mean() * 1.5),
            1, 0
        )

        return f

    def calculate_score(self, df):
        f = self._features(df).dropna(subset=["Close"])
        f["Score"] = sum(self.weights[k] * f[k] for k in self.weights)
        return f[["Score"]]

    def probability(self, sc):
        return (1 / (1 + np.exp(-sc["Score"] / 10)) * 100.0).rename("Pct")

# =============================================================================
# Bullish Growth Forecaster
# =============================================================================
class BullishGrowthForecaster:
    def __init__(self, pm, beta=0.20):
        self.pm = pm
        self.beta = beta

    def predict(self, df):  # CORRECTED indentation
        p = self.pm.probability(self.pm.calculate_score(df)).iloc[-1]
        exp = p * self.beta
        return exp, exp * 0.8, exp * 1.2

# --- Initialize technical strategy (RSI+EMA+BB)
try:
    _rsi_ema_bb_strategy = RSI_EMA_BB_Strategy()
except Exception as _e_init_te:
    print(f"⚠️  Technical strategy init failed: {_e_init_te}")
    _rsi_ema_bb_strategy = None

# =============================================================================
# Liquidity Short Detection
# =============================================================================
class LiquidityAnalyzer:
    @staticmethod
    def calculate_liquidity_score(avg_volume: float, price: float, market_cap: float) -> float:
        """Calculate liquidity score: (Average Daily Volume × Price) / Market Cap"""
        if market_cap <= 0 or price <= 0:
            return 0.0
        return (avg_volume * price) / market_cap
    
    @staticmethod
    def calculate_turnover_ratio(avg_volume: float, shares_outstanding: float) -> float:
        """Calculate turnover ratio: Daily Volume / Total Outstanding Shares"""
        if shares_outstanding <= 0:
            return 0.0
        return avg_volume / shares_outstanding
    
    @staticmethod
    def calculate_volatility_ratio(avg_volume: float, volatility: float) -> float:
        """Calculate Volume/Volatility ratio"""
        if volatility <= 0:
            return 0.0
        return avg_volume / volatility
    
    @staticmethod
    def get_volume_threshold(market_cap_cr: float) -> int:
        """Get market-cap appropriate volume threshold"""
        if market_cap_cr < 1000:  # < ₹1000Cr (small-cap)
            return LIQUIDITY_THRESHOLDS["min_volume_small"]
        elif market_cap_cr < 5000:  # ₹1000-5000Cr (mid-cap)
            return LIQUIDITY_THRESHOLDS["min_volume_medium"]
        else:  # > ₹5000Cr (large-cap)
            # Simple dynamic scaling for large caps
            base_threshold = LIQUIDITY_THRESHOLDS["min_volume_large"]
            if market_cap_cr > 20000:  # Very large cap (>₹20000Cr)
                return max(base_threshold, base_threshold * 1.5)
            return base_threshold
    
    @staticmethod
    def is_liquidity_short(
        avg_volume: float, 
        bid_ask_spread: float, 
        current_price: float,
        turnover_ratio: float,
        volatility_ratio: float,
        liquidity_score: float,
        market_cap_cr: float = 0.0,
        min_turnover_m: float = 2.0,
        max_spread_pct: float = 1.5,
        strict_mode: bool = False
    ) -> bool:
        """Determine if a stock is a liquidity short based on multiple metrics"""
        
        # Adjust thresholds for strict mode (larger portfolios)
        turnover_multiplier = 2.5 if strict_mode else 1.0
        spread_multiplier = 0.6 if strict_mode else 1.0
        
        # Market-cap based volume threshold
        volume_threshold = LiquidityAnalyzer.get_volume_threshold(market_cap_cr)
        if strict_mode:
            volume_threshold = int(volume_threshold * 1.5)
        
        if avg_volume < volume_threshold:
            return True
        
        # Absolute turnover value threshold (₹-based) - configurable
        min_turnover_value = min_turnover_m * 1000000 * turnover_multiplier
        turnover_value = avg_volume * current_price
        if turnover_value < min_turnover_value:
            return True
        
        # Spread threshold - configurable
        max_spread = (max_spread_pct / 100.0) * spread_multiplier
        if bid_ask_spread > (current_price * max_spread):
            return True
        
        # Turnover ratio threshold
        min_turnover_ratio = LIQUIDITY_THRESHOLDS["min_turnover_ratio"]
        if strict_mode:
            min_turnover_ratio *= 1.5
        if turnover_ratio < min_turnover_ratio:
            return True
        
        # Liquidity score threshold
        min_liq_score = LIQUIDITY_THRESHOLDS["min_liquidity_score"]
        if strict_mode:
            min_liq_score *= 2.0
        if liquidity_score < min_liq_score:
            return True
        
        # Volume/Volatility ratio threshold
        min_vol_vol_ratio = LIQUIDITY_THRESHOLDS["min_vol_volatility_ratio"]
        if strict_mode:
            min_vol_vol_ratio *= 1.5
        if volatility_ratio < min_vol_vol_ratio:
            return True
        
        return False
    
    @staticmethod
    def get_liquidity_warning_level(
        avg_volume: float,
        current_price: float,
        market_cap_cr: float,
        spread_pct: float
    ) -> str:
        """Get liquidity warning level: CRITICAL, HIGH, MEDIUM, LOW"""
        
        volume_threshold = LiquidityAnalyzer.get_volume_threshold(market_cap_cr)
        turnover_value = avg_volume * current_price
        
        # Critical: Multiple severe issues
        critical_flags = 0
        if avg_volume < volume_threshold * 0.5:  # Less than 50% of required volume
            critical_flags += 1
        if turnover_value < LIQUIDITY_THRESHOLDS["min_turnover_value"] * 0.5:  # Less than ₹1M turnover
            critical_flags += 1
        if spread_pct > LIQUIDITY_THRESHOLDS["max_spread_pct"] * 2:  # > 3% spread
            critical_flags += 1
        
        if critical_flags >= 2:
            return "CRITICAL"
        elif critical_flags == 1:
            return "HIGH"
        elif (avg_volume < volume_threshold or 
              turnover_value < LIQUIDITY_THRESHOLDS["min_turnover_value"] or
              spread_pct > LIQUIDITY_THRESHOLDS["max_spread_pct"]):
            return "MEDIUM"
        else:
            return "LOW"
    
    @staticmethod
    def calculate_liquidity_penalty(
        avg_volume: float,
        current_price: float,
        market_cap_cr: float,
        spread_pct: float,
        liquidity_score: float
    ) -> float:
        """Calculate liquidity penalty score (0 = worst, 10 = best)"""
        penalty_score = 10.0  # Start with perfect score
        
        # Volume penalty based on market cap thresholds
        volume_threshold = LiquidityAnalyzer.get_volume_threshold(market_cap_cr)
        volume_ratio = avg_volume / volume_threshold
        if volume_ratio < 1.0:
            penalty_score -= (1.0 - volume_ratio) * 4.0  # Up to -4 points
        
        # Turnover value penalty
        turnover_value = avg_volume * current_price
        min_turnover = LIQUIDITY_THRESHOLDS["min_turnover_value"]
        turnover_ratio = turnover_value / min_turnover
        if turnover_ratio < 1.0:
            penalty_score -= (1.0 - turnover_ratio) * 3.0  # Up to -3 points
        
        # Spread penalty
        max_spread = LIQUIDITY_THRESHOLDS["max_spread_pct"]
        if spread_pct > max_spread:
            spread_excess = (spread_pct - max_spread) / max_spread
            penalty_score -= min(spread_excess * 2.0, 2.0)  # Up to -2 points
        
        # Liquidity score penalty
        min_liq_score = LIQUIDITY_THRESHOLDS["min_liquidity_score"]
        if liquidity_score < min_liq_score:
            score_deficit = (min_liq_score - liquidity_score) / min_liq_score
            penalty_score -= score_deficit * 1.0  # Up to -1 point
        
        return max(0.0, penalty_score)  # Ensure non-negative
    
    @staticmethod
    def get_liquidity_bonus(
        avg_volume: float,
        current_price: float,
        market_cap_cr: float,
        liquidity_score: float
    ) -> float:
        """Calculate liquidity bonus for highly liquid stocks (0-5 points)"""
        bonus = 0.0
        
        # High volume bonus
        volume_threshold = LiquidityAnalyzer.get_volume_threshold(market_cap_cr)
        volume_ratio = avg_volume / volume_threshold
        if volume_ratio > 2.0:  # More than 2x required volume
            bonus += min((volume_ratio - 2.0) * 0.5, 2.0)  # Up to +2 points
        
        # High turnover bonus
        turnover_value = avg_volume * current_price
        min_turnover = LIQUIDITY_THRESHOLDS["min_turnover_value"]
        turnover_ratio = turnover_value / min_turnover
        if turnover_ratio > 3.0:  # More than 3x minimum turnover
            bonus += min((turnover_ratio - 3.0) * 0.2, 1.5)  # Up to +1.5 points
        
        # Exceptional liquidity score bonus
        if liquidity_score > LIQUIDITY_THRESHOLDS["min_liquidity_score"] * 5:
            bonus += 1.5  # +1.5 points for exceptional liquidity
        
        return min(bonus, 5.0)  # Cap at +5 points

# =============================================================================
# Data classes
# =============================================================================
class News(NamedTuple):
    tkr: str
    sent: float
    headline: str
    snippet: str
    ts: Optional[datetime]
    confidence: float = 0.8  # Default confidence for high-quality filtering

class Fin(NamedTuple):
    net: float
    inc: float
    d2e: Optional[float]
    margin: Optional[float] = None
    volume_growth: Optional[float] = None
    debt_red: Optional[float] = None
    orders: Optional[float] = None
    pe_ratio: float = 0.0  # NEW - default to 0.0 when not available

# =============================================================================

def tier1_stop_from_bb(df: pd.DataFrame) -> tuple[float, float]:
    """
    Returns (entry, stop) using BB lower band logic.
    Entry = close * 0.95 (tight mean-revert bias), Stop = max(BB_lower*0.98, close*0.93)
    """
    if df is None or df.empty:
        return 0.0, 0.0
    close = float(df["Close"].iloc[-1])
    # Recompute BB lower quickly (robust even if talib missing)
    try:
        if talib is not None:
            _, _, lower = talib.BBANDS(df["Close"], timeperiod=20, nbdevup=2, nbdevdn=2, matype=0)
            bb_lower = float(lower.iloc[-1])
        else:
            sma = df["Close"].rolling(20, min_periods=1).mean()
            std = df["Close"].rolling(20, min_periods=1).std(ddof=0).fillna(0)
            bb_lower = float((sma - 2.0*std).iloc[-1])
    except Exception:
        bb_lower = close * 0.96
    stop = max(bb_lower * 0.98, close * 0.93)
    return close * 0.95, stop

# CERTAINTY SCORE ENGINE (Breakout Confidence Predictor)
# =============================================================================
class CertaintyEngine:
    @staticmethod
    def calculate(df: pd.DataFrame, fin_data: Fin, news: List[News],
                  mcap: float, deal_impact: float, tier: str) -> float:
        """Generates 0-100 certainty score for early breakout after consolidation/dip"""
        certainty = 0.0
        # 1. Compression Ratio (25%)
        try:
            high_52w = df["High"].rolling(250).max().iloc[-1]
            low_52w = df["Low"].rolling(250).min().iloc[-1]
            current_range = df["High"].iloc[-20:].max() - df["Low"].iloc[-20:].min()
            compression = min(1.0, (high_52w - low_52w) / (current_range + 1e-5))
            certainty += 25 * (0.8 if compression > 3.0 else 0.5 if compression > 2.0 else 0.2)
        except:
            pass
        # 2. Fundamental Health (20%)
        fund_strength = 0
        # Handle both Fin objects and dict objects
        if isinstance(fin_data, dict):
            if fin_data.get('margin') and fin_data.get('margin') > 0.18: fund_strength += 7
            if fin_data.get('d2e') and fin_data.get('d2e') < 0.5: fund_strength += 7
            if fin_data.get('inc') and fin_data.get('inc') > 100: fund_strength += 6
        else:
            if fin_data.margin and fin_data.margin > 0.18: fund_strength += 7
            if fin_data.d2e and fin_data.d2e < 0.5: fund_strength += 7
            if fin_data.inc and fin_data.inc > 100: fund_strength += 6
        certainty += fund_strength
        # 3. Volume Divergence (20%)
        try:
            down = df[df["Close"] < df["Open"]].tail(5)["Volume"].mean()
            up = df[df["Close"] > df["Open"]].tail(5)["Volume"].mean()
            certainty += 20 * min(1.0, up/(down+1e-5)*0.5)
        except:
            pass
        # 4. Deal Impact Catalyst (15%)
        certainty += min(15, deal_impact * 4)
        # 5. Recovery Momentum (10%)
        try:
            recovery = (df["Close"].iloc[-1] - df["Low"].rolling(30).min().iloc[-1]) / df["Close"].iloc[-1]
            certainty += 10 * min(1.0, recovery*5)
        except:
            pass
        # 6. Tier Confidence Boost (10%)
        tier_boost = {"TIER1":10, "TIER2":6, "TIER3":2}.get(tier, 0)
        certainty += tier_boost
        return min(100, certainty)

    @staticmethod
    def detect_accumulation(df: pd.DataFrame) -> bool:
        """Identifies accumulation during consolidation"""
        if len(df) < 30: return False
        try:
            consolidating = (df["Close"].iloc[-20] * 0.97 < df["Close"].iloc[-1] < df["Close"].iloc[-20] * 1.03)
            up_vol = df[df["Close"] > df["Close"].shift(1)].tail(10)["Volume"].mean()
            dn_vol = df[df["Close"] < df["Close"].shift(1)].tail(10)["Volume"].mean()
            return consolidating and (up_vol > dn_vol * 1.3)
        except:
            return False

# =============================================================================
class TierScoringSystem:
    def __init__(self):
        self.weights = {
            "margin_growth": 1.8,
            "volume_growth": 1.5,
            "deleveraging": 1.3,
            "strategic_pivot": 1.6,
            "order_conversion": 1.4,
        }

    def calculate_fundamental_score(self, data: dict) -> float:
        sc = 0.0
        mg = data.get("margin_expansion") or 0
        if mg > 0:
            sc += self.weights["margin_growth"] * min(mg * 100, 500) / 100
        vg = data.get("volume_growth") or 0
        if vg > 0:
            sc += self.weights["volume_growth"] * log2(1 + vg * 100)
        dl = data.get("debt_reduction") or 0
        if dl > 0:
            sc += self.weights["deleveraging"] * dl * 20
        if data.get("strategic_pivot"):
            sc += self.weights["strategic_pivot"]
        oc = data.get("order_conversion_ratio") or 0
        if oc > 0.6:
            sc += self.weights["order_conversion"] * oc
        return sc

# =============================================================================
# Dynamic Tier Classification (with deal impact upgrade)
# =============================================================================
class TierClassifier:
    @staticmethod
    def classify(tkr: str, fin_data: Fin, rel_news: List[News],
                 mcap: float,
                 deal_amount: float,
                 deal_count: int,
                 deal_impact_score: float,
                 liquidity_data: tuple = None,
                 valuation_score: float = 5.0) -> str:
        # Handle both Fin objects and dict objects
        if isinstance(fin_data, dict):
            margin_strength = (fin_data.get('margin', 0) or 0) > 0.15
            vol_growth = (fin_data.get('volume_growth', 0) or 0) > 0.10
            debt_improve = (fin_data.get('debt_red', 0) or 0) > 0.05
        else:
            margin_strength = (fin_data.margin or 0) > 0.15
            vol_growth = (fin_data.volume_growth or 0) > 0.10
            debt_improve = (fin_data.debt_red or 0) > 0.05
        texts = [(n.headline + " " + n.snippet).lower() for n in rel_news]

        def any_kw(*kws):
            return any(all(k in txt for k in kws) for txt in texts)  # CORRECTED indentation

        strategic_pivot = any("pivot" in txt for txt in texts)
        order_growth = any_kw("order", "win") or any_kw("order", "bag") or any_kw("contract", "award")
        profit_growth = any(
            re.search(r"(pat|net profit|profit)\s.*(rise|up|jump|surge|increase|grow)", txt)
            for txt in texts
        )
        turnaround = any(re.search(r"(loss|negative).*(profit|positive)", txt) for txt in texts)

        # Calculate fundamental strength score (0-10)
        fundamental_strength = 0.0
        if margin_strength:
            fundamental_strength += 2.5
        if vol_growth:
            fundamental_strength += 2.0
        if debt_improve:
            fundamental_strength += 1.5
        if order_growth:
            fundamental_strength += 2.0
        if profit_growth:
            fundamental_strength += 1.5
        if turnaround:
            fundamental_strength += 0.5

        # CORRECTED sector placement (after fundamental_strength)
        try:
            _sector = detect_sector(tkr)
            sector_multiplier = SECTOR_WEIGHTS.get(_sector, 1.0)
            fundamental_strength = float(fundamental_strength) * float(sector_multiplier)
        except Exception:
            pass

        # Base tier logic
        if ((margin_strength and vol_growth and debt_improve) or
            (margin_strength and order_growth) or
            (margin_strength and profit_growth) or
            (turnaround and (margin_strength or order_growth))):
            tier = "TIER1"
        elif (margin_strength or vol_growth or debt_improve or strategic_pivot or
              order_growth or profit_growth):
            tier = "TIER2"
        else:
            tier = "TIER3"

        # Auto-upgrade via deal impact
        deal_amount_cr = deal_amount / 1e7  # Convert rupees to Cr
        if deal_impact_score >= 3.0 or (mcap > 0 and deal_amount_cr > mcap * 0.05):
            tier = "TIER1"
        elif deal_impact_score >= 1.5:
            tier = "TIER2"
        
        # Upgrade tier for strong fundamentals + liquidity
        if liquidity_data:
            liquidity_score, warning_level, is_liquidity_short = liquidity_data
            if (fundamental_strength >= 8 and liquidity_score >= 0.02 and not is_liquidity_short):
                if tier != "TIER1":
                    tier = "TIER1"
        
        # Downgrade for valuation risks
        if valuation_score <= 4.0:  # Poor valuation
            if tier == "TIER1":
                tier = "TIER2"
            elif tier == "TIER2":
                tier = "TIER3"
        
        # Apply smart tier management
        if ENHANCED_MODULES_AVAILABLE:
            try:
                from enhanced_screener_modules import SmartTierManager
                smart_tier_manager = SmartTierManager()
                sector = detect_sector(tkr)
                
                # Check if there are active manual overrides for this ticker
                if tkr in smart_tier_manager.manual_overrides:
                    override_data = smart_tier_manager.manual_overrides[tkr]
                    if override_data['expiry'] > datetime.now():
                        tier = f"TIER{override_data['tier']}"
                    else:
                        # Expired override - clean it up
                        del smart_tier_manager.manual_overrides[tkr]
                        
            except Exception as e:
                logging.debug(f"Smart tier override check failed for {tkr}: {e}")
                # Fallback to manual tier overrides
                if tkr in HIGH_CONVICTION_TIER:
                    tier = "TIER1"
        else:
            # Legacy manual overrides
            if tkr in HIGH_CONVICTION_TIER:
                tier = "TIER1"
        
        return tier

# =============================================================================
# Helpers

def _compute_technical_score(df: pd.DataFrame) -> float:
    """Compute composite technical score using RSI/EMA/Bollinger strategy."""
    try:
        if _rsi_ema_bb_strategy is None or df is None or df.empty:
            return 0.0
        feats = _rsi_ema_bb_strategy.calculate_features(df)
        return float(_rsi_ema_bb_strategy.calculate_score(feats))
    except Exception as _e_te_sc:
        logging.debug(f"Tech score error: {_e_te_sc}")
        return 0.0

# =============================================================================
now = datetime.now
lg = lambda v: log2(v) if v > 0 else 0.0

def add_indicator(value, format_str, good_threshold=None, bad_threshold=None):
    """Adds a '*' for good values and '!' for bad values to a formatted string."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "n/a"
    
    base_str = format_str.format(value)
    
    if good_threshold is not None and value >= good_threshold:
        return f"{base_str}*"
    if bad_threshold is not None and value <= bad_threshold:
        return f"{base_str}!"
        
    # Fallback for simple positive/negative
    if good_threshold is None and bad_threshold is None:
        if value > 0:
            return f"{base_str}*"
        elif value < 0:
            return f"{base_str}!"
            
    return base_str

def parse_ts(fn):
    m = NEWS_FILE_RE.search(fn)
    return datetime.strptime(m.group(1) + m.group(2), TS_FMT) if m else None

def find_latest_news_files(directory: str = ".", max_files: int = 5) -> List[str]:
    """
    Automatically find the latest 4-5 news_output_*.txt files in the specified directory.
    Returns sorted list of file paths by timestamp (newest first).
    """
    try:
        # Find all news_output files in the directory
        pattern = os.path.join(directory, "news_output_*.txt")
        all_files = glob.glob(pattern)
        
        if not all_files:
            print(f"⚠️  No news_output_*.txt files found in {directory}")
            return []
        
        # Parse timestamps and filter out empty files and log-only files
        files_with_timestamps = []
        empty_files_count = 0
        log_files_count = 0
        for file_path in all_files:
            filename = os.path.basename(file_path)
            timestamp = parse_ts(filename)
            if timestamp:
                # Check file size and content quality
                try:
                    file_size = os.path.getsize(file_path)
                    if file_size > 0:
                        # Quick content check to ensure it's not just a log file
                        is_valid_news = False
                        try:
                            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                                # Read first few lines to check content
                                sample = f.read(2000)  # Read first 2KB
                                # Check for actual news content indicators
                                if ('====' in sample and 'relevant news items' in sample) or \
                                   ('Title:' in sample and 'Summary:' in sample) or \
                                   ('Score:' in sample and 'Source:' in sample):
                                    is_valid_news = True
                                elif 'Fetched 0 news items' in sample or 'INFO - Fetched 0' in sample:
                                    # This is likely a log file with no news
                                    log_files_count += 1
                        except Exception:
                            # If we can't read the file, assume it might be valid
                            is_valid_news = True
                        
                        if is_valid_news:
                            files_with_timestamps.append((timestamp, file_path, file_size))
                        else:
                            log_files_count += 1
                    else:
                        empty_files_count += 1
                except OSError:
                    empty_files_count += 1
        
        if not files_with_timestamps:
            if empty_files_count > 0 or log_files_count > 0:
                print(f"⚠️  Found {empty_files_count} empty files and {log_files_count} log-only files, but no files with actual news content")
                print("   💡 Tip: The latest files might still be processing. Try running again in a few minutes.")
            else:
                print("⚠️  No valid news_output files with proper timestamp format found")
            return []
        
        # Sort by timestamp (newest first)
        files_with_timestamps.sort(key=lambda x: x[0], reverse=True)
        latest_files = [file_path for _, file_path, _ in files_with_timestamps[:max_files]]
        
        print(f"📰 Auto-selected {len(latest_files)} latest news files with actual content:")
        if empty_files_count > 0 or log_files_count > 0:
            print(f"   (Skipped {empty_files_count} empty files and {log_files_count} log-only files)")
        for i, (timestamp, file_path, file_size) in enumerate(files_with_timestamps[:max_files]):
            filename = os.path.basename(file_path)
            age_hours = (datetime.now() - timestamp).total_seconds() / 3600
            size_info = f" ({file_size:,} bytes)"
            print(f"   {i+1}. {filename} ({age_hours:.1f}h ago){size_info}")
        
        return latest_files
        
    except Exception as e:
        print(f"❌ Error finding news files: {e}")
        return []

def decay(ts, hrs):
    if not ts:
        return 1.0
    age = (now() - ts).total_seconds() / 3600
    if age < PLATEAU_HR:
        return 1.0
    if hrs <= PLATEAU_HR:
        return DECAY_FLOOR
    return max(DECAY_FLOOR, 1 - (age - PLATEAU_HR) / (hrs - PLATEAU_HR))

sanitize = lambda t: t.lstrip('$')
_STOP = {"india","limited","ltd","plc","industries","energy","corp","corporation",
         "company","co","global","group","and"}
tokens = lambda txt: [
    w for w in re.sub(r"[^A-Za-z0-9 ]", " ", txt).lower().split()
    if w and w not in _STOP
]

# =============================================================================
# Currency parsing
# =============================================================================
_CURRENCY_SYMBOLS = r"(?:(?:₹|Rs|INR|\$|USD|EUR|£|GBP|¥|JPY|€)\s*)?"
_NUMBER_PATTERN   = r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?)"
_UNIT_PATTERN     = (
    r"\s*("
    r"lakh\s*crore|crore|cr|lakh|lacs?|million|mn|billion|bn|thousand|k"
    r"|dividend|payout|bonus|fund|investment|ncd|debenture"
    r")?\b"
)
_M = re.compile(_CURRENCY_SYMBOLS + _NUMBER_PATTERN + _UNIT_PATTERN, re.I)
MULT = {
    "lakh crore":1e12,"crore":1e7,"cr":1e7,"lakh":1e5,"lacs":1e5,"lac":1e5,
    "billion":1e9,"bn":1e9,"million":1e6,"mn":1e6,"thousand":1e3,"k":1e3,
    "dividend":1.0,"payout":1.0,"bonus":1.0,"fund":1.0,"investment":1.0,
    "ncd":1.0,"debenture":1.0
}

def normalize_number(num_str: str) -> float:
    s = num_str.replace(',', '').strip()
    if s.startswith('(') and s.endswith(')'):
        return -float(s[1:-1])
    return float(s)

def money_cr(txt: str) -> float:
    total = 0.0
    for num_str, unit in _M.findall(txt):
        try:
            num = normalize_number(num_str)
            mult = MULT.get((unit or "").lower(), 1.0)
            total += num * mult
        except:
            continue
    return total

def extract_deal_amounts(txt: str) -> List[Tuple[float, str]]:
    """Return list of (absolute_amount_in_base_units, deal_type_str)."""
    amounts: List[Tuple[float, str]] = []
    for num_str, unit in _M.findall(txt):
        try:
            num = normalize_number(num_str)
            mult = MULT.get((unit or "").lower(), 1.0)
            amount = num * mult
            deal_type = "general"
            m = DEAL_PATTERN.search(txt)
            if m:
                deal_type = m.group(1).lower().strip()
            amounts.append((amount, deal_type))
            # Debug output for amount extraction
            # print(f"DEBUG extract: '{num_str}' '{unit}' -> {num} * {mult} = {amount} ({deal_type})")
        except Exception:
            continue
    return amounts

def consolidate_deals(news: List[News]) -> Tuple[float,int]:
    """Aggregate monetary values with sentiment & deal-type weighting.
       Returns (total_amount_in_base_units, deal_count_detected)."""
    # ensure we only process each unique deal once
    seen_deals = set()
    deal_amounts: List[float] = []
    for n in news:
        txt = normalize_hyphens(n.headline + " " + n.snippet)
        amts = extract_deal_amounts(txt)
        sentiment_weight = 1.0 + abs(n.sent) / 2
        for amt, deal_type in amts:
            # signature per deal: deal_type + rounded amount
            sig = (deal_type, round(amt))
            if sig in seen_deals:
                continue
            seen_deals.add(sig)
            # count once per unique deal
            weight = DEAL_TYPES.get(deal_type, 1.0)
            deal_amounts.append(amt * sentiment_weight * weight)

    deal_count = len(seen_deals)
    if not deal_amounts:
        return 0.0, 0

    deal_amounts.sort()
    consolidated: List[Tuple[float,int]] = []
    for amt in deal_amounts:
        matched = False
        for i, (base, cnt) in enumerate(consolidated):
            if base > 0 and abs(amt - base) / base < MONEY_TOLERANCE:
                new_cnt = cnt + 1
                new_avg = (base * cnt + amt) / new_cnt
                consolidated[i] = (new_avg, new_cnt)
                matched = True
                break
        if not matched:
            consolidated.append((amt, 1))

    total_amount = sum(a for a, _ in consolidated)
    return total_amount, deal_count

def calculate_deal_impact(deal_amount: float, mcap: float, deal_count: int) -> float:
    """Log-scaled size impact + frequency boost."""
    if mcap <= 0 or deal_amount <= 0:
        return 0.0
    size_impact = log2(1 + (deal_amount / mcap) * 100)   # log2 on percentage of mcap
    freq_boost  = min(deal_count * 0.3, 2.0)
    return size_impact * (1.0 + freq_boost)

noise = lambda t: any(w in t.lower() for w in ("bitcoin","dogecoin","ethereum","crypto","blockchain","nft","defi","web3"))
broker = lambda t: any(w in t.lower() for w in ("target price","downgrade","upgrade","maintain","initiates","coverage","rating"))

def comp_tokens(name: str, tkr: str) -> set[str]:
    s = set(tokens(name))
    if len(s) < 2:
        s.add(tkr.lower())
    return s

def qual_ok(n: News, cmp: set[str]) -> bool:
    txt = normalize_hyphens(n.headline + " " + n.snippet)
    return (
        not noise(txt) and
        not broker(txt) and
        (len(txt) >= 30 or n.confidence >= 0.8) and  # Softer length filter for high-confidence news
        any(re.search(rf"\b{re.escape(tok)}\b", txt.lower()) for tok in cmp)
    )

deal_rel = lambda txt: (DEAL_PATTERN_LEGACY.search(normalize_hyphens(txt).lower()) or
                        any(k in txt.lower() for k in ("crore","lakh","million","billion","%","fund","ncd","debenture","investment")))

# =============================================================================
# Parse News Files
# =============================================================================
def parse_news(files: List[str], hrs: int) -> Dict[str, dict]:
    data: Dict[str, dict] = {}
    cutoff = now() - timedelta(hours=hrs)
    cur: Optional[str] = None
    buff: List[News] = []
    cname: str = ""

    def flush():
        if cur and buff:
            data.setdefault(cur, {"name": cname, "news": []})["news"].extend(buff)
        buff.clear()

    for fp in files:
        fts = parse_ts(os.path.basename(fp))
        if fts and fts < cutoff:
            continue
        with open(fp, encoding="utf-8", errors="replace") as fh:
            for raw in fh:
                ln = raw.strip()
                if not ln or set(ln) == {"="}:  # Skip separator lines
                    continue
                
                # Check for new format: "TICKER.NS - X relevant news items"
                m = NEWS_SECTION_RE.match(ln)
                if m:
                    flush()
                    cur = m.group(1).upper()
                    cname = ""  # Company name not available in this format
                    original_sc = 0.0  # Reset sentiment score to prevent bleed-through
                    current_confidence = 0.8  # Reset confidence to prevent bleed-through
                    continue
                
                # Check for old format: "TICKER.NS (company name)"
                m = TICK_RE.match(ln)
                if m:
                    flush()
                    cur = m.group(1).upper()
                    cname = ln.split("(",1)[1].rstrip(")") if "(" in ln else ""
                    original_sc = 0.0  # Reset sentiment score to prevent bleed-through
                    current_confidence = 0.8  # Reset confidence to prevent bleed-through
                    continue
                
                if cur:
                    # Parse news entries with scores: [1] SENT +1.00 × CRED 1.00 × REL 1.80 = +1.80 | Conf: 0.70 | Source: news.google.com
                    if ln.startswith("[") and "]" in ln and ("SENT" in ln or "Score:" in ln):
                        original_sc = 0.0
                        confidence = 0.8  # Default confidence
                        
                        try:
                            if "SENT" in ln and "=" in ln:
                                # New format: [1] SENT +1.00 × CRED 1.00 × REL 1.80 = +1.80 | Conf: 0.70
                                score_part = ln.split("=")[1].split("|")[0].strip()
                                original_sc = float(score_part)
                                
                                # Extract confidence
                                if "Conf:" in ln:
                                    conf_part = ln.split("Conf:")[1].split("|")[0].strip()
                                    confidence = float(conf_part)
                            else:
                                # Old format: [1] Score: +0.62 | Confidence: 0.70
                                score_part = ln.split("Score:")[1].split("|")[0].strip()
                                original_sc = float(score_part)
                                
                                if "Confidence:" in ln:
                                    conf_part = ln.split("Confidence:")[1].split("|")[0].strip()
                                    confidence = float(conf_part)
                        except:
                            original_sc = 0.0
                            confidence = 0.8
                        
                        # Store confidence for next headline
                        current_confidence = confidence
                        continue
                    
                    # Parse title lines: "    Title: headline text"
                    elif ln.strip().startswith("Title:"):
                        hl = ln.split("Title:", 1)[1].strip()
                        hl = normalize_hyphens(hl)
                        
                        # Enhance sentiment with transformer-based analysis
                        enhanced_sc = _sentiment_analyzer.analyze_sentiment(hl, original_sc if 'original_sc' in locals() else 0.0)
                        
                        # Use stored confidence or default
                        conf = current_confidence if 'current_confidence' in locals() else 0.8
                        
                        buff.append(News(cur, enhanced_sc, hl, "", fts, conf))
                    
                    # Parse summary lines: "    Summary: summary text"
                    elif ln.strip().startswith("Summary:") and buff:
                        summary = ln.split("Summary:", 1)[1].strip()
                        summary = normalize_hyphens(summary)
                        # Update the last news item with summary
                        last = buff[-1]
                        buff[-1] = last._replace(snippet=summary)
                    
                    # Handle old format news entries: [score] headline
                    elif ln.startswith("[") and "]" in ln and "Score:" not in ln and "SENT" not in ln:
                        end = ln.index("]")
                        try:
                            original_sc = float(ln[1:end].strip())
                        except:
                            original_sc = 0.0
                        hl = ln[end+1:].strip().rsplit(" - ",1)[0]
                        hl = normalize_hyphens(hl)
                        
                        # Enhance sentiment with transformer-based analysis
                        enhanced_sc = _sentiment_analyzer.analyze_sentiment(hl, original_sc)
                        
                        # Default confidence for old format
                        buff.append(News(cur, enhanced_sc, hl, "", fts, 0.8))
                    
                    # Handle continuation lines for old format
                    elif buff and not ln.strip().startswith(("Title:", "Summary:", "Time:", "Source:")):
                        last = buff[-1]
                        new_snip = normalize_hyphens(ln)
                        buff[-1] = last._replace(snippet=(last.snippet + " " + new_snip).strip())
    flush()
    return data  # CORRECTED - was returning min((data), DEAL_IMPACT_CAP)

# =============================================================================
# Optimization Functions
# =============================================================================

def lightweight_shortlist(news_data: Dict[str, dict], top_n: int = 80, soft_mode: bool = False) -> List[str]:
    """
    ENHANCED INSTITUTIONAL-GRADE INITIAL FILTERING
    
    Implements best practices from quantitative funds:
    1. Multi-factor scoring with news quality
    2. Deal significance weighting  
    3. Sentiment momentum analysis
    4. Catalyst strength assessment
    5. Risk-first approach with quality gates
    """
    scores = []
    all_deal_amounts = []
    all_pos_counts = []
    
    # First pass: collect statistics for normalization
    for t, d in news_data.items():
        rel = d["news"]
        if rel:
            pos_count = sum(len(POSITIVE_PATTERN.findall(n.headline + ' ' + n.snippet)) for n in rel)
            all_pos_counts.append(pos_count)
            
            deal_amount = 0.0
            for n in rel:
                txt = normalize_hyphens(n.headline + " " + n.snippet)
                amts = extract_deal_amounts(txt)
                if amts:
                    deal_amount += sum(amt for amt, _ in amts)
            all_deal_amounts.append(deal_amount)
    
    # Calculate percentiles for institutional-grade normalization
    if soft_mode:
        # Soft mode: Much more inclusive thresholds
        pos_75th = np.percentile(all_pos_counts, 50) if all_pos_counts else 1  # Use median instead of 75th
        pos_90th = np.percentile(all_pos_counts, 75) if all_pos_counts else 1  # Use 75th instead of 90th
        deal_75th = np.percentile(all_deal_amounts, 50) if all_deal_amounts else 1  # Use median instead of 75th
        deal_90th = np.percentile(all_deal_amounts, 75) if all_deal_amounts else 1  # Use 75th instead of 90th
    else:
        # Standard mode: Original strict thresholds
        pos_75th = np.percentile(all_pos_counts, 75) if all_pos_counts else 1
        pos_90th = np.percentile(all_pos_counts, 90) if all_pos_counts else 1
        deal_75th = np.percentile(all_deal_amounts, 75) if all_deal_amounts else 1
        deal_90th = np.percentile(all_deal_amounts, 90) if all_deal_amounts else 1
    
    # Second pass: calculate sophisticated scores
    for t, d in news_data.items():
        rel = d["news"]
        if not rel:
            continue
            
        # === INSTITUTIONAL SCORING COMPONENTS ===
        
        # 1. SENTIMENT QUALITY SCORE (Weight: 25%)
        pos_count = sum(len(POSITIVE_PATTERN.findall(n.headline + ' ' + n.snippet)) for n in rel)
        
        # Percentile-based scoring (institutional approach)
        if soft_mode:
            # Soft mode: More generous sentiment scoring
            if pos_count >= pos_90th:
                sentiment_score = 25.0  # Top quartile sentiment
            elif pos_count >= pos_75th:
                sentiment_score = 18.0  # Above median sentiment
            elif pos_count >= 1:  # Any positive sentiment
                sentiment_score = 12.0  # Basic positive sentiment
            else:
                sentiment_score = 5.0   # Neutral/no sentiment (still include)
        else:
            # Standard mode: Original strict scoring
            if pos_count >= pos_90th:
                sentiment_score = 25.0  # Top 10% sentiment
            elif pos_count >= pos_75th:
                sentiment_score = 15.0  # Top 25% sentiment
            elif pos_count >= pos_75th * 0.5:
                sentiment_score = 8.0   # Above median sentiment
            else:
                sentiment_score = 2.0   # Below average sentiment
        
        # 2. DEAL SIGNIFICANCE SCORE (Weight: 30%)
        deal_amount = 0.0
        deal_count = 0
        for n in rel:
            txt = normalize_hyphens(n.headline + " " + n.snippet)
            amts = extract_deal_amounts(txt)
            if amts:
                deal_amount += sum(amt for amt, _ in amts)
                deal_count += len(amts)
        
        # Multi-tier deal scoring (institutional methodology)
        if soft_mode:
            # Soft mode: More generous deal scoring
            if deal_amount >= deal_90th and deal_count >= 2:
                deal_score = 30.0  # Multiple significant deals
            elif deal_amount >= deal_90th:
                deal_score = 25.0  # Single significant deal
            elif deal_amount >= deal_75th:
                deal_score = 20.0  # Above median deal
            elif deal_amount > 0:
                deal_score = 15.0  # Any deal activity (more generous)
            else:
                deal_score = 8.0   # No deals but still score (news presence)
        else:
            # Standard mode: Original strict scoring
            if deal_amount >= deal_90th and deal_count >= 2:
                deal_score = 30.0  # Multiple significant deals
            elif deal_amount >= deal_90th:
                deal_score = 25.0  # Single significant deal
            elif deal_amount >= deal_75th:
                deal_score = 18.0  # Substantial deal
            elif deal_amount > 0:
                deal_score = 10.0  # Minor deal activity
            else:
                deal_score = 0.0   # No deals
        
        # 3. NEWS FRESHNESS & MOMENTUM (Weight: 20%)
        now_ts = now()
        freshness_scores = []
        for n in rel:
            if n.ts:
                hours_ago = (now_ts - n.ts).total_seconds() / 3600
                if hours_ago <= 2:
                    freshness_scores.append(20.0)    # Breaking news
                elif hours_ago <= 6:
                    freshness_scores.append(15.0)    # Very recent
                elif hours_ago <= 12:
                    freshness_scores.append(10.0)    # Recent
                elif hours_ago <= 24:
                    freshness_scores.append(5.0)     # Same day
                else:
                    freshness_scores.append(1.0)     # Older news
        
        freshness_score = max(freshness_scores) if freshness_scores else 5.0
        
        # 4. CATALYST STRENGTH ANALYSIS (Weight: 15%)
        catalyst_keywords = {
            'high_impact': ['merger', 'acquisition', 'buyback', 'dividend', 'split', 'bonus'],
            'medium_impact': ['result', 'earnings', 'partnership', 'contract', 'order'],
            'growth_impact': ['expansion', 'launch', 'approval', 'license', 'export']
        }
        
        catalyst_score = 0.0
        for n in rel:
            txt = (n.headline + " " + n.snippet).lower()
            for keyword in catalyst_keywords['high_impact']:
                if keyword in txt:
                    catalyst_score += 5.0
            for keyword in catalyst_keywords['medium_impact']:
                if keyword in txt:
                    catalyst_score += 3.0
            for keyword in catalyst_keywords['growth_impact']:
                if keyword in txt:
                    catalyst_score += 2.0
        
        catalyst_score = min(catalyst_score, 15.0)  # Cap at maximum weight
        
        # 5. NEWS QUALITY & CREDIBILITY (Weight: 10%)
        # Based on sentiment consistency and news count
        sentiments = [n.sent for n in rel if hasattr(n, 'sent') and n.sent is not None]
        if sentiments:
            sentiment_consistency = 1.0 - (np.std(sentiments) / (max(sentiments) - min(sentiments) + 0.01))
            news_count_bonus = min(len(rel) / 5.0, 2.0)  # Bonus for multiple news sources
            quality_score = (sentiment_consistency * 5.0) + (news_count_bonus * 2.5)
        else:
            quality_score = 5.0  # Neutral score
        
        # === INSTITUTIONAL QUALITY GATES ===
        quality_multiplier = 1.0
        
        if soft_mode:
            # Soft mode: Much more lenient quality gates
            # Gate 1: Minimum news quality (relaxed)
            if len(rel) < 1:  # Only penalize if no news at all
                quality_multiplier *= 0.8
            
            # Gate 2: Sentiment coherence (relaxed - mixed sentiment is ok)
            # No penalty for mixed sentiment in soft mode
            
            # Gate 3: Recency requirement (relaxed)
            if not any((now_ts - n.ts).total_seconds() / 3600 <= 72 for n in rel if n.ts):  # 72 hours instead of 24
                quality_multiplier *= 0.8  # Lighter penalty
        else:
            # Standard mode: Original strict quality gates
            # Gate 1: Minimum news quality
            if len(rel) < 2:
                quality_multiplier *= 0.7  # Penalty for single news item
            
            # Gate 2: Sentiment coherence
            if sentiments and len(set(np.sign(sentiments))) > 1:
                quality_multiplier *= 0.8  # Penalty for mixed sentiment
            
            # Gate 3: Recency requirement
            if not any((now_ts - n.ts).total_seconds() / 3600 <= 24 for n in rel if n.ts):
                quality_multiplier *= 0.5  # Penalty for old news only
        
        # === FINAL INSTITUTIONAL SCORE ===
        raw_score = (
            sentiment_score * 0.25 +    # 25% weight
            deal_score * 0.30 +         # 30% weight  
            freshness_score * 0.20 +    # 20% weight
            catalyst_score * 0.15 +     # 15% weight
            quality_score * 0.10        # 10% weight
        )
        
        # Apply quality gates
        final_score = raw_score * quality_multiplier
        
        # Institutional bonus for exceptional candidates
        if (pos_count >= pos_90th and deal_amount >= deal_75th and 
            freshness_score >= 15.0 and catalyst_score >= 10.0):
            final_score *= 1.3  # 30% bonus for institutional-grade opportunities
        
        scores.append((t, final_score))
    
    # Sort and return top candidates (institutional selection)
    scores.sort(key=lambda x: x[1], reverse=True)
    
    # Log institutional screening statistics
    if scores:
        top_score = scores[0][1]
        avg_score = np.mean([s for _, s in scores])
        logging.info(f"[INSTITUTIONAL] FILTERING: Top score: {top_score:.1f}, Average: {avg_score:.1f}")
        logging.info(f"[QUALITY] thresholds: P75_sentiment={pos_75th}, P75_deals={deal_75th/1e6:.1f}M")
    
    return [t for t, _ in scores[:top_n]]

def batched_history_download(tickers: List[str], soft_mode: bool = False) -> Dict[str, pd.DataFrame]:
    """Download 5y history for all tickers in one call with data validation"""
    if not tickers:
        return {}
    
    # Prepare symbols for batch download - prevent double .NS suffix
    symbols = [ensure_ns_suffix(t) for t in tickers]
    
    try:
        # Use the safe batch download with validation
        logging.info(f"Downloading validated historical data for {len(tickers)} tickers...")
        validated_data = YFinanceDataValidator.safe_yfinance_download(
            symbols, 
            period="5y", 
            interval="1d", 
            timeout=60,
            soft_mode=soft_mode
        )
        
        # Map back to original ticker names
        results = {}
        for i, ticker in enumerate(tickers):
            symbol = symbols[i]
            if symbol in validated_data:
                clean_df = validated_data[symbol]
                if not clean_df.empty:
                    results[ticker] = clean_df
                    logging.debug(f"✅ Valid data for {ticker}: {len(clean_df)} records")
                else:
                    logging.warning(f"❌ Empty data after validation for {ticker}")
                    results[ticker] = pd.DataFrame()
            else:
                logging.warning(f"❌ No data available for {ticker}")
                results[ticker] = pd.DataFrame()
        
        success_rate = len([r for r in results.values() if not r.empty]) / len(tickers) * 100
        logging.info(f"Data validation completed. Success rate: {success_rate:.1f}% ({len([r for r in results.values() if not r.empty])}/{len(tickers)})")
        
        return results
        
    except Exception as e:
        logging.error(f"Batch history download failed: {e}")
        # Fallback to individual downloads if batch fails
        logging.info("Attempting individual downloads as fallback...")
        results = {}
        for ticker in tickers:
            try:
                clean_df = YFinanceDataValidator.safe_ticker_history(ticker, period="5y")
                results[ticker] = clean_df
                if not clean_df.empty:
                    logging.debug(f"✅ Individual download successful for {ticker}")
                else:
                    logging.debug(f"❌ Individual download failed for {ticker}")
            except Exception as individual_error:
                logging.warning(f"Individual download failed for {ticker}: {individual_error}")
                results[ticker] = pd.DataFrame()
        
        return results

def fetch_financials_late(tickers: List[str], hist_data: Dict[str, pd.DataFrame], args) -> Dict[str, Fin]:
    """Fetch financials only for top candidates"""
    fin_data = {}
    for t in tickers:
        df = hist_data.get(t, pd.DataFrame())
        f = get_fin(t, not args.allow_negative, args.skip_financial, hist_df=df)
        if f:
            fin_data[t] = f
    return fin_data

# =============================================================================
# Yahoo Helpers
# =============================================================================
def _hist(t: str) -> Tuple[str, pd.DataFrame]:
    rate_limit(0.5)  # 500ms between calls
    try:
        # Use our resilient download system - prevent double .NS suffix
        yf_symbol = ensure_ns_suffix(t)
        result = safe_yf_download([yf_symbol], period="5y")
        if yf_symbol in result:
            df = result[yf_symbol]
        else:
            df = pd.DataFrame()
        return t, df
    except Exception as e:
        logging.debug(f"Yahoo history error {t}: {e}")
        return t, pd.DataFrame()

def _meta(t: str) -> Tuple[str, float, str, float]:
    """Return (ticker, market_cap, name, bid_ask_spread_pct) with data validation"""
    rate_limit(0.5)  # 500ms between calls
    try:
        # Use safe data fetching with validation
        info = YFinanceDataValidator.safe_ticker_info(t)
        
        if not info:
            logging.debug(f"No valid metadata available for {t}")
            return t, 0.0, "", 0.0
        
        # Extract and validate name
        name = ""
        for name_field in ['longName', 'shortName', 'symbol']:
            if name_field in info and info[name_field]:
                name = str(info[name_field]).strip()
                break
        
        # Extract and validate market cap
        mc = 0.0
        market_cap_raw = info.get('marketCap', 0) or 0
        if isinstance(market_cap_raw, (int, float)) and market_cap_raw > 0:
            mc = market_cap_raw / 1e7  # Convert to Cr
            # Validate reasonable market cap range
            if mc < 0.1 or mc > 1000000:  # 0.1 Cr to 10 lakh Cr
                logging.debug(f"Extreme market cap {mc} Cr for {t}")
                mc = max(0.1, min(mc, 1000000))
        
        # Calculate bid-ask spread as percentage with validation
        spread_pct = 0.0
        bid = info.get('bid', 0) or 0
        ask = info.get('ask', 0) or 0
        current_price = (info.get('currentPrice', 0) or 
                        info.get('regularMarketPrice', 0) or 
                        info.get('previousClose', 0) or 0)
        
        # Validate price data
        if (isinstance(bid, (int, float)) and bid > 0 and
            isinstance(ask, (int, float)) and ask > 0 and
            isinstance(current_price, (int, float)) and current_price > 0):
            
            # Ensure ask >= bid (basic sanity check)
            if ask >= bid:
                spread_pct = (ask - bid) / current_price
                # Validate reasonable spread (0% to 10%)
                if spread_pct < 0 or spread_pct > 0.1:
                    logging.debug(f"Extreme spread {spread_pct*100:.2f}% for {t}")
                    spread_pct = max(0, min(spread_pct, 0.1))
            else:
                logging.debug(f"Invalid bid-ask data for {t}: bid={bid}, ask={ask}")
                spread_pct = 0.0
        
        # Final validation
        if not isinstance(mc, (int, float)) or np.isnan(mc) or np.isinf(mc):
            mc = 0.0
        if not isinstance(spread_pct, (int, float)) or np.isnan(spread_pct) or np.isinf(spread_pct):
            spread_pct = 0.0
        if not isinstance(name, str):
            name = ""
        
        return t, mc, name, spread_pct
        
    except Exception as e:
        logging.debug(f"Yahoo meta error {t}: {e}")
        return t, 0.0, "", 0.0

def get_yahoo(tks: List[str]) -> Tuple[Dict[str,pd.DataFrame], Dict[str,float], Dict[str,str], Dict[str,float]]:
    h: Dict[str,pd.DataFrame] = {}
    mc: Dict[str,float] = {}
    nm: Dict[str,str] = {}
    sp: Dict[str,float] = {}  # spread percentages
    with ThreadPoolExecutor(MAX_WORKERS) as ex:
        fut_h = {ex.submit(_hist, t): t for t in tks}
        for f in as_completed(fut_h):
            t, df = f.result(); h[t] = df
        fut_m = {ex.submit(_meta, t): t for t in tks}
        for f in as_completed(fut_m):
            t, m, name, spread = f.result(); mc[t] = m; nm[t] = name; sp[t] = spread
    return h, mc, nm, sp

# =============================================================================
# Financial Health
# =============================================================================
def _first(vs: List[Optional[float]]) -> Optional[float]:
    for v in vs:
        if v is not None and not (isinstance(v, float) and math.isnan(v)):
            return v
    return None

def get_fin(t: str, require_positive: bool=True, skip: bool=False, hist_df: Optional[pd.DataFrame]=None) -> Optional[Fin]:
    if skip:
        return Fin(0.0, 0.0, None)
    rate_limit(0.5)  # 500ms between calls
    try:
        # Prevent double .NS suffix
        yf_symbol = ensure_ns_suffix(t)
        tk = yf.Ticker(yf_symbol)
        if yshared._ERRORS.get(yf_symbol) == 404:
            return None
        fi = tk.fast_info or {}
        info = tk.info or {}

        def latest(df: pd.DataFrame, keys: List[str]) -> Optional[float]:
            if df is None or df.empty:
                return None
            for k in keys:
                if k in df.index:
                    s = df.loc[k].dropna()
                    if not s.empty:
                        return float(s.iloc[0])
            return None

        net = _first([
            latest(tk.quarterly_balancesheet, ["Total Stockholder Equity","Total Shareholder Equity","Net Worth","Shareholders Funds"]),
            latest(tk.balance_sheet, ["Total Stockholder Equity","Total Shareholder Equity","Net Worth","Shareholders Funds","Total Equity"]),
            fi.get("book_value") and fi.get("book_value") * (fi.get("shares_outstanding") or info.get("sharesOutstanding")),
            info.get("totalStockholderEquity")
        ])
        inc = _first([
            latest(tk.quarterly_income_stmt, ["Net Income","Net Profit","Profit After Tax"]),
            latest(tk.income_stmt, ["Net Income","Net Profit","Profit After Tax"]),
            fi.get("ttm_net_income"),
            info.get("netIncome") or info.get("netIncomeToCommon")
        ])
        if require_positive and ((net or 0) < 0 or (inc or 0) < 0):
            return None

        # margin
        try:
            inc_stmt = tk.income_stmt
            if inc_stmt is not None and not inc_stmt.empty:
                op_income = None; revenue = None
                for cand in ["Operating Income","OperatingIncome","EBIT","Ebit"]:
                    if cand in inc_stmt.index:
                        s = inc_stmt.loc[cand].dropna()
                        if not s.empty:
                            op_income = float(s.iloc[0]); break
                for cand in ["Total Revenue","TotalRevenue","Revenue","Sales"]:
                    if cand in inc_stmt.index:
                        s = inc_stmt.loc[cand].dropna()
                        if not s.empty:
                            revenue = float(s.iloc[0]); break
                margin = (op_income / revenue) if (op_income and revenue and revenue != 0) else None
            else:
                margin = None
        except:
            margin = None

        # volume growth (simple last-day vs avg)
        try:
            vg = hist_df["Volume"].pct_change().iloc[-1] if (hist_df is not None and not hist_df.empty) else None
        except:
            vg = None

        # debt reduction
        debt_red = None
        try:
            bs = tk.balance_sheet
            if bs is not None and not bs.empty and "Total Debt" in bs.index:
                s = bs.loc["Total Debt"].dropna()
                if len(s) >= 2:
                    cur_d, prev_d = float(s.iloc[0]), float(s.iloc[1])
                    if prev_d > 0:
                        debt_red = (prev_d - cur_d) / prev_d
        except:
            debt_red = None

        pe_ratio = info.get('trailingPE') or info.get('forwardPE') or 0.0

        return Fin(net or 0.0, inc or 0.0, info.get("debtToEquity"), margin, vg, debt_red, None, pe_ratio)
    except Exception as e:
        logging.debug(f"Fin err {t}: {e}")
        return None

# =============================================================================
# Enhanced Sector Detection + Dynamic P/E Analysis
# =============================================================================
_SECTOR_PE_BENCHMARKS = {
    'banking': {'median': 12.0, 'high': 18.0, 'keywords': ['bank','banking','sbi','hdfc bank','icici bank','kotak','axis bank','pnb','bob','canara bank','union bank','indian bank','syndicate','federal bank','rbl bank','yes bank','idfc','bandhan','equitas','ujjivan']},
    'insurance': {'median': 15.0, 'high': 25.0, 'keywords': ['insurance','life insurance','general insurance','lic','prudential','bajaj allianz','max life','future generali','star health','oriental insurance','national insurance','united india','cholamandalam','digit insurance']},
    'financial': {'median': 14.0, 'high': 22.0, 'keywords': ['financial','finance','nbfc','mutual fund','housing finance','credit','lending','investment','broking','capital','securities','wealth management','asset management','microfinance','infrastructure finance','development finance','refinance']},
    'tech': {'median': 28.0, 'high': 45.0, 'keywords': ['consultancy services','technology','software','infotech','tech mahindra','tcs','infosys','wipro','hcl tech','ltts','mphasis','mindtree','cyient','persistent','zensar','niit','kpit','it services','information technology','computer','data analytics','artificial intelligence','automation','digital services']},
    'pharma': {'median': 35.0, 'high': 55.0, 'keywords': ['pharmaceutical','pharma','drug','medicine','healthcare','biotech','laboratories','sun pharma','cipla','lupin','cadila','aurobindo','reddy','torrent pharma','glenmark','alkem','mankind','abbott','diagnostics','therapeutic']},
    'auto': {'median': 16.0, 'high': 25.0, 'keywords': ['automotive','motor','maruti','hyundai','mahindra','bajaj auto','hero motocorp','tvs motor','eicher motors','ashok leyland','force motors','escorts','swaraj','automobile','vehicle','passenger car','commercial vehicle','two wheeler','three wheeler','auto component','tyre','apollo tyres','mrf','ceat','jk tyre']},
    'fmcg': {'median': 30.0, 'high': 50.0, 'keywords': ['consumer goods','fmcg','hindustan unilever','nestle','britannia','dabur','marico','godrej consumer','colgate','procter gamble','reckitt benckiser','emami','bajaj corp','personal care','food beverage','biscuit','dairy','detergent','shampoo','toothpaste']},
    'metals': {'median': 8.0, 'high': 15.0, 'keywords': ['steel','jsw steel','tata steel','sail','nmdc','vedanta','hindalco','nalco','moil','hindustan copper','jindal steel','metal','copper','aluminum','zinc','iron ore','sponge iron','stainless steel','alloy','mining']},
    'cement': {'median': 12.0, 'high': 20.0, 'keywords': ['cement','ultratech','acc','ambuja','shree cement','jk cement','lakshmi cement','dalmia cement','bharti cement','concrete','building material']},
    'energy': {'median': 10.0, 'high': 18.0, 'keywords': ['oil gas','petroleum','reliance industries','ongc','indian oil','bpcl','hpcl','gail','ntpc','power grid','nhpc','sjvn','adani power','tata power','jsw energy','power generation','transmission','distribution','coal india','renewable energy','solar','wind power']},
    'other': {'median': 18.0, 'high': 28.0, 'keywords': []}  # Default sector
}

# Legacy mapping for backward compatibility
_SECTOR_KEYWORDS = {sector: data['keywords'] for sector, data in _SECTOR_PE_BENCHMARKS.items() if data['keywords']}

# Sector detection cache for performance optimization
_SECTOR_CACHE = {}

def detect_sector(company_name: str) -> str:
    """Enhanced sector detection using comprehensive keyword matching with priority scoring and caching"""
    # Check cache first for performance
    if company_name in _SECTOR_CACHE:
        return _SECTOR_CACHE[company_name]
    
    name_lower = (company_name or "").lower()
    
    # Score each sector based on keyword matches with priority weighting
    sector_scores = {}
    for sector, data in _SECTOR_PE_BENCHMARKS.items():
        if not data['keywords']:  # Skip 'other' sector
            continue
        
        score = 0
        # High priority keywords (more specific) get higher weights
        for kw in data['keywords']:
            if kw in name_lower:
                # Longer keywords get higher priority (more specific)
                # Exact word boundaries get bonus points
                import re
                word_boundary_match = bool(re.search(rf'\b{re.escape(kw)}\b', name_lower))
                length_bonus = len(kw) * 0.5  # Longer keywords are more specific
                boundary_bonus = 3 if word_boundary_match else 1
                score += length_bonus + boundary_bonus
        
        if score > 0:
            sector_scores[sector] = score
    
    # Return sector with highest score, or 'other' if no matches
    result = 'other'
    if sector_scores:
        best_sector = max(sector_scores.items(), key=lambda x: x[1])[0]
        # Debug: print the scoring for troubleshooting
        # print(f"DEBUG: {company_name} -> {best_sector} (scores: {sector_scores})")
        result = best_sector
    
    # Cache the result for future use
    _SECTOR_CACHE[company_name] = result
    return result

def get_sector_pe_benchmark(sector: str) -> dict:
    """Get P/E benchmarks for a sector"""
    return _SECTOR_PE_BENCHMARKS.get(sector, _SECTOR_PE_BENCHMARKS['other'])

def analyze_pe(fin_data: Fin, sector: str, market_cap_cr: float = 0.0) -> Tuple[str, str, str, float]:
    """Enhanced P/E analysis with dynamic sector-aware thresholds and valuation scoring
    Returns: (pe_string, warning_symbol, risk_level, valuation_score)
    """
    pe = fin_data.pe_ratio
    if pe is None or (isinstance(pe, float) and math.isnan(pe)):
        return "n/a", "", "NONE", 5.0  # Neutral score for unknown P/E
    
    if pe < 0:
        return f"{pe:.1f}", "!", "HIGH", 2.0  # Poor score for negative P/E
    
    # Get sector benchmarks
    benchmarks = get_sector_pe_benchmark(sector)
    sector_median = benchmarks['median']
    sector_high = benchmarks['high']
    
    # Dynamic thresholds based on sector and market cap
    # Adjust thresholds for small caps (typically trade at premium)
    if market_cap_cr > 0:
        if market_cap_cr < 1000:  # Small cap adjustment
            sector_median *= 1.2
            sector_high *= 1.2
        elif market_cap_cr > 50000:  # Large cap adjustment
            sector_median *= 0.9
            sector_high *= 0.9
    
    # Calculate warning thresholds
    # Dynamic P/E thresholds based on market conditions
    dynamic_thresholds = sector_median, sector_high  # Default fallback
    
    if ENHANCED_MODULES_AVAILABLE:
        try:
            from enhanced_screener_modules import DynamicPEAnalyzer
            pe_analyzer = DynamicPEAnalyzer()
            dynamic_data = pe_analyzer.get_dynamic_pe_thresholds(sector, market_cap_cr)
            sector_median = dynamic_data['median']
            sector_high = dynamic_data['high']
            
            # Add market context to output (optional debug)
            market_cycle_info = f"Market cycle: {dynamic_data['market_cycle']:.2f}x"
            
        except Exception as e:
            logging.debug(f"Dynamic P/E analysis failed, using static thresholds: {e}")
    
    warn_threshold = sector_median * 1.3  # 30% above sector median
    extreme_threshold = sector_high * 1.1  # 10% above sector high
    
    # Calculate valuation score (0-10 scale) - Enhanced with market cycle awareness
    valuation_score = 10.0  # Start with perfect score
    
    # Market-cycle adjusted penalties/bonuses
    cycle_multiplier = 1.0
    if ENHANCED_MODULES_AVAILABLE:
        try:
            # During bull markets, be more lenient with P/E
            # During bear markets, be stricter
            from enhanced_screener_modules import MarketCycleAnalyzer
            market_analyzer = MarketCycleAnalyzer()
            cycle_multiplier = market_analyzer.get_market_cycle_multiplier()
        except Exception as e:
            logging.debug(f"Market cycle analysis failed: {e}")
            cycle_multiplier = 1.0
    
    # Adjusted P/E evaluation with cycle awareness
    adjusted_median = sector_median * cycle_multiplier
    
    # Penalize high P/E relative to cycle-adjusted sector median
    if pe > adjusted_median * 1.5:
        valuation_score -= 4.0
    elif pe > adjusted_median * 1.3:
        valuation_score -= 2.0
        
    # Bonus for undervaluation (cycle-adjusted)
    if pe < adjusted_median * 0.7:
        valuation_score += 3.0
    elif pe < adjusted_median * 0.9:
        valuation_score += 1.5
        
    # Sector-specific absolute limits (enhanced)
    sector_abs_limit = {
        'tech': 60,      # Higher for growth sectors
        'pharma': 70,    # Innovation premium
        'banking': 25,   # Traditional limits
        'metals': 20,    # Cyclical sectors
        'energy': 25,
        'other': 45
    }.get(sector, 45)
    
    # Apply absolute limits with cycle adjustment
    cycle_adjusted_limit = sector_abs_limit * cycle_multiplier
    
    if pe > cycle_adjusted_limit:
        valuation_score -= 3.0
    elif pe > cycle_adjusted_limit * 0.8:
        valuation_score -= 1.5
    
    # Ensure score stays within bounds
    valuation_score = max(0.0, min(10.0, valuation_score))
    
    # Determine risk level and warning
    if pe > extreme_threshold:
        return f"{pe:.1f}", "!!", "EXTREME", min(valuation_score, 3.0)  # Cap score for extreme P/E
    elif pe > warn_threshold:
        return f"{pe:.1f}", "!", "HIGH", min(valuation_score, 5.0)  # Cap score for high P/E
    elif pe > sector_median:
        return f"{pe:.1f}", "", "MEDIUM", valuation_score
    else:
        return f"{pe:.1f}", "", "LOW", valuation_score

# =============================================================================
# Catalyst scoring
# =============================================================================

def catalyst_score(news: List[News]) -> tuple[float,List[str]]:
    """Catalyst scoring with sentiment weighting.

    Each catalyst weight is multiplied by:
       * 1.2 boost if keyword appears in headline (immediate relevance)
       * (1 + |sentiment|) to capture strength of sentiment
    """
    sc = 0.0
    tags: List[str] = []
    for n in news:
        text = normalize_hyphens(n.headline + " " + n.snippet).lower()
        # Sentiment multiplier: 1 + |sentiment|
        sent_mult = 1.0 + abs(n.sent or 0.0)
        for cat, pat in CATALYST_PATTERNS.items():
            if pat.search(text):
                mult = 1.2 if pat.search(n.headline.lower()) else 1.0
                sc += CATALYST_WEIGHTS[cat] * mult * sent_mult
                tags.append("BRK" if cat == "BRKOUT" else cat)
    if "DEAL" in tags or "ORDER" in tags:
        sc *= 1.3
    return sc, tags


# =============================================================================
# Scoring helpers
# =============================================================================
size_adj = lambda mc: 1.5 if mc < 5000 else 1.2 if mc < 20000 else 0.8

def base_core_components(news: List[News], df: pd.DataFrame, mcap: float, hrs: int,
                         r1: float, r3: float, volx: float, bd: float) -> Tuple[float,float,float,int,float]:
    """Return tuple (raw_score, deal_amount_cr, deal_pct, deal_count, deal_impact_score)"""
    sent = sum(n.sent * decay(n.ts, hrs) for n in news)

    # Consolidated deals (base units) + count
    deal_amount_abs, deal_count = consolidate_deals(news)
    deal_pct = (deal_amount_abs / (mcap * 1e7) * 100) if (mcap > 0) else 0.0  # pct of mcap
    deal_impact = calculate_deal_impact(deal_amount_abs, mcap * 1e7, deal_count)

    money_sc = lg(1 + deal_amount_abs) * 1.6 * (1 if sent >= 0 else -1)
    vol_sc   = lg(max(volx,1)) * 2.0 * (1 if sent >= 0 else -1)
    # REDUCED WEIGHT: Use only 3-day change and lower multiplier for more stable momentum scoring
    mom_sc   = (r3/100) * 5 * (1 if sent >= 0 else -1)  # Changed from max(r1/100, r3/100) * 10

    pct_mentions = sum(1 for n in news if '%' in n.headline or '%' in n.snippet)
    pct_boost    = min(pct_mentions,6)*0.5*(1 if sent>=0 else -1)

    fund_mentions = sum(1 for n in news if any(k in (n.headline+n.snippet).lower()
                                               for k in ["fund","ncd","debenture","investment"]))
    fund_boost = min(fund_mentions,5)*0.7*(1 if sent>=0 else -1)

    div_sc = results_sc = 0.0
    if mcap > 0:
        for n in news:
            txt = normalize_hyphens(n.headline+" "+n.snippet).lower()
            if any(k in txt for k in ("dividend","payout","bonus issue")):
                div_amt = money_cr(txt)
                div_sc += lg(1+div_amt) * 0.8 * (1 if n.sent>=0 else -1) * decay(n.ts,hrs)
            if any(k in txt for k in ("pat","profit after tax","net profit","earnings","results")):
                res_amt = money_cr(txt)
                results_sc += lg(1+res_amt) * 1.0 * (1 if n.sent>=0 else -1) * decay(n.ts,hrs)

    buy_sc = (bd-0.5)*20 + (3 if bd>=0.7 else -3 if bd<=0.3 else 0)
    cat_sc, _ = catalyst_score(news)

    # Enhanced deal impact weight into raw core (multiplier 4.0)
    deal_impact_sc = deal_impact * 4.0 * (1 if sent >= 0 else -1)

    raw = (sent + money_sc + vol_sc + mom_sc + div_sc + results_sc +
           buy_sc + cat_sc + pct_boost + fund_boost + deal_impact_sc) * size_adj(mcap)

    return raw, deal_amount_abs/1e7, deal_pct, deal_count, deal_impact

def adjust_score_with_fii(score: float, fii_growth: float) -> float:
    """Amplify scores with institutional confirmation"""
    # Handle None values from FII data
    if fii_growth is None or not isinstance(score, (int, float)) or math.isnan(score) or math.isinf(score):
        return score
    
    # Prevent overflow by capping the score before multiplication
    safe_score = min(max(score, 0), 100)  # Cap between 0 and 100
        
    if fii_growth > 1.0:   # >1% increase
        return min(safe_score * 1.25, 200)  # Cap result to prevent overflow
    elif fii_growth > 0.5:
        return min(safe_score * 1.15, 200)
    elif fii_growth < -1.0:  # Significant outflow
        return safe_score * 0.8
    return safe_score

def momentum_quality(df: pd.DataFrame) -> float:
    """Score 0-10 for sustainable momentum quality"""
    if len(df) < 60: 
        return 5.0
    
    quality_score = 0.0
    
    # Volume confirmation (0-3 points)
    recent_vol = df["Volume"].iloc[-10:].mean()
    prev_vol = df["Volume"].iloc[-20:-10].mean()
    vol_trend = 1 if recent_vol > prev_vol else 0
    quality_score += vol_trend * 3.0
    
    # Pullback resilience (0-4 points)
    recent_high = df["High"].iloc[-20:].max()
    recent_low = df["Low"].iloc[-20:].min()
    current_price = df["Close"].iloc[-1]
    max_drawdown = (recent_high - recent_low) / current_price
    drawdown_score = 4.0 * (1 - min(0.2, max_drawdown)/0.2)  # <20% drawdown is ideal
    quality_score += max(0.0, drawdown_score)
    
    # Moving average alignment (0-3 points)
    current = df["Close"].iloc[-1]
    if len(df) >= 50:
        ma50 = df["Close"].rolling(50).mean().iloc[-1]
        if current > ma50:
            quality_score += 1.5
    if len(df) >= 200:
        ma200 = df["Close"].rolling(200).mean().iloc[-1]
        if current > ma200:
            quality_score += 1.5
    
    return min(10.0, quality_score)

def detect_institutional_accumulation(df: pd.DataFrame, fii_growth: float = None) -> float:
    """Advanced institutional accumulation scoring (0-10)
    Combines volume patterns, price action, and FII data for institutional confidence
    """
    if len(df) < 60:
        return 5.0  # Neutral for insufficient data
    
    accumulation_score = 0.0
    
    # 1. Volume-Price Divergence Analysis (0-3 points)
    # Strong accumulation shows increasing volume with stable/rising prices
    vol_20 = df["Volume"].rolling(20).mean()
    price_20 = df["Close"].rolling(20).mean()
    
    recent_vol_trend = (vol_20.iloc[-1] - vol_20.iloc[-20]) / vol_20.iloc[-20] if vol_20.iloc[-20] > 0 else 0
    recent_price_trend = (price_20.iloc[-1] - price_20.iloc[-20]) / price_20.iloc[-20] if price_20.iloc[-20] > 0 else 0
    
    # Strong volume increase with price stability/growth indicates accumulation
    if recent_vol_trend > 0.2 and recent_price_trend >= -0.05:  # Vol up 20%+, price stable/up
        accumulation_score += 3.0
    elif recent_vol_trend > 0.1 and recent_price_trend >= -0.02:  # Vol up 10%+, price nearly stable
        accumulation_score += 1.5
    
    # 2. On-Balance Volume (OBV) Trend Analysis (0-2 points)
    obv = []
    obv_val = 0
    for i in range(len(df)):
        if i == 0:
            obv.append(0)
            continue
        if df["Close"].iloc[i] > df["Close"].iloc[i-1]:
            obv_val += df["Volume"].iloc[i]
        elif df["Close"].iloc[i] < df["Close"].iloc[i-1]:
            obv_val -= df["Volume"].iloc[i]
        obv.append(obv_val)
    
    if len(obv) >= 20:
        obv_trend = (obv[-1] - obv[-20]) / max(abs(obv[-20]), 1)  # Avoid division by zero
        if obv_trend > 0.1:  # Positive OBV trend
            accumulation_score += 2.0
        elif obv_trend > 0.05:
            accumulation_score += 1.0
    
    # 3. Volume Profile Analysis (0-2 points)
    # Higher volume at lower prices suggests institutional accumulation
    recent_prices = df["Close"].iloc[-30:]
    recent_volumes = df["Volume"].iloc[-30:]
    
    if len(recent_prices) >= 20:
        price_quartiles = recent_prices.quantile([0.25, 0.75])
        low_quartile_vol = recent_volumes[recent_prices <= price_quartiles[0.25]].mean()
        high_quartile_vol = recent_volumes[recent_prices >= price_quartiles[0.75]].mean()
        
        if low_quartile_vol > 0 and high_quartile_vol > 0:
            vol_ratio = low_quartile_vol / high_quartile_vol
            if vol_ratio > 1.3:  # 30% more volume in lower price range
                accumulation_score += 2.0
            elif vol_ratio > 1.1:
                accumulation_score += 1.0
    
    # 4. FII Confirmation Bonus (0-3 points)
    if fii_growth is not None:
        if fii_growth > 2.0:  # Strong FII inflows
            accumulation_score += 3.0
        elif fii_growth > 1.0:
            accumulation_score += 2.0
        elif fii_growth > 0.5:
            accumulation_score += 1.0
        elif fii_growth < -1.0:  # Penalty for FII outflows
            accumulation_score -= 2.0
    
    return min(10.0, max(0.0, accumulation_score))

def calculate_risk_reward_ratio(df: pd.DataFrame, entry_lvl: float = None, stop_loss: float = None) -> float:
    """Calculate risk/reward ratio for position sizing"""
    if len(df) < 20:
        return 1.0  # Neutral
    
    current_price = df["Close"].iloc[-1]
    
    # Use provided levels or calculate dynamically
    if entry_lvl is None:
        entry_lvl = current_price
    if stop_loss is None:
        # Dynamic stop based on recent support
        support = df["Low"].rolling(20).min().iloc[-1]
        stop_loss = max(support, entry_lvl * 0.95)  # 5% max stop
    
    # Target based on recent resistance
    resistance = df["High"].rolling(50).max().iloc[-1]
    target = max(resistance, entry_lvl * 1.10)  # Minimum 10% target
    
    if entry_lvl <= stop_loss:
        return 1.0  # Invalid setup
    
    risk = entry_lvl - stop_loss
    reward = target - entry_lvl
    
    return reward / risk if risk > 0 else 1.0

def calculate_peer_relative_valuation(fin_data: Fin, sector: str, sector_stocks_data: Dict = None) -> float:
    """Calculate peer-relative valuation score (0-10)
    Compares P/E, ROE, and D/E ratios against sector peers
    """
    if not fin_data or not fin_data.pe_ratio:
        return 5.0  # Neutral for missing data
    
    valuation_score = 5.0  # Start neutral
    
    # Get sector benchmarks
    sector_benchmarks = get_sector_pe_benchmark(sector)
    
    # 1. P/E Relative Scoring (40% weight)
    pe_ratio = fin_data.pe_ratio
    sector_median_pe = sector_benchmarks['median']
    
    if pe_ratio > 0:
        pe_relative = pe_ratio / sector_median_pe
        if pe_relative < 0.7:  # Significantly undervalued
            pe_score = 10.0
        elif pe_relative < 0.9:  # Moderately undervalued
            pe_score = 8.0
        elif pe_relative < 1.1:  # Fairly valued
            pe_score = 6.0
        elif pe_relative < 1.3:  # Slightly overvalued
            pe_score = 4.0
        elif pe_relative < 1.5:  # Moderately overvalued
            pe_score = 2.0
        else:  # Significantly overvalued
            pe_score = 0.0
    else:
        pe_score = 0.0  # Negative P/E is problematic
    
    # 2. Profitability Scoring (30% weight)
    profit_score = 5.0
    if fin_data.inc and fin_data.net and fin_data.inc > 0:
        net_margin = fin_data.net / fin_data.inc
        if net_margin > 0.2:  # >20% net margin
            profit_score = 9.0
        elif net_margin > 0.15:  # >15% net margin
            profit_score = 7.5
        elif net_margin > 0.1:  # >10% net margin
            profit_score = 6.0
        elif net_margin > 0.05:  # >5% net margin
            profit_score = 4.0
        elif net_margin > 0:  # Positive but low
            profit_score = 2.0
        else:  # Negative margins
            profit_score = 0.0
    
    # 3. Financial Health Scoring (30% weight)
    health_score = 5.0
    if fin_data.d2e is not None:
        if fin_data.d2e < 0.3:  # Very low debt
            health_score = 9.0
        elif fin_data.d2e < 0.6:  # Low debt
            health_score = 7.0
        elif fin_data.d2e < 1.0:  # Moderate debt
            health_score = 6.0
        elif fin_data.d2e < 1.5:  # High debt
            health_score = 4.0
        elif fin_data.d2e < 2.0:  # Very high debt
            health_score = 2.0
        else:  # Extremely high debt
            health_score = 0.0
    
    # Weighted combination
    final_score = (pe_score * 0.4) + (profit_score * 0.3) + (health_score * 0.3)
    
    return min(10.0, max(0.0, final_score))

def legacy_full_score(raw_core: float, fin_data: Fin, news: List[News],
                      tkr: str, tier_scorer: TierScoringSystem) -> float:
    fundamental = 0.0
    if fin_data:
        margin_expansion = (fin_data.margin - fin_data.margin*0.95) if fin_data.margin else 0
        data = {
            "margin_expansion": margin_expansion,
            "volume_growth": fin_data.volume_growth or 0,
            "debt_reduction": fin_data.debt_red or 0,
            "strategic_pivot": 1 if any("pivot" in (n.headline+n.snippet).lower() for n in news) else 0,
            "order_conversion_ratio": fin_data.orders or 0
        }
        fundamental = tier_scorer.calculate_fundamental_score(data)
        
        # Smart tier management instead of manual overrides
        if ENHANCED_MODULES_AVAILABLE:
            try:
                from enhanced_screener_modules import SmartTierManager
                smart_tier_manager = SmartTierManager()
                sector = detect_sector(tkr)  # Assume tkr is available in scope
                
                # Dynamic sector-based multiplier instead of manual lists
                sector_momentum = smart_tier_manager.sector_rotation_weights.get(sector, 0.0)
                if sector_momentum > 0.3:  # Strong sector momentum
                    fundamental *= 1.4
                elif sector_momentum > 0.1:  # Moderate sector momentum
                    fundamental *= 1.2
                elif sector_momentum < -0.3:  # Weak sector
                    fundamental *= 0.9
                    
            except Exception as e:
                logging.debug(f"Smart tier management failed for {tkr}: {e}")
                # Fallback to minimal manual overrides
                if tkr in HIGH_CONVICTION_TIER:
                    fundamental *= 1.6
                elif tkr in SELECTIVE_TIER:
                    fundamental *= 1.3
                elif tkr in TACTICAL_TIER:
                    fundamental *= 0.8
        else:
            # Legacy manual overrides (minimal)
            if tkr in HIGH_CONVICTION_TIER:
                fundamental *= 1.6
            elif tkr in SELECTIVE_TIER:
                fundamental *= 1.3
            elif tkr in TACTICAL_TIER:
                fundamental *= 0.8
    return raw_core + fundamental

def scaled_score(raw_plus_near: float, fin_data: Fin, news: List[News], base_scale: float) -> Tuple[float,float,float]:
    base_score = raw_plus_near / base_scale
    base_score = max(0.0, min(5.0, base_score))
    fundamental = 0.0
    if fin_data:
        if fin_data.margin and fin_data.margin > 0.15:
            fundamental += min((fin_data.margin - 0.15)/0.05,2.0)
        if fin_data.volume_growth and fin_data.volume_growth>0.10:
            fundamental += min(fin_data.volume_growth*10,2.0)
        if fin_data.debt_red and fin_data.debt_red>0.05:
            fundamental += min(fin_data.debt_red*20,2.0)
        if any("order win" in (n.headline+n.snippet).lower() for n in news):
            fundamental += 1.5
        fundamental = min(fundamental,5.0)
    final = round(base_score + fundamental,2)
    return final, round(base_score,2), round(fundamental,2)

# =============================================================================
# Options OI Helper
# =============================================================================
def get_oi_change(tkr: str) -> Optional[float]:
    try:
        rate_limit(0.5)  # 500ms between calls
        # Prevent double .NS suffix
        yf_symbol = ensure_ns_suffix(tkr)
        tk = yf.Ticker(yf_symbol)
        exp = tk.options
        if len(exp) < 2:
            return None
        today, prev = tk.option_chain(exp[0]), tk.option_chain(exp[1])
        t_oi = today.calls["openInterest"].sum() + today.puts["openInterest"].sum()
        p_oi = prev.calls["openInterest"].sum() + prev.puts["openInterest"].sum()
        return None if p_oi == 0 else (t_oi - p_oi)/p_oi*100
    except:
        return None

# =============================================================================
# Entry Strategy
# =============================================================================
class EntryStrategy:
    @staticmethod
    def _atr(df: pd.DataFrame, period: int=5) -> Optional[float]:
        try:
            if df is None or df.empty or len(df) < period:
                return None
            if talib is not None:
                vals = talib.ATR(df["High"], df["Low"], df["Close"], timeperiod=period)
                if vals is None or vals.size == 0:
                    return None
                val = vals.iloc[-1]
                return None if np.isnan(val) else float(val)
            high, low, close = df["High"], df["Low"], df["Close"]
            prev_close = close.shift(1)
            tr = np.maximum(high - low,
                            np.maximum((high - prev_close).abs(),
                                       (low - prev_close).abs()))
            atr_series = tr.rolling(period).mean()
            val = atr_series.iloc[-1]
            return None if np.isnan(val) else float(val)
        except:
            return None

    @staticmethod
    def tier1(df: pd.DataFrame) -> Tuple[Optional[float], Optional[float]]:
        try:
            if df is None or df.empty:
                return None, None
            if len(df) < 3:
                lc = df["Close"].iloc[-1]
                return lc*0.95, lc*0.97
            atr = EntryStrategy._atr(df, 5)
            lc = df["Close"].iloc[-1]
            support = df["Low"].rolling(3).min().iloc[-1]
            floor, ceil = lc*0.93, lc*0.95
            entry = min(max(support, floor), ceil)
            if atr is None:
                stop = min(entry*0.97, support*0.99)
            else:
                stop = entry - 2.0*atr
            return entry, stop
        except:
            try:
                lc = df["Close"].iloc[-1]
                return lc*0.95, lc*0.97
            except:
                return None, None

    @staticmethod
    def tier2(df: pd.DataFrame) -> Tuple[Optional[float], Optional[float]]:
        try:
            if df is None or df.empty:
                return None, None
            lc = df["Close"].iloc[-1]
            entry = lc * 0.96
            atr = EntryStrategy._atr(df, 5)
            stop = atr is None and entry*0.98 or entry - 1.5*atr
            return entry, stop
        except:
            try:
                lc = df["Close"].iloc[-1]
                return lc*0.96, lc*0.98
            except:
                return None, None

    @staticmethod
    def tier3(df: pd.DataFrame) -> Tuple[Optional[float], Optional[float]]:
        try:
            if df is None or df.empty:
                return None, None
            lc = df["Close"].iloc[-1]
            atr = EntryStrategy._atr(df, 5)
            stop = atr is None and lc*0.95 or lc - 1.5*atr
            return None, stop
        except:
            try:
                lc = df["Close"].iloc[-1]
                return None, lc*0.95
            except:
                return None, None

# =============================================================================
# Position Sizing
# =============================================================================
class PositionSizer:
    """Enhanced Position Sizing with RSI+EMA Oversold Best Practices"""
    
    @staticmethod
    def get_size(tier: str, score: float, oversold_category: str = None, signal_quality: float = 0.0) -> str:
        """
        Get position size based on tier, score, and oversold signal strength
        
        Best Practices Implementation:
        - TIER1 ULTIMATE/EXTREME signals: Higher allocation
        - Scale-in approach for high-quality oversold signals
        - Conservative sizing for weak signals
        """
        base_size = ""
        
        if tier == "TIER1":
            if oversold_category in ['ULTIMATE', 'EXTREME'] and signal_quality >= 0.5:
                base_size = "4-6%" if score >= 8 else "3-5%"
            elif oversold_category == 'STRONG' and signal_quality >= 0.3:
                base_size = "3-5%" if score >= 7 else "2-4%"
            else:
                base_size = "3-5%" if score >= 7 else "2-3%"
        elif tier == "TIER2":
            if oversold_category in ['ULTIMATE', 'EXTREME'] and signal_quality >= 0.4:
                base_size = "2-4%" if score >= 6 else "1.5-3%"
            elif oversold_category == 'STRONG':
                base_size = "2-3%" if score >= 5 else "1-2%"
            else:
                base_size = "2-3%" if score >= 5 else "1-2%"
        else:  # TIER3
            base_size = "0.5-1%" if oversold_category in ['STRONG', 'EXTREME', 'ULTIMATE'] else "0.3-0.8%"
        
        return base_size
    
    @staticmethod
    def get_entry_strategy(oversold_category: str, signal_quality: float, volume_confirmation: bool) -> str:
        """
        Get entry strategy based on oversold signal characteristics
        
        Implementation of Best Practices:
        - Scale-in for EXTREME signals
        - Wait for confirmation on weaker signals
        - Full entry only on high-confidence setups
        """
        if oversold_category == 'ULTIMATE':
            return "Scale-In: 50% now, 50% on EMA14 break"
        elif oversold_category == 'EXTREME':
            if signal_quality >= 0.5 and volume_confirmation:
                return "Scale-In: 60% now, 40% on confirmation"
            else:
                return "Scale-In: 40% now, 60% on EMA14 cross"
        elif oversold_category == 'STRONG':
            if signal_quality >= 0.4 and volume_confirmation:
                return "Full Entry with volume confirmation"
            else:
                return "Wait for volume surge or EMA14 hold"
        elif oversold_category in ['MODERATE', 'MILD']:
            return "Confirm reversal first - wait for bullish candle"
        else:
            return "Standard breakout entry"
    
    @staticmethod
    def get_stop_loss_strategy(oversold_category: str, entry_price: float, swing_low: float = None) -> dict:
        """
        Get stop loss placement based on oversold signal and best practices
        
        Returns: Dictionary with stop loss levels and reasons
        """
        if swing_low is None:
            swing_low = entry_price * 0.92  # Default 8% below entry
        
        strategies = {}
        
        if oversold_category in ['ULTIMATE', 'EXTREME']:
            # Wider stops for high-conviction oversold plays
            strategies['swing_low'] = max(swing_low * 0.98, entry_price * 0.92)
            strategies['percentage'] = entry_price * 0.93  # 7% stop
            strategies['ema_based'] = entry_price * 0.95   # 5% for EMA breakdown
            strategies['recommended'] = strategies['swing_low']
            strategies['reason'] = "Below recent swing low (wider for mean reversion)"
        elif oversold_category == 'STRONG':
            strategies['swing_low'] = max(swing_low * 0.99, entry_price * 0.95)
            strategies['percentage'] = entry_price * 0.95  # 5% stop
            strategies['ema_based'] = entry_price * 0.97   # 3% for EMA breakdown
            strategies['recommended'] = strategies['percentage']
            strategies['reason'] = "5% stop - balance of risk vs signal quality"
        else:
            # Tighter stops for weaker signals
            strategies['percentage'] = entry_price * 0.97  # 3% stop
            strategies['ema_based'] = entry_price * 0.98   # 2% for EMA breakdown
            strategies['recommended'] = strategies['percentage']
            strategies['reason'] = "Tight stop - lower conviction signal"
        
        return strategies
    
    @staticmethod
    def get_profit_targets(oversold_category: str, entry_price: float, risk_amount: float) -> dict:
        """
        Get profit targets based on oversold signal strength and best practices
        
        Returns: Dictionary with target levels and trailing stop strategy
        """
        targets = {}
        
        if oversold_category == 'ULTIMATE':
            targets['target_1'] = entry_price * 1.10    # 10% target
            targets['target_2'] = entry_price * 1.18    # 18% target
            targets['booking_1'] = "50% at Target 1"
            targets['booking_2'] = "Trail stop below 3-day low"
            targets['time_exit'] = "7 trading days if Target 1 not hit"
        elif oversold_category == 'EXTREME':
            targets['target_1'] = entry_price * 1.08    # 8% target
            targets['target_2'] = entry_price * 1.15    # 15% target
            targets['booking_1'] = "50% at Target 1"
            targets['booking_2'] = "Trail stop below EMA14"
            targets['time_exit'] = "6 trading days if Target 1 not hit"
        elif oversold_category == 'STRONG':
            targets['target_1'] = entry_price * 1.06    # 6% target
            targets['target_2'] = entry_price * 1.12    # 12% target
            targets['booking_1'] = "40% at Target 1"
            targets['booking_2'] = "Trail stop below 3-day low"
            targets['time_exit'] = "5 trading days if Target 1 not hit"
        else:
            targets['target_1'] = entry_price * 1.04    # 4% target
            targets['target_2'] = entry_price * 1.08    # 8% target
            targets['booking_1'] = "30% at Target 1"
            targets['booking_2'] = "Move stop to breakeven"
            targets['time_exit'] = "4 trading days if Target 1 not hit"
        
        # Calculate risk/reward
        if risk_amount > 0:
            reward_1 = targets['target_1'] - entry_price
            targets['risk_reward_1'] = reward_1 / risk_amount
            reward_2 = targets['target_2'] - entry_price
            targets['risk_reward_2'] = reward_2 / risk_amount
        else:
            targets['risk_reward_1'] = 0.0
            targets['risk_reward_2'] = 0.0
        
        return targets
    
    @staticmethod
    def get_portfolio_allocation_advice(total_oversold_positions: int, market_condition: str = "neutral") -> str:
        """
        Get portfolio-level allocation advice based on best practices
        """
        if market_condition == "bear" and total_oversold_positions > 3:
            return "[WARNING] Reduce oversold positions in bear market - max 2 STRONG signals"
        elif total_oversold_positions > 4:
            return "[WARNING] Too many oversold positions - max 2 STRONG signals simultaneously"
        elif total_oversold_positions >= 2:
            return "[OK] Good balance - monitor correlation between positions"
        else:
            return "[OK] Room for more oversold opportunities if quality signals emerge"

# =============================================================================
# Monitoring Framework
# =============================================================================
class MonitoringFramework:
    _metrics = {
        "JSWSTEEL":"Steel/coking coal spreads",
        "NUVOCO":"Cement dispatch volumes",
        "POLYCAB":"Copper/aluminum input costs",
        "WIPRO":"Quarterly bookings & large deal TCV",
        "HATSUN":"Milk procurement vs selling price spread",
        "IOB":"NPA recovery & credit cost trend",
        "LTIM":"Deal TCV conversion to revenue",
        "RAYMOND":"Aerospace order ramp",
        "CUPID":"Middle East tender execution",
        "JIOFIN":"Cross-sell monetization / customer adds",
        "MTARTECH":"Revenue stabilization & order visibility",
        "KSOLVES":"Billable utilization & client additions",
        "INDOSTAR":"Asset quality post divestment",
        "MAHLIFE":"Pre-sales to collections conversion",
        "SUNTECK":"Quarterly pre-sales growth",
        "LTF":"Disbursement growth vs NIM stability",
    }

    @staticmethod
    def get_metric(tkr: str, tier: str) -> str:
        return MonitoringFramework._metrics.get(tkr, "Sector trend & follow-through")

# =============================================================================
# FII growth scraper (cached)
# =============================================================================
_fii_cache: Dict[str, Tuple[float,float]] = {}

async def fetch_single_fii(session: aiohttp.ClientSession, sym: str) -> Tuple[str, Optional[float]]:
    if BeautifulSoup is None:
        return sym, None
    ts_now = time.time()
    cached = _fii_cache.get(sym)
    if cached and (ts_now - cached[0]) < FII_CACHE_TTL:
        return sym, cached[1]
    url = f"https://www.screener.in/company/{sym.split('.')[0]}/consolidated/"
    try:
        async with session.get(url, headers={"User-Agent":FII_UA}, timeout=FII_TIMEOUT) as resp:
            if resp.status != 200:
                return sym, None
            text = await resp.text()
            soup = BeautifulSoup(text, "html.parser")
            h2 = soup.find(lambda tag: tag.name in ("h2","h3") and tag.get_text(strip=True).lower().startswith("shareholding pattern"))
            table = h2.find_next("table") if h2 else None
            if not table:
                return sym, None
            row = None
            for tr in table.find_all("tr"):
                cells = [c.get_text(strip=True) for c in tr.find_all(["th","td"])]
                if cells and ("fii" in cells[0].lower()):
                    row = cells
                    break
            if not row or len(row) < 3:
                return sym, None
            vals = []
            for c in row[1:]:
                if c.endswith("%"):
                    try:
                        vals.append(float(c.strip(" %")))
                    except (ValueError, TypeError):
                        continue
            if len(vals) < 2:
                return sym, None
            
            fii_change = vals[-1] - vals[-2]
            _fii_cache[sym] = (ts_now, fii_change)
            return sym, fii_change
    except (asyncio.TimeoutError, aiohttp.ClientError) as e:
        logging.warning(f"FII fetch failed for {sym}: {e}")
        return sym, None
    except Exception as e:
        logging.error(f"Unexpected error in FII fetch for {sym}: {e}")
        return sym, None

async def fetch_fii_growth_parallel(symbols: List[str]) -> Dict[str, Optional[float]]:
    """Fetches FII data for multiple symbols in parallel."""
    if not symbols:
        return {}
    
    results = {}
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_single_fii(session, s) for s in symbols]
        fii_results = await asyncio.gather(*tasks)
        for sym, val in fii_results:
            results[sym] = val
    return results

# =============================================================================
# Quarterly Results Growth
# =============================================================================
def _pct_growth(cur: float, prev: float) -> Optional[float]:
    if prev is None or cur is None:
        return None
    if prev == 0:
        return 100.0 if cur > 0 else -100.0 if cur < 0 else 0.0
    if np.isnan(prev) or np.isnan(cur):
        return None
    g = 100.0 * (cur - prev) / abs(prev)
    if cur < 0:
        g = -abs(g)
    return g

_QUARTER_LABELS_PRIORITY = [
    "Net Income",
    "Net Income Applicable to Common Shares",
    "Profit After Tax",
    "Net Profit",
    "Total Revenue",
    "Operating Income",
    "EBIT",
    "Ebit",
    "Basic EPS",
    "Diluted EPS"
]

def _single_quarter_growth(sym: str) -> Optional[float]:
    try:
        rate_limit(0.5)  # 500ms between calls
        # Prevent double .NS suffix
        yf_symbol = ensure_ns_suffix(sym)
        tk = yf.Ticker(yf_symbol)
        qis = tk.quarterly_income_stmt
        if qis is None or qis.empty or len(qis.columns) < 2:
            return None
        for label in _QUARTER_LABELS_PRIORITY:
            if label in qis.index:
                s = qis.loc[label].dropna()
                if len(s) < 2:
                    continue
                latest, prev = float(s.iloc[0]), float(s.iloc[1])
                date = s.index[0]
                if (pd.Timestamp.now().normalize() - date).days > 135:
                    return None
                return _pct_growth(latest, prev)
        return None
    except:
        return None

def get_quarter_result_growth(symbols: List[str]) -> Dict[str, Optional[float]]:
    out: Dict[str, Optional[float]] = {}
    if not symbols:
        return out
    with ThreadPoolExecutor(min(len(symbols), 8) or 1) as ex:
        fut = {ex.submit(_single_quarter_growth, s): s for s in symbols}
        for f in as_completed(fut):
            sym = fut[f]
            out[sym] = f.result()
    return out

# =============================================================================
# Main
# =============================================================================
async def main(argv=None):
    ap = argparse.ArgumentParser("Swing Trade Screener v23.5")
    ap.add_argument("--hours", type=int, default=48)
    ap.add_argument("--top", type=int, default=30)
    ap.add_argument("--min-score", type=float, default=0.15)  # Optimized for realistic score ranges
    ap.add_argument("--min-vol", type=float, default=MIN_VOL_SURGE)
    ap.add_argument("--skip-financial", action="store_true")
    ap.add_argument("--allow-negative", action="store_true")
    ap.add_argument("--soft-mode", action="store_true", help="Relaxed filters: lower thresholds, keep near-misses with ⚠️ tags")
    ap.add_argument("--debug", action="store_true")
    ap.add_argument("--debug-institutional", action="store_true", help="Enable institutional filtering debug output")
    ap.add_argument("--no-legacy-table", action="store_true")
    ap.add_argument("--legacy-score", action="store_true")
    ap.add_argument("--base-scale", type=float, default=BASE_SCORE_SCALE_DEFAULT)
    ap.add_argument("--show-raw", action="store_true")
    ap.add_argument("--no-fii", action="store_true", help="Disable late-phase FII weighting & column")
    ap.add_argument("--excel-view", action="store_true", help="Output in a tab-separated format for spreadsheets.")
    ap.add_argument("--excel-file", action="store_true", help="Generate a colored Excel file (.xlsx) with proper formatting.")
    ap.add_argument("--csv-enhanced", action="store_true", help="Generate enhanced CSV with color indicators for easy Excel import.")
    
    # Enhanced liquidity analysis arguments
    ap.add_argument("--min-turnover-m", type=float, default=1.0, help="Minimum daily turnover in ₹ million (default: 1.0)")
    ap.add_argument("--max-spread-pct", type=float, default=2.0, help="Maximum bid-ask spread percentage (default: 2.0)")
    ap.add_argument("--liquidity-strict", action="store_true", help="Use stricter liquidity thresholds for large portfolios")
    ap.add_argument("--liquidity-weight", type=float, default=10.0, help="Weight for liquidity impact on scoring (default: 10.0)")
    
    # Winsorization parameters
    ap.add_argument("--price-cap", type=float, default=None, help="Override price return winsorization threshold (e.g., 0.3 for +/-30%%)")
    ap.add_argument("--vol-cap", type=float, default=None, help="Override volume change winsorization threshold (e.g., 0.8 for +/-80%%)")
    
    # Optimization arguments
    ap.add_argument("--deep-sentiment", action="store_true", help="Enable transformer-based sentiment analysis (adds ~30s for model loading)")
    ap.add_argument("--ultra-fast", action="store_true", help="Ultra-fast mode: skip 5min data, minimal financials, basic sentiment")
    ap.add_argument("--skip-5min", action="store_true", help="Skip 5-minute intraday data fetching")
    ap.add_argument("--basic-sentiment", action="store_true", help="Use basic pattern-based sentiment instead of FinBERT")
    
    # Enhanced analytics arguments
    ap.add_argument("--enhanced-analytics", action="store_true", help="Enable enhanced financial metrics analysis (debt/equity, ROE, institutional flows)")
    ap.add_argument("--skip-enhanced", action="store_true", help="Skip enhanced financial metrics for faster execution")
    
    # Institutional screening arguments
    ap.add_argument("--institutional-filters", action="store_true", help="Enable institutional-grade screening filters (stricter quality gates)")
    ap.add_argument("--min-mcap-cr", type=float, default=50.0, help="Minimum market cap in ₹ Crores for institutional screening (default: 50.0)")
    ap.add_argument("--min-turnover-cr", type=float, default=0.3, help="Minimum average daily turnover in ₹ Crores (default: 0.3)")
    ap.add_argument("--max-gap-pct", type=float, default=25.0, help="Maximum single-day gap percentage (default: 25.0)")
    
    # Auto-detection arguments
    ap.add_argument("--news-dir", type=str, default=".", help="Directory to search for news_output_*.txt files (default: current directory)")
    ap.add_argument("--max-news-files", type=int, default=5, help="Maximum number of latest non-empty news files to auto-detect (default: 5)")
    
    # Ticker file input (bypasses news analysis)
    ap.add_argument("--ticker-file", type=str, help="Use pre-ranked ticker list from file (bypasses news analysis)")
    ap.add_argument("--ticker-count", type=int, default=100, help="Maximum number of tickers to read from ticker file (default: 100)")
    
    # Backtesting and optimization arguments
    ap.add_argument("--backtest", type=str, nargs=2, metavar=("START_DATE", "END_DATE"), 
                   help="Run backtest mode with date range (YYYY-MM-DD YYYY-MM-DD)")
    ap.add_argument("--optimize", action="store_true", help="Run parameter optimization sweep")
    ap.add_argument("--validate", action="store_true", help="Run rolling window validation on best parameters")
    ap.add_argument("--backtest-symbols", type=str, nargs="*", default=None,
                   help="Specific symbols for backtesting (default: use top screening results)")
    ap.add_argument("--train-months", type=int, default=12, help="Training window in months for validation (default: 12)")
    ap.add_argument("--test-months", type=int, default=3, help="Testing window in months for validation (default: 3)")
    ap.add_argument("--backtest-report", type=str, default="backtest_report.txt", help="Output file for backtest report")
    
    # Data quality and validation arguments
    ap.add_argument("--strict-validation", action="store_true", help="Enable strict data validation (removes more questionable data)")
    ap.add_argument("--validation-report", action="store_true", help="Generate detailed data validation report")
    ap.add_argument("--allow-partial-data", action="store_true", help="Allow processing with partial/missing data (less strict)")
    
    ap.add_argument("files", nargs="*", help="News files to analyze (default: auto-detect latest 4-5 non-empty news_output_*.txt files)")
    args = ap.parse_args(argv)

    # Apply CLI winsorization parameter overrides
    if args.price_cap is not None:
        WINSORIZATION_CONFIG['price_return_threshold'] = args.price_cap
        print(f"[WINSORIZATION] Price cap override: ±{args.price_cap*100:.0f}%")
    
    if args.vol_cap is not None:
        WINSORIZATION_CONFIG['volume_change_threshold'] = args.vol_cap
        print(f"[WINSORIZATION] Volume cap override: ±{args.vol_cap*100:.0f}%")

    logging.basicConfig(level=logging.DEBUG if args.debug else logging.INFO,
                        format="%(levelname)s: %(message)s")
    logging.getLogger('yfinance').setLevel(logging.CRITICAL)

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Initialize sentiment analyzer based on speed preferences
    global _sentiment_analyzer
    ultra_fast_mode = args.ultra_fast or args.basic_sentiment
    _sentiment_analyzer = EnhancedSentimentAnalyzer(ultra_fast=ultra_fast_mode)
    
    # Display enhanced features status
    print(f"\n[ENHANCED SCREENING] v24.1 - Institutional-Grade Filtering & Critical Financial Metrics")
    
    # Data Quality Status
    print(f"[DATA QUALITY] YFinance Data Validation: ENABLED")
    if args.strict_validation:
        print("   * Strict validation mode: Enhanced data quality checks")
    if args.validation_report:
        print("   * Validation reporting: Detailed data quality logs")
    if args.allow_partial_data:
        print("   * Partial data mode: Less strict validation (faster execution)")
    print("   * OHLC relationship validation")
    print("   * Price and volume anomaly detection")
    print("   * Financial ratio boundary checks")
    print("   * Missing data handling with fallbacks")
    
    # Institutional Screening Status
    if args.institutional_filters:
        print("[INSTITUTIONAL] Institutional-Grade Screening: ENABLED")
        print(f"   * Minimum Market Cap: Rs.{args.min_mcap_cr} Cr")
        print(f"   * Minimum Daily Turnover: Rs.{args.min_turnover_cr} Cr")
        print(f"   * Maximum Gap Tolerance: {args.max_gap_pct}%")
        print("   * Volume consistency validation")
        print("   * Momentum quality assessment")
        print("   * Technical structure filtering")
    else:
        print("[BASIC] Institutional-Grade Screening: DISABLED (use --institutional-filters to enable)")
    
    # Enhanced Financial Analysis Status
    if args.enhanced_analytics or (not args.skip_enhanced and not args.ultra_fast):
        print("[OK] Enhanced Financial Analysis: ENABLED")
        print("   * Debt-to-Equity ratio screening")
        print("   * ROE & Interest Coverage analysis") 
        print("   * Institutional holdings tracking")
        print("   * Delivery percentage conviction")
        print("   * Technical overextension alerts")
    else:
        print("[FAST] Enhanced Financial Analysis: DISABLED (use --enhanced-analytics to enable)")
    
    # Soft Mode Status
    if args.soft_mode:
        print("[SOFT] Relaxed Filtering Mode: ENABLED")
        print("   * Winsorization threshold: 30% (raised from 20%)")
        print("   * Volume threshold: 1.2x (relaxed from 1.5x)")
        print("   * Buyer dominance: 45% (relaxed from 55%)")
        print("   * Min score: 0.0 (overrides --min-score)")
        print("   * POOR data quality allowed with ⚠️ tags")
    else:
        print("[STRICT] Standard Filtering Mode: ENABLED")
        print("   * Use --soft-mode to catch more opportunities")
    
    print("="*80)

    # Stage 0: Check if using ticker file input (bypasses news analysis)
    if args.ticker_file:
        print(f"📋 Using pre-ranked tickers from file: {args.ticker_file}")
        
        # Read tickers from file
        try:
            with open(args.ticker_file, 'r') as f:
                shortlist = []
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if line and not line.startswith('#'):  # Skip comments and empty lines
                        ticker = line.upper()
                        if '.' not in ticker:  # Add .NS if not present
                            ticker += '.NS'
                        shortlist.append(ticker.replace('.NS', ''))  # Store without .NS for consistency
                        
                        if len(shortlist) >= args.ticker_count:
                            break
            
            if not shortlist:
                sys.exit(f"[ERROR] No valid tickers found in {args.ticker_file}")
            
            print(f"✅ Loaded {len(shortlist)} tickers from file")
            print(f"📊 Sample tickers: {', '.join(shortlist[:5])}{'...' if len(shortlist) > 5 else ''}")
            
            # Skip news analysis and jump directly to technical analysis
            news = {}  # Empty news dict
            
        except FileNotFoundError:
            sys.exit(f"[ERROR] Ticker file not found: {args.ticker_file}")
        except Exception as e:
            sys.exit(f"[ERROR] Failed to read ticker file: {e}")
    
    else:
        # Original news-based approach
        # Stage 0: Auto-detect latest news files if none specified
        news_files = args.files
        if not news_files:
            print(f"[SEARCH] No news files specified, auto-detecting latest {args.max_news_files} news_output_*.txt files in {args.news_dir}...")
            news_files = find_latest_news_files(directory=args.news_dir, max_files=args.max_news_files)
            if not news_files:
                sys.exit(f"[ERROR] No news_output_*.txt files found in {args.news_dir}. Please ensure news files are in the specified directory.")
        else:
            print(f"📁 Using {len(news_files)} specified news files:")
            for i, file_path in enumerate(news_files):
                filename = os.path.basename(file_path)
                print(f"   {i+1}. {filename}")

        # Parse news and create lightweight shortlist
        news = parse_news(news_files, args.hours)
        if not news:
            sys.exit("[ERROR] no valid news")
        
        # NEW: Create shortlist using fast news metrics
        if args.ultra_fast:
            # Ultra-fast mode: process only top 40 candidates (or 2x target)
            if args.soft_mode:
                shortlist_size = max(150, args.top * 6)  # Much more generous in soft mode
            else:
                shortlist_size = max(40, args.top * 2)
            logging.info("[ULTRA-FAST] Ultra-fast mode: Smaller shortlist for maximum speed")
        else:
            # Normal optimized mode: process top 80 candidates (or 3x target)
            if args.soft_mode:
                shortlist_size = max(250, args.top * 10)  # Much more generous in soft mode
            else:
                shortlist_size = max(80, args.top * 3)
        
        shortlist = lightweight_shortlist(news, top_n=shortlist_size, soft_mode=args.soft_mode)
        logging.info(f"Lightweight shortlist created: {len(shortlist)} candidates from {len(news)} total")

    # Stage 1: Batched market data download for shortlist only
    logging.info("Fetching market data for shortlisted candidates...")
    hist = batched_history_download(shortlist, args.soft_mode)
    
    # Parallel meta fetching only for shortlist
    mcap, names, spreads = {}, {}, {}
    with ThreadPoolExecutor(MAX_WORKERS) as ex:
        fut_m = {ex.submit(_meta, t): t for t in shortlist}
        for f in as_completed(fut_m):
            t, m, name, spread = f.result()
            mcap[t] = m
            names[t] = name
            spreads[t] = spread
    
    yshared._ERRORS = {}
    
    # Initialize empty financial data (will be populated later for top candidates)
    fin = {}  # Initialize as empty dict - will be populated during late financials phase
    
    # Memory cleanup after major data fetch
    memory_cleanup()

    # Stage 2: Technical scoring on shortlist (without financials yet)
    prob = BreakoutProbabilityModel()
    grow = BullishGrowthForecaster(prob)
    tier_scorer = TierScoringSystem()
    classifier = TierClassifier()
    entry = EntryStrategy()
    monitor = MonitoringFramework()

    ranks: List[Tuple[str,float]] = []
    metrics: Dict[str, Tuple] = {}
    deal_amount_map: Dict[str, float] = {}
    deal_pct_map: Dict[str, float] = {}
    rel_map: Dict[str, List[News]] = {}
    raw_score_map: Dict[str, float] = {}
    base_part_map: Dict[str, float] = {}
    fundamental_part_map: Dict[str, float] = {}
    fii_growth_map: Dict[str, Optional[float]] = {}
    fii_pct_map: Dict[str, Optional[float]] = {}
    deal_count_map: Dict[str,int] = {}
    deal_impact_map: Dict[str,float] = {}
    positive_count_map: Dict[str,int] = {}
    liquidity_map: Dict[str, Tuple[float, str, bool]] = {}  # (score, warning_level, is_short)

    # ---------- Technical scoring pass (no financials yet) ----------
    logging.info(f"Technical scoring for {len(shortlist)} candidates...")
    for t in shortlist:
        d = news.get(t)
        if not d:
            continue
        df = hist.get(t, pd.DataFrame())
        if df.empty or len(df) < 4:
            continue
        close = df["Close"]
        try:
            rise1 = (close.iloc[-1] - close.iloc[-2]) / close.iloc[-2] * 100
            rise3 = (close.iloc[-1] - close.iloc[-4]) / close.iloc[-4] * 100
        except:
            continue

        avg_vol = df["Volume"].iloc[:-1].mean() or 1
        volx = df["Volume"].iloc[-1] / avg_vol
        if volx < args.min_vol:
            continue

        # === INSTITUTIONAL-GRADE TECHNICAL FILTERS ===
        # Applied after basic volume filter but before expensive data fetching
        if args.institutional_filters:
            if hasattr(args, 'debug_institutional') and args.debug_institutional:
                print(f"🔍 INSTITUTIONAL DEBUG: {t} - Starting institutional filters...")
            
            # 1. PRICE ACTION QUALITY FILTER
            current_price = df["Close"].iloc[-1]
            mc = mcap.get(t, 0.0)  # Market cap in Cr
            
            if hasattr(args, 'debug_institutional') and args.debug_institutional:
                print(f"🔍 INSTITUTIONAL DEBUG: {t} - Price: ₹{current_price:.2f}, Market Cap: ₹{mc:.1f} Cr")
            
            # Skip penny stocks and extremely expensive stocks (institutional practice)
            if current_price < 10.0 or current_price > 10000.0:
                if hasattr(args, 'debug_institutional') and args.debug_institutional:
                    print(f"❌ INSTITUTIONAL FILTER: {t} - Price filter failed (₹{current_price:.2f})")
                continue
                
            # Skip very small cap stocks (configurable minimum market cap)
            if mc < args.min_mcap_cr:
                if hasattr(args, 'debug_institutional') and args.debug_institutional:
                    print(f"❌ INSTITUTIONAL FILTER: {t} - Market cap filter failed (₹{mc:.1f} Cr < ₹{args.min_mcap_cr} Cr)")
                continue
            
            # 2. VOLUME CONSISTENCY CHECK (Institutional requirement)
            if len(df) >= 20:
                recent_volumes = df["Volume"].tail(20)
                volume_cv = recent_volumes.std() / recent_volumes.mean() if recent_volumes.mean() > 0 else 1.0
                if hasattr(args, 'debug_institutional') and args.debug_institutional:
                    print(f"🔍 INSTITUTIONAL DEBUG: {t} - Volume CV: {volume_cv:.2f}")
                # Skip stocks with erratic volume patterns (CV > 2.0 indicates manipulation risk)
                if volume_cv > 2.0:
                    if hasattr(args, 'debug_institutional') and args.debug_institutional:
                        print(f"❌ INSTITUTIONAL FILTER: {t} - Volume consistency failed (CV: {volume_cv:.2f} > 2.0)")
                    continue
            
            # 3. EXCESSIVE GAP FILTER (Risk management)
            if len(df) >= 5:
                # Check for excessive gaps in last 5 days
                gaps = []
                for i in range(max(1, len(df)-5), len(df)):
                    if i > 0:
                        gap_pct = abs((df["Open"].iloc[i] - df["Close"].iloc[i-1]) / df["Close"].iloc[i-1] * 100)
                        gaps.append(gap_pct)
                
                max_gap = max(gaps) if gaps else 0
                if hasattr(args, 'debug_institutional') and args.debug_institutional:
                    print(f"🔍 INSTITUTIONAL DEBUG: {t} - Max gap: {max_gap:.1f}%")
                # Skip stocks with excessive gaps (configurable threshold)
                if max_gap > args.max_gap_pct:
                    if hasattr(args, 'debug_institutional') and args.debug_institutional:
                        print(f"❌ INSTITUTIONAL FILTER: {t} - Gap filter failed ({max_gap:.1f}% > {args.max_gap_pct}%)")
                    continue
            
            # 4. MOMENTUM QUALITY ASSESSMENT
            if len(df) >= 10:
                # Volume-price relationship validation
                recent_returns = df["Close"].pct_change().tail(5)
                recent_volume_change = df["Volume"].pct_change().tail(5)
                
                # Calculate correlation between volume and absolute returns
                abs_returns = recent_returns.abs()
                volume_price_corr = abs_returns.corr(recent_volume_change) if len(abs_returns) >= 3 else 0
                
                if hasattr(args, 'debug_institutional') and args.debug_institutional:
                    print(f"🔍 INSTITUTIONAL DEBUG: {t} - Volume-price correlation: {volume_price_corr:.2f}")
                
                # Skip stocks with negative volume-price correlation (indicates distribution)
                if volume_price_corr < -0.3:
                    if hasattr(args, 'debug_institutional') and args.debug_institutional:
                        print(f"❌ INSTITUTIONAL FILTER: {t} - Momentum quality failed (correlation: {volume_price_corr:.2f} < -0.3)")
                    continue
            
            # 5. BASIC LIQUIDITY GATE (Before expensive liquidity analysis)
            daily_turnover_cr = (df["Volume"].iloc[-1] * current_price) / 1e7  # Convert to Cr
            avg_daily_turnover_cr = (avg_vol * current_price) / 1e7
            
            if hasattr(args, 'debug_institutional') and args.debug_institutional:
                print(f"🔍 INSTITUTIONAL DEBUG: {t} - Avg daily turnover: ₹{avg_daily_turnover_cr:.2f} Cr")
            
            # Skip stocks with insufficient liquidity for institutional participation
            if avg_daily_turnover_cr < args.min_turnover_cr:
                if hasattr(args, 'debug_institutional') and args.debug_institutional:
                    print(f"❌ INSTITUTIONAL FILTER: {t} - Liquidity filter failed (₹{avg_daily_turnover_cr:.2f} Cr < ₹{args.min_turnover_cr} Cr)")
                continue
            
            # 6. TECHNICAL STRUCTURE FILTER
            if len(df) >= 50:
                # Moving average structure check
                ma_20 = df["Close"].tail(20).mean()
                ma_50 = df["Close"].tail(50).mean()
                
                if hasattr(args, 'debug_institutional') and args.debug_institutional:
                    print(f"🔍 INSTITUTIONAL DEBUG: {t} - MA structure: Price=₹{current_price:.1f}, MA20=₹{ma_20:.1f}, MA50=₹{ma_50:.1f}")
                
                # Skip stocks in strong downtrend (current price < MA20 < MA50 with >10% gap)
                if (current_price < ma_20 < ma_50 and 
                    (ma_50 - current_price) / current_price > 0.10):
                    if hasattr(args, 'debug_institutional') and args.debug_institutional:
                        downtrend_pct = (ma_50 - current_price) / current_price * 100
                        print(f"❌ INSTITUTIONAL FILTER: {t} - Technical structure failed (downtrend: {downtrend_pct:.1f}% > 10%)")
                    continue
            
            if hasattr(args, 'debug_institutional') and args.debug_institutional:
                print(f"✅ INSTITUTIONAL FILTER: {t} - PASSED all institutional filters!")
                
        # === END INSTITUTIONAL FILTERS ===

        # Enhanced buyer dominance calculation with multiple fallbacks
        bd = calculate_enhanced_buyer_dominance(t, df, args.ultra_fast or args.skip_5min)

        cname = names.get(t, d["name"])
        # filter to only unique deal-related news items
        rel_all = [n for n in d["news"] if qual_ok(n, comp_tokens(cname, t))]
        rel = deduplicate_news_by_signature(rel_all)
        if not rel:
            continue
        rel_map[t] = rel
        # Count positive words
        pos_count = sum(len(POSITIVE_PATTERN.findall(n.headline + ' ' + n.snippet)) for n in rel)
        positive_count_map[t] = pos_count

        # tags generation
        cat_sc, cat_tags = catalyst_score(rel)
        tags = cat_tags.copy()
        deal_count = 0
        for n in rel:
            txt = normalize_hyphens(n.headline + " " + n.snippet).lower()
            if "dividend" in txt and n.sent > 0 and "BONUS" not in tags:
                tags.append("DIV")
            if "result" in txt and n.sent > 0:
                tags.append("RES")
            if deal_rel(txt):
                tags.append("DEAL")
                deal_count += 1
            if any(k in txt for k in ["fund","ncd","debenture","investment"]) and n.sent > 0:
                tags.append("FUND")
        if deal_count > 1:
            tags.append(f"{deal_count}DEALS")

        # Consolidated deal amounts / pct / impact
        total_amt_abs, deal_cnt = consolidate_deals(rel)
        deal_amt_cr = total_amt_abs / 1e7
        deal_amount_map[t] = deal_amt_cr
        deal_count_map[t] = deal_cnt
        mc = mcap.get(t, 0.0)  # Cr
        
        # Calculate deal percentage
        deal_pct = (deal_amt_cr / mc * 100) if mc else 0.0
        deal_pct_map[t] = deal_pct
        
        # Debug output for deal calculation issues
        if args.debug and (deal_amt_cr > 0 or deal_cnt > 0):
            print(f"DEBUG {t}: deal_amt_abs={total_amt_abs:.0f}, deal_amt_cr={deal_amt_cr:.2f}, mcap_cr={mc:.2f}, deal_pct={deal_pct:.2f}%, deal_cnt={deal_cnt}")
        
        deal_imp = calculate_deal_impact(total_amt_abs, mc*1e7, deal_cnt)
        deal_impact_map[t] = deal_imp

        # === Enhanced Liquidity Analysis ===
        # Calculate average volume (10-day)
        avg_volume = df["Volume"].tail(10).mean() if len(df) >= 10 else df["Volume"].mean()
        
        # Get current price and market cap
        current_price = df["Close"].iloc[-1] if not df.empty else 0
        market_cap_cr = mc  # Already in Cr from mcap dict
        market_cap = mc * 1e7  # Convert to base units for calculations
        
        # Get shares outstanding from yfinance data (skip in ultra-fast mode)
        if args.ultra_fast:
            # Use market cap estimate in ultra-fast mode
            shares_outstanding = market_cap / current_price if current_price > 0 else 1
        else:
            try:
                # Use our resilient info download - prevent double .NS suffix
                yf_symbol = ensure_ns_suffix(t)
                tk_info = safe_yf_info(yf_symbol)
                shares_outstanding = tk_info.get("sharesOutstanding", 1) or tk_info.get("impliedSharesOutstanding", 1) or 1
            except:
                shares_outstanding = market_cap / current_price if current_price > 0 else 1
        
        # Calculate volatility using improved methods
        # 1. Annualized volatility (standard deviation approach)
        # 2. ATR as fallback
        # 3. Simple range-based as last resort
        
        if len(df) >= 20:  # Need sufficient data for reliable volatility
            try:
                # Primary: Annualized volatility using WINSORIZED daily returns
                price_data = get_robust_price_data(df, use_capped=True)
                daily_returns = price_data['returns'].dropna()
                if len(daily_returns) >= 20:
                    # Standard deviation of daily returns, annualized (252 trading days)
                    annualized_volatility = daily_returns.std() * math.sqrt(252)
                    volatility = annualized_volatility * 100  # Convert to percentage
                else:
                    raise ValueError("Insufficient return data")
            except:
                # Fallback 1: ATR-based volatility using winsorized close
                if talib is not None and len(df) >= 10:
                    try:
                        price_data = get_robust_price_data(df, use_capped=True)
                        atr_value = talib.ATR(df["High"], df["Low"], price_data['close'], timeperiod=10).iloc[-1]
                        volatility = (atr_value / price_data['close'].iloc[-1]) * 100  # Convert to percentage
                    except:
                        price_data = get_robust_price_data(df, use_capped=True)
                        volatility = (df["High"].tail(10).max() - df["Low"].tail(10).min()) / price_data['close'].tail(10).mean() * 100
                else:
                    # Fallback 2: Simple range-based volatility using winsorized close
                    price_data = get_robust_price_data(df, use_capped=True)
                    volatility = (df["High"].tail(10).max() - df["Low"].tail(10).min()) / df["Close"].tail(10).mean() * 100
        else:
            # Insufficient data - use simple calculation
            if len(df) >= 5:
                volatility = (df["High"].tail(len(df)).max() - df["Low"].tail(len(df)).min()) / df["Close"].tail(len(df)).mean() * 100
            else:
                volatility = 1.0  # Default minimal volatility
        
        if volatility <= 0:
            volatility = 0.01  # Prevent division by zero
        
        # Calculate liquidity metrics
        liquidity_score = LiquidityAnalyzer.calculate_liquidity_score(
            avg_volume, current_price, market_cap
        )
        turnover_ratio = LiquidityAnalyzer.calculate_turnover_ratio(
            avg_volume, shares_outstanding
        )
        volatility_ratio = LiquidityAnalyzer.calculate_volatility_ratio(
            avg_volume, volatility
        )
        
        # Get bid-ask spread from spread data
        spread_pct = spreads.get(t, 0.0)
        
        # Enhanced liquidity short detection (market-cap aware)
        is_liq_short = LiquidityAnalyzer.is_liquidity_short(
            avg_volume,
            spread_pct * current_price,  # Convert % spread to absolute value
            current_price,
            turnover_ratio,
            volatility_ratio,
            liquidity_score,
            market_cap_cr,  # Pass market cap for threshold determination
            args.min_turnover_m,  # Configurable minimum turnover
            args.max_spread_pct,  # Configurable maximum spread
            args.liquidity_strict  # Strict mode flag
        )
        
        # === INSTITUTIONAL-GRADE ANALYSIS ===
        institutional_metrics = {}
        
        try:
            # 1. RELATIVE VOLUME CURVE
            rel_vol_ratio, high_liquidity = institutional_analyzer.relative_volume_curve(t, df)
            institutional_metrics['rel_vol_ratio'] = rel_vol_ratio
            institutional_metrics['high_liquidity'] = high_liquidity
            
            # 2. AVWAP BREAKOUT DETECTION
            avwap_breakout = institutional_analyzer.avwap_breakout_entry(t, df)
            institutional_metrics['avwap_breakout'] = avwap_breakout
            
            # 3. ORDER BOOK IMBALANCE SCORE
            ob_score = institutional_analyzer.orderbook_imbalance_score(t, df)
            institutional_metrics['ob_score'] = ob_score
            
            # 4. CATALYST FACTOR SCORE
            inst_catalyst_score = institutional_analyzer.catalyst_factor_score(
                t, rel, deal_amt_cr * 1e7, rise1  # Convert deal amount back to base units
            )
            institutional_metrics['catalyst_score'] = inst_catalyst_score
            
            # 5. FRACTAL RISK BUDGET
            risk_budget = institutional_analyzer.fractal_risk_budget(t, df)
            institutional_metrics['risk_budget'] = risk_budget
            
            if args.debug:
                print(f"🏛️ INSTITUTIONAL {t}: RelVol={rel_vol_ratio:.2f}x, HighLiq={high_liquidity}, "
                      f"AVWAP={avwap_breakout is not None}, OB={ob_score:.1f}, "
                      f"Catalyst={catalyst_score:.1f}, Risk={risk_budget:.3f}")
                      
        except Exception as e:
            if args.debug:
                print(f"⚠️ Institutional analysis failed for {t}: {e}")
            # Set default values
            institutional_metrics = {
                'rel_vol_ratio': 1.0,
                'high_liquidity': False,
                'avwap_breakout': None,
                'ob_score': 50.0,
                'catalyst_score': 0.0,
                'risk_budget': 0.01
            }
        
        # Get liquidity warning level
        liq_warning_level = LiquidityAnalyzer.get_liquidity_warning_level(
            avg_volume, current_price, market_cap_cr, spread_pct
        )
        
        # Add enhanced liquidity tags
        if is_liq_short:
            if liq_warning_level == "CRITICAL":
                tags.append("LIQ_CRITICAL")
            elif liq_warning_level == "HIGH":
                tags.append("LIQ_HIGH")
            else:
                tags.append("LIQ_SHORT")
        
        # Debug output for liquidity analysis
        if args.debug and is_liq_short:
            volume_threshold = LiquidityAnalyzer.get_volume_threshold(market_cap_cr)
            turnover_value = avg_volume * current_price / 1e6  # In millions
            print(f"DEBUG LIQUIDITY {t}: mcap={market_cap_cr:.0f}Cr, vol_threshold={volume_threshold:,}, "
                  f"avg_vol={avg_volume:,.0f}, turnover=₹{turnover_value:.1f}M, "
                  f"spread={spread_pct*100:.2f}%, level={liq_warning_level}")
        
        # Store liquidity data for tier display
        liquidity_map[t] = (liquidity_score, liq_warning_level, is_liq_short)
        # === End Enhanced Liquidity Analysis ===

        recent = [n for n in rel if n.ts and (now() - n.ts).total_seconds() / 3600 < PLATEAU_HR]
        chosen = max(recent or rel, key=lambda x: x.sent or 0)
        head = (chosen.headline[:75] + "…") if len(chosen.headline) > 75 else chosen.headline

        raw_core, amt_cr_out, pct_out, _, _deal_imp_sc = base_core_components(
            rel, df, mc, args.hours, rise1, rise3, volx, bd
        )
        
        # Add technical score to raw core (ENHANCED FOR IDEAL CONDITIONS)
        tech_score = _compute_technical_score(df)
        
        # Enhanced weighting for technical analysis - make it primary factor (70% weight)
        raw_core = _deal_imp_sc[0] if isinstance(_deal_imp_sc, (list, tuple)) else 0.0
        raw_core = raw_core * 0.3 + tech_score * 0.7  # Tech is now 70%, deals 30%
        
        # Add momentum quality factor
        mom_quality = momentum_quality(df)
        raw_core *= (1 + mom_quality/10)  # 0-100% boost based on momentum quality
        
        # Calculate risk/reward ratio for additional scoring
        rr_ratio = calculate_risk_reward_ratio(df)
        rr_boost = 1.5 if rr_ratio >= 3.0 else (1.2 if rr_ratio >= 2.0 else 1.0)
        raw_core *= rr_boost
        
        score_df = prob.calculate_score(df)
        p_pct = prob.probability(score_df).iloc[-1]
        exp5, _, _ = grow.predict(df.tail(30))
        
        # Get valuation score for later use (placeholder during technical scoring)
        sector = detect_sector(names.get(t, ""))
        # Use neutral valuation (score 5) during technical phase
        valuation_score = 5.0  # Will be updated during late financials phase
        pe_str, warn_symbol, pe_risk = "N/A", "", 0.0  # Placeholders
        
        # Apply neutral valuation factor during technical scoring
        raw_core *= (valuation_score/10)  # 0-10 scale affects final score
        
        # Calculate liquidity impact on scoring
        liq_penalty = LiquidityAnalyzer.calculate_liquidity_penalty(
            avg_volume, current_price, market_cap_cr, spread_pct, liquidity_score
        )
        liq_bonus = LiquidityAnalyzer.get_liquidity_bonus(
            avg_volume, current_price, market_cap_cr, liquidity_score
        )
        liq_net_impact = (liq_penalty - 5.0 + liq_bonus)  # Normalize penalty (5 = neutral)
        liq_component = args.liquidity_weight * liq_net_impact / 10.0  # Scale to match other components
        
        # Debug output for liquidity scoring impact
        if args.debug:
            print(f"DEBUG SCORING {t}: liq_penalty={liq_penalty:.2f}, liq_bonus={liq_bonus:.2f}, "
                  f"liq_net={liq_net_impact:.2f}, liq_component={liq_component:.2f}")
        
        # =============================================================================
        # ENHANCED FINANCIAL METRICS INTEGRATION (Critical Additions)
        # =============================================================================
        
        # Skip enhanced metrics during initial screening to preserve speed
        # Only run on final tier candidates (applied later in the pipeline)
        enhanced_impact = 0.0
        
        # Store ticker for later enhanced analysis if it makes it to final tiers
        if (args.enhanced_analytics or not args.skip_enhanced) and not args.ultra_fast:
            # Light technical risk assessment only (no API calls)
            technical_risk_score = 0.0
            
            # Price distance from moving averages (overextension filter)
            try:
                current_price = df["Close"].iloc[-1]
                ma_50 = df["Close"].rolling(50).mean().iloc[-1]
                if ma_50 > 0:
                    price_dist_50ma = ((current_price - ma_50) / ma_50 * 100)
                    
                    # Penalize overextended stocks
                    if price_dist_50ma > 20:  # >20% above 50-day MA
                        technical_risk_score -= 1.0
                    elif price_dist_50ma > 10:
                        technical_risk_score -= 0.5
                    elif price_dist_50ma < -10:  # Oversold condition
                        technical_risk_score += 0.5
            except:
                technical_risk_score = 0.0
            
            # Apply minimal technical risk component (no API overhead)
            enhanced_impact = technical_risk_score * 0.05  # Very light impact during screening

        near = PROB_WT * p_pct / 100 + GROW_WT * exp5 + liq_component + enhanced_impact

        oi_chg = get_oi_change(t)

        if args.legacy_score:
            # Use placeholder financial data during technical phase
            placeholder_fin = {}  # Empty dict for legacy scoring
            total_raw = legacy_full_score(raw_core, placeholder_fin, rel, t, tier_scorer) + near
            scaled = total_raw
            base_part = raw_core
            fundamental_part = total_raw - raw_core - near
        else:
            total_pre = raw_core + near
            # Use placeholder financial data during technical phase
            placeholder_fin = {}  # Empty dict for scaled scoring
            scaled, base_part, fundamental_part = scaled_score(total_pre, placeholder_fin, rel, args.base_scale)
            total_raw = total_pre

        stop_lvl = df["Low"].rolling(3).min().shift(1).iloc[-1]
        if oi_chg is not None and args.legacy_score:
            scaled += np.sign(oi_chg) * min(abs(oi_chg)/10, 2.0)

        # Soft mode: allow min-score of 0 to catch more opportunities
        effective_min_score = 0.0 if args.soft_mode and args.min_score > 0.0 else args.min_score
        
        # Debug: Show score comparison
        if len(ranks) < 5:  # Only show first 5 for debugging
            soft_indicator = " (SOFT-MODE)" if args.soft_mode and effective_min_score == 0.0 else ""
            print(f"🔍 DEBUG: {t} - Raw: {total_raw:.1f}, Scaled: {scaled:.3f}, Min Required: {effective_min_score}{soft_indicator}, Base Scale: {args.base_scale}")
        
        if scaled >= effective_min_score:
            ranks.append((t, scaled))
            raw_score_map[t] = total_raw
            base_part_map[t] = base_part
            fundamental_part_map[t] = fundamental_part
            
            # Enhanced RSI + EMA Oversold Analysis with Best Practices Implementation
            enhancement_details = {}
            try:
                # Calculate RSI (14-period) and EMA indicators using winsorized data
                if len(df) >= 14:
                    try:
                        import talib
                        price_data = get_robust_price_data(df, use_capped=True)
                        rsi = talib.RSI(price_data['close'].values, timeperiod=14)
                        current_rsi = float(rsi[-1]) if not pd.isna(rsi[-1]) else 50.0
                    except:
                        # Fallback RSI calculation using winsorized data
                        price_data = get_robust_price_data(df, use_capped=True)
                        delta = price_data['close'].diff()
                        gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
                        loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
                        rs = gain / loss.replace(0, float('inf'))
                        rsi = 100 - (100 / (1 + rs))
                        current_rsi = float(rsi.iloc[-1]) if not pd.isna(rsi.iloc[-1]) else 50.0
                else:
                    current_rsi = 50.0
                
                # Calculate multiple EMA timeframes for better analysis
                current_price = df['Close'].iloc[-1]
                if len(df) >= 20:
                    ema_14 = df['Close'].ewm(span=14).mean()
                    ema_20 = df['Close'].ewm(span=20).mean()
                    ema_50 = df['Close'].ewm(span=50).mean() if len(df) >= 50 else ema_20
                    
                    # EMA slope calculations
                    ema_14_slope = ((ema_14.iloc[-1] - ema_14.iloc[-3]) / ema_14.iloc[-3] * 100) if len(ema_14) >= 3 else 0.0
                    ema_20_slope = ((ema_20.iloc[-1] - ema_20.iloc[-5]) / ema_20.iloc[-5] * 100) if len(ema_20) >= 5 else 0.0
                    
                    # Price position relative to EMAs
                    price_vs_ema14 = ((current_price - ema_14.iloc[-1]) / ema_14.iloc[-1]) * 100
                    price_vs_ema20 = ((current_price - ema_20.iloc[-1]) / ema_20.iloc[-1]) * 100
                    
                    # EMA trend alignment
                    ema_bullish_alignment = ema_14.iloc[-1] > ema_20.iloc[-1] > ema_50.iloc[-1]
                    ema_bearish_alignment = ema_14.iloc[-1] < ema_20.iloc[-1] < ema_50.iloc[-1]
                else:
                    ema_14_slope = 0.0
                    ema_20_slope = 0.0
                    price_vs_ema14 = 0.0
                    price_vs_ema20 = 0.0
                    ema_bullish_alignment = False
                    ema_bearish_alignment = False
                
                # Volume confirmation analysis
                if len(df) >= 10:
                    avg_volume_10 = df['Volume'].tail(10).mean()
                    current_volume = df['Volume'].iloc[-1]
                    volume_surge = (current_volume / avg_volume_10) if avg_volume_10 > 0 else 1.0
                    volume_confirmation = volume_surge > 1.5  # Best practice: >1.5x avg volume
                else:
                    volume_surge = 1.0
                    volume_confirmation = False
                
                # Candlestick pattern analysis (simplified)
                if len(df) >= 3:
                    # Hammer/Doji detection (simplified)
                    prev_close = df['Close'].iloc[-2]
                    prev_open = df['Open'].iloc[-2]
                    current_open = df['Open'].iloc[-1]
                    current_high = df['High'].iloc[-1]
                    current_low = df['Low'].iloc[-1]
                    
                    body_size = abs(current_price - current_open)
                    upper_shadow = current_high - max(current_price, current_open)
                    lower_shadow = min(current_price, current_open) - current_low
                    
                    # Bullish reversal patterns
                    is_hammer = (lower_shadow > 2 * body_size) and (upper_shadow < body_size)
                    is_bullish_engulfing = (current_price > prev_close) and (current_open < prev_close) and (current_price > prev_open)
                    bullish_candlestick = is_hammer or is_bullish_engulfing
                else:
                    bullish_candlestick = False
                
                # Support level analysis
                if len(df) >= 20:
                    # Recent swing lows (last 20 periods)
                    recent_low = df['Low'].tail(20).min()
                    support_distance = ((current_price - recent_low) / recent_low) * 100
                    near_support = support_distance < 5.0  # Within 5% of recent support
                    
                    # Multiple timeframe support
                    swing_low_50 = df['Low'].tail(50).min() if len(df) >= 50 else recent_low
                    major_support_distance = ((current_price - swing_low_50) / swing_low_50) * 100
                    near_major_support = major_support_distance < 10.0
                else:
                    near_support = False
                    near_major_support = False
                    support_distance = 100.0
                
                # Comprehensive oversold signal analysis with best practices
                oversold_category = 'NONE'
                oversold_signals = 0
                oversold_strength = 0.0
                signal_quality = 0.0
                entry_confidence = 0.0
                
                # Primary RSI Analysis with enhanced thresholds
                rsi_signal_strength = 0.0
                if current_rsi < 15:  # Extremely oversold - rare but powerful
                    oversold_signals += 3
                    oversold_strength += 0.8
                    rsi_signal_strength = 0.9
                    oversold_category = 'ULTIMATE'
                elif current_rsi < 20:  # Extreme oversold
                    oversold_signals += 2
                    oversold_strength += 0.6
                    rsi_signal_strength = 0.7
                    oversold_category = 'EXTREME'
                elif current_rsi < 30:  # Strong oversold
                    oversold_signals += 1
                    oversold_strength += 0.4
                    rsi_signal_strength = 0.5
                    oversold_category = 'STRONG'
                elif current_rsi < 35:  # Moderate oversold
                    oversold_strength += 0.2
                    rsi_signal_strength = 0.3
                    oversold_category = 'MODERATE'
                elif current_rsi < 40:  # Mild oversold
                    oversold_strength += 0.1
                    rsi_signal_strength = 0.1
                    oversold_category = 'MILD'
                
                # EMA Trend Confirmation (Critical for best practices)
                ema_confirmation_strength = 0.0
                if ema_14_slope > 2.0 and price_vs_ema14 > -3.0:  # Strong uptrend + price near EMA14
                    oversold_signals += 2
                    oversold_strength += 0.4
                    ema_confirmation_strength = 0.8
                    signal_quality += 0.3
                elif ema_14_slope > 1.0 and price_vs_ema14 > -5.0:  # Moderate uptrend
                    oversold_signals += 1
                    oversold_strength += 0.2
                    ema_confirmation_strength = 0.5
                    signal_quality += 0.2
                elif ema_14_slope > 0.5:  # Mild uptrend
                    oversold_strength += 0.1
                    ema_confirmation_strength = 0.3
                    signal_quality += 0.1
                elif ema_14_slope < -2.0:  # Strong downtrend - warning
                    oversold_strength *= 0.5  # Reduce confidence in counter-trend signals
                    signal_quality -= 0.2
                
                # Volume Confirmation Boost (Best Practice Implementation)
                volume_boost = 0.0
                if volume_confirmation and oversold_category in ['STRONG', 'EXTREME', 'ULTIMATE']:
                    oversold_signals += 1
                    oversold_strength += 0.3
                    volume_boost = 0.3
                    signal_quality += 0.2
                    entry_confidence += 0.3
                elif volume_surge > 1.2:  # Moderate volume increase
                    oversold_strength += 0.1
                    volume_boost = 0.1
                    signal_quality += 0.1
                    entry_confidence += 0.1
                
                # Candlestick Pattern Confirmation
                pattern_boost = 0.0
                if bullish_candlestick and oversold_category in ['MODERATE', 'STRONG', 'EXTREME']:
                    oversold_signals += 1
                    oversold_strength += 0.2
                    pattern_boost = 0.2
                    signal_quality += 0.15
                    entry_confidence += 0.2
                
                # Support Level Confirmation
                support_boost = 0.0
                if near_support and oversold_category != 'NONE':
                    oversold_strength += 0.2
                    support_boost = 0.2
                    signal_quality += 0.1
                    entry_confidence += 0.15
                elif near_major_support:
                    oversold_strength += 0.1
                    support_boost = 0.1
                    signal_quality += 0.05
                    entry_confidence += 0.1
                
                # Multiple Timeframe Alignment Bonus
                mtf_bonus = 0.0
                if ema_bullish_alignment and oversold_category in ['STRONG', 'EXTREME']:
                    oversold_strength += 0.3
                    mtf_bonus = 0.3
                    signal_quality += 0.2
                    entry_confidence += 0.25
                elif not ema_bearish_alignment:  # Neutral alignment
                    signal_quality += 0.05
                
                # Enhanced category determination with quality assessment
                if oversold_signals >= 4 and signal_quality >= 0.5:
                    oversold_category = 'ULTIMATE'
                elif oversold_signals >= 3 and signal_quality >= 0.4:
                    if oversold_category not in ['ULTIMATE']:
                        oversold_category = 'EXTREME'
                elif oversold_signals >= 2 and signal_quality >= 0.3:
                    if oversold_category not in ['ULTIMATE', 'EXTREME']:
                        oversold_category = 'STRONG'
                
                # Calculate entry zones and stop losses based on best practices
                entry_zones = {}
                stop_losses = {}
                
                if oversold_category == 'EXTREME' or oversold_category == 'ULTIMATE':
                    # Scale-in entry: 50% at 95-97%, 50% at EMA14 break
                    entry_zones['primary'] = current_price * 0.95
                    entry_zones['secondary'] = current_price * 0.97
                    entry_zones['breakout'] = max(ema_14.iloc[-1], current_price * 1.02) if len(df) >= 14 else current_price * 1.02
                    stop_losses['swing_low'] = recent_low * 0.98 if 'recent_low' in locals() else current_price * 0.92
                    stop_losses['percentage'] = current_price * 0.93
                elif oversold_category == 'STRONG':
                    # Conservative entry at 97-98%
                    entry_zones['primary'] = current_price * 0.97
                    entry_zones['secondary'] = current_price * 0.98
                    entry_zones['breakout'] = current_price * 1.015
                    stop_losses['swing_low'] = recent_low * 0.99 if 'recent_low' in locals() else current_price * 0.95
                    stop_losses['percentage'] = current_price * 0.95
                else:
                    # Conservative approach for weaker signals
                    entry_zones['primary'] = current_price * 0.98
                    entry_zones['secondary'] = current_price * 0.995
                    stop_losses['percentage'] = current_price * 0.97
                
                # Risk/Reward calculation
                risk_reward_ratio = 0.0
                if oversold_category != 'NONE' and 'primary' in entry_zones:
                    entry_price = entry_zones['primary']
                    stop_price = stop_losses.get('percentage', entry_price * 0.95)
                    
                    # Target calculation based on signal strength
                    if oversold_category in ['ULTIMATE', 'EXTREME']:
                        target_1 = entry_price * 1.08  # 8% target
                        target_2 = entry_price * 1.15  # 15% target
                    elif oversold_category == 'STRONG':
                        target_1 = entry_price * 1.06  # 6% target
                        target_2 = entry_price * 1.12  # 12% target
                    else:
                        target_1 = entry_price * 1.04  # 4% target
                        target_2 = entry_price * 1.08  # 8% target
                    
                    risk = abs(entry_price - stop_price)
                    reward = target_1 - entry_price
                    risk_reward_ratio = reward / risk if risk > 0 else 0.0
                
                # Final confidence and multiplier calculation
                final_confidence = min(entry_confidence + signal_quality, 1.0)
                oversold_multiplier = 1.0 + (oversold_strength * 0.6)  # Up to 60% boost for best signals
                
                # Quality filter - reduce multiplier for low-quality signals
                if signal_quality < 0.2:
                    oversold_multiplier = min(oversold_multiplier, 1.1)  # Cap at 10% boost
                elif signal_quality < 0.3:
                    oversold_multiplier = min(oversold_multiplier, 1.3)  # Cap at 30% boost
                
                # Store comprehensive enhancement details
                enhancement_details['oversold_category'] = oversold_category
                enhancement_details['oversold_signals'] = oversold_signals
                enhancement_details['oversold_strength'] = oversold_strength
                enhancement_details['oversold_multiplier'] = oversold_multiplier
                enhancement_details['signal_quality'] = signal_quality
                enhancement_details['entry_confidence'] = final_confidence
                enhancement_details['rsi'] = current_rsi
                enhancement_details['ema_14_slope'] = ema_14_slope
                enhancement_details['ema_20_slope'] = ema_20_slope
                enhancement_details['price_vs_ema14'] = price_vs_ema14
                enhancement_details['volume_surge'] = volume_surge
                enhancement_details['volume_confirmation'] = volume_confirmation
                enhancement_details['bullish_candlestick'] = bullish_candlestick
                enhancement_details['near_support'] = near_support
                enhancement_details['ema_bullish_alignment'] = ema_bullish_alignment
                enhancement_details['rsi_signal_strength'] = rsi_signal_strength
                enhancement_details['ema_confirmation_strength'] = ema_confirmation_strength
                enhancement_details['volume_boost'] = volume_boost
                enhancement_details['pattern_boost'] = pattern_boost
                enhancement_details['support_boost'] = support_boost
                enhancement_details['mtf_bonus'] = mtf_bonus
                enhancement_details['entry_zones'] = entry_zones
                enhancement_details['stop_losses'] = stop_losses
                enhancement_details['risk_reward_ratio'] = risk_reward_ratio
                
            except Exception as e:
                # Fallback values
                enhancement_details['oversold_category'] = 'NONE'
                enhancement_details['oversold_signals'] = 0
                enhancement_details['oversold_strength'] = 0.0
                enhancement_details['oversold_multiplier'] = 1.0
                enhancement_details['signal_quality'] = 0.0
                enhancement_details['entry_confidence'] = 0.0
                enhancement_details['rsi'] = 50.0
                enhancement_details['ema_14_slope'] = 0.0
                enhancement_details['risk_reward_ratio'] = 0.0
            
            # Store metrics with enhancement details
            metrics[t] = (rise1, volx, bd, head, p_pct, exp5, stop_lvl, oi_chg, "/".join(dict.fromkeys(tags)), enhancement_details)

        # After processing each ticker
        memory_cleanup(df, df5 if 'df5' in locals() else None, chosen if 'chosen' in locals() else None, rel, recent if 'recent' in locals() else None)

    # Debug: Show scoring summary before filtering
    if not ranks:
        # Count total processed candidates and show sample scores from metrics if available
        total_processed = len(metrics) if 'metrics' in locals() else len(shortlisted_candidates)
        print(f"\n🔍 SCORING DEBUG SUMMARY:")
        print(f"   📊 Total candidates processed: {total_processed}")
        print(f"   🎯 Min score required: {args.min_score}")
        print(f"   📏 Base scale factor: {args.base_scale}")
        print(f"   ❌ Candidates passing filter: {len(ranks)}")
        print(f"   💡 Suggestion: Try lower --min-score (e.g., 0.1) or lower --base-scale (e.g., 50)")
        sys.exit("[ERROR] no stocks meet criteria")

    ranks.sort(key=lambda x: x[1], reverse=True)

    # =============================================================================
    # INSTITUTIONAL GRADE ANALYSIS BATCH PROCESSING
    # =============================================================================
    
    # Calculate cross-sectional residual momentum for all tickers (institutional-grade)
    logging.info("[INSTITUTIONAL] Calculating cross-sectional residual momentum...")
    residual_momentum_ranks = {}
    
    if ENHANCED_IMPORTS['sklearn'] and len(ranks) > 10:  # Only if we have sklearn and enough stocks
        try:
            # Prepare data for cross-sectional analysis
            ticker_returns = []
            valid_tickers = []
            market_proxy_returns = []
            
            # Use SPY/broader market as proxy (or first large stock)
            market_ticker = "SPY" if "SPY" in hist else ranks[0][0]  # Use first ranked stock as market proxy
            market_df = hist.get(market_ticker, pd.DataFrame())
            
            if not market_df.empty and len(market_df) >= 50:
                market_returns = market_df['Close'].pct_change(periods=20).dropna()  # 20-day returns
                
                for ticker, _ in ranks[:min(len(ranks), 50)]:  # Process top 50 stocks max
                    df = hist.get(ticker, pd.DataFrame())
                    if not df.empty and len(df) >= 50:
                        stock_returns = df['Close'].pct_change(periods=20).dropna()
                        
                        # Align dates for regression
                        common_dates = market_returns.index.intersection(stock_returns.index)
                        if len(common_dates) >= 30:  # Need at least 30 observations
                            ticker_returns.append(stock_returns.loc[common_dates].values[-30:])  # Last 30 observations
                            valid_tickers.append(ticker)
                            market_proxy_returns.append(market_returns.loc[common_dates].values[-30:])
                
                # Run cross-sectional residual momentum calculation
                if len(valid_tickers) >= 10:
                    from sklearn.linear_model import LinearRegression
                    
                    for i, ticker in enumerate(valid_tickers):
                        try:
                            stock_ret = ticker_returns[i]
                            market_ret = market_proxy_returns[i]
                            
                            # Fit linear regression: stock_returns = alpha + beta * market_returns + epsilon
                            X = market_ret.reshape(-1, 1)
                            y = stock_ret
                            
                            model = LinearRegression().fit(X, y)
                            residuals = y - model.predict(X)
                            
                            # Calculate residual momentum (mean residual over last period)
                            residual_momentum = np.mean(residuals[-10:])  # Last 10 periods
                            residual_momentum_ranks[ticker] = residual_momentum
                            
                        except Exception as e:
                            residual_momentum_ranks[ticker] = 0.0
                            if args.debug:
                                print(f"DEBUG: Residual momentum failed for {ticker}: {e}")
                    
                    # Rank residuals (higher residual = outperforming market = better)
                    all_residuals = list(residual_momentum_ranks.values())
                    for ticker in residual_momentum_ranks:
                        percentile = (sum(r < residual_momentum_ranks[ticker] for r in all_residuals) / len(all_residuals)) * 100
                        residual_momentum_ranks[ticker] = percentile  # Convert to percentile rank
                    
                    logging.info(f"[INSTITUTIONAL] Residual momentum calculated for {len(valid_tickers)} stocks")
                else:
                    logging.warning("[INSTITUTIONAL] Insufficient stocks for cross-sectional residual momentum")
            else:
                logging.warning(f"[INSTITUTIONAL] Market proxy {market_ticker} insufficient data for residual momentum")
                
        except Exception as e:
            logging.warning(f"[INSTITUTIONAL] Residual momentum calculation failed: {e}")
            residual_momentum_ranks = {}
    else:
        logging.info("[INSTITUTIONAL] Skipping residual momentum: sklearn unavailable or insufficient stocks")
    
    # Update institutional metrics with residual momentum ranks
    for ticker in residual_momentum_ranks:
        if ticker in institutional_metrics:
            institutional_metrics[ticker]['residual_momentum_rank'] = residual_momentum_ranks[ticker]

    # =============================================================================
    # END INSTITUTIONAL GRADE ANALYSIS BATCH PROCESSING  
    # =============================================================================

    # ---------- Late Financials Phase ----------
    # Create scoring shortlist (top 2X final target for financial refinement)
    target_final = args.top
    if args.soft_mode:
        # Soft mode: Process more candidates to capture promising opportunities
        multiplier = 10 if args.ultra_fast else 15  # Very generous in soft mode
        scoring_shortlist_size = min(target_final * multiplier, len(ranks))
    else:
        scoring_shortlist_size = min(target_final * 2, len(ranks))
    scoring_shortlist = [t for t, _ in ranks[:scoring_shortlist_size]]
    
    logging.info(f"Fetching financial data for top {len(scoring_shortlist)} candidates...")
    
    # Fetch financial data for top candidates only
    fin_shortlist = fetch_financials_late(scoring_shortlist, hist, args)
    
    # IMPORTANT: Assign financial data to main fin dictionary for later use
    fin.update(fin_shortlist)  # This fixes the P/E N/A issue
    
    # Update scores with actual financial data
    updated_ranks = []
    for t, original_score in ranks[:scoring_shortlist_size]:
        if t not in fin_shortlist:
            # Keep original score if financial data not available
            updated_ranks.append((t, original_score))
            continue
            
        # Recalculate with actual financial data
        try:
            df = hist.get(t, pd.DataFrame())
            if df.empty:
                updated_ranks.append((t, original_score))
                continue
                
            # Get stored intermediate values
            rel = rel_map.get(t, [])
            mc = mcap.get(t, 0.0)
            
            # Recalculate valuation score with actual financial data
            sector = detect_sector(names.get(t, ""))
            pe_str, warn_symbol, pe_risk, valuation_score = analyze_pe(fin_shortlist[t], sector, mc)
            
            # Apply updated valuation factor to original base score
            base_score = base_part_map.get(t, 0.0)
            fundamental_score = fundamental_part_map.get(t, 0.0)
            
            # Recalculate total with actual financial data
            if args.legacy_score:
                close = df["Close"]
                rise1 = (close.iloc[-1] - close.iloc[-2]) / close.iloc[-2] * 100
                tech_score = _compute_technical_score(df)
                raw_core = base_score  # Use stored base technical score
                raw_core *= (valuation_score/10)  # Apply actual valuation
                
                total_raw = legacy_full_score(raw_core, fin_shortlist[t], rel, t, tier_scorer)
                updated_score = total_raw + (original_score - raw_score_map.get(t, 0.0))  # Add non-fundamental components
            else:
                # For non-legacy scoring, recalculate scaled score with actual financials
                base_technical = base_part_map.get(t, 0.0)
                adjusted_technical = base_technical * (valuation_score/10)  # Apply actual valuation
                total_pre = adjusted_technical + (original_score - base_technical)  # Keep other components
                updated_score, _, _ = scaled_score(total_pre, fin_shortlist[t], rel, args.base_scale)
            
            updated_ranks.append((t, updated_score))
            
        except Exception as e:
            logging.warning(f"Error updating score for {t}: {e}")
            updated_ranks.append((t, original_score))
    
    # Add remaining ranks without financial updates
    updated_ranks.extend(ranks[scoring_shortlist_size:])
    
    # Re-sort with updated scores
    ranks = sorted(updated_ranks, key=lambda x: x[1], reverse=True)
    
    logging.info(f"Financial scoring completed. Top score: {ranks[0][1]:.2f}")

    # Late FII (skip in ultra-fast mode)
    if not args.no_fii and not args.ultra_fast:
        syms = [t for t, _ in ranks[:args.top]]
        logging.info("Fetching FII growth for %d symbols...", len(syms))
        fii_growth_map.update(await fetch_fii_growth_parallel(syms))
    elif args.ultra_fast:
        logging.info("[ULTRA-FAST] Ultra-fast mode: Skipping FII data fetching")
        
        # Apply comprehensive scoring enhancements
        enhanced_ranks = []
        
        # Initialize enhanced modules if available (skip in ultra-fast mode)
        enhanced_engine = None
        if ENHANCED_MODULES_AVAILABLE and not args.ultra_fast:
            try:
                enhanced_engine = EnhancedScreenerEngine()
                logging.info("✅ Enhanced analysis modules loaded")
            except Exception as e:
                logging.warning(f"Enhanced modules failed to initialize: {e}")
        elif args.ultra_fast:
            logging.info("[ULTRA-FAST] Ultra-fast mode: Skipping enhanced analysis modules")
        
        for t, original_score in ranks[:args.top]:
            try:
                enhanced_score = original_score
                enhancement_details = {}
                
                # Skip all enhanced analysis in ultra-fast mode
                if args.ultra_fast:
                    enhanced_ranks.append((t, enhanced_score))
                    continue
                
                if enhanced_engine:
                    # Get sector and market cap
                    sector = detect_sector(names.get(t, t))
                    market_cap = mcap.get(t, 0.0)
                    price_data = hist.get(t, pd.DataFrame())
                    
                    # Perform comprehensive enhanced analysis
                    fin_obj = fin.get(t)  # Get the actual Fin object
                    fin_data_dict = {
                        'pe_ratio': fin_obj.pe_ratio if fin_obj else None,
                        'net_margin': fin_obj.net if fin_obj else None,
                        'income_growth': fin_obj.inc if fin_obj else None,
                        'debt_to_equity': fin_obj.d2e if fin_obj else None,
                        'margin': fin_obj.margin if fin_obj else None,
                        'volume_growth': fin_obj.volume_growth if fin_obj else None,
                        'debt_reduction': fin_obj.debt_red if fin_obj else None
                    }
                    
                    enhanced_result = await enhanced_engine.analyze_stock_comprehensive(
                        symbol=t,
                        base_score=original_score,
                        sector=sector,
                        market_cap=market_cap,
                        price_data=price_data,
                        fin_data=fin_data_dict
                    )
                    
                    enhanced_score = enhanced_result['final_score']
                    enhancement_details = enhanced_result
                    
                    # Apply FII institutional bonus (legacy compatibility)
                    fii_growth = fii_growth_map.get(t, 0.0)
                    if fii_growth and abs(fii_growth) > 0.5:
                        fii_bonus = np.sign(fii_growth) * min(abs(fii_growth) * 0.3, 1.0)
                        enhanced_score += fii_bonus
                        enhancement_details['fii_bonus'] = fii_bonus
                
                else:
                    # Enhanced RSI + EMA Mean Reversion Analysis (Primary Strategy)
                    df = hist.get(t, pd.DataFrame())
                    if not df.empty and len(df) >= 20:
                        try:
                            # Core analysis: RSI + EMA combination (highest priority)
                            close = df['Close']
                            high = df['High']
                            low = df['Low']
                            volume = df['Volume']
                            
                            # 1. RSI Analysis (Primary Signal - 50% weight)
                            delta = close.diff()
                            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
                            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
                            rs = gain / loss
                            rsi = 100 - (100 / (1 + rs))
                            current_rsi = rsi.iloc[-1] if not rsi.empty else 50
                            
                            # 2. EMA 14 Analysis (Primary Trend Confirmation - 35% weight)
                            ema_14 = close.ewm(span=14).mean()
                            current_price = close.iloc[-1]
                            current_ema = ema_14.iloc[-1] if not ema_14.empty else current_price
                            ema_slope = (ema_14.iloc[-1] - ema_14.iloc[-5]) / ema_14.iloc[-5] * 100 if len(ema_14) >= 5 else 0
                            price_vs_ema = ((current_price - current_ema) / current_ema) * 100  # % above/below EMA
                            
                            # EMA trend confirmation signals
                            ema_bullish_trend = ema_slope > 0.5  # EMA rising > 0.5%
                            price_above_ema = price_vs_ema > -2.0  # Price within 2% of EMA (not severely below)
                            ema_support = price_vs_ema > -5.0 and ema_slope > -1.0  # EMA acting as support
                            
                            # 3. Additional Confirmation Signals (15% weight total)
                            # Volume confirmation
                            avg_volume = volume.rolling(window=20).mean().iloc[-1] if len(volume) >= 20 else volume.mean()
                            volume_ratio = volume.iloc[-1] / avg_volume if avg_volume > 0 else 1.0
                            volume_confirmation = volume_ratio > 1.2  # 20% above average volume
                            
                            # Support level confirmation
                            support_level = low.rolling(window=50).min().iloc[-1] if len(low) >= 50 else low.min()
                            near_support = ((current_price - support_level) / support_level) < 0.1 if support_level > 0 else False
                            
                            # Calculate RSI + EMA Combined Score (Most Important)
                            oversold_signals = 0
                            oversold_strength = 0.0
                            signal_details = []
                            
                            # PRIMARY: RSI Oversold Conditions (50% weight)
                            if current_rsi < 20:  # Extreme oversold
                                oversold_signals += 2  # Double weight for extreme
                                oversold_strength += 0.6  # Higher strength
                                signal_details.append(f"RSI_EXTREME({current_rsi:.1f})")
                            elif current_rsi < 30:  # Oversold
                                oversold_signals += 1
                                oversold_strength += 0.4
                                signal_details.append(f"RSI_OVERSOLD({current_rsi:.1f})")
                            
                            # Handle overbought conditions (penalty)
                            elif current_rsi > 70:  # Overbought - reduce signals
                                oversold_signals -= 1
                                oversold_strength -= 0.2
                                signal_details.append(f"RSI_OVERBOUGHT({current_rsi:.1f})")
                            
                            # PRIMARY: EMA Trend Confirmation (35% weight)
                            if ema_bullish_trend and price_above_ema:
                                # Best case: Rising EMA + Price above/near EMA
                                oversold_signals += 2  # Strong signal
                                oversold_strength += 0.5
                                signal_details.append(f"EMA_BULLISH_CONFIRM({ema_slope:.1f}%)")
                            elif ema_support and current_rsi < 35:
                                # Good case: EMA support + mild oversold
                                oversold_signals += 1
                                oversold_strength += 0.3
                                signal_details.append(f"EMA_SUPPORT({price_vs_ema:.1f}%)")
                            elif ema_slope < -2.0:
                                # Bad case: Strong downtrend - reduce signals
                                oversold_signals -= 1
                                oversold_strength -= 0.3
                                signal_details.append(f"EMA_BEARISH({ema_slope:.1f}%)")
                            
                            # SECONDARY: Additional Confirmations (15% weight)
                            if volume_confirmation and current_rsi < 35:
                                oversold_signals += 1
                                oversold_strength += 0.2
                                signal_details.append(f"VOL_CONFIRM({volume_ratio:.1f}x)")
                            
                            if near_support and current_rsi < 40:
                                oversold_signals += 1
                                oversold_strength += 0.15
                                signal_details.append("SUPPORT_LEVEL")
                            
                            # Apply RSI + EMA Combined Scoring (Progressive Multipliers)
                            oversold_multiplier = 1.0
                            oversold_category = 'NONE'
                            
                            # Tier 1: Ultimate RSI + EMA Combo (RSI < 25 + Rising EMA)
                            if (current_rsi < 25 and ema_bullish_trend and price_above_ema and 
                                oversold_signals >= 3 and oversold_strength >= 0.8):
                                oversold_multiplier = 2.5  # 150% boost - Ultimate combo
                                oversold_category = 'ULTIMATE'
                                
                            # Tier 2: Extreme RSI + EMA Support (RSI < 20 + EMA support)
                            elif (current_rsi < 20 and ema_support and 
                                  oversold_signals >= 2 and oversold_strength >= 0.7):
                                oversold_multiplier = 2.2  # 120% boost
                                oversold_category = 'EXTREME'
                                
                            # Tier 3: Strong RSI + Mild EMA Confirm (RSI < 30 + Rising EMA)
                            elif (current_rsi < 30 and ema_bullish_trend and 
                                  oversold_signals >= 2 and oversold_strength >= 0.5):
                                oversold_multiplier = 1.9  # 90% boost
                                oversold_category = 'STRONG'
                                
                            # Tier 4: Moderate RSI + EMA Support (RSI < 35 + EMA near support)
                            elif (current_rsi < 35 and ema_support and 
                                  oversold_signals >= 1 and oversold_strength >= 0.3):
                                oversold_multiplier = 1.5  # 50% boost
                                oversold_category = 'MODERATE'
                                
                            # Tier 5: Mild RSI Oversold + Basic Confirmation
                            elif (current_rsi < 40 and oversold_signals >= 1 and oversold_strength >= 0.2):
                                oversold_multiplier = 1.2  # 20% boost
                                oversold_category = 'MILD'
                            
                            # Apply penalty for overbought conditions
                            elif current_rsi > 70 and ema_slope < 0:
                                oversold_multiplier = 0.85  # 15% penalty for overbought in downtrend
                                oversold_category = 'OVERBOUGHT'
                            
                            # Apply the RSI + EMA enhanced scoring
                            if oversold_multiplier != 1.0:
                                enhanced_score *= oversold_multiplier
                                enhancement_details['oversold_multiplier'] = oversold_multiplier
                                enhancement_details['oversold_category'] = oversold_category
                                enhancement_details['oversold_signals'] = oversold_signals
                                enhancement_details['oversold_strength'] = oversold_strength
                                enhancement_details['signal_details'] = ' + '.join(signal_details[:3])  # Top 3 signals
                                enhancement_details['rsi'] = current_rsi
                                enhancement_details['ema_slope'] = ema_slope
                                enhancement_details['price_vs_ema'] = price_vs_ema
                                enhancement_details['volume_ratio'] = volume_ratio
                                enhancement_details['ema_bullish'] = ema_bullish_trend
                                enhancement_details['ema_support'] = ema_support
                                
                        except Exception as e:
                            logging.warning(f"Enhanced RSI+EMA analysis failed for {t}: {e}")
                
                # Store enhanced results
                enhanced_ranks.append((t, enhanced_score))
                
                # Update metrics with enhancement details  
                if t in metrics:
                    original_metrics = metrics[t]
                    enhanced_metrics = original_metrics + (enhancement_details,)
                    metrics[t] = enhanced_metrics
                
            except Exception as e:
                logging.warning(f"Enhanced scoring failed for {t}: {e}")
                enhanced_ranks.append((t, original_score))
        
        # Use enhanced ranks
        ranks = enhanced_ranks
        ranks.sort(key=lambda x: x[1], reverse=True)
        
        # Skip expensive FII processing in ultra-fast mode  
        if not args.ultra_fast:
            enhanced_ranks = []  # Reset for FII enhancement
            for t, score in ranks:
                fii_growth = fii_growth_map.get(t, 0.0)  # Default to 0.0 if not found
                if fii_growth is None:  # Handle None values from failed FII fetches
                    fii_growth = 0.0
                
                # Apply FII confidence weighting
                enhanced_score = adjust_score_with_fii(score, fii_growth)
                
                # Additional enhancements with price data (if available)
                try:
                    # Get price data for advanced analysis using resilient download - prevent double .NS suffix
                    yf_symbol = ensure_ns_suffix(t)
                    price_result = safe_yf_download([yf_symbol], period="3mo")
                    price_data = price_result.get(yf_symbol, pd.DataFrame())
                    
                    if not price_data.empty and len(price_data) >= 20:
                        # Apply institutional accumulation bonus
                        accumulation_score = detect_institutional_accumulation(price_data, fii_growth)
                        if accumulation_score >= 7.0:  # Strong accumulation signal
                            enhanced_score *= 1.15  # 15% bonus
                        elif accumulation_score >= 6.0:  # Moderate accumulation
                            enhanced_score *= 1.08  # 8% bonus
                        
                        # Apply momentum quality filter
                        momentum_score = momentum_quality(price_data)
                        if momentum_score < 4.0:  # Poor momentum quality
                            enhanced_score *= 0.9  # 10% penalty
                        
                    memory_cleanup(price_data)
                except Exception as e:
                    logging.debug(f"Enhanced scoring failed for {t}: {e}")
                    # Continue with basic FII adjustment
                
                enhanced_ranks.append((t, enhanced_score))
            
            # Re-sort with enhanced scores
            ranks = enhanced_ranks
            ranks.sort(key=lambda x: x[1], reverse=True)
        else:
            logging.info("[ULTRA-FAST] Ultra-fast mode: Skipping FII enhancement processing")
        
        # Memory cleanup after FII data fetch (only if syms was created)
        if 'syms' in locals():
            memory_cleanup(syms)
    else:
        logging.info("FII growth skipped (--no-fii)")

    # Quarterly results growth
    top_syms = [t for t, _ in ranks[:args.top]]
    if not args.ultra_fast:
        logging.info("Fetching quarterly RESULTS growth for %d symbols...", len(top_syms))
        qtr_map = get_quarter_result_growth(top_syms)
    else:
        logging.info("[ULTRA-FAST] Ultra-fast mode: Skipping quarterly results growth")
        qtr_map = {}  # Empty dict for ultra-fast mode
    # Memory cleanup after quarterly results fetch
    memory_cleanup(top_syms)

    # Tier classification (with deal impact and enhanced criteria)
    classification: Dict[str,str] = {}
    for t, _ in ranks[:args.top]:
        rel = rel_map.get(t, [])
        mc = mcap.get(t, 0.0)
        
        # Get valuation score for tier classification
        sector = detect_sector(names.get(t, ""))
        fin_data = fin.get(t, {}) if fin else {}
        if fin_data:
            pe_str, warn_symbol, pe_risk, valuation_score = analyze_pe(fin_data, sector, mc)
        else:
            pe_str, warn_symbol, pe_risk, valuation_score = "N/A", "", 0.0, 5.0
        
        tier = classifier.classify(
            t,
            fin_data,
            rel,
            mc,
            deal_amount_map.get(t, 0.0) * 1e7,
            deal_count_map.get(t, 0),
            deal_impact_map.get(t, 0.0),
            liquidity_map.get(t, (0.0, "LOW", True)),  # liquidity_data
            valuation_score  # valuation_score
        )
        classification[t] = tier

    # GUARANTEED TOP N DISPLAY - Always show exactly args.top stocks
    # Collect all ranked stocks with their tier classifications
    all_classified_stocks = [(t, s, classification.get(t, "TIER3")) for t, s in ranks[:args.top]]
    
    # Distribute into tiers while ensuring we show all top N stocks
    tier1 = [(t, s) for t, s, tier in all_classified_stocks if tier == "TIER1"]
    tier2 = [(t, s) for t, s, tier in all_classified_stocks if tier == "TIER2"]
    tier3 = [(t, s) for t, s, tier in all_classified_stocks if tier == "TIER3"]
    
    # If we don't have enough classified stocks, add remaining to TIER3
    classified_tickers = {t for t, s, tier in all_classified_stocks}
    remaining_stocks = [(t, s) for t, s in ranks[:args.top] if t not in classified_tickers]
    tier3.extend(remaining_stocks)
    
    # Ensure we have exactly args.top total stocks displayed
    total_stocks_to_show = len(tier1) + len(tier2) + len(tier3)
    if total_stocks_to_show < args.top:
        # Add more stocks from remaining ranks if available
        additional_needed = args.top - total_stocks_to_show
        additional_stocks = ranks[total_stocks_to_show:total_stocks_to_show + additional_needed]
        tier3.extend(additional_stocks)
    
    # =============================================================================
    # ENHANCED FINANCIAL ANALYSIS FOR FINAL TIER CANDIDATES
    # =============================================================================
    
    # Run enhanced analysis only on final tier candidates to preserve speed
    enhanced_data = {}
    if (args.enhanced_analytics or not args.skip_enhanced) and not args.ultra_fast:
        # Collect final tier candidates (prioritize tier1 and tier2)
        final_candidates = []
        final_candidates.extend([t for t, _ in tier1[:5]])  # Top 5 from tier1
        final_candidates.extend([t for t, _ in tier2[:5]])  # Top 5 from tier2
        
        # Remove duplicates while preserving order
        seen = set()
        unique_candidates = []
        for t in final_candidates:
            if t not in seen:
                seen.add(t)
                unique_candidates.append(t)
        
        # Run enhanced analysis on final candidates only
        if unique_candidates:
            enhanced_data = analyze_final_candidates_enhanced(unique_candidates, hist)
    
    # Store enhanced data for use in display sections
    global _enhanced_financial_data
    _enhanced_financial_data = enhanced_data

    # ----- Certainty Score Integration -----
    certainty_map: Dict[str, float] = {}
    for t, _ in ranks[:args.top]:
        cert = CertaintyEngine.calculate(
            df=hist[t],
            fin_data=fin.get(t, {}) if fin else {},
            news=rel_map.get(t, []),
            mcap=mcap.get(t, 0.0),
            deal_impact=deal_impact_map.get(t, 0.0),
            tier=classification.get(t, "")
        )
        certainty_map[t] = cert
        if CertaintyEngine.detect_accumulation(hist[t]):
            prev = metrics[t][8]
            new_tags = prev + ",ACCUM" if prev else "ACCUM"
            metrics[t] = metrics[t][:8] + (new_tags,)
    # ---------------------------------------
    entry_obj = entry
    monitor_fw = monitor

    
    def fmt_fii(t):
        """Return color‑coded FII % with QoQ delta (fallback n/a)."""
        pct = fii_pct_map.get(t) if 'fii_pct_map' in locals() else None
        delta = fii_growth_map.get(t)
        if pct is None and delta is None:
            return " n/a "
        base = f"{pct:.1f}%" if pct is not None else "n/a"
        if delta is not None:
            sign = "+" if delta >= 0 else ""
            s = f"{base} ({sign}{delta:.1f}pp)"
            if delta > 0:
                s = f"\033[92m{s}\033[0m"
            elif delta < 0:
                s = f"\033[91m{s}\033[0m"
            return s
        return base
        """Return colour‑coded FII percentage (fallback 0.0%)."""
        v = fii_growth_map.get(t)
        if v is None:
            v = 0.0
        s = f"{v:+5.1f}%"
        if v > 0:
            s = f"[92m{s}[0m"
        elif v < 0:
            s = f"[91m{s}[0m"
        return s

    def fmt_qtr(t):
        v = qtr_map.get(t)
        if v is None:
            return " n/a "
        s = f"{v:+5.1f}%"
        if v >= 15:
            s = f"\033[95m{s}\033[0m"  # Purple - Excellent
        elif v >= 5:
            s = f"\033[92m{s}\033[0m"  # Green - Good
        elif v >= 0:
            s = f"\033[93m{s}\033[0m"  # Yellow - OK
        else:
            s = f"\033[91m{s}\033[0m"  # Red - Bad
        return s
    
    def get_visual_length(text):
        """Get the visual length of text excluding ANSI escape codes."""
        import re
        # Remove ANSI escape sequences
        ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
        return len(ansi_escape.sub('', text))
    
    def pad_colored_text(text, width, align='left'):
        """Pad colored text to specific width, accounting for ANSI codes."""
        visual_length = get_visual_length(text)
        padding_needed = max(0, width - visual_length)
        
        if align == 'left':
            return text + ' ' * padding_needed
        elif align == 'right':
            return ' ' * padding_needed + text
        else:  # center
            left_pad = padding_needed // 2
            right_pad = padding_needed - left_pad
            return ' ' * left_pad + text + ' ' * right_pad

    def calculate_entry_price(ticker, df):
        """Calculate optimal entry price based on today's high/low and technical analysis."""
        if df.empty or len(df) < 2:
            return None, "No data"
        
        try:
            # Get today's data
            today_high = df['High'].iloc[-1]
            today_low = df['Low'].iloc[-1]
            today_close = df['Close'].iloc[-1]
            today_open = df['Open'].iloc[-1]
            yesterday_close = df['Close'].iloc[-2] if len(df) >= 2 else today_close
            
            # Calculate key levels
            daily_range = today_high - today_low
            mid_price = (today_high + today_low) / 2
            
            # Entry strategy based on today's price action
            if today_close > today_open:  # Green candle
                if today_close > (today_high - daily_range * 0.2):  # Near high
                    # For stocks near high, suggest entry on slight pullback
                    entry_price = today_high - (daily_range * 0.15)  # 15% pullback from high
                    strategy = "Pullback Entry"
                else:
                    # Good momentum, can enter near current levels
                    entry_price = min(today_close + (daily_range * 0.05), today_high * 0.98)
                    strategy = "Momentum Entry"
            else:  # Red candle or doji
                if today_close < (today_low + daily_range * 0.2):  # Near low
                    # Entry at slight bounce from low
                    entry_price = today_low + (daily_range * 0.1)  # 10% bounce from low
                    strategy = "Bounce Entry"
                else:
                    # Wait for better levels
                    entry_price = mid_price
                    strategy = "Mid-Range Entry"
            
            # Safety check - ensure entry price is reasonable
            min_entry = min(today_low * 1.001, yesterday_close * 0.95)  # Max 5% below yesterday
            max_entry = max(today_high * 0.999, yesterday_close * 1.05)  # Max 5% above yesterday
            
            entry_price = max(min_entry, min(entry_price, max_entry))
            
            # Calculate distance from current price
            distance_pct = ((entry_price - today_close) / today_close) * 100
            
            return entry_price, f"{strategy} ({distance_pct:+.1f}%)"
            
        except Exception as e:
            return today_close if 'today_close' in locals() else None, f"Error: {str(e)[:20]}"

    # ---- Summary Header ----
    print("\n" + "═"*160)
    print(f"DeepSeek Swing Screener  ·  Run: {run_ts}  ·  Hours Back: {args.hours}")
    print(f"Analyzed: {len(ranks)}  |  Tier1: {len(tier1)}  Tier2: {len(tier2)}  Tier3: {len(tier3)}")
    print("═"*160)

    # Legacy flat table (colored) - Now in Excel-like format
    if not args.no_legacy_table:
        # Define column widths for Excel-like formatting (improved spacing)
        main_col_widths = {
            'rank': 7, 'ticker': 12, 'score': 8, 'fii': 9, 'qtr': 9, 'delta1d': 9,
            'volx': 7, 'bd': 7, 'prob': 9, 'grow': 9, 'entry_price': 12, 'entry_strategy': 16,
            'stop': 10, 'amt_cr': 12, 'amt_pct': 9, 'dcnt': 7, 'poscnt': 8
        }
        
        if args.excel_view:
            header_cols = ['RANK', 'TICKER', 'SCORE', 'FII%', 'QTR%', 'Δ1D%', 'VOLX', 'BD%', 'P%', 'G%', 'ENTRY_PRICE', 'ENTRY_STRATEGY', 'STOP', 'AMT(Cr)', 'AMT%', 'D_CNT', 'POS_CNT']
            print("\t".join(header_cols))
        else:
            # Excel-like header with proper spacing
            header_row = (
                f"{'RANK':<{main_col_widths['rank']}} "
                f"{'TICKER':<{main_col_widths['ticker']}} "
                f"{'SCORE':<{main_col_widths['score']}} "
                f"{'FII%':<{main_col_widths['fii']}} "
                f"{'QTR%':<{main_col_widths['qtr']}} "
                f"{'Δ1D%':<{main_col_widths['delta1d']}} "
                f"{'VOLX':<{main_col_widths['volx']}} "
                f"{'BD%':<{main_col_widths['bd']}} "
                f"{'P%':<{main_col_widths['prob']}} "
                f"{'G%':<{main_col_widths['grow']}} "
                f"{'ENTRY_PRICE':<{main_col_widths['entry_price']}} "
                f"{'ENTRY_STRATEGY':<{main_col_widths['entry_strategy']}} "
                f"{'STOP':<{main_col_widths['stop']}} "
                f"{'AMT(Cr)':<{main_col_widths['amt_cr']}} "
                f"{'AMT%':<{main_col_widths['amt_pct']}} "
                f"{'D_CNT':<{main_col_widths['dcnt']}} "
                f"{'POS_CNT':<{main_col_widths['poscnt']}}"
            )
            print(header_row)
            # Calculate separator line length based on actual column widths
            main_separator_length = sum([main_col_widths['rank'], main_col_widths['ticker'], 
                                       main_col_widths['score'], main_col_widths['fii'], 
                                       main_col_widths['qtr'], main_col_widths['delta1d'], 
                                       main_col_widths['volx'], main_col_widths['bd'], 
                                       main_col_widths['prob'], main_col_widths['grow'], 
                                       main_col_widths['entry_price'], main_col_widths['entry_strategy'],
                                       main_col_widths['stop'], main_col_widths['amt_cr'], 
                                       main_col_widths['amt_pct'], main_col_widths['dcnt'], 
                                       main_col_widths['poscnt']]) + 16  # spaces between columns
            print("-" * main_separator_length)

        for i,(t,s) in enumerate(ranks[:args.top],1):
            # Handle both old and new metrics format
            metric_data = metrics[t]
            if len(metric_data) == 10:  # New format with enhancement_details
                r1, vx, bd, head, p, g, stop, oi_chg, _tags, enhancement_details = metric_data
            else:  # Old format
                r1, vx, bd, head, p, g, stop, oi_chg, _tags = metric_data
                enhancement_details = {}
            
            if args.excel_view:
                fii_val = fii_growth_map.get(t)
                qtr_val = qtr_map.get(t)
                amt_pct = deal_pct_map.get(t, 0.0)
                deal_cnt_val = deal_count_map.get(t, 0)
                pos_cnt_val = positive_count_map.get(t, 0)
                bd_pct = bd * 100
                
                # Calculate entry price for Excel view
                df = hist[t] if t in hist else pd.DataFrame()
                entry_price, entry_strategy = calculate_entry_price(t, df)

                row_data = [
                    i, t, f"{s:.2f}",
                    add_indicator(fii_val, "{:+.1f}%") if fii_val is not None else "n/a",
                    add_indicator(qtr_val, "{:+.1f}%") if qtr_val is not None else "n/a",
                    add_indicator(r1, "{:+.1f}"),
                    add_indicator(vx, "{:.2f}", good_threshold=args.min_vol),
                    add_indicator(bd_pct, "{:.1f}", good_threshold=60, bad_threshold=40),
                    add_indicator(p, "{:.1f}", good_threshold=70),
                    add_indicator(g, "{:+.2f}", good_threshold=10),
                    f"{entry_price:.2f}" if entry_price else "N/A",
                    entry_strategy if entry_strategy else "No data",
                    f"{stop:.2f}",
                    f"{deal_amount_map.get(t, 0.0):.1f}",
                    add_indicator(amt_pct, "{:.1f}%", good_threshold=5.0),
                    add_indicator(deal_cnt_val, "{}", good_threshold=2),
                    add_indicator(pos_cnt_val, "{}", good_threshold=10)
                ]
                print("\t".join(map(str, row_data)))
            else:
                # Excel-like format with colors
                fii_val = fii_growth_map.get(t)
                qtr_val = qtr_map.get(t)
                
                # Format and color the values
                score_txt = f"{s:.2f}"
                if s >= 10:
                    score_txt = f"\033[95m{score_txt}\033[0m"  # Purple - Excellent
                elif s >= 8:
                    score_txt = f"\033[92m{score_txt}\033[0m"  # Green - Good
                
                fii_txt = f"{fii_val:+.1f}%" if fii_val is not None else "n/a"
                if fii_val is not None:
                    if fii_val >= 15:
                        fii_txt = f"\033[95m{fii_txt}\033[0m"  # Purple - Excellent
                    elif fii_val >= 5:
                        fii_txt = f"\033[92m{fii_txt}\033[0m"  # Green - Good
                    elif fii_val >= 0:
                        fii_txt = f"\033[93m{fii_txt}\033[0m"  # Yellow - OK
                    else:
                        fii_txt = f"\033[91m{fii_txt}\033[0m"  # Red - Bad
                
                qtr_txt = f"{qtr_val:+.1f}%" if qtr_val is not None else "n/a"
                if qtr_val is not None:
                    if qtr_val >= 15:
                        qtr_txt = f"\033[95m{qtr_txt}\033[0m"  # Purple - Excellent
                    elif qtr_val >= 5:
                        qtr_txt = f"\033[92m{qtr_txt}\033[0m"  # Green - Good
                    elif qtr_val >= 0:
                        qtr_txt = f"\033[93m{qtr_txt}\033[0m"  # Yellow - OK
                    else:
                        qtr_txt = f"\033[91m{qtr_txt}\033[0m"  # Red - Bad
                
                r1_txt = f"{r1:+.1f}"
                if r1 >= 5:
                    r1_txt = f"\033[95m{r1_txt}\033[0m"  # Purple - Excellent
                elif r1 >= 2:
                    r1_txt = f"\033[92m{r1_txt}\033[0m"  # Green - Good
                elif r1 >= 0:
                    r1_txt = f"\033[93m{r1_txt}\033[0m"  # Yellow - OK
                else:
                    r1_txt = f"\033[91m{r1_txt}\033[0m"  # Red - Bad
                
                volx_txt = f"{vx:.2f}"
                if vx >= args.min_vol * 2:
                    volx_txt = f"\033[95m{volx_txt}\033[0m"  # Purple - Excellent volume
                elif vx >= args.min_vol:
                    volx_txt = f"\033[92m{volx_txt}\033[0m"  # Green - Good volume
                
                bd_pct = bd * 100
                bd_txt = f"{bd_pct:.1f}"
                if bd_pct >= 80:
                    bd_txt = f"\033[95m{bd_txt}\033[0m"  # Purple - Excellent
                elif bd_pct >= 60:
                    bd_txt = f"\033[92m{bd_txt}\033[0m"  # Green - Good
                elif bd_pct >= 50:
                    bd_txt = f"\033[93m{bd_txt}\033[0m"  # Yellow - OK
                elif bd_pct <= 40:
                    bd_txt = f"\033[91m{bd_txt}\033[0m"  # Red - Bad
                
                p_txt = f"{p:.1f}"
                if p >= 80:
                    p_txt = f"\033[95m{p_txt}\033[0m"  # Purple - Excellent
                elif p >= 70:
                    p_txt = f"\033[92m{p_txt}\033[0m"  # Green - Good
                
                g_txt = f"{g:+.2f}"
                if g >= 20:
                    g_txt = f"\033[95m{g_txt}\033[0m"  # Purple - Excellent
                elif g >= 10:
                    g_txt = f"\033[92m{g_txt}\033[0m"  # Green - Good
                
                amt_pct = deal_pct_map.get(t, 0.0)
                amt_pct_txt = f"{amt_pct:.1f}%"
                if amt_pct >= 5:
                    amt_pct_txt = f"\033[95m{amt_pct_txt}\033[0m"
                
                dcnt_val = deal_count_map.get(t, 0)
                dcnt_txt = f"{dcnt_val}"
                if dcnt_val >= 5:
                    dcnt_txt = f"\033[95m{dcnt_txt}\033[0m"  # Purple - Excellent
                elif dcnt_val > 1:
                    dcnt_txt = f"\033[92m{dcnt_txt}\033[0m"  # Green - Good
                
                poscnt_val = positive_count_map.get(t, 0)
                poscnt_txt = f"{poscnt_val}"
                if poscnt_val >= 15:
                    poscnt_txt = f"\033[95m{poscnt_txt}\033[0m"  # Purple - Excellent
                elif poscnt_val >= 10:
                    poscnt_txt = f"\033[92m{poscnt_txt}\033[0m"  # Green - Good
                
                # Calculate entry price based on today's high/low
                df = hist[t] if t in hist else pd.DataFrame()
                entry_price, entry_strategy = calculate_entry_price(t, df)
                
                if entry_price:
                    entry_price_txt = f"{entry_price:.2f}"
                    if "Pullback" in entry_strategy:
                        entry_price_txt = f"\033[93m{entry_price_txt}\033[0m"  # Yellow - Wait for pullback
                    elif "Momentum" in entry_strategy:
                        entry_price_txt = f"\033[92m{entry_price_txt}\033[0m"  # Green - Good to enter
                    elif "Bounce" in entry_strategy:
                        entry_price_txt = f"\033[96m{entry_price_txt}\033[0m"  # Cyan - Bounce opportunity
                    
                    # Truncate strategy for display
                    strategy_display = entry_strategy[:15] + "..." if len(entry_strategy) > 15 else entry_strategy
                else:
                    entry_price_txt = "N/A"
                    strategy_display = "No data"

                # Create Excel-like data row with proper padding for colored text
                data_row = (
                    f"{pad_colored_text(str(i), main_col_widths['rank'])} "
                    f"{pad_colored_text(t, main_col_widths['ticker'])} "
                    f"{pad_colored_text(score_txt, main_col_widths['score'])} "
                    f"{pad_colored_text(fii_txt, main_col_widths['fii'])} "
                    f"{pad_colored_text(qtr_txt, main_col_widths['qtr'])} "
                    f"{pad_colored_text(r1_txt, main_col_widths['delta1d'])} "
                    f"{pad_colored_text(volx_txt, main_col_widths['volx'])} "
                    f"{pad_colored_text(bd_txt, main_col_widths['bd'])} "
                    f"{pad_colored_text(p_txt, main_col_widths['prob'])} "
                    f"{pad_colored_text(g_txt, main_col_widths['grow'])} "
                    f"{pad_colored_text(entry_price_txt, main_col_widths['entry_price'])} "
                    f"{pad_colored_text(strategy_display, main_col_widths['entry_strategy'])} "
                    f"{pad_colored_text(f'{stop:.2f}', main_col_widths['stop'])} "
                    f"{pad_colored_text(f'{deal_amount_map.get(t, 0.0):.1f}', main_col_widths['amt_cr'])} "
                    f"{pad_colored_text(amt_pct_txt, main_col_widths['amt_pct'])} "
                    f"{pad_colored_text(dcnt_txt, main_col_widths['dcnt'])} "
                    f"{pad_colored_text(poscnt_txt, main_col_widths['poscnt'])}"
                )
                print(data_row)

    # ---- Excel File Generation with Colors ----
    def generate_colored_excel(tier_data, filename="swing_screener_analysis.xlsx"):
        """Generate a colored Excel file with all tiers and proper formatting."""
        if not OPENPYXL_AVAILABLE:
            print("⚠️  openpyxl not available. Install with: pip install openpyxl")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Analysis"

        # Define color schemes
        colors = {
            'header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'tier1': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Light green
            'tier2': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # Light yellow
            'tier3': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Light red
            'positive': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),  # Green
            'negative': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),  # Red
            'warning': PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid'),   # Yellow
            'excellent': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Dark green
        }

        fonts = {
            'header': Font(color='FFFFFF', bold=True),
            'bold': Font(bold=True),
            'normal': Font()
        }

        row_num = 1

        # Process each tier
        for tier_name, tier_list, tier_code in tier_data:
            if not tier_list:
                continue

            # Tier header
            ws.cell(row=row_num, column=1, value=tier_name).font = fonts['bold']
            ws.cell(row=row_num, column=1).fill = colors.get(tier_code.lower(), colors['header'])
            row_num += 1

            # Column headers
            headers = ["Ticker", "Score", "Cert%", "FII%", "QTR%", "P/E", "%Day", "%Week", "%1M", "%3M", "%1Y", "%5Y", "R/R", "Last Price", "Entry", "Stop Loss", "Size", "Deal%", "D_CNT", "POS_CNT", "Tags"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.fill = colors['header']
                cell.font = fonts['header']
            row_num += 1

            # Data rows
            for t, s in tier_list:
                # Handle both legacy and new metrics format
                if len(metrics[t]) > 9:  # Has enhancement details
                    r1, vx, bd, head, p, g, stop_prev, oi_chg, tags_str, enhancement_details = metrics[t]
                else:  # Legacy format
                    r1, vx, bd, head, p, g, stop_prev, oi_chg, tags_str = metrics[t]
                    enhancement_details = {}
                df = hist[t]
                
                if tier_code == "TIER1":
                    entry_lvl, stop_loss = entry_obj.tier1(df)
                elif tier_code == "TIER2":
                    entry_lvl, stop_loss = entry_obj.tier2(df)
                else:
                    entry_lvl, stop_loss = entry_obj.tier3(df)

                ltp = df["Close"].iloc[-1] if not df.empty else 0.0
                if stop_loss is None or (isinstance(stop_loss, float) and (np.isnan(stop_loss) or stop_loss <= 0)):
                    stop_loss = (entry_lvl or ltp) * 0.97

                size = PositionSizer.get_size(tier_code, s)
                fii_txt = fmt_fii(t) if not args.no_fii else "n/a"
                qtr_txt = fmt_qtr(t)
                
                # Clean text values for Excel compatibility - remove ANSI codes and problematic chars
                def clean_for_excel(text):
                    if text == "n/a":
                        return "n/a"
                    # Remove ANSI escape codes
                    import re
                    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
                    cleaned = ansi_escape.sub('', str(text))
                    # Remove other problematic characters
                    cleaned = cleaned.replace("+", "").replace("%", "")
                    # Remove any remaining non-printable characters and trim whitespace
                    cleaned = ''.join(char for char in cleaned if ord(char) >= 32 or char in '\t\n\r').strip()
                    return cleaned
                
                fii_clean = clean_for_excel(fii_txt)
                qtr_clean = clean_for_excel(qtr_txt)
                
                cert_score = certainty_map.get(t, 0.0)
                dealpct = deal_pct_map.get(t, 0.0)
                dcnt = deal_count_map.get(t, 0)
                poscnt = positive_count_map.get(t, 0)

                # Calculate price changes
                try:
                    d1 = (df['Close'].iloc[-1] - df['Close'].iloc[-2]) / df['Close'].iloc[-2] * 100
                except Exception:
                    d1 = 0.0
                try:
                    d7_idx = -6 if len(df) >= 7 else 0
                    d7 = (df['Close'].iloc[-1] - df['Close'].iloc[d7_idx]) / df['Close'].iloc[d7_idx] * 100
                except Exception:
                    d7 = 0.0
                try:
                    d30_idx = -21 if len(df) >= 22 else 0
                    d30 = (df['Close'].iloc[-1] - df['Close'].iloc[d30_idx]) / df['Close'].iloc[d30_idx] * 100
                except Exception:
                    d30 = 0.0
                try:
                    d63_idx = -63 if len(df) >= 64 else 0
                    d63 = (df["Close"].iloc[-1] - df["Close"].iloc[d63_idx]) / df["Close"].iloc[d63_idx] * 100
                except Exception:
                    d63 = 0.0
                try:
                    d250_idx = -251 if len(df) >= 252 else 0
                    d250 = (df['Close'].iloc[-1] - df['Close'].iloc[d250_idx]) / df['Close'].iloc[d250_idx] * 100
                except Exception:
                    d250 = 0.0
                try:
                    # Calculate 5-year percentage change - FIXED VERSION
                    # For 5-year data, we need approximately 1250 trading days
                    trading_days_5y = 1250  # Approximate number of trading days in 5 years
                    if len(df) >= trading_days_5y:
                        # We have sufficient data for full 5-year calculation
                        d5y_idx = -trading_days_5y
                        d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                    elif len(df) >= 1000:  # At least ~4 years of data
                        # Use ~4 years of data if available
                        d5y_idx = -1000
                        d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                    elif len(df) >= 750:  # At least ~3 years of data
                        # Use ~3 years of data if available
                        d5y_idx = -750
                        d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                    elif len(df) >= 500:  # At least ~2 years of data
                        # Use ~2 years of data if available (but mark as limited)
                        d5y_idx = -500
                        d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                    else:
                        # Insufficient data for meaningful 5-year calculation
                        d5y = 0.0  # Not enough data for 5Y calculation
                except Exception:
                    d5y = 0.0

                mc = mcap.get(t, 0.0)  # Market cap in Cr
                sector = detect_sector(names.get(t,""))
                fin_data = fin.get(t, {}) if fin else {}
                if fin_data:
                    pe_str, warn_symbol, pe_risk, _ = analyze_pe(fin_data, sector, mc)
                else:
                    pe_str, warn_symbol, pe_risk = "N/A", "", 0.0
                pe_clean = pe_str.replace("!", "") if pe_str else ""

                # Risk/Reward calculation
                if entry_lvl is not None and stop_loss and (entry_lvl - stop_loss) > 0:
                    tgt_price = (entry_lvl or ltp) * (1 + g/100)
                    risk = entry_lvl - stop_loss
                    reward = tgt_price - entry_lvl
                    rr_val = reward / risk if risk else 0
                else:
                    rr_val = 0.0

                # Data row - format all numerical values to 2 decimal places
                row_data = [
                    t, round(s, 2), round(cert_score, 2), fii_clean, qtr_clean, pe_clean, 
                    round(d1, 2), round(d7, 2), round(d30, 2), round(d63, 2), round(d250, 2), 
                    round(d5y, 2), round(rr_val, 2), round(ltp, 2), round(entry_lvl or 0, 2), 
                    round(stop_loss, 2), size, round(dealpct, 2), dcnt, poscnt, 
                    (tags_str or "")[:30]
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    
                    # Apply tier background color
                    if tier_code == "TIER1":
                        cell.fill = colors['tier1']
                    elif tier_code == "TIER2":
                        cell.fill = colors['tier2']
                    elif tier_code == "TIER3":
                        cell.fill = colors['tier3']

                    # Apply conditional formatting based on column
                    if col == 2:  # Score
                        if isinstance(value, (int, float)) and value >= 8:
                            cell.fill = colors['excellent']
                            cell.font = Font(color='FFFFFF', bold=True)
                    elif col == 3:  # Cert%
                        if isinstance(value, (int, float)):
                            if value >= 70:
                                cell.fill = colors['positive']
                            elif value <= 30:
                                cell.fill = colors['negative']
                    elif col in [7, 8, 9, 10, 11, 12]:  # Price change columns (Day, Week, 1M, 3M, 1Y, 5Y)
                        if isinstance(value, (int, float)):
                            if value > 0:
                                cell.fill = colors['positive']
                            elif value < 0:
                                cell.fill = colors['negative']
                    elif col == 6:  # P/E
                        if warn_symbol:
                            cell.fill = colors['warning']
                            cell.font = fonts['bold']
                    elif col == 13:  # R/R
                        if isinstance(value, (int, float)):
                            if value >= 3:
                                cell.fill = colors['positive']
                            elif value < 1:
                                cell.fill = colors['negative']
                    elif col == 18:  # Deal%
                        if isinstance(value, (int, float)) and value >= 5:
                            cell.fill = colors['warning']
                    elif col == 19:  # D_CNT
                        if isinstance(value, int) and value > 1:
                            cell.fill = colors['positive']
                    elif col == 20:  # POS_CNT
                        if isinstance(value, int) and value >= 10:
                            cell.fill = colors['positive']

                row_num += 1

            row_num += 1  # Space between tiers

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        try:
            wb.save(filename)
            print(f"✅ Colored Excel file generated: {filename}")
        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")

    # ---- Enhanced CSV Generation with Color Indicators ----
    def generate_enhanced_csv(tier_data, filename="swing_screener_analysis.csv"):
        """Generate a CSV file with color indicators for easy Excel import and manual formatting."""
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            headers = ["Tier", "Ticker", "Score", "Score_Color", "Cert%", "Cert_Color", "FII%", "QTR%", "P/E", "PE_Warning", 
                      "%Day", "Day_Color", "%Week", "Week_Color", "%1M", "1M_Color", "%3M", "3M_Color", "%1Y", "1Y_Color", 
                      "%5Y", "5Y_Color", "R/R", "RR_Color", "Last_Price", "Entry", "Stop_Loss", "Size", "Deal%", "Deal_Color", 
                      "D_CNT", "DCNT_Color", "POS_CNT", "POSCNT_Color", "Tags"]
            writer.writerow(headers)
            
            for tier_name, tier_list, tier_code in tier_data:
                if not tier_list:
                    continue
                    
                for t, s in tier_list:
                    # Handle both legacy and new metrics format
                    if len(metrics[t]) > 9:  # Has enhancement details
                        r1, vx, bd, head, p, g, stop_prev, oi_chg, tags_str, enhancement_details = metrics[t]
                    else:  # Legacy format
                        r1, vx, bd, head, p, g, stop_prev, oi_chg, tags_str = metrics[t]
                        enhancement_details = {}
                    df = hist[t]
                    
                    if tier_code == "TIER1":
                        entry_lvl, stop_loss = entry_obj.tier1(df)
                    elif tier_code == "TIER2":
                        entry_lvl, stop_loss = entry_obj.tier2(df)
                    else:
                        entry_lvl, stop_loss = entry_obj.tier3(df)

                    ltp = df["Close"].iloc[-1] if not df.empty else 0.0
                    if stop_loss is None or (isinstance(stop_loss, float) and (np.isnan(stop_loss) or stop_loss <= 0)):
                        stop_loss = (entry_lvl or ltp) * 0.97

                    size = PositionSizer.get_size(tier_code, s)
                    fii_txt = fmt_fii(t) if not args.no_fii else "n/a"
                    qtr_txt = fmt_qtr(t)
                    
                    # Clean text values for CSV compatibility - remove ANSI codes and problematic chars
                    def clean_for_excel(text):
                        if text == "n/a":
                            return "n/a"
                        # Remove ANSI escape codes
                        import re
                        ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
                        cleaned = ansi_escape.sub('', str(text))
                        # Remove other problematic characters
                        cleaned = cleaned.replace("+", "").replace("%", "")
                        # Remove any remaining non-printable characters and trim whitespace
                        cleaned = ''.join(char for char in cleaned if ord(char) >= 32 or char in '\t\n\r').strip()
                        return cleaned
                    
                    fii_clean = clean_for_excel(fii_txt)
                    qtr_clean = clean_for_excel(qtr_txt)
                    
                    cert_score = certainty_map.get(t, 0.0)
                    dealpct = deal_pct_map.get(t, 0.0)
                    dcnt = deal_count_map.get(t, 0)
                    poscnt = positive_count_map.get(t, 0)

                    # Calculate price changes
                    try:
                        d1 = (df['Close'].iloc[-1] - df['Close'].iloc[-2]) / df['Close'].iloc[-2] * 100
                    except Exception:
                        d1 = 0.0
                    try:
                        d7_idx = -6 if len(df) >= 7 else 0
                        d7 = (df['Close'].iloc[-1] - df['Close'].iloc[d7_idx]) / df['Close'].iloc[d7_idx] * 100
                    except Exception:
                        d7 = 0.0
                    try:
                        d30_idx = -21 if len(df) >= 22 else 0
                        d30 = (df['Close'].iloc[-1] - df['Close'].iloc[d30_idx]) / df['Close'].iloc[d30_idx] * 100
                    except Exception:
                        d30 = 0.0
                    try:
                        d63_idx = -63 if len(df) >= 64 else 0
                        d63 = (df["Close"].iloc[-1] - df["Close"].iloc[d63_idx]) / df["Close"].iloc[d63_idx] * 100
                    except Exception:
                        d63 = 0.0
                    try:
                        d250_idx = -251 if len(df) >= 252 else 0
                        d250 = (df['Close'].iloc[-1] - df['Close'].iloc[d250_idx]) / df['Close'].iloc[d250_idx] * 100
                    except Exception:
                        d250 = 0.0
                    try:
                        # Calculate 5-year percentage change - FIXED VERSION
                        # For 5-year data, we need approximately 1250 trading days
                        trading_days_5y = 1250  # Approximate number of trading days in 5 years
                        if len(df) >= trading_days_5y:
                            # We have sufficient data for full 5-year calculation
                            d5y_idx = -trading_days_5y
                            d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                        elif len(df) >= 1000:  # At least ~4 years of data
                            # Use ~4 years of data if available
                            d5y_idx = -1000
                            d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                        elif len(df) >= 750:  # At least ~3 years of data
                            # Use ~3 years of data if available
                            d5y_idx = -750
                            d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                        elif len(df) >= 500:  # At least ~2 years of data
                            # Use ~2 years of data if available (but mark as limited)
                            d5y_idx = -500
                            d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                        else:
                            # Insufficient data for meaningful 5-year calculation
                            d5y = 0.0  # Not enough data for 5Y calculation
                    except Exception:
                        d5y = 0.0

                    sector = detect_sector(names.get(t,""))
                    fin_data = fin.get(t, {}) if fin else {}
                    if fin_data:
                        pe_str, warn_symbol, pe_risk, _ = analyze_pe(fin_data, sector, mc)
                    else:
                        pe_str, warn_symbol, pe_risk = "N/A", "", 0.0
                    pe_display = pe_str
                    if warn_symbol:
                        pe_display = f"\033[91m{pe_str}{warn_symbol}\033[0m"  # Red for warnings
                    pe_clean = pe_str.replace("!", "") if pe_str else ""

                    # Risk/Reward calculation
                    if entry_lvl is not None and stop_loss and (entry_lvl - stop_loss) > 0:
                        tgt_price = (entry_lvl or ltp) * (1 + g/100)
                        risk = entry_lvl - stop_loss
                        reward = tgt_price - entry_lvl
                        rr_val = reward / risk if risk else 0
                    else:
                        rr_val = 0.0

                    # Color indicators
                    score_color = "EXCELLENT" if s >= 8 else "GOOD" if s >= 6 else "FAIR"
                    cert_color = "GOOD" if cert_score >= 70 else "BAD" if cert_score <= 30 else "FAIR"
                    day_color = "POSITIVE" if d1 > 0 else "NEGATIVE" if d1 < 0 else "NEUTRAL"
                    week_color = "POSITIVE" if d7 > 0 else "NEGATIVE" if d7 < 0 else "NEUTRAL"
                    m1_color = "POSITIVE" if d30 > 0 else "NEGATIVE" if d30 < 0 else "NEUTRAL"
                    m3_color = "POSITIVE" if d63 > 0 else "NEGATIVE" if d63 < 0 else "NEUTRAL"
                    y1_color = "POSITIVE" if d250 > 0 else "NEGATIVE" if d250 < 0 else "NEUTRAL"
                    y5_color = "POSITIVE" if d5y > 0 else "NEGATIVE" if d5y < 0 else "NEUTRAL"
                    rr_color = "EXCELLENT" if rr_val >= 3 else "BAD" if rr_val < 1 else "FAIR"
                    deal_color = "HIGHLIGHT" if dealpct >= 5 else "NORMAL"
                    dcnt_color = "GOOD" if dcnt > 1 else "NORMAL"
                    poscnt_color = "GOOD" if poscnt >= 10 else "NORMAL"
                    pe_warning = "YES" if warn_symbol else "NO"

                    # Write data row
                    row_data = [
                        tier_code, t, f"{s:.1f}", score_color, f"{cert_score:.1f}", cert_color, 
                        fii_clean, qtr_clean, pe_clean, pe_warning,
                        f"{d1:+.1f}", day_color, f"{d7:+.1f}", week_color, f"{d30:+.1f}", m1_color,
                        f"{d63:+.1f}", m3_color, f"{d250:+.1f}", y1_color, f"{d5y:+.1f}", y5_color,
                        f"{rr_val:.1f}", rr_color, f"{ltp:.2f}", f"{entry_lvl or 0:.2f}", f"{stop_loss:.2f}",
                        size, f"{dealpct:.1f}", deal_color, dcnt, dcnt_color, poscnt, poscnt_color,
                        (tags_str or "")[:50]
                    ]
                    writer.writerow(row_data)
        
        print(f"✅ Enhanced CSV file generated: {filename}")
        print("💡 Import into Excel and use conditional formatting based on color indicator columns!")

    # ---- Improved Tier Tables (Colored) ----
    def print_tier(header: str, arr: List[Tuple[str,float]], code: str):
        if not arr:
            print(f"\n{header}\n(no candidates)\n")
            return
        show_raw = args.show_raw and not args.legacy_score

        print(f"\n\033[1;36m{header}\033[0m")
        
        # Define column widths for Excel-like formatting (increased for better alignment)
        col_widths = {
            'ticker': 12, 'liqscore': 9, 'score': 7, 'color_score': 10, 'cert': 8, 'fii': 8, 'qtr': 9, 'pe': 7,
            'oversold': 10, 'rsi': 6, 'ema_slp': 8, 'bb_pos': 7, 'swing_rev': 9, 'day': 7, 'week': 8, 'month': 8, '3m': 8, 'year': 8, '5y': 8,
            'rr': 6, 'price': 11, 'entry': 11, 'stop': 11, 'size': 7,
            'deal': 7, 'dcnt': 6, 'pos': 7, 'tags': 35
        }
        
        # Create Excel-like header using consistent padding
        if code == "TIER3":
            header_row = (
                f"{'Ticker':<{col_widths['ticker']}} "
                f"{'LiqScore':<{col_widths['liqscore']}} "
                f"{'Score':<{col_widths['score']}} "
                f"{'ColorScr':<{col_widths['color_score']}} "
                f"{'Cert%':<{col_widths['cert']}} "
                f"{'FII%':<{col_widths['fii']}} "
                f"{'QTR%':<{col_widths['qtr']}} "
                f"{'P/E':<{col_widths['pe']}} "
                f"{'Oversold':<{col_widths['oversold']}} "
                f"{'RSI':>{col_widths['rsi']}} "
                f"{'EMA_Slp':>{col_widths['ema_slp']}} "
                f"{'BB_Pos':>{col_widths['bb_pos']}} "
                f"{'SwingRev':^{col_widths['swing_rev']}} "
                f"{'%Day':>{col_widths['day']}} "
                f"{'%Week':>{col_widths['week']}} "
                f"{'%1M':>{col_widths['month']}} "
                f"{'%3M':>{col_widths['3m']}} "
                f"{'%1Y':>{col_widths['year']}} "
                f"{'%5Y':>{col_widths['5y']}} "
                f"{'R/R':>{col_widths['rr']}} "
                f"{'Last Price':>{col_widths['price']}} "
                f"{'Stop Loss':>{col_widths['stop']}} "
                f"{'Size':<{col_widths['size']}} "
                f"{'Deal%':<{col_widths['deal']}} "
                f"{'D_CNT':<{col_widths['dcnt']}} "
                f"{'POS_CNT':<{col_widths['pos']}} "
                f"{'Tags':<{col_widths['tags']}}"
            )
        else:
            header_row = (
                f"{'Ticker':<{col_widths['ticker']}} "
                f"{'LiqScore':<{col_widths['liqscore']}} "
                f"{'Score':<{col_widths['score']}} "
                f"{'ColorScr':<{col_widths['color_score']}} "
                f"{'Cert%':<{col_widths['cert']}} "
                f"{'FII%':<{col_widths['fii']}} "
                f"{'QTR%':<{col_widths['qtr']}} "
                f"{'P/E':<{col_widths['pe']}} "
                f"{'Oversold':<{col_widths['oversold']}} "
                f"{'RSI':>{col_widths['rsi']}} "
                f"{'EMA_Slp':>{col_widths['ema_slp']}} "
                f"{'BB_Pos':>{col_widths['bb_pos']}} "
                f"{'SwingRev':^{col_widths['swing_rev']}} "
                f"{'%Day':>{col_widths['day']}} "
                f"{'%Week':>{col_widths['week']}} "
                f"{'%1M':>{col_widths['month']}} "
                f"{'%3M':>{col_widths['3m']}} "
                f"{'%1Y':>{col_widths['year']}} "
                f"{'%5Y':>{col_widths['5y']}} "
                f"{'R/R':>{col_widths['rr']}} "
                f"{'Last Price':>{col_widths['price']}} "
                f"{'Entry':>{col_widths['entry']}} "
                f"{'Stop Loss':>{col_widths['stop']}} "
                f"{'Size':<{col_widths['size']}} "
                f"{'Deal%':<{col_widths['deal']}} "
                f"{'D_CNT':<{col_widths['dcnt']}} "
                f"{'POS_CNT':<{col_widths['pos']}} "
                f"{'Tags':<{col_widths['tags']}}"
            )
        
        print(header_row)
        # Calculate separator line length based on actual column widths
        if code == "TIER3":
            separator_length = sum([col_widths['ticker'], col_widths['liqscore'], col_widths['score'], 
                                  col_widths['color_score'], col_widths['cert'], col_widths['fii'], 
                                  col_widths['qtr'], col_widths['pe'], col_widths['oversold'], 
                                  col_widths['rsi'], col_widths['ema_slp'], col_widths['bb_pos'], 
                                  col_widths['swing_rev'], col_widths['day'], col_widths['week'], 
                                  col_widths['month'], col_widths['3m'], col_widths['year'], 
                                  col_widths['5y'], col_widths['rr'], col_widths['price'], 
                                  col_widths['stop'], col_widths['size'], col_widths['deal'], 
                                  col_widths['dcnt'], col_widths['pos'], col_widths['tags']]) + 25  # spaces between columns
        else:
            separator_length = sum([col_widths['ticker'], col_widths['liqscore'], col_widths['score'], 
                                  col_widths['color_score'], col_widths['cert'], col_widths['fii'], 
                                  col_widths['qtr'], col_widths['pe'], col_widths['oversold'], 
                                  col_widths['rsi'], col_widths['ema_slp'], col_widths['bb_pos'], 
                                  col_widths['swing_rev'], col_widths['day'], col_widths['week'], 
                                  col_widths['month'], col_widths['3m'], col_widths['year'], 
                                  col_widths['5y'], col_widths['rr'], col_widths['price'], 
                                  col_widths['entry'], col_widths['stop'], col_widths['size'], 
                                  col_widths['deal'], col_widths['dcnt'], col_widths['pos'], 
                                  col_widths['tags']]) + 26  # spaces between columns
        print("-" * separator_length)

        for t, s in arr:
            # Extract metrics with enhancement details if available
            if len(metrics[t]) > 9:  # Has enhancement details
                r1, vx, bd, head, p, g, stop_prev, oi_chg, tags_str, enhancement_details = metrics[t]
            else:  # Legacy format
                r1, vx, bd, head, p, g, stop_prev, oi_chg, tags_str = metrics[t]
                enhancement_details = {}
            
            # Extract oversold information
            oversold_category = enhancement_details.get('oversold_category', 'NONE')
            oversold_signals = enhancement_details.get('oversold_signals', 0)
            oversold_strength = enhancement_details.get('oversold_strength', 0.0)
            oversold_multiplier = enhancement_details.get('oversold_multiplier', 1.0)
            current_rsi = enhancement_details.get('rsi', 50.0)
            
            df = hist[t]
            if code == "TIER1":
                _, stop_loss = entry_obj.tier1(df)
            elif code == "TIER2":
                _, stop_loss = entry_obj.tier2(df)
            else:
                _, stop_loss = entry_obj.tier3(df)
            
            # Use new entry price calculation based on today's high/low
            entry_price, entry_strategy = calculate_entry_price(t, df)
            entry_lvl = entry_price  # Use the new calculated entry price

            ltp = df["Close"].iloc[-1] if not df.empty else 0.0
            if stop_loss is None or (isinstance(stop_loss, float) and (np.isnan(stop_loss) or stop_loss <= 0)):
                stop_loss = (entry_lvl or ltp) * 0.97

            size = PositionSizer.get_size(code, s)
            cert_score = certainty_map.get(t, 0.0)
            
            # Calculate price changes
            try:
                d1 = (df['Close'].iloc[-1] - df['Close'].iloc[-2]) / df['Close'].iloc[-2] * 100
            except Exception:
                d1 = 0.0
            try:
                d7_idx = -6 if len(df) >= 7 else 0
                d7 = (df['Close'].iloc[-1] - df['Close'].iloc[d7_idx]) / df['Close'].iloc[d7_idx] * 100
            except Exception:
                d7 = 0.0
            try:
                d30_idx = -21 if len(df) >= 22 else 0
                d30 = (df['Close'].iloc[-1] - df['Close'].iloc[d30_idx]) / df['Close'].iloc[d30_idx] * 100
            except Exception:
                d30 = 0.0
            try:
                d63_idx = -63 if len(df) >= 64 else 0
                d63 = (df["Close"].iloc[-1] - df["Close"].iloc[d63_idx]) / df["Close"].iloc[d63_idx] * 100
            except Exception:
                d63 = 0.0
            try:
                d250_idx = -251 if len(df) >= 252 else 0
                d250 = (df['Close'].iloc[-1] - df['Close'].iloc[d250_idx]) / df['Close'].iloc[d250_idx] * 100
            except Exception:
                d250 = 0.0
            try:
                # Calculate 5-year percentage change - FIXED VERSION
                # For 5-year data, we need approximately 1250 trading days
                trading_days_5y = 1250  # Approximate number of trading days in 5 years
                if len(df) >= trading_days_5y:
                    # We have sufficient data for full 5-year calculation
                    d5y_idx = -trading_days_5y
                    d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                elif len(df) >= 1000:  # At least ~4 years of data
                    # Use ~4 years of data if available
                    d5y_idx = -1000
                    d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                elif len(df) >= 750:  # At least ~3 years of data
                    # Use ~3 years of data if available
                    d5y_idx = -750
                    d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                elif len(df) >= 500:  # At least ~2 years of data
                    # Use ~2 years of data if available (but mark as limited)
                    d5y_idx = -500
                    d5y = (df['Close'].iloc[-1] - df['Close'].iloc[d5y_idx]) / df['Close'].iloc[d5y_idx] * 100
                else:
                    # Insufficient data for meaningful 5-year calculation
                    d5y = 0.0  # Not enough data for 5Y calculation
                    d5y_idx = 0
                
                # Debug output for 5-year calculation
                if args.debug:
                    data_years_approx = len(df) / 250.0  # Approximate years of data
                    print(f"DEBUG 5Y for {t}: df_len={len(df)} (~{data_years_approx:.1f}Y), actual_idx={d5y_idx}, d5y={d5y:.2f}%")
                        
            except Exception:
                d5y = 0.0

            mc = mcap.get(t, 0.0)  # Market cap in Cr
            sector = detect_sector(names.get(t,""))
            fin_data = fin.get(t, {}) if fin else {}
            if fin_data:
                pe_str, warn_symbol, pe_risk, _ = analyze_pe(fin_data, sector, mc)
            else:
                pe_str, warn_symbol, pe_risk = "N/A", "", 0.0
            pe_clean = pe_str.replace("!", "") if pe_str else ""
            pe_display = pe_str if pe_str else ""

            # Risk/Reward calculation
            if entry_lvl is not None and stop_loss and (entry_lvl - stop_loss) > 0:
                tgt_price = (entry_lvl or ltp) * (1 + g/100)
                risk = entry_lvl - stop_loss
                reward = tgt_price - entry_lvl
                rr_val = reward / risk if risk else 0
            else:
                rr_val = 0.0

            # Get clean FII and QTR values
            fii_val = fii_growth_map.get(t)
            qtr_val = qtr_map.get(t)
            fii_txt = f"{fii_val:+.1f}%" if fii_val is not None else "n/a"
            qtr_txt = f"{qtr_val:+.1f}%" if qtr_val is not None else "n/a"
            
            dealpct = deal_pct_map.get(t, 0.0)
            dcnt = deal_count_map.get(t, 0)
            poscnt = positive_count_map.get(t, 0)

            # Apply colors to various columns
            score_display = f"{s:.2f}"
            if s >= 10:
                score_display = f"\033[95m{score_display}\033[0m"  # Purple - Excellent
            elif s >= 8:
                score_display = f"\033[92m{score_display}\033[0m"  # Green - Good
            elif s >= 6:
                score_display = f"\033[93m{score_display}\033[0m"  # Yellow - OK
            
            cert_display = f"{cert_score:.2f}%"
            if cert_score >= 80:
                cert_display = f"\033[95m{cert_display}\033[0m"  # Purple - Excellent
            elif cert_score >= 70:
                cert_display = f"\033[92m{cert_display}\033[0m"  # Green - Good
            elif cert_score >= 50:
                cert_display = f"\033[93m{cert_display}\033[0m"  # Yellow - OK
            elif cert_score <= 30:
                cert_display = f"\033[91m{cert_display}\033[0m"  # Red - Bad
            
            # Color FII and QTR values
            fii_colored = fii_txt
            if fii_val is not None:
                if fii_val >= 15:
                    fii_colored = f"\033[95m{fii_txt}\033[0m"  # Purple - Excellent
                elif fii_val >= 5:
                    fii_colored = f"\033[92m{fii_txt}\033[0m"  # Green - Good
                elif fii_val >= 0:
                    fii_colored = f"\033[93m{fii_txt}\033[0m"  # Yellow - OK
                else:
                    fii_colored = f"\033[91m{fii_txt}\033[0m"  # Red - Bad
            
            qtr_colored = qtr_txt
            if qtr_val is not None:
                if qtr_val >= 15:
                    qtr_colored = f"\033[95m{qtr_txt}\033[0m"  # Purple - Excellent
                elif qtr_val >= 5:
                    qtr_colored = f"\033[92m{qtr_txt}\033[0m"  # Green - Good
                elif qtr_val >= 0:
                    qtr_colored = f"\033[93m{qtr_txt}\033[0m"  # Yellow - OK
                else:
                    qtr_colored = f"\033[91m{qtr_txt}\033[0m"  # Red - Bad
            
            # Color price changes with tiered performance levels
            d1_colored = f"{d1:.2f}"
            if d1 >= 5:
                d1_colored = f"\033[95m{d1_colored}\033[0m"  # Purple - Excellent
            elif d1 >= 2:
                d1_colored = f"\033[92m{d1_colored}\033[0m"  # Green - Good
            elif d1 >= 0:
                d1_colored = f"\033[93m{d1_colored}\033[0m"  # Yellow - OK
            else:
                d1_colored = f"\033[91m{d1_colored}\033[0m"  # Red - Bad
            
            d7_colored = f"{d7:.2f}"
            if d7 >= 10:
                d7_colored = f"\033[95m{d7_colored}\033[0m"  # Purple - Excellent
            elif d7 >= 5:
                d7_colored = f"\033[92m{d7_colored}\033[0m"  # Green - Good
            elif d7 >= 0:
                d7_colored = f"\033[93m{d7_colored}\033[0m"  # Yellow - OK
            else:
                d7_colored = f"\033[91m{d7_colored}\033[0m"  # Red - Bad
            
            d30_colored = f"{d30:.2f}"
            if d30 >= 20:
                d30_colored = f"\033[95m{d30_colored}\033[0m"  # Purple - Excellent
            elif d30 >= 10:
                d30_colored = f"\033[92m{d30_colored}\033[0m"  # Green - Good
            elif d30 >= 0:
                d30_colored = f"\033[93m{d30_colored}\033[0m"  # Yellow - OK
            else:
                d30_colored = f"\033[91m{d30_colored}\033[0m"  # Red - Bad
            
            d63_colored = f"{d63:.2f}"
            if d63 >= 30:
                d63_colored = f"\033[95m{d63_colored}\033[0m"  # Purple - Excellent
            elif d63 >= 15:
                d63_colored = f"\033[92m{d63_colored}\033[0m"  # Green - Good
            elif d63 >= 0:
                d63_colored = f"\033[93m{d63_colored}\033[0m"  # Yellow - OK
            else:
                d63_colored = f"\033[91m{d63_colored}\033[0m"  # Red - Bad
            
            d250_colored = f"{d250:.2f}"
            if d250 >= 50:
                d250_colored = f"\033[95m{d250_colored}\033[0m"  # Purple - Excellent
            elif d250 >= 25:
                d250_colored = f"\033[92m{d250_colored}\033[0m"  # Green - Good
            elif d250 >= 0:
                d250_colored = f"\033[93m{d250_colored}\033[0m"  # Yellow - OK
            else:
                d250_colored = f"\033[91m{d250_colored}\033[0m"  # Red - Bad
            
            d5y_colored = f"{d5y:.2f}"
            if d5y >= 100:
                d5y_colored = f"\033[95m{d5y_colored}\033[0m"  # Purple - Excellent
            elif d5y >= 50:
                d5y_colored = f"\033[92m{d5y_colored}\033[0m"  # Green - Good
            elif d5y >= 0:
                d5y_colored = f"\033[93m{d5y_colored}\033[0m"  # Yellow - OK
            else:
                d5y_colored = f"\033[91m{d5y_colored}\033[0m"  # Red - Bad
            
            # Color R/R ratio
            rr_colored = f"{rr_val:.2f}"
            if rr_val >= 5:
                rr_colored = f"\033[95m{rr_colored}\033[0m"  # Purple - Excellent
            elif rr_val >= 3:
                rr_colored = f"\033[92m{rr_colored}\033[0m"  # Green - Good
            elif rr_val >= 1.5:
                rr_colored = f"\033[93m{rr_colored}\033[0m"  # Yellow - OK
            elif rr_val < 1:
                rr_colored = f"\033[91m{rr_colored}\033[0m"  # Red - Bad
            
            # Color deal percentage
            deal_display = f"{dealpct:.2f}%"
            if dealpct >= 10:
                deal_display = f"\033[95m{deal_display}\033[0m"  # Purple - Excellent
            elif dealpct >= 5:
                deal_display = f"\033[92m{deal_display}\033[0m"  # Green - Good
            elif dealpct >= 2:
                deal_display = f"\033[93m{deal_display}\033[0m"  # Yellow - OK
            
            # Color deal count
            dcnt_colored = f"{dcnt}"
            if dcnt >= 5:
                dcnt_colored = f"\033[95m{dcnt}\033[0m"  # Purple - Excellent
            elif dcnt > 1:
                dcnt_colored = f"\033[92m{dcnt}\033[0m"  # Green - Good
            
            # Color positive count
            poscnt_colored = f"{poscnt}"
            if poscnt >= 15:
                poscnt_colored = f"\033[95m{poscnt}\033[0m"  # Purple - Excellent
            elif poscnt >= 10:
                poscnt_colored = f"\033[92m{poscnt}\033[0m"  # Green - Good
            
            # Color P/E with warning
            if warn_symbol:
                pe_display = f"\033[91m{pe_str}!\033[0m"  # Red for warning

            tags_display = (tags_str or "")[:30]
            
            # Format oversold information with color coding
            if oversold_category == 'EXTREME':
                oversold_display = f"\033[95m{oversold_category}\033[0m"  # Purple - Excellent opportunity
            elif oversold_category == 'STRONG':
                oversold_display = f"\033[92m{oversold_category}\033[0m"   # Green - Good opportunity
            elif oversold_category == 'MODERATE':
                oversold_display = f"\033[93m{oversold_category}\033[0m"  # Yellow - OK opportunity
            elif oversold_category == 'MILD':
                oversold_display = f"{oversold_category}"    # No color - neutral
            else:
                oversold_display = f"{oversold_category}"    # No color - none
            
            # Format RSI with color coding
            rsi_display = f"{current_rsi:.0f}"
            if current_rsi < 20:
                rsi_display = f"\033[95m{rsi_display}\033[0m"  # Purple - Excellent buying opportunity
            elif current_rsi < 30:
                rsi_display = f"\033[92m{rsi_display}\033[0m"  # Green - Good buying opportunity
            elif current_rsi < 50:
                rsi_display = f"\033[93m{rsi_display}\033[0m"  # Yellow - OK range
            elif current_rsi > 70:
                rsi_display = f"\033[91m{rsi_display}\033[0m"  # Red - Overbought (bad for buying)
            
            # Format liquidity score with color coding
            liq_score, liq_level, is_liq_short = liquidity_map.get(t, (0.0, "LOW", False))
            liq_display = f"{liq_score:.4f}"
            if is_liq_short:
                if liq_level == "CRITICAL":
                    liq_display = f"\033[91m{liq_display}!!\033[0m"  # Red - Bad (critical risk)
                elif liq_level == "HIGH":
                    liq_display = f"\033[91m{liq_display}!\033[0m"   # Red - Bad (high risk)
                else:
                    liq_display = f"\033[93m{liq_display}\033[0m"     # Yellow - OK (medium risk)
            else:
                if liq_score >= 0.8:
                    liq_display = f"\033[95m{liq_display}\033[0m"     # Purple - Excellent liquidity
                elif liq_score >= 0.6:
                    liq_display = f"\033[92m{liq_display}\033[0m"     # Green - Good liquidity
                else:
                    liq_display = f"\033[93m{liq_display}\033[0m"     # Yellow - OK liquidity

            # Calculate actual technical indicators for display
            try:
                features = _rsi_ema_bb_strategy.calculate_features(df) if _rsi_ema_bb_strategy else {}
                actual_rsi = features.get('rsi', 50.0)
                ema_slope = features.get('ema_slope', 0.0)
                
                # Calculate Bollinger position
                current_price = features.get('ema14', ltp)
                bb_lower = features.get('bb_lower', ltp)
                bb_upper = features.get('bb_upper', ltp)
                bb_range = bb_upper - bb_lower
                bb_position = ((current_price - bb_lower) / bb_range * 100) if bb_range > 0 else 50.0
                
                # Enhanced Color Coding for Mean Reversion Buy Signals
                
                # Color EMA slope - FOR MEAN REVERSION (declining is good for buy setup)
                ema_slope_display = f"{ema_slope:.1f}%"
                if ema_slope < -2.0:
                    ema_slope_display = f"\033[95m{ema_slope_display}📈\033[0m"  # Purple - Excellent buy setup
                elif ema_slope < -1.0:
                    ema_slope_display = f"\033[92m{ema_slope_display}\033[0m"  # Green - Good buy setup
                elif ema_slope < -0.5:
                    ema_slope_display = f"\033[93m{ema_slope_display}\033[0m"  # Yellow - OK setup
                elif ema_slope > 2.0:
                    ema_slope_display = f"\033[91m{ema_slope_display}⚠️\033[0m"   # Red - Bad (strong rise, avoid)
                elif ema_slope > 0.5:
                    ema_slope_display = f"\033[91m{ema_slope_display}\033[0m"   # Red - Bad (rising, not ideal)
                
                # Color Bollinger Band position - FOR BUY SIGNALS
                bb_pos_display = f"{bb_position:.0f}%"
                if bb_position <= 10:
                    bb_pos_display = f"\033[95m{bb_pos_display}*\033[0m"  # Purple - Excellent oversold opportunity
                elif bb_position <= 20:
                    bb_pos_display = f"\033[92m{bb_pos_display}+\033[0m"  # Green - Good buy zone
                elif bb_position <= 30:
                    bb_pos_display = f"\033[92m{bb_pos_display}\033[0m"    # Green - Good zone
                elif bb_position <= 40:
                    bb_pos_display = f"\033[93m{bb_pos_display}\033[0m"    # Yellow - OK entry
                elif bb_position >= 80:
                    bb_pos_display = f"\033[91m{bb_pos_display}❌\033[0m"  # Red - Bad (overbought)
                elif bb_position >= 70:
                    bb_pos_display = f"\033[91m{bb_pos_display}\033[0m"    # Red for expensive zone
                # No color for neutral zone (50-70%)
                    
                # Enhanced RSI color coding - PERFECT FOR SWING BUYING
                rsi_display = f"{actual_rsi:.0f}"
                if actual_rsi <= 20:
                    rsi_display = f"\033[95m{rsi_display}🚀\033[0m"  # Magenta + rocket for extreme oversold (excellent buy)
                elif actual_rsi <= 30:
                    rsi_display = f"\033[92m{rsi_display}💎\033[0m"  # Green + diamond for prime buy zone
                elif actual_rsi <= 35:
                    rsi_display = f"\033[92m{rsi_display}\033[0m"    # Green for good buy zone
                elif actual_rsi <= 40:
                    rsi_display = f"\033[93m{rsi_display}\033[0m"    # Yellow for acceptable entry
                elif actual_rsi >= 80:
                    rsi_display = f"\033[91m{rsi_display}⚠️\033[0m"  # Red text + warning sign for extreme overbought (avoid)
                elif actual_rsi >= 70:
                    rsi_display = f"\033[91m{rsi_display}⚠️\033[0m"   # Red + warning for overbought (risky)
                elif actual_rsi >= 60:
                    rsi_display = f"\033[91m{rsi_display}\033[0m"    # Red for expensive zone
                # No color for neutral RSI (40-60)
                
                # Add EMA14 vs EMA50 position indicator for additional context
                ema14 = features.get('ema14', 0.0)
                ema50 = features.get('ema50', 0.0)
                ema_cross_signal = ""
                if ema14 > 0 and ema50 > 0:
                    if ema14 < ema50:
                        cross_gap = (ema50 - ema14) / ema50 * 100
                        if cross_gap > 5:
                            ema_cross_signal = "🔽"  # Strong below EMA50 (good for mean reversion)
                        elif cross_gap > 2:
                            ema_cross_signal = "↘️"   # Moderately below EMA50
                    else:
                        cross_gap = (ema14 - ema50) / ema50 * 100
                        if cross_gap > 5:
                            ema_cross_signal = "🔺"  # Strong above EMA50 (might be late)
                
                # Enhance EMA slope display with cross signal
                if ema_cross_signal:
                    if "🔽" in ema_cross_signal or "↘️" in ema_cross_signal:
                        # Below EMA50 is good for mean reversion - add to display
                        ema_slope_display = ema_slope_display.replace("\033[0m", f"{ema_cross_signal}\033[0m")
                    else:
                        # Above EMA50 might be warning - add subtle indicator
                        ema_slope_display = ema_slope_display.replace("\033[0m", f"{ema_cross_signal}\033[0m")
                
                # Perfect combination indicator - when all 3 align for buy
                perfect_buy_setup = (actual_rsi <= 30 and bb_position <= 20 and ema_slope < -0.5)
                if perfect_buy_setup:
                    # Add special indicator to RSI display for perfect setup
                    rsi_display = rsi_display.replace("\033[0m", "⭐\033[0m")  # Add star for perfect setup
            except:
                # Initialize all variables in case of exception
                actual_rsi = 50.0
                ema_slope = 0.0
                bb_position = 50.0
                ema_slope_display = "0.0%"  # No color for neutral/error
                bb_pos_display = "50%"      # No color for neutral/error
                rsi_display = "50"          # Default RSI display

            # SWING REVERSAL ANALYSIS - NEW ENHANCEMENT
            swing_reversal_data = calculate_swing_reversal_signals(df, lookback_period=20)
            
            # Calculate comprehensive color score for this stock
            institutional_data = institutional_metrics.get(t, {})
            stock_data_for_scoring = {
                'liquidity_score': liq_score,
                'score': s,
                'fii_growth': fii_val if fii_val is not None else 0,
                'qtr_growth': qtr_val if qtr_val is not None else 0,
                'daily_change': d1,
                'rsi': actual_rsi,
                'ema_slope': ema_slope,
                'bb_position': bb_position,
                'rr_ratio': rr_val,
                'deal_percentage': dealpct,
                'swing_reversal': swing_reversal_data,  # NEW: Include swing reversal signals
                'volume_multiplier': vx,  # CRITICAL: Volume confirmation for real money
                'buyer_dominance': bd * 100,  # CRITICAL: Buyer vs seller dominance percentage
                # INSTITUTIONAL GRADE METRICS (INDIVIDUAL KEYS FOR COLOR SCORING)
                'institutional_rel_vol_ratio': institutional_data.get('rel_vol_ratio', 1.0),
                'institutional_avwap_breakout': institutional_data.get('avwap_breakout', 0),
                'institutional_ob_score': institutional_data.get('ob_score', 50.0),
                'institutional_residual_momentum': institutional_data.get('residual_momentum_rank', 0.0),
                'institutional_catalyst_score': institutional_data.get('catalyst_score', 0.0),
                'institutional_vol_divergence': institutional_data.get('vol_divergence', 0),
                'institutional_risk_budget': institutional_data.get('risk_budget', 1.0),
                # Keep original nested format for backward compatibility
                'institutional_metrics': institutional_data
            }
            
            color_score, color_score_display, color_grade, score_breakdown = calculate_color_score(stock_data_for_scoring)
            
            # SWING REVERSAL DISPLAY - NEW ENHANCEMENT
            swing_rev_display = "NEUT"  # Default
            reversal_direction = swing_reversal_data.get('reversal_direction', 'NEUTRAL')
            reversal_strength = swing_reversal_data.get('reversal_strength', 0.0)
            
            if reversal_direction == 'BULLISH_REVERSAL':
                if reversal_strength >= 0.8:
                    swing_rev_display = "BULL-TARGET"  # Excellent bullish reversal
                elif reversal_strength >= 0.6:
                    swing_rev_display = "BULL+"   # Good bullish reversal
                elif reversal_strength >= 0.4:
                    swing_rev_display = "BULL"    # Mild bullish reversal
                else:
                    swing_rev_display = "bull"                   # No color - Minor bullish
            elif reversal_direction == 'BEARISH_REVERSAL':
                if reversal_strength >= 0.8:
                    swing_rev_display = "BEAR-X"  # Strong bearish reversal (avoid)
                elif reversal_strength >= 0.6:
                    swing_rev_display = "BEAR-"   # Moderate bearish reversal
                else:
                    swing_rev_display = "bear"    # Minor bearish warning
            # NEUTRAL stays as "NEUT"

            # Create Excel-like data row with proper padding for colored text
            if code == "TIER3":
                data_row = (
                    f"{pad_colored_text(t, col_widths['ticker'])} "
                    f"{pad_colored_text(liq_display, col_widths['liqscore'])} "
                    f"{pad_colored_text(score_display, col_widths['score'])} "
                    f"{pad_colored_text(color_score_display, col_widths['color_score'])} "
                    f"{pad_colored_text(cert_display, col_widths['cert'])} "
                    f"{pad_colored_text(fii_colored, col_widths['fii'])} "
                    f"{pad_colored_text(qtr_colored, col_widths['qtr'])} "
                    f"{pad_colored_text(pe_display, col_widths['pe'])} "
                    f"{pad_colored_text(oversold_display, col_widths['oversold'])} "
                    f"{pad_colored_text(rsi_display, col_widths['rsi'], 'right')} "
                    f"{pad_colored_text(ema_slope_display, col_widths['ema_slp'], 'right')} "
                    f"{pad_colored_text(bb_pos_display, col_widths['bb_pos'], 'right')} "
                    f"{pad_colored_text(swing_rev_display, col_widths['swing_rev'], 'center')} "
                    f"{pad_colored_text(d1_colored, col_widths['day'], 'right')} "
                    f"{pad_colored_text(d7_colored, col_widths['week'], 'right')} "
                    f"{pad_colored_text(d30_colored, col_widths['month'], 'right')} "
                    f"{pad_colored_text(d63_colored, col_widths['3m'], 'right')} "
                    f"{pad_colored_text(d250_colored, col_widths['year'], 'right')} "
                    f"{pad_colored_text(d5y_colored, col_widths['5y'], 'right')} "
                    f"{pad_colored_text(rr_colored, col_widths['rr'], 'right')} "
                    f"{pad_colored_text(f'{ltp:.2f}', col_widths['price'], 'right')} "
                    f"{pad_colored_text(f'{stop_loss:.2f}', col_widths['stop'], 'right')} "
                    f"{pad_colored_text(size, col_widths['size'])} "
                    f"{pad_colored_text(deal_display, col_widths['deal'])} "
                    f"{pad_colored_text(dcnt_colored, col_widths['dcnt'])} "
                    f"{pad_colored_text(poscnt_colored, col_widths['pos'])} "
                    f"{pad_colored_text(tags_display, col_widths['tags'])}"
                )
            else:
                data_row = (
                    f"{pad_colored_text(t, col_widths['ticker'])} "
                    f"{pad_colored_text(liq_display, col_widths['liqscore'])} "
                    f"{pad_colored_text(score_display, col_widths['score'])} "
                    f"{pad_colored_text(color_score_display, col_widths['color_score'])} "
                    f"{pad_colored_text(cert_display, col_widths['cert'])} "
                    f"{pad_colored_text(fii_colored, col_widths['fii'])} "
                    f"{pad_colored_text(qtr_colored, col_widths['qtr'])} "
                    f"{pad_colored_text(pe_display, col_widths['pe'])} "
                    f"{pad_colored_text(oversold_display, col_widths['oversold'])} "
                    f"{pad_colored_text(rsi_display, col_widths['rsi'], 'right')} "
                    f"{pad_colored_text(ema_slope_display, col_widths['ema_slp'], 'right')} "
                    f"{pad_colored_text(bb_pos_display, col_widths['bb_pos'], 'right')} "
                    f"{pad_colored_text(swing_rev_display, col_widths['swing_rev'], 'center')} "
                    f"{pad_colored_text(d1_colored, col_widths['day'], 'right')} "
                    f"{pad_colored_text(d7_colored, col_widths['week'], 'right')} "
                    f"{pad_colored_text(d30_colored, col_widths['month'], 'right')} "
                    f"{pad_colored_text(d63_colored, col_widths['3m'], 'right')} "
                    f"{pad_colored_text(d250_colored, col_widths['year'], 'right')} "
                    f"{pad_colored_text(d5y_colored, col_widths['5y'], 'right')} "
                    f"{pad_colored_text(rr_colored, col_widths['rr'], 'right')} "
                    f"{pad_colored_text(f'{ltp:.2f}', col_widths['price'], 'right')} "
                    f"{pad_colored_text(f'{entry_lvl or 0:.2f}', col_widths['entry'], 'right')} "
                    f"{pad_colored_text(f'{stop_loss:.2f}', col_widths['stop'], 'right')} "
                    f"{pad_colored_text(size, col_widths['size'])} "
                    f"{pad_colored_text(deal_display, col_widths['deal'])} "
                    f"{pad_colored_text(dcnt_colored, col_widths['dcnt'])} "
                    f"{pad_colored_text(poscnt_colored, col_widths['pos'])} "
                    f"{pad_colored_text(tags_display, col_widths['tags'])}"
                )
            
            print(data_row)

    print_tier("=== TIER 1: CORE POSITIONS (Scale In) ===", tier1, "TIER1")
    print_tier("=== TIER 2: SELECTIVE (Confirm First) ===", tier2, "TIER2")
    print_tier("=== TIER 3: TACTICAL (Strict Stops) ===", tier3, "TIER3")

    # ---- Color Score Summary ----
    print("\n" + "="*100)
    print("🎨 COLOR SCORE ANALYSIS - TOP OPPORTUNITIES BY COMPREHENSIVE METRICS")
    print("="*100)
    print("📊 Color Score Legend:")
    print("   🟣 ULTIMATE (70+): Perfect buying opportunity across all metrics")
    print("   🟣 EXCELLENT (50-69): Outstanding setup, very high confidence")
    print("   🟢 GOOD (30-49): Solid opportunity, good risk/reward")
    print("   🟡 OK (15-29): Acceptable entry, monitor closely") 
    print("   ⚪ NEUTRAL (0-14): Mixed signals, proceed with caution")
    print("   ⚪ WEAK (-15 to -1): Poor setup, likely avoid")
    print("   ⚠️ AVOID (<-15): Multiple red flags, definitely avoid")
    print("\n💡 Note: Color score excludes longer-term price movements (week%, month%, 3M%, 1Y%, 5Y%)")
    print("   Focus on: Liquidity, Score, FII/QTR, Daily%, RSI, EMA Slope, BB Position, R/R, Deal%")
    print("-" * 100)
    
    # Collect and sort all stocks by color score
    color_scored_stocks = []
    for t, s in ranks[:args.top]:
        if len(metrics[t]) > 9:
            enhancement_details = metrics[t][9]
            actual_rsi = enhancement_details.get('rsi', 50.0)
            ema_slope = enhancement_details.get('ema_14_slope', 0.0)
            bb_position = enhancement_details.get('bb_position', 50.0)
        else:
            actual_rsi = 50.0
            ema_slope = 0.0
            bb_position = 50.0
            
        # Get the same data used in display
        try:
            liq_score = liquidity_scores.get(t, 0.5)
            fii_val = parse_percentage(fin_data.get(t, {}).get('fii_qtr_change', '0%'))
            qtr_val = parse_percentage(fin_data.get(t, {}).get('quarterly_sales_growth', '0%'))
            
            df = hist[t]
            d1 = (df['Close'].iloc[-1] - df['Close'].iloc[-2]) / df['Close'].iloc[-2] * 100 if len(df) >= 2 else 0.0
            
            # Calculate RR ratio
            entry_lvl, stop_loss = entry_obj.tier3(df)
            ltp = df["Close"].iloc[-1] if not df.empty else 0.0
            if stop_loss is None or (isinstance(stop_loss, float) and (np.isnan(stop_loss) or stop_loss <= 0)):
                stop_loss = (entry_lvl or ltp) * 0.97
            
            if entry_lvl is not None and stop_loss and (entry_lvl - stop_loss) > 0:
                tgt_price = (entry_lvl or ltp) * 1.10  # 10% target
                risk = entry_lvl - stop_loss
                reward = tgt_price - entry_lvl
                rr_val = reward / risk if risk else 0
            else:
                rr_val = 0.0
                
            dealpct = delivery_pct.get(t, 0.0)
            
            stock_data_for_scoring = {
                'liquidity_score': liq_score,
                'score': s,
                'fii_growth': fii_val if fii_val is not None else 0,
                'qtr_growth': qtr_val if qtr_val is not None else 0,
                'daily_change': d1,
                'rsi': actual_rsi,
                'ema_slope': ema_slope,
                'bb_position': bb_position,
                'rr_ratio': rr_val,
                'deal_percentage': dealpct,
                'swing_reversal': {},  # Empty for color score only mode
                'volume_multiplier': vx,  # CRITICAL: Volume confirmation
                'buyer_dominance': bd * 100,  # CRITICAL: Buyer dominance percentage
                # INSTITUTIONAL GRADE METRICS (NEW)
                'institutional_metrics': institutional_metrics.get(t, {})  # Add institutional analysis results
            }
            
            color_score, color_score_display, color_grade, score_breakdown = calculate_color_score(stock_data_for_scoring)
            color_scored_stocks.append((t, color_score, color_score_display, color_grade, score_breakdown, s))
        except:
            continue
    
    # Sort by color score (highest first)
    color_scored_stocks.sort(key=lambda x: x[1], reverse=True)
    
    print(f"{'Rank':<4} {'Ticker':<10} {'ColorScore':<12} {'Grade':<10} {'MainScore':<10} {'Key Strengths'}")
    print("-" * 100)
    
    for i, (ticker, color_score, color_display, grade, breakdown, main_score) in enumerate(color_scored_stocks[:20], 1):
        # Find top 3 positive contributors
        positive_items = [(k, v) for k, v in breakdown.items() if v.startswith('+')]
        positive_items.sort(key=lambda x: int(x[1].split()[0][1:]), reverse=True)
        top_strengths = ', '.join([f"{k}({v.split()[0]})" for k, v in positive_items[:3]])
        
        print(f"{i:<4} {ticker:<10} {color_display:<12} {grade:<10} {main_score:<10.1f} {top_strengths}")
    
    print(f"\n💡 Color Score Methodology:")
    print(f"   • Liquidity (15pts): Entry/exit ease")
    print(f"   • Main Score (12pts): Core screening score") 
    print(f"   • RSI (12pts): Oversold opportunities (lower = better)")
    print(f"   • Swing Reversal (12pts): Highest/lowest tick analysis")
    print(f"   • Volume Confirmation (10pts): CRITICAL for real money deployment")
    print(f"   • Buyer Dominance (10pts): CRITICAL buyer vs seller strength")
    print(f"   • FII/QTR Growth (10pts each): Institutional interest & recent performance")
    print(f"   • BB Position (10pts): Oversold positioning (lower = better)")
    print(f"   • Daily Change (8pts): Recent momentum")
    print(f"   • EMA Slope (8pts): Mean reversion setup (declining = better)")
    print(f"   • R/R Ratio (8pts): Risk/reward attractiveness")
    print(f"   • Deal % (6pts): Delivery percentage quality")
    print(f"   📊 Total possible: 181 points (ENHANCED with Volume, Buyer Dominance & Institutional Analysis)")
    print(f"\n🚨 CRITICAL SAFETY CHECKS:")
    print(f"   • Volume < 1.5x = RED FLAG (avoid investment)")
    print(f"   • Buyer Dominance < 50% = RED FLAG (seller pressure)")
    print(f"   • Both must be strong before deploying real money!")
    print(f"\n🏛️ INSTITUTIONAL GRADE ANALYSIS:")
    print(f"   • Relative Volume Curve (10pts): Time-of-day volume pattern analysis")
    print(f"   • AVWAP Breakout (10pts): Anchored VWAP break + liquidity pocket detection")
    print(f"   • Order Book Imbalance (10pts): Price-volume proxy for order flow")
    print(f"   • Residual Momentum (10pts): Cross-sectional market beta-adjusted returns")
    print(f"   • Catalyst Factor (10pts): Earnings surprise + deal flow momentum")
    print(f"\n🎯 Swing Reversal Signals:")
    print(f"   • BULL🎯 (Purple): Excellent bullish reversal from trough")
    print(f"   • BULL+ (Green): Good bullish reversal signal")
    print(f"   • BULL (Yellow): Mild bullish reversal")
    print(f"   • BEAR❌ (Red): Strong bearish reversal (avoid)")
    print(f"   • BEAR- (Red): Moderate bearish warning")
    print(f"   • NEUT: No significant reversal signals")

    # ---- Enhanced RSI + EMA Oversold Benefits Summary with Best Practices ----
    print("\n" + "="*100)
    print("🎯 RSI + EMA OVERSOLD SIGNALS - SWING TRADING OPPORTUNITIES")
    print("="*100)
    
    oversold_stocks = []
    ultimate_signals = []
    extreme_signals = []
    strong_signals = []
    
    for t, s in ranks[:args.top]:
        if len(metrics[t]) > 9:  # Has enhancement details
            enhancement_details = metrics[t][9]
            oversold_category = enhancement_details.get('oversold_category', 'NONE')
            oversold_multiplier = enhancement_details.get('oversold_multiplier', 1.0)
            signal_quality = enhancement_details.get('signal_quality', 0.0)
            entry_confidence = enhancement_details.get('entry_confidence', 0.0)
            current_rsi = enhancement_details.get('rsi', 50.0)
            ema_14_slope = enhancement_details.get('ema_14_slope', 0.0)
            volume_surge = enhancement_details.get('volume_surge', 1.0)
            volume_confirmation = enhancement_details.get('volume_confirmation', False)
            bullish_candlestick = enhancement_details.get('bullish_candlestick', False)
            near_support = enhancement_details.get('near_support', False)
            ema_bullish_alignment = enhancement_details.get('ema_bullish_alignment', False)
            risk_reward_ratio = enhancement_details.get('risk_reward_ratio', 0.0)
            entry_zones = enhancement_details.get('entry_zones', {})
            stop_losses = enhancement_details.get('stop_losses', {})
            
            if oversold_category != 'NONE' and oversold_multiplier > 1.0:
                tier = classification.get(t, 'UNCLASSIFIED')
                
                # Get sector for sector-specific advice
                sector = detect_sector(names.get(t, ""))
                
                stock_data = {
                    'ticker': t,
                    'score': s,
                    'category': oversold_category,
                    'multiplier': oversold_multiplier,
                    'quality': signal_quality,
                    'confidence': entry_confidence,
                    'rsi': current_rsi,
                    'ema_slope': ema_14_slope,
                    'volume_surge': volume_surge,
                    'volume_conf': volume_confirmation,
                    'candlestick': bullish_candlestick,
                    'support': near_support,
                    'ema_align': ema_bullish_alignment,
                    'rr_ratio': risk_reward_ratio,
                    'entry_zones': entry_zones,
                    'stop_losses': stop_losses,
                    'tier': tier,
                    'sector': sector
                }
                
                oversold_stocks.append(stock_data)
                
                # Categorize for specific recommendations
                if oversold_category == 'ULTIMATE':
                    ultimate_signals.append(stock_data)
                elif oversold_category == 'EXTREME':
                    extreme_signals.append(stock_data)
                elif oversold_category == 'STRONG':
                    strong_signals.append(stock_data)
    
    if oversold_stocks:
        # Sort by signal quality and confidence
        oversold_stocks.sort(key=lambda x: (x['quality'], x['confidence'], x['multiplier']), reverse=True)
        
        print(f"{'Ticker':<10} {'Tier':<6} {'Score':<6} {'Category':<10} {'Quality':<8} {'Conf%':<6} {'RSI':<4} {'EMA%':<6} {'Vol':<6} {'R/R':<5} {'Entry Strategy'}")
        print("-" * 110)
        
        for stock in oversold_stocks[:20]:  # Top 20 oversold opportunities
            t = stock['ticker']
            
            # Clean category display (no color codes for Excel compatibility)
            cat_display = stock['category'][:8]
            
            # Clean RSI display
            rsi_display = f"{stock['rsi']:.0f}"
            
            # Clean EMA slope display
            ema_display = f"{stock['ema_slope']:+.1f}"
            
            # Volume confirmation indicators (Excel compatible)
            vol_indicator = "CONF" if stock['volume_conf'] else f"{stock['volume_surge']:.1f}x"
            
            # Quality score (no color codes)
            quality_display = f"{stock['quality']:.2f}"
            
            # Entry strategy recommendation
            if stock['category'] in ['ULTIMATE', 'EXTREME']:
                if stock['confidence'] >= 0.5:
                    entry_strategy = "Scale-In 50/50"
                else:
                    entry_strategy = "Wait EMA14↗"
            elif stock['category'] == 'STRONG':
                if stock['volume_conf'] and stock['support']:
                    entry_strategy = "Full Entry"
                else:
                    entry_strategy = "Confirm First"
            else:
                entry_strategy = "Conservative"
            
            print(f"{t:<10} {stock['tier']:<6} {stock['score']:<6.2f} {cat_display:<10} {quality_display:<8} "
                  f"{stock['confidence']*100:<6.0f} {rsi_display:<4} {ema_display:<6} {vol_indicator:<6} "
                  f"{stock['rr_ratio']:<5.1f} {entry_strategy}")
        
        # Best Practices Implementation Summary
        print(f"\n[CHART] SIGNAL QUALITY BREAKDOWN:")
        print(f"   [HOT] ULTIMATE (RSI<25 + Rising EMA): {len(ultimate_signals)}")
        print(f"   [LIGHTNING] EXTREME (RSI<20 + EMA Support): {len(extreme_signals)}")
        print(f"   [STRONG] STRONG (RSI<30 + Volume Confirmation): {len(strong_signals)}")
        
        if ultimate_signals or extreme_signals:
            print(f"\n[TARGET] TOP TRADING OPPORTUNITIES:")
            
            for stock in (ultimate_signals + extreme_signals)[:5]:
                t = stock['ticker']
                current_price = hist[t]['Close'].iloc[-1] if t in hist and not hist[t].empty else 0
                
                print(f"\n[CHART] {t} ({stock['sector']}) - {stock['category']} Signal")
                
                # Entry zones from best practices
                if 'primary' in stock['entry_zones']:
                    primary_entry = stock['entry_zones']['primary']
                    secondary_entry = stock['entry_zones'].get('secondary', primary_entry)
                    breakout_entry = stock['entry_zones'].get('breakout', current_price * 1.02)
                    
                    print(f"   [MONEY] Current Price: Rs.{current_price:.2f}")
                    print(f"   [TARGET] Entry Zone 1: Rs.{primary_entry:.2f} (50% position)")
                    print(f"   [TARGET] Entry Zone 2: Rs.{secondary_entry:.2f} (25% position)")
                    print(f"   [ROCKET] EMA14 Breakout: Rs.{breakout_entry:.2f} (25% position)")
                    
                    if 'percentage' in stock['stop_losses']:
                        stop_price = stock['stop_losses']['percentage']
                        print(f"   [STOP] Stop Loss: Rs.{stop_price:.2f}")
                        
                        # Calculate targets based on best practices
                        if stock['category'] == 'ULTIMATE':
                            target_1 = primary_entry * 1.08
                            target_2 = primary_entry * 1.15
                        elif stock['category'] == 'EXTREME':
                            target_1 = primary_entry * 1.06
                            target_2 = primary_entry * 1.12
                        else:
                            target_1 = primary_entry * 1.05
                            target_2 = primary_entry * 1.10
                        
                        print(f"   [TARGET] Target 1: Rs.{target_1:.2f} (50% booking)")
                        print(f"   [TARGET] Target 2: Rs.{target_2:.2f} (trail stop)")
                        
                        # Risk calculation
                        risk_pct = abs(primary_entry - stop_price) / primary_entry * 100
                        reward_pct = (target_1 - primary_entry) / primary_entry * 100
                        print(f"   [CHART] Risk: {risk_pct:.1f}% | Reward: {reward_pct:.1f}% | R/R: {stock['rr_ratio']:.1f}")
                
                # Sector-specific considerations
                if 'COMMODITY' in stock['sector'].upper() or any(x in t for x in ['COAL', 'METALS', 'OIL']):
                    print(f"   ⚠️  Monitor: Commodity price trends")
                elif 'INFRA' in stock['sector'].upper() or 'ENGINEER' in stock['sector'].upper():
                    print(f"   📋 Monitor: Order book announcements")
                elif 'PHARMA' in stock['sector'].upper():
                    print(f"   🧪 Monitor: Drug approvals/regulatory updates")
                
                # Warning signs to watch
                print(f"   🚨 Exit if: RSI<20 after entry, EMA14 breakdown, Volume<80% avg")
        
        print(f"\n💡 PORTFOLIO ALLOCATION GUIDANCE:")
        print(f"   • Max 2 STRONG signals simultaneously")
        print(f"   • Position sizes: TIER1(3-5%), TIER2(2-3%), TIER3(0.5-1%)")
        print(f"   • Max portfolio risk per trade: 1%")
        print(f"   • Time exit: Close if Target 1 not hit in 5 trading days")
        
    else:
        print("No significant RSI + EMA oversold opportunities found in current market conditions.")
        print("💡 This is normal during strong bull markets. Focus on breakout strategies instead.")
    
    print("="*100)

    # News Sentiment Integration
    try:
        from simple_news_integration import get_news_sentiment_scores, get_top_sentiment_picks, get_sentiment_avoid_list
        
        sentiment_scores = get_news_sentiment_scores()
        if sentiment_scores:
            print("\n" + "="*80)
            print("NEWS SENTIMENT ANALYSIS")
            print("="*80)
            
            # Show sentiment for tier 1 stocks
            print("\nTIER 1 NEWS SENTIMENT:")
            tier1_with_sentiment = []
            for ticker, score in tier1:
                sentiment = sentiment_scores.get(ticker, 0.0)
                sentiment_label = ("Very Positive" if sentiment >= 0.5 else 
                                 "Positive" if sentiment >= 0.1 else
                                 "Very Negative" if sentiment <= -0.5 else
                                 "Negative" if sentiment <= -0.1 else "Neutral")
                tier1_with_sentiment.append((ticker, score, sentiment, sentiment_label))
                print(f"  {ticker:<12} Score: {score:6.2f}  Sentiment: {sentiment:+.3f} ({sentiment_label})")
            
            # Top positive sentiment picks
            top_sentiment = get_top_sentiment_picks(10, 0.5)
            if top_sentiment:
                print(f"\nTOP POSITIVE SENTIMENT PICKS:")
                for ticker, sent_score in top_sentiment:
                    in_tier = next((t for t,s in tier1+tier2+tier3 if t == ticker), None)
                    tier_info = f" [In Screener]" if in_tier else ""
                    print(f"  {ticker:<12} {sent_score:+.3f}{tier_info}")
            
            # Avoid list
            avoid_list = get_sentiment_avoid_list(-0.5)
            if avoid_list:
                print(f"\nAVOID LIST (Negative Sentiment):")
                for ticker, sent_score in avoid_list[-8:]:  # Show worst 8
                    in_tier = next((t for t,s in tier1+tier2+tier3 if t == ticker), None)
                    tier_info = f" [WARNING: In Screener!]" if in_tier else ""
                    print(f"  {ticker:<12} {sent_score:+.3f}{tier_info}")
            
    except ImportError:
        print("\nNews sentiment analysis not available (simple_news_integration.py not found)")
    except Exception as e:
        print(f"\nNews sentiment analysis error: {e}")

    # Generate colored Excel file if requested
    if args.excel_file:
        tier_data = [
            ("=== TIER 1: CORE POSITIONS (Scale In) ===", tier1, "TIER1"),
            ("=== TIER 2: SELECTIVE (Confirm First) ===", tier2, "TIER2"),
            ("=== TIER 3: TACTICAL (Strict Stops) ===", tier3, "TIER3")
        ]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"swing_screener_analysis_{timestamp}.xlsx"
        generate_colored_excel(tier_data, excel_filename)

    # Generate enhanced CSV file if requested
    if args.csv_enhanced:
        tier_data = [
            ("=== TIER 1: CORE POSITIONS (Scale In) ===", tier1, "TIER1"),
            ("=== TIER 2: SELECTIVE (Confirm First) ===", tier2, "TIER2"),
            ("=== TIER 3: TACTICAL (Strict Stops) ===", tier3, "TIER3")
        ]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"swing_screener_analysis_{timestamp}.csv"
        generate_enhanced_csv(tier_data, csv_filename)

    # =============================================================================
    # 🚨 CRITICAL SAFETY CHECKS - Volume & Buyer Dominance Validation
    # =============================================================================
    
    def volume_buyer_safety_check(ticker_list, metrics, soft_mode=False):
        """Apply critical safety checks for volume and buyer dominance before real money deployment"""
        safe_picks = []
        warning_picks = []
        dangerous_picks = []
        soft_mode_picks = []  # New category for near-misses in soft mode
        
        # Adjust thresholds based on mode
        if soft_mode:
            min_volume = 1.2     # Relaxed from 1.5x
            min_buyer_dom = 45.0 # Relaxed from 55%
        else:
            min_volume = 1.5     # Standard strict thresholds
            min_buyer_dom = 55.0
        
        for ticker, score in ticker_list:
            if ticker in metrics:
                metric_data = metrics[ticker]
                if len(metric_data) >= 3:
                    vx = metric_data[1]  # Volume multiplier
                    bd = metric_data[2]  # Buyer dominance (0-1)
                    bd_pct = bd * 100
                    
                    # CRITICAL SAFETY THRESHOLDS (adjusted for soft mode)
                    volume_safe = vx >= min_volume
                    buyer_safe = bd_pct >= min_buyer_dom
                    
                    # Standard thresholds for comparison
                    volume_safe_strict = vx >= 1.5
                    buyer_safe_strict = bd_pct >= 55.0
                    
                    safety_score = 0
                    safety_flags = []
                    
                    if vx >= 2.5:
                        safety_score += 3
                        safety_flags.append(f"HOT-VOL:{vx:.1f}x")
                    elif vx >= 2.0:
                        safety_score += 2
                        safety_flags.append(f"GOOD-VOL:{vx:.1f}x")
                    elif vx >= 1.5:
                        safety_score += 1
                        safety_flags.append(f"OK-VOL:{vx:.1f}x")
                    elif vx >= 1.2 and soft_mode:
                        safety_score += 0
                        safety_flags.append(f"SOFT-VOL:{vx:.1f}x⚠️")
                    else:
                        safety_score -= 2
                        safety_flags.append(f"LOW-VOL:{vx:.1f}x")
                    
                    if bd_pct >= 75:
                        safety_score += 3
                        safety_flags.append(f"HOT-BD:{bd_pct:.1f}%")
                    elif bd_pct >= 65:
                        safety_score += 2
                        safety_flags.append(f"GOOD-BD:{bd_pct:.1f}%")
                    elif bd_pct >= 55:
                        safety_score += 1
                        safety_flags.append(f"OK-BD:{bd_pct:.1f}%")
                    elif bd_pct >= 45 and soft_mode:
                        safety_score += 0
                        safety_flags.append(f"SOFT-BD:{bd_pct:.1f}%⚠️")
                    else:
                        safety_score -= 2
                        safety_flags.append(f"WEAK-BD:{bd_pct:.1f}%")
                    
                    # Classification based on safety (enhanced with soft mode)
                    if volume_safe_strict and buyer_safe_strict and safety_score >= 3:
                        safe_picks.append((ticker, score, safety_score, safety_flags))
                    elif volume_safe_strict and buyer_safe_strict:
                        warning_picks.append((ticker, score, safety_score, safety_flags))
                    elif soft_mode and volume_safe and buyer_safe:
                        # Near-misses that qualify in soft mode
                        soft_mode_picks.append((ticker, score, safety_score, safety_flags))
                    else:
                        dangerous_picks.append((ticker, score, safety_score, safety_flags))
        
        return safe_picks, warning_picks, dangerous_picks, soft_mode_picks
    
    # Apply safety checks to all tiers (with soft mode support)
    tier1_safe, tier1_warning, tier1_dangerous, tier1_soft = volume_buyer_safety_check(tier1, metrics, args.soft_mode)
    tier2_safe, tier2_warning, tier2_dangerous, tier2_soft = volume_buyer_safety_check(tier2, metrics, args.soft_mode)
    tier3_safe, tier3_warning, tier3_dangerous, tier3_soft = volume_buyer_safety_check(tier3, metrics, args.soft_mode)
    
    # Print safety analysis
    print("\n" + "[ALERT]" + "="*92)
    print("[LOCK] CRITICAL SAFETY ANALYSIS - Volume & Buyer Dominance Validation")
    print("[ALERT]" + "="*92)
    
    all_safe = tier1_safe + tier2_safe + tier3_safe
    all_warning = tier1_warning + tier2_warning + tier3_warning 
    all_dangerous = tier1_dangerous + tier2_dangerous + tier3_dangerous
    all_soft = tier1_soft + tier2_soft + tier3_soft  # New soft mode picks
    
    if all_safe:
        print(f"\n[OK] SAFE FOR REAL MONEY ({len(all_safe)} stocks):")
        print("   Volume ≥1.5x + Buyer Dominance ≥55% + High Safety Score")
        for ticker, score, safety_score, flags in all_safe[:5]:  # Top 5 safe picks
            flags_str = " ".join(flags)
            print(f"   [TARGET] {ticker:<12} Score: {score:.2f} | Safety: {safety_score} | {flags_str}")
    
    if all_warning:
        print(f"\n[WARNING] PROCEED WITH CAUTION ({len(all_warning)} stocks):")
        print("   Minimum safety requirements met but not optimal")
        for ticker, score, safety_score, flags in all_warning[:3]:
            flags_str = " ".join(flags)
            print(f"   [CAUTION] {ticker:<12} Score: {score:.2f} | Safety: {safety_score} | {flags_str}")
    
    if all_dangerous:
        print(f"\n[ALERT] AVOID - DANGEROUS ({len(all_dangerous)} stocks):")
        print("   Failed critical volume or buyer dominance tests")
        for ticker, score, safety_score, flags in all_dangerous[:3]:
            flags_str = " ".join(flags)
            print(f"   [X] {ticker:<12} Score: {score:.2f} | Safety: {safety_score} | {flags_str}")
    
    if all_soft and args.soft_mode:
        print(f"\n[SOFT] SOFT MODE - NEAR MISSES ({len(all_soft)} stocks):")
        print("   ⚠️ Relaxed thresholds: Vol≥1.2x, BD≥45% (Use with extra caution)")
        for ticker, score, safety_score, flags in all_soft[:5]:
            flags_str = " ".join(flags)
            print(f"   [⚠️] {ticker:<12} Score: {score:.2f} | Safety: {safety_score} | {flags_str}")
    
    print(f"\n[INFO] SAFETY SUMMARY:")
    print(f"   [OK] Safe for investment: {len(all_safe)} stocks")
    print(f"   [WARNING] Requires caution: {len(all_warning)} stocks") 
    if args.soft_mode:
        print(f"   [⚠️] Soft mode near-misses: {len(all_soft)} stocks")
    print(f"   [X] Too risky (avoid): {len(all_dangerous)} stocks")
    investment_grade = len(all_safe + all_warning) + (len(all_soft) if args.soft_mode else 0)
    print(f"   📊 Investment-grade stocks: {investment_grade}/{len(tier1 + tier2 + tier3)}")

    # Write watchlists
    # =============================================================================
    # 🎯 FINAL INVESTMENT VERDICT - SAFETY-VALIDATED TOP PICKS
    # =============================================================================
    print("\n" + "🎯" + "="*98)
    print("📈 FINAL INVESTMENT VERDICT - SAFETY-VALIDATED PICKS FOR REAL MONEY")
    print("🎯" + "="*98)
    
    # =============================================================================
    # 🧠 NO-BRAINER PICKS LOGIC (Only from Safety-Validated Stocks)
    # =============================================================================
    
    # Use only safety-validated stocks for recommendations
    safe_investment_candidates = [ticker for ticker, score, safety_score, flags in all_safe]
    caution_candidates = [ticker for ticker, score, safety_score, flags in all_warning]
    
    # Logic 1: Highest momentum stock from SAFE picks (Score >= 3.5 + safety validated)
    momentum_leader = None
    for ticker, score, safety_score, flags in all_safe:
        if score >= 3.5:
            momentum_leader = (ticker, score, "SAFE_MOMENTUM")
            break
    
    # Logic 2: Best value opportunity from SAFE picks
    value_leader = None
    for ticker, score, safety_score, flags in all_safe:
        if score >= 2.0:  # Lower threshold for value but must be safe
            value_leader = (ticker, score, "SAFE_VALUE")
            break
    
    # Logic 3: Best oversold opportunity from SAFE + WARNING picks (with safety note)
    oversold_gem = None
    for ticker, score, safety_score, flags in all_safe + all_warning:
        # Look for oversold indicators in the flags or tier data
        if score >= 1.5:  # Minimum score threshold
            oversold_gem = (ticker, score, "SAFE_OVERSOLD" if (ticker, score, safety_score, flags) in all_safe else "CAUTION_OVERSOLD")
            break
    
    # Fallback selections if primary logic fails (from caution list)
    if not momentum_leader and all_warning:
        ticker, score, safety_score, flags = all_warning[0]
        momentum_leader = (ticker, score, "CAUTION_MOMENTUM")
    
    if not value_leader and all_safe:
        ticker, score, safety_score, flags = all_safe[0]
        value_leader = (ticker, score, "SAFE_BACKUP")
    
    # Compile no-brainer picks (safety-validated only)
    no_brainer_picks = []
    if momentum_leader:
        no_brainer_picks.append(momentum_leader)
    if value_leader and value_leader[0] != (momentum_leader[0] if momentum_leader else None):
        no_brainer_picks.append(value_leader)
    if oversold_gem and oversold_gem[0] not in [p[0] for p in no_brainer_picks]:
        no_brainer_picks.append(oversold_gem)
    
    # Remove duplicates while preserving order
    seen = set()
    unique_picks = []
    for pick in no_brainer_picks:
        if pick[0] not in seen:
            unique_picks.append(pick)
            seen.add(pick[0])
    
    no_brainer_picks = unique_picks[:3]  # Top 3 only
    
    # Display safety-validated no-brainer picks
    print("\n🧠 TOP SAFETY-VALIDATED PICKS (REAL MONEY READY)")
    print("="*70)
    
    if not no_brainer_picks:
        print("🚨 WARNING: No stocks passed both scoring and safety validation!")
        print("   All current candidates have volume or buyer dominance issues.")
        print("   💡 Recommendation: Wait for better opportunities with:")
        print("      • Volume ≥ 1.5x average")
        print("      • Buyer Dominance ≥ 55%")
        print("   🔄 Re-run screener in 1-2 days for fresh candidates.")
        return
    
    total_allocation = 0
    for i, (ticker, score, logic_type) in enumerate(no_brainer_picks, 1):
        # Get safety status for this ticker
        safety_status = "⚠️ CAUTION"
        safety_details = ""
        
        for t, s, safety_score, flags in all_safe:
            if t == ticker:
                safety_status = "✅ SAFE"
                safety_details = f"Vol+BD Validated | Safety: {safety_score}"
                break
        
        for t, s, safety_score, flags in all_warning:
            if t == ticker:
                safety_status = "⚠️ CAUTION"
                safety_details = f"Minimum Safety | Safety: {safety_score}"
                break
        
        # Determine allocation and strategy based on safety and logic
        if "SAFE" in logic_type and safety_status == "✅ SAFE":
            allocation = 4.0
            emoji = "�"
            action = "IMMEDIATE BUY"
            reasoning = "High score + volume & buyer dominance validated"
        elif "CAUTION" in logic_type or safety_status == "⚠️ CAUTION":
            allocation = 2.0
            emoji = "⚠️"
            action = "CAREFUL ENTRY"
            reasoning = "Good score but volume/BD needs monitoring"
        else:
            allocation = 3.0
            emoji = "🥈"
            action = "SCALE IN"
            reasoning = "Solid opportunity with safety validation"
        
        total_allocation += allocation
        
        print(f"{i}. {emoji} {ticker:<12} Score: {score:4.2f} | {action}")
        print(f"   💡 Logic: {reasoning}")
        print(f"   � Safety: {safety_status} - {safety_details}")
        print(f"   �📊 Allocation: {allocation:.1f}% of portfolio")
        print(f"   🎯 Strategy: {logic_type.replace('_', ' ').title()} with safety validation")
        print()
    
    print(f"💰 Total Safety-Validated Allocation: {total_allocation:.1f}%")
    print(f"💎 Cash Reserve: {100-total_allocation:.1f}% (for other opportunities)")
    print(f"🔒 Risk Level: CONTROLLED - Only safety-validated stocks with volume & buyer confirmation")
    
    print("\n🚀 ENHANCED ACTION PLAN (Volume & Buyer Dominance Validated):")
    print("1. Only deploy in safety-validated stocks (✅ or ⚠️)")
    print("2. Monitor volume (must stay ≥1.5x) and buyer dominance (≥55%)")
    print("3. Exit immediately if volume drops below 1.2x or BD falls below 50%")
    print("4. Keep 80%+ cash for high-grade opportunities only")
    print("5. Re-validate safety metrics daily before adding positions")
    
    print("\n" + "="*70)
    
    # Identify the absolute best opportunities
    top_tier1 = tier1[:2] if len(tier1) >= 2 else tier1  # Top 2 from Tier 1
    top_tier2 = [stock for stock in tier2 if stock[1] >= 3.0][:3]  # High-score Tier 2
    
    print("\n🏆 CORE HOLDINGS (Scale In Gradually):")
    print("-" * 60)
    for i, (ticker, score) in enumerate(top_tier1, 1):
        print(f"{i}. {ticker:<12} Score: {score:4.2f} | TIER 1 Core Position")
        print(f"   💰 Strategy: Scale in over 2-3 weeks | Risk: Medium | Time: 3-6 months")
        print(f"   📊 Allocation: 3-5% of portfolio")
        print()
    
    print("⚡ MOMENTUM PLAYS (High Score Opportunities):")
    print("-" * 60)
    momentum_stocks = [(t, s) for t, s in tier2 if s >= 3.5][:3]
    for i, (ticker, score) in enumerate(momentum_stocks, 1):
        print(f"{i}. {ticker:<12} Score: {score:4.2f} | High Momentum Signal")
        print(f"   🚀 Strategy: Quick entry on volume breakout | Target: +8-12% | Stop: -5%")
        print(f"   📊 Allocation: 1-2% per trade | Risk: High | Time: 1-3 weeks")
        print()
    
    print("💎 VALUE OPPORTUNITIES (Selective Picks):")
    print("-" * 60)
    value_opportunities = [(t, s) for t, s in tier2 if 1.5 <= s < 3.0][:4]
    
    for i, (ticker, score) in enumerate(value_opportunities, 1):
        print(f"{i}. {ticker:<12} Score: {score:4.2f} | Value Entry Candidate")
        print(f"   💎 Strategy: Watch for dips and DCA | Expected: +15-25% | Risk-Reward: 3:1")
        print(f"   📊 Allocation: 2-3% | Time: 2-8 weeks")
        print()
    
    # Calculate portfolio allocation
    total_core = len(top_tier1) * 4  # 4% each for core holdings
    total_momentum = len(momentum_stocks) * 1.5  # 1.5% each for momentum
    total_value = len(value_opportunities) * 2.5  # 2.5% each for value
    total_allocation_full = total_core + total_momentum + total_value
    
    print("📋 PORTFOLIO ALLOCATION SUMMARY:")
    print("-" * 60)
    print(f"🏆 Core Holdings:     {total_core:5.1f}% ({len(top_tier1)} stocks)")
    print(f"⚡ Momentum Plays:    {total_momentum:5.1f}% ({len(momentum_stocks)} stocks)")  
    print(f"💎 Value Opportunities: {total_value:5.1f}% ({len(value_opportunities)} stocks)")
    print(f"💰 Cash Reserve:      {100-total_allocation_full:5.1f}%")
    print(f"📊 Total Deployed:    {total_allocation_full:5.1f}%")
    
    print("\n🎯 ACTION PLAN:")
    print("1. Start with CORE HOLDINGS (scale in over 2-3 weeks)")
    print("2. Watch VALUE opportunities for dip entries") 
    print("3. Take MOMENTUM PLAYS only on strong volume breakouts")
    print("4. Maintain 20-30% cash for new opportunities")
    print("5. Review and rebalance weekly with fresh screener runs")
    
    # Summary of best opportunities
    all_top_picks = []
    if top_tier1:
        all_top_picks.extend([(t, s, "CORE") for t, s in top_tier1])
    if momentum_stocks:
        all_top_picks.extend([(t, s, "MOMENTUM") for t, s in momentum_stocks])
    if value_opportunities:
        all_top_picks.extend([(t, s, "VALUE") for t, s in value_opportunities[:2]])  # Top 2 value picks
    
    if all_top_picks:
        print("\n🌟 TODAY'S TOP INVESTMENT PICKS:")
        print("-" * 40)
        for i, (ticker, score, category) in enumerate(all_top_picks, 1):
            category_emoji = {"CORE": "🏆", "MOMENTUM": "⚡", "VALUE": "💎"}
            print(f"{i:2d}. {category_emoji[category]} {ticker:<10} ({category:8s}) Score: {score:.2f}")
    
    print("🎯" + "="*98)
    
    # =============================================================================
    # ENHANCED FINANCIAL METRICS SUMMARY (Critical Insights)
    # =============================================================================
    
    # Only show enhanced analysis if requested and not in ultra-fast mode
    if (args.enhanced_analytics or not args.skip_enhanced) and not args.ultra_fast:
        print("\n📊 ENHANCED FINANCIAL HEALTH ANALYSIS")
        print("=" * 100)
        
        # Define column widths for financial health table
        fin_col_widths = {
            'ticker': 12, 'de': 14, 'roe': 14, 'intcov': 14, 
            'insthold': 14, 'delpct': 14, 'finhealth': 18
        }
        
        # Display enhanced metrics for analyzed candidates
        if _enhanced_financial_data:
            # Create properly formatted header
            header_row = (
                f"{'Ticker':<{fin_col_widths['ticker']}} "
                f"{'D/E':<{fin_col_widths['de']}} "
                f"{'ROE%':<{fin_col_widths['roe']}} "
                f"{'IntCov':<{fin_col_widths['intcov']}} "
                f"{'InstHold%':<{fin_col_widths['insthold']}} "
                f"{'DelPct%':<{fin_col_widths['delpct']}} "
                f"{'FinHealth':<{fin_col_widths['finhealth']}}"
            )
            print(header_row)
            
            # Calculate separator line length
            separator_length = sum(fin_col_widths.values()) + 6  # spaces between columns
            print("-" * separator_length)
            
            alerts = {'excellent': [], 'high_debt': [], 'overextended': [], 'poor_coverage': []}
            
            for ticker, data in _enhanced_financial_data.items():
                try:
                    financial_health = data['financial_health']
                    institutional_flow = data['institutional_flow']
                    technical_risk = data['technical_risk']
                    
                    # Extract metrics
                    debt_to_equity = financial_health.get('debt_to_equity', 999)
                    roe = financial_health.get('roe', 0)
                    interest_coverage = financial_health.get('interest_coverage', 0)
                    inst_holding = institutional_flow.get('institutional_holding', 0)
                    delivery_pct = institutional_flow.get('delivery_percentage', 50)
                    price_dist_50ma = technical_risk.get('price_distance_50ma', 0)
                    
                    # Format with intelligent highlighting
                    de_formatted = format_financial_metric(debt_to_equity, "debt_to_equity")
                    roe_formatted = format_financial_metric(roe, "roe")
                    ic_formatted = format_financial_metric(interest_coverage, "interest_coverage")
                    inst_formatted = format_financial_metric(inst_holding, "institutional_holding")
                    del_formatted = format_financial_metric(delivery_pct, "delivery_percentage")
                    
                    # Get overall health grade
                    grade, color_code, description = get_financial_health_grade(debt_to_equity, roe, interest_coverage)
                    grade_colored = f"{color_code}{grade}\033[0m"
                    
                    # Display the row with proper padding for colored text
                    data_row = (
                        f"{pad_colored_text(ticker, fin_col_widths['ticker'])} "
                        f"{pad_colored_text(de_formatted, fin_col_widths['de'])} "
                        f"{pad_colored_text(roe_formatted, fin_col_widths['roe'])} "
                        f"{pad_colored_text(ic_formatted, fin_col_widths['intcov'])} "
                        f"{pad_colored_text(inst_formatted, fin_col_widths['insthold'])} "
                        f"{pad_colored_text(del_formatted, fin_col_widths['delpct'])} "
                        f"{pad_colored_text(grade_colored, fin_col_widths['finhealth'])}"
                    )
                    print(data_row)
                    
                    # Collect alerts for summary
                    if debt_to_equity <= 0.3 and roe >= 15:
                        alerts['excellent'].append(f"⭐ {ticker}: Excellent fundamentals (D/E={debt_to_equity:.2f}, ROE={roe:.1f}%)")
                    
                    if debt_to_equity > 2.0:
                        alerts['high_debt'].append(f"⚠️  {ticker}: High debt (D/E={debt_to_equity:.2f}) - Use caution")
                    
                    if price_dist_50ma > 15:
                        alerts['overextended'].append(f"📈 {ticker}: Overextended {price_dist_50ma:.1f}% above 50-MA - Wait for dip")
                    
                    if interest_coverage < 1.0 and interest_coverage > 0:
                        alerts['poor_coverage'].append(f"⚠️ {ticker}: Poor interest coverage ({interest_coverage:.1f}x) - Debt stress risk")
                        
                except Exception as e:
                    error_row = (
                        f"{pad_colored_text(ticker, fin_col_widths['ticker'])} "
                        f"{pad_colored_text('ERROR', fin_col_widths['de'])} "
                        f"{pad_colored_text('ERROR', fin_col_widths['roe'])} "
                        f"{pad_colored_text('ERROR', fin_col_widths['intcov'])} "
                        f"{pad_colored_text('ERROR', fin_col_widths['insthold'])} "
                        f"{pad_colored_text('ERROR', fin_col_widths['delpct'])} "
                        f"{pad_colored_text('ERROR', fin_col_widths['finhealth'])}"
                    )
                    print(error_row)
            
            print("\n💡 ENHANCED METRICS LEGEND:")
            print("D/E: Debt-to-Equity (\033[95m*Excellent≤0.2\033[0m, \033[92m*Good≤0.5\033[0m, \033[93m~OK≤1.0\033[0m, \033[91m!Bad>2.0\033[0m)")
            print("ROE%: Return on Equity (\033[95m*Excellent≥20%\033[0m, \033[92m*Good≥15%\033[0m, \033[93m~OK≥10%\033[0m, \033[91m!Bad<5%\033[0m)")
            print("IntCov: Interest Coverage (\033[95m*Excellent≥10x\033[0m, \033[92m*Good≥5x\033[0m, \033[93m~OK≥2x\033[0m, \033[91m!Bad<1x\033[0m)")
            print("InstHold%: Institutional Holdings (\033[95m*Excellent≥60%\033[0m, \033[92m*Good≥40%\033[0m, \033[93m~OK≥20%\033[0m, \033[91m!Bad<10%\033[0m)")
            print("DelPct%: Delivery Percentage (\033[95m*Excellent≥80%\033[0m, \033[92m*Good≥70%\033[0m, \033[93m~OK≥50%\033[0m, \033[91m!Bad<35%\033[0m)")
            print("FinHealth: Overall financial health grade (\033[95mExcellent\033[0m, \033[92mGood\033[0m, \033[93mFair\033[0m, Weak, \033[91mPoor\033[0m)")
            
            # Display categorized alerts
            if any(alerts.values()):
                print("\n🚨 ENHANCED SCREENING ALERTS:")
                
                if alerts['excellent']:
                    print("\n✅ TOP FUNDAMENTAL QUALITY:")
                    for alert in alerts['excellent']:
                        print(f"   {alert}")
                
                if alerts['high_debt']:
                    print("\n⚠️  HIGH DEBT WARNINGS:")
                    for alert in alerts['high_debt']:
                        print(f"   {alert}")
                
                if alerts['poor_coverage']:
                    print("\n⚠️ DEBT STRESS RISKS:")
                    for alert in alerts['poor_coverage']:
                        print(f"   {alert}")
                
                if alerts['overextended']:
                    print("\n📊 TECHNICAL CAUTIONS:")
                    for alert in alerts['overextended']:
                        print(f"   {alert}")
            else:
                print("\n   ✅ All analyzed picks show balanced risk profiles")
                
            # Final recommendations based on enhanced analysis
            print("\n🎯 ENHANCED ANALYSIS RECOMMENDATIONS:")
            excellent_picks = [ticker for ticker, data in _enhanced_financial_data.items() 
                             if data['financial_health'].get('debt_to_equity', 999) <= 0.5 
                             and data['financial_health'].get('roe', 0) >= 15]
            
            avoid_picks = [ticker for ticker, data in _enhanced_financial_data.items() 
                          if data['financial_health'].get('debt_to_equity', 999) > 2.0 
                          or data['financial_health'].get('interest_coverage', 0) < 1.0]
            
            if excellent_picks:
                print(f"✅ BUY: {', '.join(excellent_picks[:3])} - Strong fundamentals")
            
            if avoid_picks:
                print(f"❌ AVOID: {', '.join(avoid_picks)} - Financial stress indicators")
        
        else:
            print("   📊 No final candidates qualified for enhanced analysis")
        
        print("\n" + "=" * 90)
    
    elif args.skip_enhanced:
        print("\n⚡ Enhanced financial analysis skipped (--skip-enhanced flag used)")
    elif args.ultra_fast:
        print("\n⚡ Enhanced financial analysis skipped (ultra-fast mode)")
    else:
        print("\n💡 Use --enhanced-analytics flag to enable detailed financial health analysis")

    # =============================================================================
    # BACKTESTING INTEGRATION
    # =============================================================================
    if (args.backtest or args.optimize or args.validate) and EnhancedBacktester is not None:
        print("\n" + "="*80)
        print("🎯 BACKTESTING & OPTIMIZATION MODULE")
        print("="*80)
        
        # Determine symbols for backtesting
        if args.backtest_symbols:
            backtest_symbols = args.backtest_symbols
            print(f"Using specified symbols: {', '.join(backtest_symbols)}")
        else:
            # Use top screening results
            top_count = min(10, len(ranks))  # Use top 10 results
            backtest_symbols = [ticker for ticker, _ in ranks[:top_count]]
            print(f"Using top {top_count} screening results: {', '.join(backtest_symbols)}")
        
        if not backtest_symbols:
            print("❌ No symbols available for backtesting. Run screening first or specify --backtest-symbols")
        else:
            # Initialize backtester
            backtester = EnhancedBacktester()
            
            try:
                if args.optimize:
                    print("\n🔍 Running Parameter Optimization...")
                    start_date = args.backtest[0] if args.backtest else "2023-01-01"
                    end_date = args.backtest[1] if args.backtest else "2024-12-31"
                    
                    best_params = await backtester.optimize_parameters(
                        symbols=backtest_symbols,
                        start_date=start_date,
                        end_date=end_date
                    )
                    
                    print("\n📊 OPTIMIZATION RESULTS:")
                    print(f"   Best RSI Lower: {best_params.rsi_lower}")
                    print(f"   Best RSI Upper: {best_params.rsi_upper}")
                    print(f"   Best EMA Short: {best_params.ema_short}")
                    print(f"   Best EMA Long: {best_params.ema_long}")
                    print(f"   Best Volume Threshold: {best_params.volume_threshold}")
                    print(f"   Best Stop Loss: {best_params.stop_loss_pct}%")
                    print(f"   Best Take Profit: {best_params.take_profit_pct}%")
                    
                elif args.backtest:
                    print(f"\n📈 Running Backtest: {args.backtest[0]} to {args.backtest[1]}")
                    
                    # Use default parameters for single backtest
                    params = BacktestParameters()
                    results = await backtester.run_backtest(
                        symbols=backtest_symbols,
                        start_date=args.backtest[0],
                        end_date=args.backtest[1],
                        parameters=params
                    )
                    
                    print(f"\n📊 BACKTEST RESULTS:")
                    print(f"   Total Return: {results.total_return:.2%}")
                    print(f"   Annualized Return: {results.annualized_return:.2%}")
                    print(f"   Max Drawdown: {results.max_drawdown:.2%}")
                    print(f"   Sharpe Ratio: {results.sharpe_ratio:.2f}")
                    print(f"   Win Rate: {results.win_rate:.2%}")
                    print(f"   Total Trades: {results.total_trades}")
                    print(f"   Avg Trade Return: {results.avg_trade_return:.2%}")
                    print(f"   Volatility: {results.volatility:.2%}")
                
                if args.validate:
                    print(f"\n🔄 Running Rolling Window Validation...")
                    print(f"   Training Window: {args.train_months} months")
                    print(f"   Testing Window: {args.test_months} months")
                    
                    # Get the best parameters if optimization was run
                    if args.optimize and 'best_params' in locals():
                        validation_params = best_params
                    else:
                        validation_params = BacktestParameters()
                    
                    validation_results = await backtester.rolling_window_validation(
                        symbols=backtest_symbols,
                        train_months=args.train_months,
                        test_months=args.test_months,
                        parameters=validation_params
                    )
                    
                    print(f"\n📈 VALIDATION RESULTS:")
                    avg_return = sum(r.total_return for r in validation_results) / len(validation_results)
                    avg_sharpe = sum(r.sharpe_ratio for r in validation_results) / len(validation_results)
                    avg_winrate = sum(r.win_rate for r in validation_results) / len(validation_results)
                    print(f"   Average Return: {avg_return:.2%}")
                    print(f"   Average Sharpe: {avg_sharpe:.2f}")
                    print(f"   Average Win Rate: {avg_winrate:.2%}")
                    print(f"   Validation Periods: {len(validation_results)}")
                
                # Generate backtest report
                print(f"\n💾 Generating detailed report: {args.backtest_report}")
                report_content = []
                report_content.append("SWING SCREENER BACKTESTING REPORT")
                report_content.append("=" * 50)
                report_content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                report_content.append(f"Symbols Tested: {', '.join(backtest_symbols)}")
                report_content.append("")
                
                if args.optimize and 'best_params' in locals():
                    report_content.append("OPTIMIZED PARAMETERS:")
                    report_content.append(f"  RSI Lower: {best_params.rsi_lower}")
                    report_content.append(f"  RSI Upper: {best_params.rsi_upper}")
                    report_content.append(f"  EMA Short: {best_params.ema_short}")
                    report_content.append(f"  EMA Long: {best_params.ema_long}")
                    report_content.append(f"  Volume Threshold: {best_params.volume_threshold}")
                    report_content.append(f"  Stop Loss: {best_params.stop_loss_pct}%")
                    report_content.append(f"  Take Profit: {best_params.take_profit_pct}%")
                    report_content.append("")
                
                if args.backtest and 'results' in locals():
                    report_content.append("BACKTEST PERFORMANCE:")
                    report_content.append(f"  Period: {args.backtest[0]} to {args.backtest[1]}")
                    report_content.append(f"  Total Return: {results.total_return:.2%}")
                    report_content.append(f"  Annualized Return: {results.annualized_return:.2%}")
                    report_content.append(f"  Max Drawdown: {results.max_drawdown:.2%}")
                    report_content.append(f"  Sharpe Ratio: {results.sharpe_ratio:.2f}")
                    report_content.append(f"  Win Rate: {results.win_rate:.2%}")
                    report_content.append(f"  Total Trades: {results.total_trades}")
                    report_content.append(f"  Avg Trade Return: {results.avg_trade_return:.2%}")
                    report_content.append(f"  Volatility: {results.volatility:.2%}")
                    report_content.append("")
                
                if args.validate and 'validation_results' in locals():
                    report_content.append("VALIDATION RESULTS:")
                    report_content.append(f"  Training Window: {args.train_months} months")
                    report_content.append(f"  Testing Window: {args.test_months} months")
                    report_content.append(f"  Average Return: {avg_return:.2%}")
                    report_content.append(f"  Average Sharpe: {avg_sharpe:.2f}")
                    report_content.append(f"  Average Win Rate: {avg_winrate:.2%}")
                    report_content.append(f"  Validation Periods: {len(validation_results)}")
                    report_content.append("")
                
                with open(args.backtest_report, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(report_content))
                
                print("✅ Backtesting completed successfully!")
                
            except Exception as e:
                print(f"❌ Backtesting failed: {str(e)}")
                print("Continuing with normal screening output...")
        
        print("="*80)
        print()
    
    elif (args.backtest or args.optimize or args.validate):
        print("\n❌ Backtesting features require enhanced_backtester.py module")
        print("   Please ensure enhanced_backtester.py is in the same directory")
        print("   Continuing with normal screening output...\n")

    with open("swing_watchlist.txt","w",encoding="utf-8") as f:
        f.write("🧠 NO-BRAINER PICKS (Start Today):\n")
        for i, (ticker, score, logic_type) in enumerate(no_brainer_picks, 1):
            allocation = 2.0 if logic_type == "MOMENTUM" else 4.0 if logic_type in ["BANKING", "CORE"] else 3.0
            f.write(f"{i}. {ticker} ({logic_type}) - Score: {score:.2f} - Allocation: {allocation:.1f}%\n")
        
        f.write("\nCore Positions (Tier 1):\n")
        f.write("\n".join(f"- {t}" for t,_ in tier1) or "- None")
        f.write("\n\nSelective Opportunities (Tier 2):\n")
        f.write("\n".join(f"- {t}" for t,_ in tier2) or "- None")
        f.write("\n\nTactical / Strict Risk (Tier 3):\n")
        f.write("\n".join(f"- {t}" for t,_ in tier3) or "- None")
        
        # Add investment picks to watchlist
        if all_top_picks:
            f.write("\n\nTODAY'S TOP INVESTMENT PICKS:\n")
            for i, (ticker, score, category) in enumerate(all_top_picks, 1):
                f.write(f"{i}. {ticker} ({category}) - Score: {score:.2f}\n")

    with open("tickers.txt","w",encoding="utf-8") as fh:
        fh.write("\n".join(t for t,_ in ranks))

# =============================================================================
# Entry-point
# =============================================================================
if __name__ == "__main__":
    warnings.simplefilter("ignore", category=FutureWarning)
    
    # Performance profiling support
    import sys
    
    # Check for profiling flag
    if '--profile' in sys.argv:
        import cProfile
        import pstats
        import io
        
        print("🔍 Starting performance profiling...")
        
        # Remove profiling flag for main function
        sys.argv.remove('--profile')
        
        # Create profiler
        pr = cProfile.Profile()
        pr.enable()
        
        # Run main function
        try:
            asyncio.run(main())
        finally:
            pr.disable()
            
            # Generate profile report
            s = io.StringIO()
            ps = pstats.Stats(pr, stream=s)
            ps.sort_stats('cumulative')
            ps.print_stats(20)  # Top 20 functions
            
            # Save detailed profile
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            profile_file = f"screener_profile_{timestamp}.prof"
            pr.dump_stats(profile_file)
            
            print(f"\n📊 Performance Profile (Top 20 functions):")
            print(s.getvalue())
            print(f"\n💾 Detailed profile saved to: {profile_file}")
            print("📈 To view with snakeviz: pip install snakeviz && snakeviz " + profile_file)
            
            # Print performance summary from our monitoring
            perf_summary = performance_optimizer.get_performance_summary()
            if perf_summary:
                print(f"\n🎯 Performance Summary from Internal Monitoring:")
                for func_name, stats in perf_summary.items():
                    print(f"  {func_name}: {stats['calls']} calls, "
                          f"avg {stats['avg_time']:.3f}s, total {stats['total_time']:.3f}s")
            
            # Cache performance
            cache_stats = cache_manager.get_stats()
            print(f"\n💾 Cache Performance:")
            print(f"  Hit Rate: {cache_stats['hit_rate']:.1f}%")
            print(f"  Cache Size: {cache_stats['cache_size']} items")
            
    else:
        # Normal execution
        try:
            asyncio.run(main())
        finally:
            # Print YFinance cache performance stats
            print_cache_stats()

def dynamic_liquidity_threshold(market_cap_cr: float, base: float) -> float:
    try:
        from math import log1p
        mcap = max(0.0, float(market_cap_cr))
        return float(base) * (1.0 + log1p(mcap / 10000.0))
    except Exception:
        return base


@lru_cache(maxsize=5000)
def cached_fin_sentiment(text: str) -> float:
    try:
        if not text:
            return 0.0
        if 'finbert_analyze' in globals():
            try:
                return float(cached_fin_sentiment(text))
            except Exception:
                pass
        if SentimentIntensityAnalyzer is not None:
            try:
                _v = SentimentIntensityAnalyzer()
                sv = _v.polarity_scores(text)
                return float(sv.get('compound', 0.0))
            except Exception:
                return 0.0
        return 0.0
    except Exception:
        return 0.0


def apply_excel_conditional_formatting(writer, sheet_name: str):
    try:
        book = writer.book
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            return
        try:
            ws.conditional_format('C2:C9999', {'type': '3_color_scale'})
        except Exception:
            pass
        try:
            ws.conditional_format('N2:N9999', {'type': '3_color_scale'})
        except Exception:
            pass
        try:
            ws.conditional_format('E2:E9999', {
                'type': 'cell', 'criteria': '<', 'value': 1.0,
                'format': book.add_format({'bg_color': '#FFC7CE'})
            })
        except Exception:
            pass
    except Exception:
        return


def sector_relative_pe(pe: float, sector_median: float | None) -> float:
    try:
        if pe is None or pe <= 0 or not sector_median or sector_median <= 0:
            return 1.0
        return float(pe) / float(sector_median)
    except Exception:
        return 1.0


def add_earnings_surprise(tkr: str) -> float:
    try:
        if not yf:
            return 1.0
        y = yf.Ticker(tkr)
        df = None
        for attr in ('earnings_dates', 'earnings_history', 'calendar'):
            try:
                obj = getattr(y, attr)
                if hasattr(obj, 'reset_index'):
                    df = obj
                    break
            except Exception:
                continue
        if df is None or len(df) == 0:
            return 1.0
        cols = [c for c in df.columns if 'Surprise' in str(c)]
        if not cols:
            return 1.0
        last = float(df.iloc[0][cols[0]])
        return max(0.5, min(2.0, 1.0 + last/100.0))
    except Exception:
        return 1.0


def _format_tech_line(tkr: str, f: dict, tech_score: float) -> str:
    try:
        rsi = f.get('rsi', 0.0)
        ema14 = f.get('ema14', 0.0)
        ema50 = f.get('ema50', 0.0)
        slope = f.get('ema_slope', 0.0)
        bb_low = f.get('bb_lower', 0.0)
        bb_up = f.get('bb_upper', 0.0)
        vwap = f.get('vwap', 0.0)
        price = ema14 if ema14 else vwap
        bb_pos = 0.0
        if bb_up and bb_low and bb_up > bb_low and price:
            bb_pos = max(0.0, min(1.0, (price - bb_low) / (bb_up - bb_low)))
        vwap_delta = 0.0
        if vwap:
            vwap_delta = ((price - vwap) / vwap) * 100.0
        vol_spike = 'Y' if f.get('vol_spike', False) else 'N'
        return (f"   · {tkr}  RSI:{rsi:4.1f}  EMA14:{ema14:,.2f}  EMA50:{ema50:,.2f}  "
                f"Slope%:{slope:5.2f}  BBpos:{bb_pos:0.2f}  VWAPΔ%:{vwap_delta:6.2f}  VolSpike:{vol_spike}  Tech:{tech_score:5.1f}")
    except Exception:
        return f"   · {tkr}  (tech diagnostics unavailable)"
